"""
Attendance blueprint — comprehensive test suite for check-in/out flows.

Covers (all gaps NOT already in test_leave_attendance.py):
  /location, /attendance (kiosk), /api/attendance/checkin,
  /api/employee/checkin, /api/employee/qr-face-checkin,
  /api/employee/attendance, shift swap lifecycle,
  /monthly_report_export, /send_absentee_report, /my_attendance_pdf,
  helper functions, correct_attendance edge cases,
  bulk_assign_shift edge cases, assign_shift, edit_shift edge cases,
  delete_shift_form, employee_attendance_detail edge cases,
  update_default_shift NameError bug.

NOTE: blueprints/attendance.py is a migration-in-progress stub.  It references
globals that are defined in app.py but not imported (GRACE_MINUTES, OFFICE_LAT,
OFFICE_LON, OFFICE_RADIUS_M, _face_recognition_available, etc.).  The
_inject_att_globals fixture below injects them from app.py into the blueprint
module before each test so routes run without NameErrors.
"""
import base64
import datetime
import io
import pytest


# ── Module-level fixture: inject missing globals from app.py → blueprint ─────

@pytest.fixture(autouse=True)
def _inject_att_globals():
    """Inject app.py globals referenced by blueprints/attendance.py but never
    imported there.  Without this, most routes raise NameError at runtime."""
    import app as _app
    import blueprints.attendance as _att
    import time as _time_mod
    import io as _io_mod
    try:
        import openpyxl as _openpyxl
        from openpyxl.styles import (
            PatternFill as _PatternFill, Font as _Font,
            Alignment as _Alignment, Side as _Side, Border as _Border,
        )
    except ImportError:
        _openpyxl = _PatternFill = _Font = _Alignment = _Side = _Border = None

    injected = {
        "GRACE_MINUTES":              getattr(_app, "GRACE_MINUTES", 15),
        "OFFICE_LAT":                 getattr(_app, "OFFICE_LAT", 17.494664737165042),
        "OFFICE_LON":                 getattr(_app, "OFFICE_LON", 78.40496618113566),
        "OFFICE_RADIUS_M":            getattr(_app, "OFFICE_RADIUS_M", 300),
        "_face_recognition_available": getattr(_app, "_face_recognition_available", False),
        "_face_enc_cache":            getattr(_app, "_face_enc_cache", {}),
        "face_recognition":           getattr(_app, "face_recognition", None),
        "UPLOAD_FOLDER":              getattr(_app, "UPLOAD_FOLDER", "/tmp"),
        "_WA_FP_VERIFY_WINDOW_SEC":   getattr(_app, "_WA_FP_VERIFY_WINDOW_SEC", 120),
        "_MOBILE_BIO_VERIFY_WINDOW_SEC": getattr(_app, "_MOBILE_BIO_VERIFY_WINDOW_SEC", 120),
        "load_default_shift":         getattr(_app, "load_default_shift", lambda: None),
        "_safe_redirect":             getattr(_app, "_safe_redirect", lambda d, f="/admin": d),
        "time":                       _time_mod,
        "_io":                        _io_mod,
        "openpyxl":                   _openpyxl,
        "PatternFill":                _PatternFill,
        "Font":                       _Font,
        "Alignment":                  _Alignment,
        "Side":                       _Side,
        "Border":                     _Border,
    }
    for name, val in injected.items():
        setattr(_att, name, val)
    yield
    for name in injected:
        try:
            delattr(_att, name)
        except AttributeError:
            pass


# ── Session / token helpers ──────────────────────────────────────────────────

def _admin_session(client, seed_admin):
    client.post("/admin_login", data={
        "identifier": seed_admin["username"],
        "password":   seed_admin["password"],
    })
    return client


def _emp_session(client, seed_employee):
    with client.session_transaction() as sess:
        sess["employee_id"]   = seed_employee["employee_id"]
        sess["employee_name"] = seed_employee["name"]
    return client


def _admin_token(client, seed_admin):
    return client.post("/api/login", json={
        "username": seed_admin["username"],
        "password": seed_admin["password"],
    }).get_json()["token"]


def _emp_token(client, seed_employee):
    return client.post("/api/employee/login", json={
        "employee_id": seed_employee["employee_id"],
        "password":    seed_employee["password"],
    }).get_json()["token"]


# ── Fixtures ─────────────────────────────────────────────────────────────────

@pytest.fixture
def shift_a(db_engine):
    cur = db_engine.cursor()
    cur.execute(
        "INSERT INTO shifts (name, start_time, half_time, end_time) "
        "VALUES ('CI Morning Shift', '09:00', '13:00', '18:00') RETURNING id"
    )
    sid = cur.fetchone()[0]
    yield {"id": sid, "name": "CI Morning Shift"}
    cur.execute("UPDATE employees SET shift_id=NULL WHERE shift_id=%s", (sid,))
    cur.execute("UPDATE break_config SET shift_id=NULL WHERE shift_id=%s", (sid,))
    cur.execute("DELETE FROM shifts WHERE id=%s", (sid,))
    cur.close()


@pytest.fixture
def shift_b(db_engine):
    cur = db_engine.cursor()
    cur.execute(
        "INSERT INTO shifts (name, start_time, half_time, end_time) "
        "VALUES ('CI Evening Shift', '14:00', '17:00', '22:00') RETURNING id"
    )
    sid = cur.fetchone()[0]
    yield {"id": sid, "name": "CI Evening Shift"}
    cur.execute("UPDATE employees SET shift_id=NULL WHERE shift_id=%s", (sid,))
    cur.execute("UPDATE break_config SET shift_id=NULL WHERE shift_id=%s", (sid,))
    cur.execute("DELETE FROM shifts WHERE id=%s", (sid,))
    cur.close()


@pytest.fixture
def seed_employee2(db_engine):
    """Second test employee for shift-swap tests."""
    from utils.auth import generate_password_hash
    cur = db_engine.cursor()
    cur.execute(
        "INSERT INTO employees (employee_id, name, email, password, force_pin_change) "
        "VALUES (%s,%s,%s,%s,0) ON CONFLICT (employee_id) DO NOTHING",
        ("TST002", "Test Employee 2", "emp2@test.local", generate_password_hash("EmpPass@2")),
    )
    yield {"employee_id": "TST002", "password": "EmpPass@2", "name": "Test Employee 2"}
    cur.execute("DELETE FROM shift_swap_requests WHERE requester_id='TST002' OR target_id='TST002'")
    cur.execute("DELETE FROM attendance WHERE employee_id='TST002'")
    cur.execute("DELETE FROM api_tokens WHERE identity='TST002'")
    cur.execute("DELETE FROM employees WHERE employee_id='TST002'")
    cur.close()


@pytest.fixture
def attendance_today(db_engine, seed_employee):
    """Seed a login-only record for today — next checkin becomes a logout."""
    cur = db_engine.cursor()
    today = datetime.date.today()
    cur.execute("DELETE FROM attendance WHERE employee_id=%s AND date=%s",
                (seed_employee["employee_id"], today))
    cur.execute(
        "INSERT INTO attendance (employee_id, date, login_time, status) "
        "VALUES (%s, %s, '09:00:00', 'Full Day Login')",
        (seed_employee["employee_id"], today),
    )
    yield {"employee_id": seed_employee["employee_id"], "date": today}
    cur.execute("DELETE FROM attendance WHERE employee_id=%s AND date=%s",
                (seed_employee["employee_id"], today))
    cur.close()


@pytest.fixture
def attendance_completed(db_engine, seed_employee):
    """Seed a completed (login + logout) record — next checkin becomes a relogin."""
    cur = db_engine.cursor()
    today = datetime.date.today()
    cur.execute("DELETE FROM attendance WHERE employee_id=%s AND date=%s",
                (seed_employee["employee_id"], today))
    cur.execute(
        "INSERT INTO attendance (employee_id, date, login_time, logout_time, "
        "status, logout_status, attendance_type, worked_minutes) "
        "VALUES (%s, %s, '09:00:00', '18:00:00', "
        "'Full Day Login', 'Completed', 'Full Day', 540)",
        (seed_employee["employee_id"], today),
    )
    yield {"employee_id": seed_employee["employee_id"], "date": today}
    cur.execute("DELETE FROM attendance WHERE employee_id=%s AND date=%s",
                (seed_employee["employee_id"], today))
    cur.close()


@pytest.fixture
def swap_pending(db_engine, seed_employee, seed_employee2, shift_a, shift_b):
    """Seed a Pending_Target swap request from TST001 → TST002."""
    cur = db_engine.cursor()
    cur.execute("UPDATE employees SET shift_id=%s WHERE employee_id='TST001'", (shift_a["id"],))
    cur.execute("UPDATE employees SET shift_id=%s WHERE employee_id='TST002'", (shift_b["id"],))
    cur.execute(
        "INSERT INTO shift_swap_requests "
        "(requester_id, target_id, requester_shift_id, target_shift_id, reason, status) "
        "VALUES ('TST001','TST002',%s,%s,'CI swap reason','Pending_Target') RETURNING id",
        (shift_a["id"], shift_b["id"]),
    )
    req_id = cur.fetchone()[0]
    yield {
        "id": req_id,
        "requester_id": "TST001",
        "target_id":    "TST002",
        "req_shift_id": shift_a["id"],
        "tgt_shift_id": shift_b["id"],
    }
    cur.execute("DELETE FROM shift_swap_requests WHERE id=%s", (req_id,))
    cur.execute("UPDATE employees SET shift_id=NULL WHERE employee_id IN ('TST001','TST002')")
    cur.close()


@pytest.fixture
def swap_pending_admin(db_engine, seed_employee, seed_employee2, shift_a, shift_b):
    """Seed a Pending_Admin swap request (target already accepted)."""
    cur = db_engine.cursor()
    cur.execute("UPDATE employees SET shift_id=%s WHERE employee_id='TST001'", (shift_a["id"],))
    cur.execute("UPDATE employees SET shift_id=%s WHERE employee_id='TST002'", (shift_b["id"],))
    cur.execute(
        "INSERT INTO shift_swap_requests "
        "(requester_id, target_id, requester_shift_id, target_shift_id, reason, status, target_response) "
        "VALUES ('TST001','TST002',%s,%s,'CI swap reason','Pending_Admin','Accepted') RETURNING id",
        (shift_a["id"], shift_b["id"]),
    )
    req_id = cur.fetchone()[0]
    yield {
        "id": req_id,
        "requester_id": "TST001",
        "target_id":    "TST002",
        "req_shift_id": shift_a["id"],
        "tgt_shift_id": shift_b["id"],
    }
    cur.execute("DELETE FROM shift_swap_requests WHERE id=%s", (req_id,))
    cur.execute("UPDATE employees SET shift_id=NULL WHERE employee_id IN ('TST001','TST002')")
    cur.close()


# ═══════════════════════════════════════════════════════════════════════════════
# Helper function unit tests
# ═══════════════════════════════════════════════════════════════════════════════

class TestIsWithinRange:
    """Unit tests for the haversine geo-fence helper."""

    def test_same_point_is_within(self):
        from blueprints.attendance import is_within_range
        assert is_within_range(12.9716, 77.5946, 12.9716, 77.5946) is True

    def test_very_close_distance(self):
        from blueprints.attendance import is_within_range
        # ~1 m apart — should be within any reasonable office radius
        assert is_within_range(12.971600, 77.594600, 12.971609, 77.594600) is True

    def test_different_cities_out_of_range(self):
        from blueprints.attendance import is_within_range
        # Bangalore vs Mumbai (~980 km)
        assert is_within_range(12.9716, 77.5946, 19.0760, 72.8777) is False


class TestFmtT:
    """Unit tests for the _fmt_t time/timedelta formatter."""

    def test_none_returns_none(self):
        from blueprints.attendance import _fmt_t
        assert _fmt_t(None) is None

    def test_time_object(self):
        from blueprints.attendance import _fmt_t
        assert _fmt_t(datetime.time(9, 5, 30)) == "09:05:30"

    def test_timedelta_hours_minutes(self):
        from blueprints.attendance import _fmt_t
        td = datetime.timedelta(hours=14, minutes=30, seconds=0)
        assert _fmt_t(td) == "14:30:00"

    def test_timedelta_seconds_only(self):
        from blueprints.attendance import _fmt_t
        td = datetime.timedelta(seconds=3661)  # 1h 1m 1s
        assert _fmt_t(td) == "01:01:01"

    def test_midnight(self):
        from blueprints.attendance import _fmt_t
        assert _fmt_t(datetime.time(0, 0, 0)) == "00:00:00"


# ═══════════════════════════════════════════════════════════════════════════════
# /location endpoint (no auth, POST JSON)
# ═══════════════════════════════════════════════════════════════════════════════

class TestLocationEndpoint:

    def test_returns_ok(self, client):
        rv = client.post("/location", json={"lat": 12.9716, "lon": 77.5946})
        assert rv.status_code == 200
        assert rv.get_json()["status"] == "ok"

    def test_stores_coordinates_in_session(self, client):
        client.post("/location", json={"lat": 12.9716, "lon": 77.5946})
        with client.session_transaction() as sess:
            assert sess["lat"] == 12.9716
            assert sess["lon"] == 77.5946

    def test_overwrites_previous_coordinates(self, client):
        client.post("/location", json={"lat": 1.0, "lon": 1.0})
        client.post("/location", json={"lat": 12.9716, "lon": 77.5946})
        with client.session_transaction() as sess:
            assert sess["lat"] == 12.9716

    def test_missing_lat_key_raises(self, client):
        rv = client.post("/location", json={"lon": 77.5946})
        # Production code does data["lat"] — KeyError propagates as 500
        assert rv.status_code in (400, 500)


# ═══════════════════════════════════════════════════════════════════════════════
# /attendance (kiosk check-in, no auth, POST JSON)
# ═══════════════════════════════════════════════════════════════════════════════

class TestKioskAttendance:

    def test_invalid_auth_combo(self, client, seed_employee):
        rv = client.post("/attendance", json={
            "employee_id": seed_employee["employee_id"],
            "auth_combo":  "bad_combo",
        })
        data = rv.get_json()
        assert data["ok"] is False
        assert "Invalid auth combination" in data["msg"]

    def test_missing_emp_id_qr_only(self, client):
        rv = client.post("/attendance", json={"auth_combo": "qr_only", "employee_id": ""})
        assert rv.get_json()["ok"] is False

    def test_missing_emp_id_fingerprint_only(self, client):
        rv = client.post("/attendance", json={"auth_combo": "fingerprint_only", "employee_id": ""})
        data = rv.get_json()
        assert data["ok"] is False
        assert "Employee ID is required" in data["msg"]

    def test_unknown_employee(self, client):
        rv = client.post("/attendance", json={
            "employee_id": "GHOST99", "auth_combo": "qr_only",
        })
        data = rv.get_json()
        assert data["ok"] is False
        assert "not found" in data["msg"].lower()

    def test_qr_only_login_creates_attendance_record(self, client, seed_employee, db_engine, mocker):
        mocker.patch("blueprints.attendance.get_auth_config", return_value={
            "fingerprint_enabled": False, "qr_enabled": True,
            "face_enabled": True, "location_enabled": False,
        })
        today = datetime.date.today()
        cur = db_engine.cursor()
        cur.execute("DELETE FROM attendance WHERE employee_id=%s AND date=%s",
                    (seed_employee["employee_id"], today))
        cur.close()

        rv = client.post("/attendance", json={
            "employee_id": seed_employee["employee_id"],
            "auth_combo":  "qr_only",
        })
        data = rv.get_json()
        assert data["ok"] is True, data
        assert data["type"] == "login"
        assert data["status"] in ("Full Day Login", "Late Login", "Half Day Login")

        cur = db_engine.cursor()
        cur.execute("SELECT login_time, logout_time FROM attendance WHERE employee_id=%s AND date=%s",
                    (seed_employee["employee_id"], today))
        row = cur.fetchone()
        cur.execute("DELETE FROM attendance WHERE employee_id=%s AND date=%s",
                    (seed_employee["employee_id"], today))
        cur.close()

        assert row is not None
        assert row[0] is not None   # login_time set
        assert row[1] is None       # logout_time still NULL

    def test_qr_only_logout_sets_logout_time(self, client, seed_employee, attendance_today, db_engine, mocker):
        mocker.patch("blueprints.attendance.get_auth_config", return_value={
            "fingerprint_enabled": False, "qr_enabled": True,
            "face_enabled": True, "location_enabled": False,
        })
        rv = client.post("/attendance", json={
            "employee_id": seed_employee["employee_id"],
            "auth_combo":  "qr_only",
        })
        data = rv.get_json()
        assert data["ok"] is True, data
        assert data["type"] == "logout"
        assert data["status"] in ("Half Day Logout", "Early Logout", "Completed")

        cur = db_engine.cursor()
        cur.execute("SELECT logout_time FROM attendance WHERE employee_id=%s AND date=%s",
                    (seed_employee["employee_id"], attendance_today["date"]))
        assert cur.fetchone()[0] is not None
        cur.close()

    def test_qr_only_relogin_clears_logout_time(self, client, seed_employee, attendance_completed, db_engine, mocker):
        mocker.patch("blueprints.attendance.get_auth_config", return_value={
            "fingerprint_enabled": False, "qr_enabled": True,
            "face_enabled": True, "location_enabled": False,
        })
        rv = client.post("/attendance", json={
            "employee_id": seed_employee["employee_id"],
            "auth_combo":  "qr_only",
        })
        data = rv.get_json()
        assert data["ok"] is True, data
        assert data["type"] == "relogin"

        cur = db_engine.cursor()
        cur.execute("SELECT logout_time FROM attendance WHERE employee_id=%s AND date=%s",
                    (seed_employee["employee_id"], attendance_completed["date"]))
        assert cur.fetchone()[0] is None
        cur.close()

    def test_qr_face_empty_face_image(self, client, seed_employee):
        rv = client.post("/attendance", json={
            "employee_id": seed_employee["employee_id"],
            "auth_combo":  "qr_face",
            "face_image":  "",
        })
        data = rv.get_json()
        assert data["ok"] is False
        assert "Face photo not captured" in data["msg"]

    def test_qr_face_invalid_base64(self, client, seed_employee):
        rv = client.post("/attendance", json={
            "employee_id": seed_employee["employee_id"],
            "auth_combo":  "qr_face",
            "face_image":  "!!!not_valid_base64!!!",
        })
        data = rv.get_json()
        assert data["ok"] is False
        assert "Invalid face image data" in data["msg"]

    def test_qr_face_recognition_unavailable(self, client, seed_employee, mocker):
        from PIL import Image as _PIL
        buf = io.BytesIO()
        _PIL.new("RGB", (100, 100)).save(buf, format="JPEG")
        face_b64 = base64.b64encode(buf.getvalue()).decode()

        mocker.patch("blueprints.attendance.get_auth_config", return_value={
            "fingerprint_enabled": False, "qr_enabled": True,
            "face_enabled": True, "location_enabled": False,
        })
        mocker.patch("blueprints.attendance._face_recognition_available", False)
        rv = client.post("/attendance", json={
            "employee_id": seed_employee["employee_id"],
            "auth_combo":  "qr_face",
            "face_image":  face_b64,
        })
        data = rv.get_json()
        assert data["ok"] is False
        assert "unavailable" in data["msg"].lower()

    def test_fingerprint_not_enabled_returns_403(self, client, seed_employee, mocker):
        mocker.patch("blueprints.attendance.get_auth_config", return_value={
            "fingerprint_enabled": False, "location_enabled": False,
            "face_enabled": True, "qr_enabled": True,
        })
        rv = client.post("/attendance", json={
            "employee_id": seed_employee["employee_id"],
            "auth_combo":  "fingerprint_only",
        })
        assert rv.status_code == 403
        assert rv.get_json()["ok"] is False

    def test_fingerprint_not_verified_returns_401(self, client, seed_employee, mocker):
        mocker.patch("blueprints.attendance.get_auth_config", return_value={
            "fingerprint_enabled": True, "location_enabled": False,
            "face_enabled": True, "qr_enabled": True,
        })
        mocker.patch("blueprints.attendance._wa_fingerprint_recently_verified", return_value=False)
        rv = client.post("/attendance", json={
            "employee_id": seed_employee["employee_id"],
            "auth_combo":  "fingerprint_only",
        })
        assert rv.status_code == 401
        assert rv.get_json()["ok"] is False

    def test_location_required_but_missing(self, client, seed_employee, mocker):
        mocker.patch("blueprints.attendance.get_auth_config", return_value={
            "fingerprint_enabled": False, "location_enabled": True,
            "face_enabled": False, "qr_enabled": True,
        })
        rv = client.post("/attendance", json={
            "employee_id": seed_employee["employee_id"],
            "auth_combo":  "qr_only",
        })
        data = rv.get_json()
        assert data["ok"] is False
        assert "Location not captured" in data["msg"]

    def test_location_outside_office_rejected(self, client, seed_employee, mocker):
        mocker.patch("blueprints.attendance.get_auth_config", return_value={
            "fingerprint_enabled": False, "location_enabled": True,
            "face_enabled": False, "qr_enabled": True,
        })
        mocker.patch("blueprints.attendance.is_within_range", return_value=False)
        rv = client.post("/attendance", json={
            "employee_id": seed_employee["employee_id"],
            "auth_combo":  "qr_only",
            "lat": 19.0760, "lon": 72.8777,
        })
        data = rv.get_json()
        assert data["ok"] is False
        assert "outside" in data["msg"].lower()


# ═══════════════════════════════════════════════════════════════════════════════
# /api/attendance/checkin (admin Bearer token, POST JSON)
# ═══════════════════════════════════════════════════════════════════════════════

class TestApiAdminCheckin:

    def test_requires_admin_token(self, client):
        rv = client.post("/api/attendance/checkin", json={"employee_id": "TST001"})
        assert rv.status_code == 401

    def test_missing_employee_id_returns_400(self, client, seed_admin):
        token = _admin_token(client, seed_admin)
        rv = client.post("/api/attendance/checkin", json={},
                         headers={"Authorization": f"Bearer {token}"})
        assert rv.status_code == 400
        assert rv.get_json()["ok"] is False

    def test_unknown_employee_returns_error(self, client, seed_admin):
        token = _admin_token(client, seed_admin)
        rv = client.post("/api/attendance/checkin", json={"employee_id": "GHOST99"},
                         headers={"Authorization": f"Bearer {token}"})
        data = rv.get_json()
        assert data["ok"] is False
        assert "not found" in data["msg"].lower()

    def test_login_creates_record(self, client, seed_admin, seed_employee, db_engine):
        today = datetime.date.today()
        cur = db_engine.cursor()
        cur.execute("DELETE FROM attendance WHERE employee_id=%s AND date=%s",
                    (seed_employee["employee_id"], today))
        cur.close()

        token = _admin_token(client, seed_admin)
        rv = client.post("/api/attendance/checkin",
                         json={"employee_id": seed_employee["employee_id"]},
                         headers={"Authorization": f"Bearer {token}"})
        data = rv.get_json()
        assert data["ok"] is True
        assert data["type"] == "login"
        assert data["status"] in ("Full Day Login", "Late Login", "Half Day Login")

        cur = db_engine.cursor()
        cur.execute("DELETE FROM attendance WHERE employee_id=%s AND date=%s",
                    (seed_employee["employee_id"], today))
        cur.close()

    def test_logout_updates_record(self, client, seed_admin, seed_employee, attendance_today):
        token = _admin_token(client, seed_admin)
        rv = client.post("/api/attendance/checkin",
                         json={"employee_id": seed_employee["employee_id"]},
                         headers={"Authorization": f"Bearer {token}"})
        data = rv.get_json()
        assert data["ok"] is True
        assert data["type"] == "logout"
        assert "att_type" in data

    def test_relogin_clears_logout_time(self, client, seed_admin, seed_employee, attendance_completed):
        token = _admin_token(client, seed_admin)
        rv = client.post("/api/attendance/checkin",
                         json={"employee_id": seed_employee["employee_id"]},
                         headers={"Authorization": f"Bearer {token}"})
        data = rv.get_json()
        assert data["ok"] is True
        assert data["type"] == "relogin"

    def test_geo_fence_rejection(self, client, seed_admin, seed_employee, db_engine, mocker):
        today = datetime.date.today()
        cur = db_engine.cursor()
        cur.execute("DELETE FROM attendance WHERE employee_id=%s AND date=%s",
                    (seed_employee["employee_id"], today))
        cur.close()

        mocker.patch("blueprints.attendance.is_within_range", return_value=False)
        token = _admin_token(client, seed_admin)
        rv = client.post("/api/attendance/checkin",
                         json={"employee_id": seed_employee["employee_id"],
                               "lat": 19.0760, "lon": 72.8777},
                         headers={"Authorization": f"Bearer {token}"})
        data = rv.get_json()
        assert data["ok"] is False
        assert "outside" in data["msg"].lower()


# ═══════════════════════════════════════════════════════════════════════════════
# /api/employee/checkin (employee Bearer token, POST JSON)
# ═══════════════════════════════════════════════════════════════════════════════

class TestApiEmployeeCheckin:

    def test_requires_employee_token(self, client):
        rv = client.post("/api/employee/checkin", json={})
        assert rv.status_code == 401

    def test_login_creates_attendance(self, client, seed_employee, db_engine):
        today = datetime.date.today()
        cur = db_engine.cursor()
        cur.execute("DELETE FROM attendance WHERE employee_id=%s AND date=%s",
                    (seed_employee["employee_id"], today))
        cur.close()

        token = _emp_token(client, seed_employee)
        rv = client.post("/api/employee/checkin", json={},
                         headers={"Authorization": f"Bearer {token}"})
        data = rv.get_json()
        assert data["ok"] is True
        assert data["action"] == "login"
        assert data["status"] in ("Full Day Login", "Late Login", "Half Day Login")

        cur = db_engine.cursor()
        cur.execute("DELETE FROM attendance WHERE employee_id=%s AND date=%s",
                    (seed_employee["employee_id"], today))
        cur.close()

    def test_logout_updates_existing_login(self, client, seed_employee, attendance_today):
        token = _emp_token(client, seed_employee)
        rv = client.post("/api/employee/checkin", json={},
                         headers={"Authorization": f"Bearer {token}"})
        data = rv.get_json()
        assert data["ok"] is True
        assert data["action"] == "logout"
        assert "att_type" in data

    def test_relogin_after_completed(self, client, seed_employee, attendance_completed):
        token = _emp_token(client, seed_employee)
        rv = client.post("/api/employee/checkin", json={},
                         headers={"Authorization": f"Bearer {token}"})
        data = rv.get_json()
        assert data["ok"] is True
        assert data["action"] == "relogin"

    def test_offline_punch_within_24h_accepted(self, client, seed_employee, db_engine):
        today = datetime.date.today()
        cur = db_engine.cursor()
        cur.execute("DELETE FROM attendance WHERE employee_id=%s AND date=%s",
                    (seed_employee["employee_id"], today))
        cur.close()

        two_hours_ago = datetime.datetime.now() - datetime.timedelta(hours=2)
        token = _emp_token(client, seed_employee)
        rv = client.post("/api/employee/checkin",
                         json={"punched_at": two_hours_ago.isoformat()},
                         headers={"Authorization": f"Bearer {token}"})
        data = rv.get_json()
        assert data["ok"] is True
        assert data["action"] == "login"

        cur = db_engine.cursor()
        cur.execute("DELETE FROM attendance WHERE employee_id=%s AND date=%s",
                    (seed_employee["employee_id"], today))
        cur.close()

    def test_offline_punch_over_24h_rejected(self, client, seed_employee, db_engine):
        today = datetime.date.today()
        cur = db_engine.cursor()
        cur.execute("DELETE FROM attendance WHERE employee_id=%s AND date=%s",
                    (seed_employee["employee_id"], today))
        cur.close()

        two_days_ago = datetime.datetime.now() - datetime.timedelta(days=2)
        token = _emp_token(client, seed_employee)
        rv = client.post("/api/employee/checkin",
                         json={"punched_at": two_days_ago.isoformat()},
                         headers={"Authorization": f"Bearer {token}"})
        assert rv.status_code == 400
        data = rv.get_json()
        assert data["ok"] is False
        assert "24" in data["msg"]

    def test_invalid_lat_lon_returns_400(self, client, seed_employee, db_engine):
        today = datetime.date.today()
        cur = db_engine.cursor()
        cur.execute("DELETE FROM attendance WHERE employee_id=%s AND date=%s",
                    (seed_employee["employee_id"], today))
        cur.close()

        token = _emp_token(client, seed_employee)
        rv = client.post("/api/employee/checkin",
                         json={"lat": "not_a_float", "lon": "also_bad"},
                         headers={"Authorization": f"Bearer {token}"})
        assert rv.status_code == 400
        assert rv.get_json()["ok"] is False

    def test_geo_fence_rejection(self, client, seed_employee, db_engine, mocker):
        today = datetime.date.today()
        cur = db_engine.cursor()
        cur.execute("DELETE FROM attendance WHERE employee_id=%s AND date=%s",
                    (seed_employee["employee_id"], today))
        cur.close()

        mocker.patch("blueprints.attendance.is_within_range", return_value=False)
        token = _emp_token(client, seed_employee)
        rv = client.post("/api/employee/checkin",
                         json={"lat": 19.0760, "lon": 72.8777},
                         headers={"Authorization": f"Bearer {token}"})
        data = rv.get_json()
        assert data["ok"] is False
        assert "outside" in data["msg"].lower()


# ═══════════════════════════════════════════════════════════════════════════════
# /api/employee/qr-face-checkin (public kiosk, multipart/form-data)
# ═══════════════════════════════════════════════════════════════════════════════

class TestApiQrFaceCheckin:

    def test_invalid_auth_combo_returns_400(self, client, seed_employee):
        rv = client.post("/api/employee/qr-face-checkin", data={
            "employee_id": seed_employee["employee_id"],
            "auth_combo":  "unknown_combo",
        })
        assert rv.status_code == 400
        assert rv.get_json()["ok"] is False

    def test_missing_employee_id_returns_400(self, client):
        rv = client.post("/api/employee/qr-face-checkin", data={
            "employee_id": "",
            "auth_combo":  "qr_face",
        })
        assert rv.status_code == 400

    def test_qr_disabled_returns_403(self, client, seed_employee, mocker):
        mocker.patch("blueprints.attendance.get_auth_config", return_value={
            "qr_enabled": False, "face_enabled": True,
            "fingerprint_enabled": False, "location_enabled": False,
        })
        rv = client.post("/api/employee/qr-face-checkin", data={
            "employee_id": seed_employee["employee_id"],
            "auth_combo":  "qr_face",
        })
        assert rv.status_code == 403

    def test_face_disabled_returns_403(self, client, seed_employee, mocker):
        mocker.patch("blueprints.attendance.get_auth_config", return_value={
            "qr_enabled": True, "face_enabled": False,
            "fingerprint_enabled": False, "location_enabled": False,
        })
        rv = client.post("/api/employee/qr-face-checkin", data={
            "employee_id": seed_employee["employee_id"],
            "auth_combo":  "qr_face",
        })
        assert rv.status_code == 403

    def test_fingerprint_disabled_returns_403(self, client, seed_employee, mocker):
        mocker.patch("blueprints.attendance.get_auth_config", return_value={
            "qr_enabled": True, "face_enabled": True,
            "fingerprint_enabled": False, "location_enabled": False,
        })
        rv = client.post("/api/employee/qr-face-checkin", data={
            "employee_id": seed_employee["employee_id"],
            "auth_combo":  "qr_fingerprint",
        })
        assert rv.status_code == 403

    def test_fingerprint_not_verified_returns_401(self, client, seed_employee, mocker):
        mocker.patch("blueprints.attendance.get_auth_config", return_value={
            "qr_enabled": True, "face_enabled": True,
            "fingerprint_enabled": True, "location_enabled": False,
        })
        mocker.patch("blueprints.attendance._wa_fingerprint_recently_verified", return_value=False)
        mocker.patch("blueprints.attendance._mobile_biometric_recently_verified", return_value=False)
        rv = client.post("/api/employee/qr-face-checkin", data={
            "employee_id": seed_employee["employee_id"],
            "auth_combo":  "qr_fingerprint",
        })
        assert rv.status_code == 401

    def test_unknown_employee_returns_404(self, client, mocker):
        mocker.patch("blueprints.attendance.get_auth_config", return_value={
            "qr_enabled": True, "face_enabled": True,
            "fingerprint_enabled": False, "location_enabled": False,
        })
        rv = client.post("/api/employee/qr-face-checkin", data={
            "employee_id": "GHOST99",
            "auth_combo":  "qr_face",
        })
        assert rv.status_code == 404

    def test_qr_fingerprint_login_with_verified_fingerprint(
            self, client, seed_employee, db_engine, mocker):
        """qr_fingerprint with a verified fingerprint should create a login record."""
        today = datetime.date.today()
        cur = db_engine.cursor()
        cur.execute("DELETE FROM attendance WHERE employee_id=%s AND date=%s",
                    (seed_employee["employee_id"], today))
        cur.close()

        mocker.patch("blueprints.attendance.get_auth_config", return_value={
            "qr_enabled": True, "face_enabled": True,
            "fingerprint_enabled": True, "location_enabled": False,
        })
        mocker.patch("blueprints.attendance._wa_fingerprint_recently_verified", return_value=True)

        rv = client.post("/api/employee/qr-face-checkin", data={
            "employee_id": seed_employee["employee_id"],
            "auth_combo":  "qr_fingerprint",
        })
        data = rv.get_json()
        assert data["ok"] is True
        assert data["action"] == "login"

        cur = db_engine.cursor()
        cur.execute("DELETE FROM attendance WHERE employee_id=%s AND date=%s",
                    (seed_employee["employee_id"], today))
        cur.close()

    def test_qr_face_without_photo_returns_400(self, client, seed_employee, mocker):
        mocker.patch("blueprints.attendance.get_auth_config", return_value={
            "qr_enabled": True, "face_enabled": True,
            "fingerprint_enabled": False, "location_enabled": False,
        })
        rv = client.post("/api/employee/qr-face-checkin", data={
            "employee_id": seed_employee["employee_id"],
            "auth_combo":  "qr_face",
        })
        assert rv.status_code == 400
        assert "face photo required" in rv.get_json()["msg"].lower()

    def test_qr_face_with_face_recognition_unavailable(self, client, seed_employee, mocker):
        from PIL import Image as _PIL
        buf = io.BytesIO()
        _PIL.new("RGB", (100, 100)).save(buf, format="JPEG")
        buf.seek(0)

        mocker.patch("blueprints.attendance.get_auth_config", return_value={
            "qr_enabled": True, "face_enabled": True,
            "fingerprint_enabled": False, "location_enabled": False,
        })
        mocker.patch("blueprints.attendance._face_recognition_available", False)

        rv = client.post("/api/employee/qr-face-checkin", data={
            "employee_id": seed_employee["employee_id"],
            "auth_combo":  "qr_face",
            "face_photo":  (buf, "face.jpg", "image/jpeg"),
        }, content_type="multipart/form-data")
        assert rv.status_code == 503
        assert "unavailable" in rv.get_json()["msg"].lower()

    def test_qr_fingerprint_relogin_after_completed(
            self, client, seed_employee, attendance_completed, mocker):
        mocker.patch("blueprints.attendance.get_auth_config", return_value={
            "qr_enabled": True, "face_enabled": True,
            "fingerprint_enabled": True, "location_enabled": False,
        })
        mocker.patch("blueprints.attendance._wa_fingerprint_recently_verified", return_value=True)

        rv = client.post("/api/employee/qr-face-checkin", data={
            "employee_id": seed_employee["employee_id"],
            "auth_combo":  "qr_fingerprint",
        })
        data = rv.get_json()
        assert data["ok"] is True
        assert data["action"] == "relogin"


# ═══════════════════════════════════════════════════════════════════════════════
# /api/employee/attendance (employee Bearer token, GET)
# ═══════════════════════════════════════════════════════════════════════════════

class TestApiEmployeeAttendanceGet:

    def test_requires_employee_token(self, client):
        rv = client.get("/api/employee/attendance")
        assert rv.status_code == 401

    def test_returns_required_json_keys(self, client, seed_employee):
        token = _emp_token(client, seed_employee)
        rv = client.get("/api/employee/attendance",
                        headers={"Authorization": f"Bearer {token}"})
        data = rv.get_json()
        assert data["ok"] is True
        assert "records" in data
        assert "summary" in data
        assert "year" in data
        assert "month" in data
        assert "month_name" in data

    def test_custom_year_month(self, client, seed_employee):
        token = _emp_token(client, seed_employee)
        rv = client.get("/api/employee/attendance?year=2025&month=3",
                        headers={"Authorization": f"Bearer {token}"})
        data = rv.get_json()
        assert data["ok"] is True
        assert data["year"] == 2025
        assert data["month"] == 3
        assert "March 2025" in data["month_name"]

    def test_invalid_year_returns_400(self, client, seed_employee):
        token = _emp_token(client, seed_employee)
        rv = client.get("/api/employee/attendance?year=notanumber",
                        headers={"Authorization": f"Bearer {token}"})
        assert rv.status_code == 400
        assert rv.get_json()["ok"] is False

    def test_records_contain_expected_fields(self, client, seed_employee, db_engine):
        today = datetime.date.today()
        cur = db_engine.cursor()
        cur.execute("DELETE FROM attendance WHERE employee_id=%s AND date=%s",
                    (seed_employee["employee_id"], today))
        cur.execute(
            "INSERT INTO attendance (employee_id, date, login_time, logout_time, "
            "status, logout_status, attendance_type, worked_minutes) "
            "VALUES (%s,%s,'09:00:00','18:00:00','Full Day Login','Completed','Full Day',540)",
            (seed_employee["employee_id"], today),
        )
        cur.close()

        token = _emp_token(client, seed_employee)
        rv = client.get(f"/api/employee/attendance?year={today.year}&month={today.month}",
                        headers={"Authorization": f"Bearer {token}"})
        data = rv.get_json()
        assert data["ok"] is True
        rec = next((r for r in data["records"] if r["date"] == str(today)), None)
        assert rec is not None
        assert rec["attendance_type"] == "Full Day"
        assert rec["login_time"] == "09:00:00"
        assert "logout_time" in rec
        assert "worked_minutes" in rec

        cur = db_engine.cursor()
        cur.execute("DELETE FROM attendance WHERE employee_id=%s AND date=%s",
                    (seed_employee["employee_id"], today))
        cur.close()

    def test_summary_counts_full_and_late_days(self, client, seed_employee, db_engine):
        today = datetime.date.today()
        cur = db_engine.cursor()
        cur.execute("DELETE FROM attendance WHERE employee_id=%s AND date=%s",
                    (seed_employee["employee_id"], today))
        cur.execute(
            "INSERT INTO attendance (employee_id, date, login_time, attendance_type) "
            "VALUES (%s,%s,'09:00:00','Full Day')",
            (seed_employee["employee_id"], today),
        )
        cur.close()

        token = _emp_token(client, seed_employee)
        rv = client.get(f"/api/employee/attendance?year={today.year}&month={today.month}",
                        headers={"Authorization": f"Bearer {token}"})
        data = rv.get_json()
        assert data["summary"]["full_days"] >= 1

        cur = db_engine.cursor()
        cur.execute("DELETE FROM attendance WHERE employee_id=%s AND date=%s",
                    (seed_employee["employee_id"], today))
        cur.close()


# ═══════════════════════════════════════════════════════════════════════════════
# Shift swap lifecycle
# ═══════════════════════════════════════════════════════════════════════════════

class TestShiftSwapLifecycle:

    # ── submit_shift_swap ────────────────────────────────────────────────────

    def test_submit_requires_employee_session(self, client):
        rv = client.post("/submit_shift_swap", data={"target_id": "TST002"})
        assert rv.status_code in (302, 401)

    def test_submit_missing_target_id(self, client, seed_employee):
        _emp_session(client, seed_employee)
        rv = client.post("/submit_shift_swap", data={"target_id": ""})
        assert rv.status_code == 302
        assert "invalid_target" in rv.headers["Location"]

    def test_submit_self_as_target(self, client, seed_employee):
        _emp_session(client, seed_employee)
        rv = client.post("/submit_shift_swap",
                         data={"target_id": seed_employee["employee_id"]})
        assert rv.status_code == 302
        assert "invalid_target" in rv.headers["Location"]

    def test_submit_no_shift_assigned(self, client, seed_employee, seed_employee2, db_engine):
        """When employees have no shift assigned, swap is rejected."""
        cur = db_engine.cursor()
        cur.execute("UPDATE employees SET shift_id=NULL WHERE employee_id IN ('TST001','TST002')")
        cur.close()

        _emp_session(client, seed_employee)
        rv = client.post("/submit_shift_swap",
                         data={"target_id": seed_employee2["employee_id"], "reason": "test"})
        assert rv.status_code == 302
        assert "no_shift" in rv.headers["Location"]

    def test_submit_same_shift_rejected(self, client, seed_employee, seed_employee2, shift_a, db_engine):
        """Both employees on same shift → same_shift error."""
        cur = db_engine.cursor()
        cur.execute("UPDATE employees SET shift_id=%s WHERE employee_id IN ('TST001','TST002')",
                    (shift_a["id"],))
        cur.close()

        _emp_session(client, seed_employee)
        rv = client.post("/submit_shift_swap",
                         data={"target_id": seed_employee2["employee_id"], "reason": "test"})
        assert rv.status_code == 302
        assert "same_shift" in rv.headers["Location"]

        cur = db_engine.cursor()
        cur.execute("UPDATE employees SET shift_id=NULL WHERE employee_id IN ('TST001','TST002')")
        cur.close()

    def test_submit_happy_path_creates_pending_target(
            self, client, seed_employee, seed_employee2, shift_a, shift_b, db_engine):
        cur = db_engine.cursor()
        cur.execute("UPDATE employees SET shift_id=%s WHERE employee_id='TST001'", (shift_a["id"],))
        cur.execute("UPDATE employees SET shift_id=%s WHERE employee_id='TST002'", (shift_b["id"],))
        cur.execute("DELETE FROM shift_swap_requests WHERE requester_id='TST001' AND target_id='TST002'")
        cur.close()

        _emp_session(client, seed_employee)
        rv = client.post("/submit_shift_swap",
                         data={"target_id": seed_employee2["employee_id"], "reason": "CI reason"})
        assert rv.status_code == 302
        assert "swap_sent=1" in rv.headers["Location"]

        cur = db_engine.cursor()
        cur.execute(
            "SELECT status FROM shift_swap_requests "
            "WHERE requester_id='TST001' AND target_id='TST002'"
        )
        row = cur.fetchone()
        assert row is not None
        assert row[0] == "Pending_Target"
        cur.execute("DELETE FROM shift_swap_requests WHERE requester_id='TST001' AND target_id='TST002'")
        cur.execute("UPDATE employees SET shift_id=NULL WHERE employee_id IN ('TST001','TST002')")
        cur.close()

    def test_submit_duplicate_request_rejected(self, client, seed_employee, swap_pending):
        _emp_session(client, seed_employee)
        rv = client.post("/submit_shift_swap",
                         data={"target_id": swap_pending["target_id"], "reason": "again"})
        assert rv.status_code == 302
        assert "duplicate" in rv.headers["Location"]

    # ── respond_shift_swap ───────────────────────────────────────────────────

    def test_respond_accept_moves_to_pending_admin(
            self, client, seed_employee2, swap_pending, db_engine):
        with client.session_transaction() as sess:
            sess["employee_id"]   = seed_employee2["employee_id"]
            sess["employee_name"] = seed_employee2["name"]

        rv = client.post(f"/respond_shift_swap/{swap_pending['id']}",
                         data={"action": "accept", "response": "Happy to swap"})
        assert rv.status_code == 302
        assert "swap_responded=1" in rv.headers["Location"]

        cur = db_engine.cursor()
        cur.execute("SELECT status FROM shift_swap_requests WHERE id=%s", (swap_pending["id"],))
        assert cur.fetchone()[0] == "Pending_Admin"
        cur.close()

    def test_respond_reject_sets_rejected(self, client, seed_employee2, swap_pending, db_engine):
        with client.session_transaction() as sess:
            sess["employee_id"]   = seed_employee2["employee_id"]
            sess["employee_name"] = seed_employee2["name"]

        rv = client.post(f"/respond_shift_swap/{swap_pending['id']}",
                         data={"action": "reject", "response": "Cannot swap"})
        assert rv.status_code == 302

        cur = db_engine.cursor()
        cur.execute("SELECT status FROM shift_swap_requests WHERE id=%s", (swap_pending["id"],))
        assert cur.fetchone()[0] == "Rejected"
        cur.close()

    def test_respond_nonexistent_request(self, client, seed_employee2):
        with client.session_transaction() as sess:
            sess["employee_id"]   = seed_employee2["employee_id"]
            sess["employee_name"] = seed_employee2["name"]

        rv = client.post("/respond_shift_swap/99999", data={"action": "accept"})
        assert rv.status_code == 302
        assert "not_found" in rv.headers["Location"]

    def test_respond_wrong_employee_is_rejected(self, client, seed_employee, swap_pending):
        """Requester cannot respond to their own swap request."""
        _emp_session(client, seed_employee)
        rv = client.post(f"/respond_shift_swap/{swap_pending['id']}",
                         data={"action": "accept"})
        assert rv.status_code == 302
        assert "not_found" in rv.headers["Location"]

    # ── admin_shift_swap ─────────────────────────────────────────────────────

    def test_admin_approve_swaps_shift_assignments(
            self, client, seed_admin, swap_pending_admin, db_engine):
        _admin_session(client, seed_admin)
        rv = client.post(f"/admin_shift_swap/{swap_pending_admin['id']}",
                         data={"action": "approve", "admin_response": "Approved by CI"})
        assert rv.status_code == 302

        cur = db_engine.cursor()
        cur.execute("SELECT shift_id FROM employees WHERE employee_id='TST001'")
        tst001_shift = cur.fetchone()[0]
        cur.execute("SELECT shift_id FROM employees WHERE employee_id='TST002'")
        tst002_shift = cur.fetchone()[0]
        cur.close()

        assert tst001_shift == swap_pending_admin["tgt_shift_id"]
        assert tst002_shift == swap_pending_admin["req_shift_id"]

    def test_admin_reject_sets_rejected_admin(
            self, client, seed_admin, swap_pending_admin, db_engine):
        _admin_session(client, seed_admin)
        rv = client.post(f"/admin_shift_swap/{swap_pending_admin['id']}",
                         data={"action": "reject", "admin_response": "Rejected by CI"})
        assert rv.status_code == 302

        cur = db_engine.cursor()
        cur.execute("SELECT status FROM shift_swap_requests WHERE id=%s",
                    (swap_pending_admin["id"],))
        assert cur.fetchone()[0] == "Rejected_Admin"
        cur.close()

    def test_admin_approve_nonexistent_redirects(self, client, seed_admin):
        _admin_session(client, seed_admin)
        rv = client.post("/admin_shift_swap/99999", data={"action": "approve"})
        assert rv.status_code == 302
        assert "not_found" in rv.headers["Location"]


# ═══════════════════════════════════════════════════════════════════════════════
# /monthly_report_export (admin, GET → xlsx download)
# ═══════════════════════════════════════════════════════════════════════════════

class TestMonthlyReportExport:

    def test_requires_admin(self, client):
        rv = client.get("/monthly_report_export")
        assert rv.status_code == 302

    def test_returns_xlsx_content_type(self, client, seed_admin):
        _admin_session(client, seed_admin)
        rv = client.get("/monthly_report_export?year=2025&month=1")
        assert rv.status_code == 200
        ct = rv.headers.get("Content-Type", "")
        assert "openxmlformats" in ct or "spreadsheet" in ct

    def test_filename_in_content_disposition(self, client, seed_admin):
        _admin_session(client, seed_admin)
        rv = client.get("/monthly_report_export?year=2025&month=3")
        cd = rv.headers.get("Content-Disposition", "")
        assert "attendance_2025_03.xlsx" in cd

    def test_response_has_xlsx_magic_bytes(self, client, seed_admin):
        _admin_session(client, seed_admin)
        rv = client.get("/monthly_report_export?year=2025&month=1")
        # XLSX is a ZIP archive — starts with PK magic bytes
        assert rv.data[:2] == b"PK"

    def test_default_params_uses_current_month(self, client, seed_admin):
        _admin_session(client, seed_admin)
        rv = client.get("/monthly_report_export")
        assert rv.status_code == 200
        assert rv.data[:2] == b"PK"


# ═══════════════════════════════════════════════════════════════════════════════
# /send_absentee_report (admin, POST → JSON)
# ═══════════════════════════════════════════════════════════════════════════════

class TestSendAbsenteeReport:

    def test_requires_admin(self, client):
        rv = client.post("/send_absentee_report")
        assert rv.status_code == 302

    def test_no_email_config_returns_error_json(self, client, seed_admin, mocker):
        mocker.patch("blueprints.attendance.get_email_config", return_value=None)
        _admin_session(client, seed_admin)
        rv = client.post("/send_absentee_report")
        data = rv.get_json()
        assert data["ok"] is False
        assert "Email not configured" in data["msg"]

    def test_send_success_returns_ok(self, client, seed_admin, mocker):
        mocker.patch("blueprints.attendance.get_email_config", return_value={
            "user": "smtp@test.com",
            "from_email": "smtp@test.com",
            "password": "pass",
            "host": "smtp.test.com",
            "port": 587,
        })
        mock_send = mocker.patch("blueprints.attendance.send_email_smtp", return_value=None)
        _admin_session(client, seed_admin)
        rv = client.post("/send_absentee_report")
        data = rv.get_json()
        assert data["ok"] is True
        assert "Report sent" in data["msg"]
        mock_send.assert_called_once()

    def test_smtp_failure_returns_error_json(self, client, seed_admin, mocker):
        mocker.patch("blueprints.attendance.get_email_config", return_value={
            "user": "smtp@test.com", "from_email": "smtp@test.com",
        })
        mocker.patch("blueprints.attendance.send_email_smtp",
                     side_effect=Exception("SMTP connection refused"))
        _admin_session(client, seed_admin)
        rv = client.post("/send_absentee_report")
        data = rv.get_json()
        assert data["ok"] is False
        assert "Failed" in data["msg"]

    def test_email_sent_to_from_email_field(self, client, seed_admin, mocker):
        """send_email_smtp is called with the from_email address as recipient."""
        mocker.patch("blueprints.attendance.get_email_config", return_value={
            "user": "user@smtp.com",
            "from_email": "reports@smtp.com",
        })
        mock_send = mocker.patch("blueprints.attendance.send_email_smtp", return_value=None)
        _admin_session(client, seed_admin)
        client.post("/send_absentee_report")
        called_to = mock_send.call_args[0][0]
        assert called_to == "reports@smtp.com"


# ═══════════════════════════════════════════════════════════════════════════════
# /my_attendance_pdf (employee, GET → HTML page)
# ═══════════════════════════════════════════════════════════════════════════════

class TestMyAttendancePdf:

    def test_requires_employee_session(self, client):
        rv = client.get("/my_attendance_pdf")
        assert rv.status_code == 302

    def test_returns_html_200(self, client, seed_employee):
        _emp_session(client, seed_employee)
        rv = client.get("/my_attendance_pdf")
        assert rv.status_code == 200
        assert b"Attendance Report" in rv.data

    def test_employee_name_appears_in_page(self, client, seed_employee):
        _emp_session(client, seed_employee)
        rv = client.get("/my_attendance_pdf")
        assert seed_employee["name"].encode() in rv.data

    def test_custom_year_month_in_page(self, client, seed_employee):
        _emp_session(client, seed_employee)
        rv = client.get("/my_attendance_pdf?year=2025&month=1")
        assert rv.status_code == 200
        assert b"January 2025" in rv.data

    def test_print_button_present(self, client, seed_employee):
        _emp_session(client, seed_employee)
        rv = client.get("/my_attendance_pdf")
        assert b"window.print()" in rv.data


# ═══════════════════════════════════════════════════════════════════════════════
# correct_attendance — edge cases beyond basic auth guard
# ═══════════════════════════════════════════════════════════════════════════════

class TestCorrectAttendanceEdgeCases:

    def test_missing_emp_id_redirects(self, client, seed_admin):
        _admin_session(client, seed_admin)
        rv = client.post("/correct_attendance", data={
            "emp_id": "", "date": "2025-03-10", "attendance_type": "Full Day",
        })
        assert rv.status_code == 302

    def test_missing_attendance_type_redirects(self, client, seed_admin, seed_employee):
        _admin_session(client, seed_admin)
        rv = client.post("/correct_attendance", data={
            "emp_id": seed_employee["employee_id"],
            "date":   "2025-03-10",
            "attendance_type": "",
        })
        assert rv.status_code == 302

    def test_invalid_date_format_redirects(self, client, seed_admin, seed_employee):
        _admin_session(client, seed_admin)
        rv = client.post("/correct_attendance", data={
            "emp_id":          seed_employee["employee_id"],
            "date":            "not-a-date",
            "attendance_type": "Full Day",
        })
        assert rv.status_code == 302

    def test_insert_when_no_existing_record(self, client, seed_admin, seed_employee, db_engine):
        date_str = "2025-01-15"
        cur = db_engine.cursor()
        cur.execute("DELETE FROM attendance WHERE employee_id=%s AND date=%s",
                    (seed_employee["employee_id"], date_str))
        cur.close()

        _admin_session(client, seed_admin)
        rv = client.post("/correct_attendance", data={
            "emp_id":          seed_employee["employee_id"],
            "date":            date_str,
            "login_time":      "09:00",
            "logout_time":     "18:00",
            "attendance_type": "Full Day",
            "year":            "2025",
            "month":           "1",
        })
        assert rv.status_code == 302

        cur = db_engine.cursor()
        cur.execute("SELECT attendance_type, status FROM attendance WHERE employee_id=%s AND date=%s",
                    (seed_employee["employee_id"], date_str))
        row = cur.fetchone()
        cur.execute("DELETE FROM attendance WHERE employee_id=%s AND date=%s",
                    (seed_employee["employee_id"], date_str))
        cur.close()

        assert row is not None
        assert row[0] == "Full Day"
        assert row[1] == "Manual"

    def test_update_when_existing_record(self, client, seed_admin, seed_employee, db_engine):
        date_str = "2025-01-15"
        cur = db_engine.cursor()
        cur.execute("DELETE FROM attendance WHERE employee_id=%s AND date=%s",
                    (seed_employee["employee_id"], date_str))
        cur.execute(
            "INSERT INTO attendance (employee_id, date, login_time, attendance_type, status) "
            "VALUES (%s,%s,'08:00:00','Half Day','Late Login')",
            (seed_employee["employee_id"], date_str),
        )
        cur.close()

        _admin_session(client, seed_admin)
        rv = client.post("/correct_attendance", data={
            "emp_id":          seed_employee["employee_id"],
            "date":            date_str,
            "login_time":      "09:00",
            "logout_time":     "18:00",
            "attendance_type": "Full Day",
            "year":            "2025",
            "month":           "1",
        })
        assert rv.status_code == 302

        cur = db_engine.cursor()
        cur.execute("SELECT attendance_type, status FROM attendance WHERE employee_id=%s AND date=%s",
                    (seed_employee["employee_id"], date_str))
        row = cur.fetchone()
        cur.execute("DELETE FROM attendance WHERE employee_id=%s AND date=%s",
                    (seed_employee["employee_id"], date_str))
        cur.close()

        assert row[0] == "Full Day"
        assert row[1] == "Manual"


# ═══════════════════════════════════════════════════════════════════════════════
# bulk_assign_shift — edge cases beyond basic happy path
# ═══════════════════════════════════════════════════════════════════════════════

class TestBulkAssignShiftEdgeCases:

    def test_no_emp_ids_no_dept_updates_all_employees(
            self, client, seed_admin, seed_employee, shift_a, db_engine):
        """Known behavior: no emp_ids + no dept_filter → UPDATE employees SET shift_id=X (no WHERE)."""
        _admin_session(client, seed_admin)
        rv = client.post("/bulk_assign_shift", data={"shift_id": str(shift_a["id"])})
        assert rv.status_code == 302

        cur = db_engine.cursor()
        cur.execute("SELECT shift_id FROM employees WHERE employee_id=%s",
                    (seed_employee["employee_id"],))
        row = cur.fetchone()
        cur.execute("UPDATE employees SET shift_id=NULL WHERE employee_id=%s",
                    (seed_employee["employee_id"],))
        cur.close()

        assert row[0] == shift_a["id"]

    def test_dept_filter_updates_only_that_department(
            self, client, seed_admin, seed_employee, shift_a, db_engine):
        cur = db_engine.cursor()
        cur.execute("UPDATE employees SET department='Engineering' WHERE employee_id=%s",
                    (seed_employee["employee_id"],))
        cur.close()

        _admin_session(client, seed_admin)
        rv = client.post("/bulk_assign_shift", data={
            "shift_id":    str(shift_a["id"]),
            "dept_filter": "Engineering",
        })
        assert rv.status_code == 302

        cur = db_engine.cursor()
        cur.execute("SELECT shift_id FROM employees WHERE employee_id=%s",
                    (seed_employee["employee_id"],))
        row = cur.fetchone()
        cur.execute("UPDATE employees SET shift_id=NULL, department=NULL WHERE employee_id=%s",
                    (seed_employee["employee_id"],))
        cur.close()

        assert row[0] == shift_a["id"]

    def test_emp_ids_list_updates_specific_employees(
            self, client, seed_admin, seed_employee, shift_a, db_engine):
        _admin_session(client, seed_admin)
        rv = client.post("/bulk_assign_shift", data={
            "shift_id": str(shift_a["id"]),
            "emp_ids":  [seed_employee["employee_id"]],
        })
        assert rv.status_code == 302

        cur = db_engine.cursor()
        cur.execute("SELECT shift_id FROM employees WHERE employee_id=%s",
                    (seed_employee["employee_id"],))
        row = cur.fetchone()
        cur.execute("UPDATE employees SET shift_id=NULL WHERE employee_id=%s",
                    (seed_employee["employee_id"],))
        cur.close()

        assert row[0] == shift_a["id"]


# ═══════════════════════════════════════════════════════════════════════════════
# assign_shift — JSON response and NULL assignment
# ═══════════════════════════════════════════════════════════════════════════════

class TestAssignShift:

    def test_returns_json_ok(self, client, seed_admin, seed_employee, shift_a, db_engine):
        _admin_session(client, seed_admin)
        rv = client.post("/assign_shift", data={
            "emp_id":   seed_employee["employee_id"],
            "shift_id": str(shift_a["id"]),
        })
        assert rv.status_code == 200
        assert rv.get_json()["ok"] is True

        cur = db_engine.cursor()
        cur.execute("UPDATE employees SET shift_id=NULL WHERE employee_id=%s",
                    (seed_employee["employee_id"],))
        cur.close()

    def test_empty_shift_id_clears_assignment(self, client, seed_admin, seed_employee, db_engine):
        _admin_session(client, seed_admin)
        rv = client.post("/assign_shift", data={
            "emp_id":   seed_employee["employee_id"],
            "shift_id": "",
        })
        assert rv.get_json()["ok"] is True

        cur = db_engine.cursor()
        cur.execute("SELECT shift_id FROM employees WHERE employee_id=%s",
                    (seed_employee["employee_id"],))
        assert cur.fetchone()[0] is None
        cur.close()


# ═══════════════════════════════════════════════════════════════════════════════
# edit_shift — exception path when sid cannot be parsed
# ═══════════════════════════════════════════════════════════════════════════════

class TestEditShiftEdgeCases:

    def test_no_sid_in_form_redirects_to_schedule(self, client, seed_admin):
        """Unparseable shift_id triggers the except branch → redirect."""
        _admin_session(client, seed_admin)
        rv = client.post("/edit_shift", data={
            "shift_name": "Morning",
            "start_time": "09:00",
            "half_time":  "13:00",
            "end_time":   "18:00",
            "shift_id":   "not_an_int",
        })
        assert rv.status_code == 302
        assert "schedule" in rv.headers["Location"]

    def test_missing_name_skips_db_update(self, client, seed_admin, shift_a, db_engine):
        """Empty name causes early redirect without updating the DB."""
        _admin_session(client, seed_admin)
        rv = client.post(f"/edit_shift/{shift_a['id']}", data={
            "shift_name": "",
            "start_time": "08:00",
            "half_time":  "12:00",
            "end_time":   "17:00",
        })
        assert rv.status_code == 302

        cur = db_engine.cursor()
        cur.execute("SELECT name FROM shifts WHERE id=%s", (shift_a["id"],))
        assert cur.fetchone()[0] == shift_a["name"]  # unchanged
        cur.close()


# ═══════════════════════════════════════════════════════════════════════════════
# delete_shift_form — POST /delete_shift (form-based, without path param)
# ═══════════════════════════════════════════════════════════════════════════════

class TestDeleteShiftForm:

    def test_missing_shift_id_just_redirects(self, client, seed_admin):
        _admin_session(client, seed_admin)
        rv = client.post("/delete_shift", data={"shift_id": ""})
        assert rv.status_code == 302

    def test_deletes_shift_and_clears_employee_assignments(self, client, seed_admin, db_engine):
        cur = db_engine.cursor()
        cur.execute(
            "INSERT INTO shifts (name, start_time, half_time, end_time) "
            "VALUES ('Temp Delete Shift','10:00','14:00','19:00') RETURNING id"
        )
        sid = cur.fetchone()[0]
        cur.close()

        _admin_session(client, seed_admin)
        rv = client.post("/delete_shift", data={"shift_id": str(sid)})
        assert rv.status_code == 302

        cur = db_engine.cursor()
        cur.execute("SELECT id FROM shifts WHERE id=%s", (sid,))
        assert cur.fetchone() is None
        cur.close()


# ═══════════════════════════════════════════════════════════════════════════════
# employee_attendance_detail — edge cases
# ═══════════════════════════════════════════════════════════════════════════════

class TestEmployeeAttendanceDetailEdgeCases:

    def test_unknown_employee_id_returns_404(self, client, seed_admin):
        _admin_session(client, seed_admin)
        rv = client.get("/employee_attendance_detail/NOTEXIST/2025/3")
        assert rv.status_code == 404

    def test_valid_employee_returns_200_with_name(self, client, seed_admin, seed_employee):
        _admin_session(client, seed_admin)
        rv = client.get(f"/employee_attendance_detail/{seed_employee['employee_id']}/2025/3")
        assert rv.status_code == 200
        assert seed_employee["name"].encode() in rv.data

    def test_month_name_in_response(self, client, seed_admin, seed_employee):
        _admin_session(client, seed_admin)
        rv = client.get(f"/employee_attendance_detail/{seed_employee['employee_id']}/2025/6")
        assert rv.status_code == 200
        assert b"June 2025" in rv.data


# ═══════════════════════════════════════════════════════════════════════════════
# update_default_shift — missing fields guard and known NameError bug
# ═══════════════════════════════════════════════════════════════════════════════

class TestUpdateDefaultShift:

    def test_missing_fields_redirects_with_error(self, client, seed_admin):
        _admin_session(client, seed_admin)
        rv = client.post("/update_default_shift", data={
            "shift_start": "", "shift_half": "", "shift_end": "",
        })
        assert rv.status_code == 302
        assert "error" in rv.headers["Location"]

    def test_valid_fields_nameerror_bug_returns_500(self, client, seed_admin, monkeypatch):
        """Known bug: load_default_shift() is not imported in attendance.py (line 313).
        Without the injection, NameError → Flask 500 (not propagated to test client).
        Remove the injected helper to simulate the raw broken state."""
        import blueprints.attendance as att
        monkeypatch.delattr(att, "load_default_shift", raising=False)
        _admin_session(client, seed_admin)
        rv = client.post("/update_default_shift", data={
            "shift_start": "09:00",
            "shift_half":  "13:00",
            "shift_end":   "18:00",
        })
        assert rv.status_code == 500

    def test_valid_fields_with_mocked_load_fn_redirects(self, client, seed_admin, mocker):
        """When load_default_shift is mocked (bug fixed), the route redirects."""
        import blueprints.attendance as att_mod
        mocker.patch.object(att_mod, "load_default_shift", create=True, return_value=None)
        _admin_session(client, seed_admin)
        rv = client.post("/update_default_shift", data={
            "shift_start": "09:00",
            "shift_half":  "13:00",
            "shift_end":   "18:00",
        })
        assert rv.status_code == 302
        assert "default_saved=1" in rv.headers["Location"]
