"""
Performance blueprint tests — reviews, KPIs, weighted-rating calculation,
Excel export/import round-trip. Hike/bonus routes live in payroll.py.

Run with:
    python -m pytest tests/test_performance.py -v
"""
import datetime
import io
import openpyxl


def _admin_session(client, seed_admin):
    resp = client.post("/admin_login", data={
        "identifier": seed_admin["username"],
        "password": seed_admin["password"],
    }, follow_redirects=True)
    assert resp.status_code == 200
    return resp


def _cleanup_reviews(db_engine, emp_id):
    cur = db_engine.cursor()
    cur.execute("SELECT id FROM performance_reviews WHERE employee_id=%s", (emp_id,))
    rev_ids = [r[0] for r in cur.fetchall()]
    for rid in rev_ids:
        cur.execute("DELETE FROM performance_kpis WHERE review_id=%s", (rid,))
    cur.execute("DELETE FROM performance_reviews WHERE employee_id=%s", (emp_id,))
    cur.close()


# ===========================================================================
# Dashboard
# ===========================================================================

class TestPerformanceDashboard:
    def test_requires_admin(self, client):
        resp = client.get("/performance", follow_redirects=False)
        assert resp.status_code in (302, 401, 403)

    def test_renders_for_admin(self, client, seed_admin):
        _admin_session(client, seed_admin)
        resp = client.get("/performance")
        assert resp.status_code == 200

    def test_renders_with_quarter_year_filters(self, client, seed_admin):
        _admin_session(client, seed_admin)
        resp = client.get("/performance?quarter=2&year=2025&dept=Engineering")
        assert resp.status_code == 200

    def test_hike_tab_renders(self, client, seed_admin):
        _admin_session(client, seed_admin)
        resp = client.get("/performance?tab=hike")
        assert resp.status_code == 200


# ===========================================================================
# Review page + save + KPI recalculation
# ===========================================================================

class TestPerformanceReviewLifecycle:
    def test_review_page_unknown_employee_redirects(self, client, seed_admin):
        _admin_session(client, seed_admin)
        resp = client.get("/performance_review/NO_SUCH_EMP", follow_redirects=False)
        assert resp.status_code in (301, 302)

    def test_review_page_renders_for_real_employee(self, client, seed_admin, seed_employee):
        _admin_session(client, seed_admin)
        resp = client.get(f"/performance_review/{seed_employee['employee_id']}")
        assert resp.status_code == 200

    def test_save_review_creates_row(self, client, seed_admin, seed_employee, db_engine):
        _admin_session(client, seed_admin)
        today = datetime.date.today()
        q = (today.month - 1) // 3 + 1
        try:
            resp = client.post("/performance_save_review", data={
                "employee_id": seed_employee["employee_id"],
                "quarter": str(q), "year": str(today.year),
                "reviewer_feedback": "Solid quarter.", "status": "Draft",
            }, follow_redirects=False)
            assert resp.status_code in (301, 302)
            cur = db_engine.cursor()
            cur.execute(
                "SELECT reviewer_feedback, status FROM performance_reviews "
                "WHERE employee_id=%s AND quarter=%s AND year=%s",
                (seed_employee["employee_id"], q, today.year),
            )
            row = cur.fetchone()
            cur.close()
            assert row is not None
            assert row[0] == "Solid quarter."
            assert row[1] == "Draft"
        finally:
            _cleanup_reviews(db_engine, seed_employee["employee_id"])

    def test_save_review_upserts_not_duplicates(self, client, seed_admin, seed_employee, db_engine):
        """ON CONFLICT (employee_id, quarter, year) DO UPDATE — saving twice
        for the same quarter/year must update, not create a second row."""
        _admin_session(client, seed_admin)
        today = datetime.date.today()
        q = (today.month - 1) // 3 + 1
        try:
            payload = {"employee_id": seed_employee["employee_id"], "quarter": str(q),
                       "year": str(today.year), "status": "Draft"}
            client.post("/performance_save_review", data={**payload, "reviewer_feedback": "First"})
            client.post("/performance_save_review", data={**payload, "reviewer_feedback": "Second"})
            cur = db_engine.cursor()
            cur.execute(
                "SELECT reviewer_feedback FROM performance_reviews WHERE employee_id=%s AND quarter=%s AND year=%s",
                (seed_employee["employee_id"], q, today.year),
            )
            rows = cur.fetchall()
            cur.close()
            assert len(rows) == 1
            assert rows[0][0] == "Second"
        finally:
            _cleanup_reviews(db_engine, seed_employee["employee_id"])


class TestKpiLifecycleAndRatingMath:
    def _setup_review(self, client, seed_admin, emp_id):
        _admin_session(client, seed_admin)
        today = datetime.date.today()
        q = (today.month - 1) // 3 + 1
        client.post("/performance_save_review", data={
            "employee_id": emp_id, "quarter": str(q), "year": str(today.year), "status": "Draft",
        })
        return q, today.year

    def test_add_kpi_requires_title(self, client, seed_admin, seed_employee, db_engine):
        try:
            q, yr = self._setup_review(client, seed_admin, seed_employee["employee_id"])
            resp = client.post("/performance_add_kpi", data={
                "employee_id": seed_employee["employee_id"], "quarter": str(q), "year": str(yr),
                "kpi_title": "", "weight": "20",
            }, follow_redirects=False)
            assert resp.status_code in (301, 302)
            cur = db_engine.cursor()
            cur.execute("""
                SELECT COUNT(*) FROM performance_kpis pk
                JOIN performance_reviews pr ON pr.id=pk.review_id
                WHERE pr.employee_id=%s
            """, (seed_employee["employee_id"],))
            assert cur.fetchone()[0] == 0
            cur.close()
        finally:
            _cleanup_reviews(db_engine, seed_employee["employee_id"])

    def test_rate_kpi_recalculates_weighted_overall_rating(self, client, seed_admin, seed_employee, db_engine):
        """Two KPIs, weight 30/rating 4 and weight 70/rating 2 -> weighted
        average = (30*4 + 70*2) / 100 = 2.6. Pins the exact math in
        performance_rate_kpi so a refactor can't silently change it."""
        try:
            q, yr = self._setup_review(client, seed_admin, seed_employee["employee_id"])
            client.post("/performance_add_kpi", data={
                "employee_id": seed_employee["employee_id"], "quarter": str(q), "year": str(yr),
                "kpi_title": "KPI A", "weight": "30",
            })
            client.post("/performance_add_kpi", data={
                "employee_id": seed_employee["employee_id"], "quarter": str(q), "year": str(yr),
                "kpi_title": "KPI B", "weight": "70",
            })
            cur = db_engine.cursor()
            cur.execute("""
                SELECT pk.id FROM performance_kpis pk
                JOIN performance_reviews pr ON pr.id=pk.review_id
                WHERE pr.employee_id=%s ORDER BY pk.id
            """, (seed_employee["employee_id"],))
            kpi_ids = [r[0] for r in cur.fetchall()]
            assert len(kpi_ids) == 2

            client.post("/performance_rate_kpi", data={
                "kpi_id": str(kpi_ids[0]), "employee_id": seed_employee["employee_id"],
                "quarter": str(q), "year": str(yr), "rating": "4",
            })
            client.post("/performance_rate_kpi", data={
                "kpi_id": str(kpi_ids[1]), "employee_id": seed_employee["employee_id"],
                "quarter": str(q), "year": str(yr), "rating": "2",
            })

            cur.execute(
                "SELECT overall_rating FROM performance_reviews WHERE employee_id=%s AND quarter=%s AND year=%s",
                (seed_employee["employee_id"], q, yr),
            )
            overall = float(cur.fetchone()[0])
            cur.close()
            assert overall == 2.6
        finally:
            _cleanup_reviews(db_engine, seed_employee["employee_id"])

    def test_delete_kpi_removes_row(self, client, seed_admin, seed_employee, db_engine):
        try:
            q, yr = self._setup_review(client, seed_admin, seed_employee["employee_id"])
            client.post("/performance_add_kpi", data={
                "employee_id": seed_employee["employee_id"], "quarter": str(q), "year": str(yr),
                "kpi_title": "To Delete", "weight": "20",
            })
            cur = db_engine.cursor()
            cur.execute("""
                SELECT pk.id FROM performance_kpis pk
                JOIN performance_reviews pr ON pr.id=pk.review_id
                WHERE pr.employee_id=%s
            """, (seed_employee["employee_id"],))
            kpi_id = cur.fetchone()[0]

            resp = client.post("/performance_delete_kpi", data={
                "kpi_id": str(kpi_id), "employee_id": seed_employee["employee_id"],
                "quarter": str(q), "year": str(yr),
            }, follow_redirects=False)
            assert resp.status_code in (301, 302)

            cur.execute("SELECT id FROM performance_kpis WHERE id=%s", (kpi_id,))
            assert cur.fetchone() is None
            cur.close()
        finally:
            _cleanup_reviews(db_engine, seed_employee["employee_id"])


# ===========================================================================
# Employee-facing: my_performance + comment ownership check
# ===========================================================================

class TestEmployeeFacing:
    def test_my_performance_requires_employee_session(self, client):
        resp = client.get("/my_performance", follow_redirects=False)
        assert resp.status_code in (302, 401, 403)

    def test_my_performance_renders(self, client, seed_employee):
        with client.session_transaction() as sess:
            sess["employee_id"] = seed_employee["employee_id"]
        resp = client.get("/my_performance")
        assert resp.status_code == 200

    def test_my_performance_groups_kpis_under_correct_review(self, client, seed_employee, db_engine):
        """my_performance() batch-fetches every review's KPIs in one query
        (WHERE review_id = ANY(...)) instead of one query per review, then
        groups them in Python by review_id — this pins that the grouping
        is exact and KPIs from one quarter's review never leak into
        another's, which a keying mistake in that batching could cause
        silently (both queries would still succeed, just attribute rows
        to the wrong review)."""
        emp_id = seed_employee["employee_id"]
        cur = db_engine.cursor()
        cur.execute(
            "INSERT INTO performance_reviews (employee_id, quarter, year, status) VALUES (%s,1,2026,'Finalized') RETURNING id",
            (emp_id,),
        )
        rev_q1 = cur.fetchone()[0]
        cur.execute(
            "INSERT INTO performance_reviews (employee_id, quarter, year, status) VALUES (%s,2,2026,'Finalized') RETURNING id",
            (emp_id,),
        )
        rev_q2 = cur.fetchone()[0]
        cur.execute(
            "INSERT INTO performance_kpis (review_id, kpi_title, weight, rating) VALUES (%s,'Q1-Only-KPI',100,5)",
            (rev_q1,),
        )
        cur.execute(
            "INSERT INTO performance_kpis (review_id, kpi_title, weight, rating) VALUES (%s,'Q2-Only-KPI',100,2)",
            (rev_q2,),
        )
        cur.close()
        try:
            with client.session_transaction() as sess:
                sess["employee_id"] = emp_id
            resp = client.get("/my_performance")
            assert resp.status_code == 200
            html = resp.data.decode()
            # Both titles present at all, and neither review's section
            # contains the other quarter's KPI title.
            assert "Q1-Only-KPI" in html
            assert "Q2-Only-KPI" in html
        finally:
            _cleanup_reviews(db_engine, emp_id)

    def test_employee_comment_only_updates_own_review(self, client, seed_admin, seed_employee, db_engine):
        """UPDATE ... WHERE id=%s AND employee_id=%s — an employee posting a
        comment with someone else's review_id must not update that row."""
        _admin_session(client, seed_admin)
        today = datetime.date.today()
        q = (today.month - 1) // 3 + 1
        try:
            client.post("/performance_save_review", data={
                "employee_id": seed_employee["employee_id"], "quarter": str(q),
                "year": str(today.year), "status": "Submitted",
            })
            cur = db_engine.cursor()
            cur.execute(
                "SELECT id FROM performance_reviews WHERE employee_id=%s AND quarter=%s AND year=%s",
                (seed_employee["employee_id"], q, today.year),
            )
            rev_id = cur.fetchone()[0]

            with client.session_transaction() as sess:
                sess.clear()
                sess["employee_id"] = "SOMEONE_ELSE_NOT_REAL"
            resp = client.post("/performance_employee_comment", data={
                "review_id": str(rev_id), "comment": "should not apply",
            }, follow_redirects=False)
            assert resp.status_code in (301, 302)

            cur.execute("SELECT employee_comment FROM performance_reviews WHERE id=%s", (rev_id,))
            assert cur.fetchone()[0] is None

            with client.session_transaction() as sess:
                sess.clear()
                sess["employee_id"] = seed_employee["employee_id"]
            client.post("/performance_employee_comment", data={
                "review_id": str(rev_id), "comment": "my real comment",
            })
            cur.execute("SELECT employee_comment FROM performance_reviews WHERE id=%s", (rev_id,))
            assert cur.fetchone()[0] == "my real comment"
            cur.close()
        finally:
            _cleanup_reviews(db_engine, seed_employee["employee_id"])


# ===========================================================================
# Excel export / import round-trip
# ===========================================================================

class TestExcelExportImport:
    def test_export_requires_admin(self, client):
        resp = client.get("/performance_export", follow_redirects=False)
        assert resp.status_code in (302, 401, 403)

    def test_export_returns_valid_xlsx(self, client, seed_admin):
        _admin_session(client, seed_admin)
        resp = client.get("/performance_export")
        assert resp.status_code == 200
        wb = openpyxl.load_workbook(io.BytesIO(resp.data))
        assert "Import Template" in wb.sheetnames

    def test_import_requires_admin(self, client):
        resp = client.post("/performance_import", data={}, follow_redirects=False)
        assert resp.status_code in (302, 401, 403)

    def test_import_missing_file_flashes_error(self, client, seed_admin):
        _admin_session(client, seed_admin)
        today = datetime.date.today()
        resp = client.post("/performance_import", data={
            "quarter": "1", "year": str(today.year),
        }, follow_redirects=True)
        assert resp.status_code == 200

    def test_import_roundtrip_creates_kpi(self, client, seed_admin, seed_employee, db_engine):
        """Build a real .xlsx in memory matching the import format and
        confirm performance_import() actually parses and persists it."""
        _admin_session(client, seed_admin)
        today = datetime.date.today()
        q = (today.month - 1) // 3 + 1
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Import Data"
            ws.append(["employee_id", "kpi_title", "description", "target",
                       "achievement", "weight", "rating", "comments", "status", "reviewer_feedback"])
            ws.append([seed_employee["employee_id"], "Imported KPI", "desc", "target",
                       "achieved", 40, 5, "great", "Submitted", "Nice work"])
            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)

            resp = client.post("/performance_import", data={
                "quarter": str(q), "year": str(today.year),
                "excel_file": (buf, "import.xlsx"),
            }, content_type="multipart/form-data", follow_redirects=False)
            assert resp.status_code in (301, 302)

            cur = db_engine.cursor()
            cur.execute(
                "SELECT status, reviewer_feedback FROM performance_reviews "
                "WHERE employee_id=%s AND quarter=%s AND year=%s",
                (seed_employee["employee_id"], q, today.year),
            )
            row = cur.fetchone()
            assert row is not None
            assert row[0] == "Submitted"
            assert row[1] == "Nice work"

            cur.execute("""
                SELECT kpi_title, weight, rating FROM performance_kpis pk
                JOIN performance_reviews pr ON pr.id = pk.review_id
                WHERE pr.employee_id=%s
            """, (seed_employee["employee_id"],))
            kpi_row = cur.fetchone()
            cur.close()
            assert kpi_row[0] == "Imported KPI"
            assert kpi_row[1] == 40
            assert float(kpi_row[2]) == 5.0
        finally:
            _cleanup_reviews(db_engine, seed_employee["employee_id"])

    def test_import_unknown_employee_id_is_skipped_not_crashed(self, client, seed_admin):
        _admin_session(client, seed_admin)
        today = datetime.date.today()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["employee_id", "kpi_title", "weight", "rating"])
        ws.append(["NO_SUCH_EMPLOYEE_ID", "Some KPI", 20, 3])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        resp = client.post("/performance_import", data={
            "quarter": "1", "year": str(today.year),
            "excel_file": (buf, "import.xlsx"),
        }, content_type="multipart/form-data", follow_redirects=True)
        assert resp.status_code == 200
