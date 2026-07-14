import sys
try:
    sys.stdout.reconfigure(encoding='utf-8')
    sys.stderr.reconfigure(encoding='utf-8')
except Exception:
    pass

from flask import Flask, render_template, request, session, jsonify, redirect, url_for, flash, send_from_directory, current_app
import uuid
from werkzeug.utils import secure_filename
from urllib.parse import urlparse
import datetime
import html as _html
try:
    import face_recognition
    _face_recognition_available = True
except Exception as _fr_err:
    face_recognition = None
    _face_recognition_available = False
    print(f"⚠  face_recognition unavailable ({_fr_err}). Face features disabled.")

# Cache known face encodings by (employee_id, file_mtime) to avoid recomputing on every punch
_face_enc_cache: dict = {}

def _get_known_face_encoding(emp_id: str, face_path: str):
    """Return the cached face encoding for an employee, recomputing only when the file changes."""
    try:
        mtime = os.path.getmtime(face_path)
    except OSError:
        return None
    cached = _face_enc_cache.get(emp_id)
    if cached and cached[0] == mtime:
        return cached[1]
    img  = face_recognition.load_image_file(face_path)
    encs = face_recognition.face_encodings(img)
    enc  = encs[0] if encs else None
    _face_enc_cache[emp_id] = (mtime, enc)
    return enc
try:
    # typing.Literal was added in Python 3.8; backport it for 3.7 so webauthn imports cleanly
    import typing as _typing
    if not hasattr(_typing, "Literal"):
        from typing_extensions import Literal as _Literal
        _typing.Literal = _Literal
    import webauthn
    from webauthn.helpers.structs import (
        AuthenticatorSelectionCriteria, AuthenticatorAttachment, UserVerificationRequirement,
        ResidentKeyRequirement, PublicKeyCredentialDescriptor, AuthenticatorTransport,
        COSEAlgorithmIdentifier, AttestationConveyancePreference,
    )
    _webauthn_available = True
except Exception as _wa_err:
    webauthn = None
    _webauthn_available = False
    print(f"⚠  webauthn unavailable ({_wa_err}). Fingerprint features disabled. (Needs Python 3.8+; "
          f"runs fine in the production Podman image, which uses Python 3.11.)")
from database import get_db_connection
from qr_generator import generate_qr
import bcrypt as _bcrypt
from werkzeug.security import check_password_hash as _wz_check_pw

def generate_password_hash(pw: str, **_) -> str:
    """Hash a password with bcrypt (work factor 12)."""
    return _bcrypt.hashpw(pw.encode(), _bcrypt.gensalt(rounds=12)).decode()

def check_password_hash(pw_hash: str, pw: str) -> bool:
    """Verify a password against bcrypt or legacy werkzeug hash."""
    if not pw_hash:
        return False
    if pw_hash.startswith("$2b$") or pw_hash.startswith("$2a$"):
        try:
            return _bcrypt.checkpw(pw.encode(), pw_hash.encode())
        except Exception:
            return False
    # Legacy pbkdf2/scrypt hash from werkzeug — still valid on upgrade
    return _wz_check_pw(pw_hash, pw)
from functools import wraps
from contextlib import contextmanager
import os
import math
import re
import calendar
import psycopg2
import smtplib
import ssl
import secrets
import json
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.application import MIMEApplication
from email import encoders
import threading
import io as _io
import hashlib
import time
import base64
from werkzeug.exceptions import HTTPException
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from dotenv import load_dotenv

load_dotenv()

import logging

# ── Startup: warn if critical env vars are missing ──
_missing_env = [k for k in ("DB_HOST", "DB_USER", "DB_PASS", "DB_NAME") if not os.environ.get(k)]
if _missing_env:
    import warnings
    warnings.warn(
        f"Missing required environment variables: {', '.join(_missing_env)}. "
        "Copy .env.example to .env and fill in the values.",
        stacklevel=2
    )

from extensions import app, limiter, app_log, _allowed_origins

# ── Trusted base URL for email links (avoids Host-header injection) ───────────
# Set APP_URL=https://yourdomain.com in .env for production.
# Falls back to request.host_url only when the env var is absent (local dev).
_APP_URL = os.environ.get("APP_URL", "").rstrip("/")

def _safe_app_url() -> str:
    """Return a trusted base URL, never derived from the Host header."""
    return _APP_URL if _APP_URL else request.host_url.rstrip("/")

def _safe_redirect(dest: str, fallback: str = "/admin") -> str:
    """Validate that a redirect target is a relative path (prevents open redirect)."""
    if dest and dest.startswith("/") and not dest.startswith("//"):
        return dest
    return fallback

def _safe_referrer_redirect(referrer: str, fallback: str) -> str:
    """Like _safe_redirect, but also accepts an absolute Referer header as long
    as it points back at this same app (scheme+host), reducing it to a
    relative path first. Referer is client-supplied and can be forged by
    non-browser HTTP clients, so it's never trusted as-is."""
    if not referrer:
        return fallback
    from urllib.parse import urlparse as _urlparse
    p = _urlparse(referrer)
    if not p.scheme and not p.netloc:
        return _safe_redirect(referrer, fallback)
    if p.netloc == request.host:
        path = p.path or "/"
        return _safe_redirect(path + (("?" + p.query) if p.query else ""), fallback)
    return fallback

@app.context_processor
def inject_common_vars():
    return dict(
        shift_start=SHIFT_START.strftime("%I:%M %p"),
        shift_end=SHIFT_END.strftime("%I:%M %p"),
    )

@app.template_filter('qr_url')
def _qr_url_filter(p):
    """Normalize QR code paths — old code stored absolute OS paths; extract just static/qrcodes/<file>."""
    import re
    if not p:
        return ''
    m = re.search(r'static[/\\]qrcodes[/\\]([^/\\]+\.png)', str(p))
    return f'static/qrcodes/{m.group(1)}' if m else str(p)

# /favicon.ico and /healthz are served by blueprints/health.py


# Jinja2 filter: handles both datetime.time and datetime.timedelta.
# psycopg2 returns TIME columns as datetime.time (hits the strftime branch
# below); the timedelta branch is a defensive fallback kept from when this
# ran against mysql-connector, which returned TIME columns as timedelta.
@app.template_filter("fmt_time")
def fmt_time_filter(value):
    if value is None:
        return "--"
    if isinstance(value, str):
        return value
    if hasattr(value, "strftime"):
        return value.strftime("%H:%M:%S")
    # timedelta fallback — see comment above
    total = int(value.total_seconds())
    return "{:02d}:{:02d}:{:02d}".format(total // 3600, (total % 3600) // 60, total % 60)

# Templates that need arithmetic on a TIME value (elapsed-time math, HH/MM/SS
# breakdowns) used to rely on mysql-connector's timedelta.seconds. psycopg2
# returns datetime.time instead, which has no .seconds — this filter gives
# templates a type-agnostic "total seconds" so that math still works.
@app.template_filter("time_seconds")
def time_seconds_filter(value):
    if value is None:
        return 0
    if hasattr(value, "hour"):
        return value.hour * 3600 + value.minute * 60 + value.second
    return int(value.total_seconds())

# ---------------- CONFIG ----------------
# secret_key, session cookie flags, and PERMANENT_SESSION_LIFETIME are
# authoritative in extensions.py. Do not duplicate them here.

UPLOAD_FOLDER = os.path.join(os.path.dirname(os.path.abspath(__file__)), "dataset")
app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER

if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

# ---------------- CSRF PROTECTION ----------------
_EMP_ID_RE = re.compile(r'^[A-Za-z0-9_\-]+$')

def _csrf_token():
    if "_csrf" not in session:
        session["_csrf"] = secrets.token_hex(32)
    return session["_csrf"]

app.jinja_env.globals["csrf_token"] = _csrf_token
app.jinja_env.globals["timedelta"]  = datetime.timedelta

@app.context_processor
def inject_companies_context():
    """Inject active company and companies list into every admin template."""
    if not session.get("admin_logged_in"):
        return {}
    try:
        db = get_db_connection(); cur = db.cursor(buffered=True)
        cur.execute("""
            SELECT id, name, COALESCE(code,''), COALESCE(pin,'')
            FROM companies ORDER BY name
        """)
        rows = cur.fetchall()
        cur.close(); db.close()
        active_cid = session.get("active_company_id")
        active_company = None
        for r in rows:
            if r[0] == active_cid:
                active_company = {"id": r[0], "name": r[1], "code": r[2]}
                break
        return {
            "all_companies": [{"id": r[0], "name": r[1], "code": r[2], "has_pin": bool(r[3])} for r in rows],
            "active_company": active_company,
        }
    except Exception:
        return {"all_companies": [], "active_company": None}

@app.context_processor
def inject_overdue_onboardings():
    if not session.get("admin_logged_in"):
        return {}
    try:
        db = get_db_connection(); cur = db.cursor()
        today = datetime.date.today()
        cur.execute("""
            SELECT COUNT(*) FROM employee_onboarding
            WHERE status != 'Completed' AND due_date < %s
        """, (today,))
        count = cur.fetchone()[0]
        cur.close(); db.close()
        return {"overdue_onboardings": count}
    except Exception:
        return {"overdue_onboardings": 0}

_SESSION_MAX_AGE = 8 * 3600  # 8 hours absolute — stolen cookie cannot be used indefinitely

@app.before_request
def _enforce_session_lifetime():
    """Expire sessions that are older than the absolute max age, regardless of activity."""
    if request.path.startswith("/static/") or request.path == "/healthz":
        return
    created = session.get("_session_created")
    if created and (time.time() - created) > _SESSION_MAX_AGE:
        session.clear()
        if request.path.startswith("/api/"):
            from flask import jsonify as _jfy
            return _jfy({"ok": False, "msg": "Session expired. Please log in again."}), 401
        flash("Your session expired. Please log in again.", "warning")
        return redirect(url_for("auth.admin_login"))


@app.before_request
def _enforce_csrf():
    if request.method != "POST":
        return
    if current_app.testing:
        return  # CSRF disabled in test mode; Bearer-token tests handle auth separately
    if request.path.startswith("/api/"):
        return  # API routes use Bearer-token auth — no session/CSRF needed
    # NOTE: We intentionally do NOT skip JSON requests here. The auto-inject
    # script (_inject_csrf_meta) adds X-CSRF-Token to every fetch() call, so
    # legitimate JSON POSTs from the web UI already carry the token.
    # Skipping CSRF for is_json would allow XSS payloads to forge state-changing
    # JSON requests without a token.
    token = session.get("_csrf")
    submitted = (request.form.get("_csrf_token")
                 or request.headers.get("X-CSRF-Token")
                 or request.headers.get("X-CSRFToken"))
    if not token or not submitted or not secrets.compare_digest(str(token), str(submitted)):
        # Browser form submissions: redirect to login so the user gets a fresh session+token
        if request.accept_mimetypes.accept_html and not request.headers.get("X-Requested-With"):
            flash("Your session expired. Please log in again.", "warning")
            login_url = url_for("auth.employee_login") if request.path.startswith("/employee") or "employee" in request.path else url_for("auth.admin_login")
            return redirect(login_url)
        return jsonify({"ok": False, "msg": "Session expired. Please refresh and try again."}), 403


@app.before_request
def _resolve_tenant():
    """Determine the tenant database for this request and store it in g.tenant_db."""
    from flask import g as _g

    # Skip for static files and special paths
    skip_prefixes = ("/static/", "/healthz", "/create_org", "/super_admin")
    if any(request.path.startswith(p) for p in skip_prefixes):
        return

    # 1. Already resolved in this session
    if session.get("tenant_db"):
        _g.tenant_db = session["tenant_db"]
        return

    # 2. Subdomain resolution
    host = request.host.split(":")[0]  # strip port
    parts = host.split(".")
    if len(parts) >= 3:
        subdomain = parts[0]
        try:
            from database import get_master_db
            conn = get_master_db()
            cur = conn.cursor(buffered=True)
            cur.execute(
                "SELECT db_name FROM tenants WHERE subdomain=%s AND status='active'",
                (subdomain,)
            )
            row = cur.fetchone()
            cur.close(); conn.close()
            if row:
                _g.tenant_db = row[0]
                session["tenant_db"] = row[0]
                return
        except Exception:
            pass  # master DB not yet set up — fall through to default

    # 3. Default single-tenant fallback
    _g.tenant_db = os.environ.get("DB_NAME", "employee_attendance")


_CSRF_HEAD_RE    = re.compile(rb'</head>', re.IGNORECASE)
_CSRF_BODY_RE    = re.compile(rb'</body>', re.IGNORECASE)
# Matches <script>/<style> tags without a nonce — used to inject CSP nonces
_SCRIPT_TAG_RE   = re.compile(rb'<script(?!\s[^>]*\bnonce\b)(?=[\s>])', re.IGNORECASE)
_STYLE_TAG_RE    = re.compile(rb'<style(?!\s[^>]*\bnonce\b)(?=[\s>])',  re.IGNORECASE)
# Capture inline event-handler values for dynamic CSP sha256 hash generation.
# Two patterns: double-quoted and single-quoted attribute values.
_CSP_EV_DQ = re.compile(
    rb'\bon(?:animationend|blur|change|click|contextmenu|copy|cut|dblclick|drag|dragend'
    rb'|dragenter|dragleave|dragover|dragstart|drop|error|focus|input|invalid|keydown|keypress'
    rb'|keyup|load|mousedown|mousemove|mouseout|mouseover|mouseup|paste|pointerdown'
    rb'|pointermove|pointerup|reset|scroll|select|submit|touchend|touchmove|touchstart'
    rb'|transitionend|wheel)\s*=\s*"([^"]*)"',
    re.IGNORECASE,
)
_CSP_EV_SQ = re.compile(
    rb"\bon(?:animationend|blur|change|click|contextmenu|copy|cut|dblclick|drag|dragend"
    rb"|dragenter|dragleave|dragover|dragstart|drop|error|focus|input|invalid|keydown|keypress"
    rb"|keyup|load|mousedown|mousemove|mouseout|mouseover|mouseup|paste|pointerdown"
    rb"|pointermove|pointerup|reset|scroll|select|submit|touchend|touchmove|touchstart"
    rb"|transitionend|wheel)\s*=\s*'([^']*)'",
    re.IGNORECASE,
)
_CSRF_SCRIPT  = (
    b'<script>(function(){'
    b'var m=document.querySelector(\'meta[name="csrf-token"]\');'
    b'if(!m)return;'
    b'window._csrfToken=function(){return m.content;};'
    b'var _of=window.fetch;'
    b'window.fetch=function(u,o){'
    b'o=o||{};'
    b'var mt=(o.method||"GET").toUpperCase();'
    b'if(mt==="POST"||mt==="PUT"||mt==="PATCH"||mt==="DELETE"){'
    b'if(o.headers instanceof Headers){'
    b'if(!o.headers.has("X-CSRF-Token"))o.headers.set("X-CSRF-Token",m.content);'
    b'}else{o.headers=Object.assign({},o.headers||{});'
    b'if(!o.headers["X-CSRF-Token"])o.headers["X-CSRF-Token"]=m.content;}}'
    b'return _of.call(this,u,o);};'
    b'document.addEventListener("DOMContentLoaded",function(){'
    b'document.querySelectorAll("form").forEach(function(f){'
    b'if(f.method.toLowerCase()==="post"&&!f.querySelector(\'[name="_csrf_token"]\')){'
    b'var i=document.createElement("input");'
    b'i.type="hidden";i.name="_csrf_token";i.value=m.content;'
    b'f.prepend(i);}});});})();</script>'
)

@app.before_request
def _set_csp_nonce():
    from flask import g
    g.csp_nonce = secrets.token_urlsafe(16)

_SETTINGS_PATHS = {"/settings", "/setup", "/admin_set_recovery_email",
                   "/save_security_settings", "/toggle_auth_feature",
                   "/toggle_fingerprint", "/save_company_code", "/save_geo_settings",
                   "/save_company_info", "/toggle_feature"}

@app.after_request
def _security_headers(response):
    from flask import g
    response.headers["X-Frame-Options"] = "DENY"
    response.headers["X-Content-Type-Options"] = "nosniff"
    response.headers["Referrer-Policy"] = "strict-origin-when-cross-origin"
    response.headers["Permissions-Policy"] = "camera=(self), microphone=(), geolocation=(self)"
    response.headers["Cross-Origin-Opener-Policy"] = "same-origin"
    response.headers["Cross-Origin-Resource-Policy"] = "same-origin"
    response.headers["Server"] = "AttendanceApp"
    if request.scheme == "https":
        response.headers["Strict-Transport-Security"] = "max-age=31536000; includeSubDomains; preload"
    ct = response.content_type or ""
    if "text/html" in ct:
        nonce = getattr(g, "csp_nonce", "")
        # _inject_csrf_meta runs before this hook (Flask reverses registration order),
        # so response.get_data() is the final HTML with nonces already injected.
        # Scan for inline event-handler values and compute sha256 hashes so they
        # pass CSP without needing 'unsafe-inline'.
        try:
            data = response.get_data()
            _ev_hashes: set = set()
            for _pat in (_CSP_EV_DQ, _CSP_EV_SQ):
                for _m in _pat.finditer(data):
                    _body = _html.unescape(_m.group(1).decode("utf-8", errors="replace"))
                    _ev_hashes.add(
                        "'sha256-" + base64.b64encode(
                            hashlib.sha256(_body.encode("utf-8")).digest()
                        ).decode() + "'"
                    )
        except Exception:
            _ev_hashes = set()
        _unsafe_hashes = " 'unsafe-hashes'" if _ev_hashes else ""
        _hash_src = (" " + " ".join(sorted(_ev_hashes))) if _ev_hashes else ""
        response.headers["Content-Security-Policy"] = (
            "default-src 'self'; "
            f"script-src 'self' 'nonce-{nonce}'{_unsafe_hashes}{_hash_src}; "
            f"style-src-elem 'self' 'nonce-{nonce}'; "
            "style-src-attr 'unsafe-inline'; "
            f"style-src 'self' 'nonce-{nonce}'; "
            "img-src 'self' data: blob:; "
            "font-src 'self' data:; "
            "connect-src 'self'; "
            "frame-ancestors 'none'; "
            "report-uri /csp-report;"
        )
    return response


@app.route("/csp-report", methods=["POST"])
def csp_report():
    """Receives Content-Security-Policy violation reports from browsers."""
    try:
        report = request.get_json(force=True, silent=True) or {}
        violation = report.get("csp-report", report)
        app_log.warning(
            "CSP violation",
            extra={
                "blocked_uri": violation.get("blocked-uri", ""),
                "violated_directive": violation.get("violated-directive", ""),
                "document_uri": violation.get("document-uri", ""),
                "source_file": violation.get("source-file", ""),
            },
        )
    except Exception:
        pass
    return "", 204


@app.after_request
def _bust_settings_cache(response):
    if request.method == "POST" and request.path in _SETTINGS_PATHS:
        invalidate_settings_cache()
    return response

@app.after_request
def _inject_csrf_meta(response):
    """Inject CSRF meta tag and auto-inject script into every HTML page."""
    if response.status_code >= 300 or not response.content_type.startswith("text/html"):
        return response
    try:
        from flask import g
        token = _csrf_token()
        meta  = f'<meta name="csrf-token" content="{token}" />'.encode()
        data  = response.get_data()
        data  = _CSRF_HEAD_RE.sub(meta + b'</head>', data, count=1)
        data  = _CSRF_BODY_RE.sub(_CSRF_SCRIPT + b'</body>', data, count=1)
        nonce = getattr(g, "csp_nonce", None)
        if nonce:
            nb = nonce.encode()
            data = _SCRIPT_TAG_RE.sub(b'<script nonce="' + nb + b'"', data)
            data = _STYLE_TAG_RE.sub(b'<style nonce="' + nb + b'"', data)
        response.set_data(data)
    except Exception:
        pass
    return response

# ---------------- AUDIT LOGGING ----------------
def _audit(action, table=None, record_id=None, detail=None):
    """Write one row to audit_logs. Never raises — audit must never break main flow."""
    try:
        actor      = session.get("admin_username") or session.get("employee_id") or "system"
        actor_type = "admin" if session.get("admin_logged_in") else "employee"
        ip         = request.remote_addr or ""
        db = get_db_connection(); cursor = db.cursor()
        cursor.execute("""INSERT INTO audit_logs (actor, actor_type, action, target_table, target_id, detail, ip_address)
                          VALUES (%s,%s,%s,%s,%s,%s,%s)""",
                       (actor, actor_type, action, table, str(record_id) if record_id is not None else None, detail, ip))
        db.commit(); cursor.close(); db.close()
    except Exception:
        pass

# ---------------- MALWARE SCANNING (ClamAV) ----------------
try:
    import clamd as _clamd_lib
    _clamav_available = True
except ImportError:
    _clamd_lib = None
    _clamav_available = False

_CLAMAV_HOST = os.environ.get("CLAMAV_HOST", "clamav")
_CLAMAV_PORT = int(os.environ.get("CLAMAV_PORT", "3310"))
_MALWARE_SCAN_ENABLED = os.environ.get("MALWARE_SCAN_ENABLED", "true").strip().lower() not in ("false", "0", "no")

def _scan_for_malware(file_storage):
    """Scan an uploaded file with ClamAV before it's saved. Returns (is_clean, error_msg).
    Fails closed (rejects the upload) in production if the scanner is unavailable
    or unreachable; fails open with a logged warning in development, so a missing
    local ClamAV instance doesn't block day-to-day dev work.

    Set MALWARE_SCAN_ENABLED=false to turn this off deliberately (e.g. a
    memory-constrained deployment that can't run ClamAV) — that's a clean
    skip, not a failure, so it doesn't trigger the fail-closed behavior
    below and permanently block uploads."""
    if not _MALWARE_SCAN_ENABLED:
        return True, None
    _dev = os.environ.get("APP_ENV", "production") == "development"
    if not _clamav_available:
        app_log.error("clamd package not installed — malware scanning skipped")
        return (True, None) if _dev else (False, "Malware scanning is unavailable — upload rejected.")
    try:
        cd = _clamd_lib.ClamdNetworkSocket(host=_CLAMAV_HOST, port=_CLAMAV_PORT, timeout=15)
        pos = file_storage.stream.tell()
        file_storage.stream.seek(0)
        result = cd.instream(file_storage.stream)
        file_storage.stream.seek(pos)
        status, signature = result.get("stream", (None, None))
        if status == "FOUND":
            app_log.warning("Malware detected in upload %r: %s", file_storage.filename, signature)
            return False, "This file was flagged by malware scanning and cannot be uploaded."
        return True, None
    except Exception as _e:
        app_log.error("ClamAV scan failed (%s): %s", type(_e).__name__, _e)
        return (True, None) if _dev else (False, "File could not be scanned for malware — please try again shortly.")


# ---------------- FILE UPLOAD VALIDATION ----------------
_ALLOWED_MIME_MAP = {
    'pdf':  {'application/pdf'},
    'jpg':  {'image/jpeg'},
    'jpeg': {'image/jpeg'},
    'png':  {'image/png'},
    'doc':  {'application/msword'},
    'docx': {'application/vnd.openxmlformats-officedocument.wordprocessingml.document'},
    'xls':  {'application/vnd.ms-excel'},
    'xlsx': {'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'},
}
_MAX_DOC_SIZE_MB = 10

def _validate_upload(file_storage, allowed_exts=None):
    """Returns (ok, error_message). Checks extension + MIME type + size."""
    if not file_storage or not file_storage.filename:
        return False, "No file selected."
    ext = file_storage.filename.rsplit('.', 1)[-1].lower() if '.' in file_storage.filename else ''
    if allowed_exts and ext not in allowed_exts:
        return False, f"File type .{ext} not allowed. Allowed: {', '.join(sorted(allowed_exts))}"
    ct = (file_storage.content_type or '').split(';')[0].strip().lower()
    if ct and ext in _ALLOWED_MIME_MAP and ct not in _ALLOWED_MIME_MAP[ext]:
        return False, f"File content does not match its extension."
    # Check magic bytes for PDF and images
    header = file_storage.stream.read(8)
    file_storage.stream.seek(0)
    if ext == 'pdf' and not header.startswith(b'%PDF'):
        return False, "Invalid PDF file."
    if ext == 'png' and not header.startswith(b'\x89PNG'):
        return False, "Invalid PNG file."
    if ext in ('jpg', 'jpeg') and not header.startswith(b'\xff\xd8'):
        return False, "Invalid JPEG file."
    file_storage.stream.seek(0, 2)
    size_mb = file_storage.stream.tell() / (1024 * 1024)
    file_storage.stream.seek(0)
    if size_mb > _MAX_DOC_SIZE_MB:
        return False, f"File too large ({size_mb:.1f} MB). Maximum: {_MAX_DOC_SIZE_MB} MB."
    clean, scan_err = _scan_for_malware(file_storage)
    if not clean:
        return False, scan_err
    return True, None

# ---------------- COMPANY SETTINGS (with 60-second TTL cache) ----------------
_co_cache      = {"data": None, "expires": None}
_auth_cache    = {"data": None, "expires": None}
_settings_lock = threading.Lock()
_CO_CACHE_TTL  = 60  # seconds

def _co_expired(cache):
    return cache["data"] is None or datetime.datetime.now() >= cache["expires"]

def invalidate_settings_cache():
    with _settings_lock:
        _co_cache["data"]    = None
        _auth_cache["data"]  = None

def get_company_settings():
    with _settings_lock:
        if not _co_expired(_co_cache):
            return dict(_co_cache["data"])
    try:
        db = get_db_connection()
        cursor = db.cursor(buffered=True)
        cursor.execute("SELECT company_name, company_tagline, company_logo, currency_symbol, timezone, setup_done, COALESCE(company_code,'') FROM company_settings LIMIT 1")
        row = cursor.fetchone()
        cursor.close(); db.close()
        if row:
            result = {"company_name": row[0], "company_tagline": row[1],
                      "company_logo": row[2], "currency_symbol": row[3],
                      "company_code": row[6],
                      "timezone": row[4], "setup_done": bool(row[5])}
            with _settings_lock:
                _co_cache["data"]    = result
                _co_cache["expires"] = datetime.datetime.now() + datetime.timedelta(seconds=_CO_CACHE_TTL)
            return dict(result)
    except Exception:
        pass
    return {"company_name": "My Company", "company_tagline": "Employee Attendance System",
            "company_logo": None, "currency_symbol": "₹", "timezone": "Asia/Kolkata",
            "setup_done": False, "company_code": ""}

_AUTH_CONFIG_DEFAULTS = {
    "fingerprint_enabled": False,
    "qr_enabled": True,
    "face_enabled": True,
    "location_enabled": True,
    "employee_password_auth": True,
}

def get_auth_config():
    with _settings_lock:
        if not _co_expired(_auth_cache):
            return dict(_auth_cache["data"])
    try:
        db = get_db_connection()
        cursor = db.cursor(buffered=True)
        cursor.execute("""
            SELECT COALESCE(fingerprint_enabled, 0),
                   COALESCE(qr_enabled, 1),
                   COALESCE(face_enabled, 1),
                   COALESCE(location_enabled, 1),
                   COALESCE(employee_password_auth, 1)
            FROM company_settings LIMIT 1
        """)
        row = cursor.fetchone()
        cursor.close(); db.close()
        if row:
            result = {
                "fingerprint_enabled": bool(row[0]),
                "qr_enabled":          bool(row[1]),
                "face_enabled":        bool(row[2]),
                "location_enabled":    bool(row[3]),
                "employee_password_auth": bool(row[4]),
            }
            with _settings_lock:
                _auth_cache["data"]    = result
                _auth_cache["expires"] = datetime.datetime.now() + datetime.timedelta(seconds=_CO_CACHE_TTL)
            return dict(result)
        return dict(_AUTH_CONFIG_DEFAULTS)
    except Exception:
        return dict(_AUTH_CONFIG_DEFAULTS)

def get_fingerprint_enabled():
    return get_auth_config()["fingerprint_enabled"]

# ── Per-company feature settings ──────────────────────────────────────────────
def _read_global_features():
    """Read global company_settings feature flags as dict."""
    try:
        db = get_db_connection(); cur = db.cursor(buffered=True)
        cur.execute("""
            SELECT face_auth_enabled, geo_enabled, COALESCE(geo_radius,300), qr_enabled,
                   pin_enabled, COALESCE(fingerprint_enabled,0), COALESCE(biometric_enabled,0),
                   COALESCE(notify_leave,1), COALESCE(notify_payslip,1),
                   COALESCE(notify_resignation,1), COALESCE(notify_doc_expiry,1),
                   COALESCE(session_timeout,30),
                   COALESCE(late_deduction_pct,10), COALESCE(half_day_deduction_pct,50),
                   COALESCE(grace_minutes,15), COALESCE(holiday_pay,'paid'),
                   COALESCE(leave_pay,'exclude'),
                   COALESCE(shift_start,'09:00:00'), COALESCE(shift_half,'13:00:00'),
                   COALESCE(shift_end,'18:00:00')
            FROM company_settings LIMIT 1
        """)
        r = cur.fetchone(); cur.close(); db.close()
        if r:
            return {
                "face_auth_enabled": bool(r[0]), "geo_enabled": bool(r[1]),
                "geo_radius": r[2], "qr_enabled": bool(r[3]), "pin_enabled": bool(r[4]),
                "fingerprint_enabled": bool(r[5]), "biometric_enabled": bool(r[6]),
                "notify_leave": bool(r[7]), "notify_payslip": bool(r[8]),
                "notify_resignation": bool(r[9]), "notify_doc_expiry": bool(r[10]),
                "session_timeout": r[11],
                "late_deduction_pct": float(r[12]), "half_day_deduction_pct": float(r[13]),
                "grace_minutes": int(r[14]), "holiday_pay": r[15], "leave_pay": r[16],
                "shift_start": r[17], "shift_half": r[18], "shift_end": r[19],
            }
    except Exception:
        pass
    return {
        "face_auth_enabled": True, "geo_enabled": False, "geo_radius": 300,
        "qr_enabled": True, "pin_enabled": True, "fingerprint_enabled": False,
        "biometric_enabled": False, "notify_leave": True, "notify_payslip": True,
        "notify_resignation": True, "notify_doc_expiry": True, "session_timeout": 30,
        "late_deduction_pct": 10.0, "half_day_deduction_pct": 50.0, "grace_minutes": 15,
        "holiday_pay": "paid", "leave_pay": "exclude",
        "shift_start": "09:00:00", "shift_half": "13:00:00", "shift_end": "18:00:00",
    }

def get_co_features(company_id=None):
    """Return feature settings for a company, falling back to global defaults."""
    if not company_id:
        return _read_global_features()
    try:
        db = get_db_connection(); cur = db.cursor(buffered=True)
        cur.execute("""
            SELECT face_auth_enabled, geo_enabled, geo_radius, qr_enabled,
                   pin_enabled, fingerprint_enabled, biometric_enabled,
                   notify_leave, notify_payslip, notify_resignation, notify_doc_expiry,
                   session_timeout, late_deduction_pct, half_day_deduction_pct,
                   grace_minutes, holiday_pay, leave_pay, shift_start, shift_half, shift_end
            FROM company_feature_settings WHERE company_id=%s
        """, (company_id,))
        r = cur.fetchone(); cur.close(); db.close()
        if r:
            return {
                "face_auth_enabled": bool(r[0]), "geo_enabled": bool(r[1]),
                "geo_radius": r[2], "qr_enabled": bool(r[3]), "pin_enabled": bool(r[4]),
                "fingerprint_enabled": bool(r[5]), "biometric_enabled": bool(r[6]),
                "notify_leave": bool(r[7]), "notify_payslip": bool(r[8]),
                "notify_resignation": bool(r[9]), "notify_doc_expiry": bool(r[10]),
                "session_timeout": r[11],
                "late_deduction_pct": float(r[12]), "half_day_deduction_pct": float(r[13]),
                "grace_minutes": int(r[14]), "holiday_pay": r[15], "leave_pay": r[16],
                "shift_start": r[17], "shift_half": r[18], "shift_end": r[19],
            }
    except Exception:
        pass
    return _read_global_features()

_VALID_CFS_COLS = frozenset({
    "face_auth_enabled", "geo_enabled", "geo_radius", "qr_enabled", "pin_enabled",
    "fingerprint_enabled", "biometric_enabled", "notify_leave", "notify_payslip",
    "notify_resignation", "notify_doc_expiry", "session_timeout",
    "late_deduction_pct", "half_day_deduction_pct", "grace_minutes",
    "shift_start", "shift_half", "shift_end", "holiday_pay", "leave_pay",
})

def _upsert_co_feature(company_id, field, value):
    """Insert or update a single field in company_feature_settings."""
    if not company_id:
        return
    if field not in _VALID_CFS_COLS:
        app_log.error("_upsert_co_feature: rejected unknown column %s", field)
        return
    try:
        db = get_db_connection(); cur = db.cursor(buffered=True)
        cur.execute(f"""
            INSERT INTO company_feature_settings (company_id, {field})
            VALUES (%s, %s)
            ON CONFLICT (company_id) DO UPDATE SET {field}=EXCLUDED.{field}
        """, (company_id, value))
        db.commit(); cur.close(); db.close()
    except Exception:
        pass

def _upsert_co_features(company_id, fields_dict):
    """Insert or update multiple fields in company_feature_settings."""
    if not company_id or not fields_dict:
        return
    if not all(k in _VALID_CFS_COLS for k in fields_dict.keys()):
        app_log.error("_upsert_co_features: rejected unknown columns %s", list(fields_dict.keys()))
        return
    try:
        cols   = ", ".join(fields_dict.keys())
        vals   = list(fields_dict.values())
        placeholders = ", ".join(["%s"] * len(vals))
        updates = ", ".join(f"{k}=EXCLUDED.{k}" for k in fields_dict.keys())
        db = get_db_connection(); cur = db.cursor(buffered=True)
        cur.execute(f"""
            INSERT INTO company_feature_settings (company_id, {cols})
            VALUES (%s, {placeholders})
            ON CONFLICT (company_id) DO UPDATE SET {updates}
        """, [company_id] + vals)
        db.commit(); cur.close(); db.close()
    except Exception:
        pass

@app.context_processor
def inject_company():
    return {"co": get_company_settings()}

# Office location — read from .env so no restart needed for coord changes
OFFICE_LAT = float(os.environ.get("OFFICE_LAT", "17.494664737165042"))
OFFICE_LON = float(os.environ.get("OFFICE_LON", "78.40496618113566"))
OFFICE_RADIUS_M = 300   # metres — 300 m radius as per policy

# Shift timings (overridden by DB on startup via load_default_shift())
SHIFT_START = datetime.time(9, 0)    # Full Day Login cutoff
SHIFT_HALF  = datetime.time(13, 0)   # Half Day threshold
SHIFT_END   = datetime.time(18, 0)   # Full Day Logout cutoff

def load_default_shift():
    global SHIFT_START, SHIFT_HALF, SHIFT_END
    try:
        db = get_db_connection()
        cur = db.cursor(buffered=True)
        cur.execute("SELECT shift_start, shift_half, shift_end FROM company_settings LIMIT 1")
        row = cur.fetchone()
        cur.close(); db.close()
        if row and row[0]:
            def _to_time(v):
                if isinstance(v, datetime.timedelta):
                    total = int(v.total_seconds())
                    return datetime.time(total // 3600, (total % 3600) // 60)
                if isinstance(v, datetime.time):
                    return v
                return datetime.time(9, 0)
            SHIFT_START = _to_time(row[0])
            SHIFT_HALF  = _to_time(row[1])
            SHIFT_END   = _to_time(row[2])
    except Exception:
        pass

# Deduction rates and salary rules (loaded from DB on startup)
LATE_DEDUCTION_RATE = 0.10
HALF_DAY_RATE       = 0.50
GRACE_MINUTES       = 15       # grace period after shift start for full-day login
HOLIDAY_PAY         = 'paid'   # 'paid' = full day pay | 'unpaid' = no pay
LEAVE_PAY           = 'exclude' # 'exclude' = not a working day | 'absent' = count as absent

def load_salary_rules():
    global LATE_DEDUCTION_RATE, HALF_DAY_RATE, GRACE_MINUTES, HOLIDAY_PAY, LEAVE_PAY
    try:
        db     = get_db_connection()
        cursor = db.cursor(buffered=True)
        cursor.execute(
            "SELECT COALESCE(late_deduction_pct,10), COALESCE(half_day_deduction_pct,50), "
            "       COALESCE(grace_minutes,15), COALESCE(holiday_pay,'paid'), COALESCE(leave_pay,'exclude') "
            "FROM company_settings LIMIT 1"
        )
        row = cursor.fetchone()
        cursor.close(); db.close()
        if row:
            LATE_DEDUCTION_RATE = float(row[0]) / 100.0
            HALF_DAY_RATE       = float(row[1]) / 100.0
            GRACE_MINUTES       = int(row[2])
            HOLIDAY_PAY         = str(row[3])
            LEAVE_PAY           = str(row[4])
    except Exception:
        pass

def load_deduction_rates():
    load_salary_rules()

with app.app_context():
    try:
        load_default_shift()
        load_salary_rules()
    except Exception:
        pass

# ---------------- IMAGE UPLOAD VALIDATION ----------------
_ALLOWED_IMG_EXT  = {'.jpg', '.jpeg', '.png', '.webp', '.bmp'}
_ALLOWED_IMG_MIME = {'image/jpeg', 'image/png', 'image/webp', 'image/bmp', 'image/gif'}

_MAX_PHOTO_SIZE_MB = 5
_IMG_MAGIC = {
    '.jpg':  (b'\xff\xd8',),
    '.jpeg': (b'\xff\xd8',),
    '.png':  (b'\x89PNG',),
    '.webp': (b'RIFF',),
    '.bmp':  (b'BM',),
}

def _validate_image_file(file):
    """Return (ok, error_msg). Checks extension, MIME type, magic bytes, and size."""
    if not file or not file.filename:
        return False, "No file selected."
    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in _ALLOWED_IMG_EXT:
        return False, f"Invalid file type '{ext}'. Only JPG, PNG, WEBP or BMP allowed."
    ct = (file.content_type or "").lower().split(";")[0].strip()
    if ct and ct not in _ALLOWED_IMG_MIME:
        return False, f"Invalid content type '{ct}'. Only image files accepted."
    header = file.stream.read(8)
    file.stream.seek(0)
    for magic in _IMG_MAGIC.get(ext, ()):
        if not header.startswith(magic):
            return False, "File content does not match its extension. Upload a real image."
    file.stream.seek(0, 2)
    size_mb = file.stream.tell() / (1024 * 1024)
    file.stream.seek(0)
    if size_mb > _MAX_PHOTO_SIZE_MB:
        return False, f"Photo too large ({size_mb:.1f} MB). Maximum: {_MAX_PHOTO_SIZE_MB} MB."
    clean, scan_err = _scan_for_malware(file)
    if not clean:
        return False, scan_err
    return True, ""


# ── PII Encryption ────────────────────────────────────────────────
# Set ENCRYPTION_KEY in .env: python -c "from cryptography.fernet import Fernet; print(Fernet.generate_key().decode())"
from cryptography.fernet import Fernet, InvalidToken as _FernetInvalid

_ENCRYPTION_KEY = os.environ.get("ENCRYPTION_KEY", "").encode()
_fernet = None
if _ENCRYPTION_KEY:
    try:
        _fernet = Fernet(_ENCRYPTION_KEY)
    except Exception:
        app_log.critical(
            "ENCRYPTION_KEY is set but invalid — PII fields will be stored in plaintext. "
            "Regenerate with: python -c \"from cryptography.fernet import Fernet; print(Fernet.generate_key().decode())\""
        )
else:
    if os.environ.get("APP_ENV", "production") != "development":
        raise RuntimeError(
            "ENCRYPTION_KEY is not set — refusing to start. "
            "Aadhaar, PAN, and bank account numbers require encryption in production. "
            "Generate a key: python -c \"from cryptography.fernet import Fernet; print(Fernet.generate_key().decode())\""
        )
    else:
        app_log.critical(
            "ENCRYPTION_KEY is not set — PII fields will be stored in plaintext. "
            "Generate a key: python -c \"from cryptography.fernet import Fernet; print(Fernet.generate_key().decode())\""
        )

def encrypt_pii(value: str) -> str:
    """Encrypt a PII string. Returns original value if encryption not configured."""
    if not value or not _fernet:
        return value
    return _fernet.encrypt(value.encode()).decode()

def decrypt_pii(value: str) -> str:
    """Decrypt a PII string. Returns original value if decryption fails (handles legacy plaintext)."""
    if not value or not _fernet:
        return value
    try:
        return _fernet.decrypt(value.encode()).decode()
    except (_FernetInvalid, Exception):
        return value  # legacy plaintext — return as-is


# ---------------- DB CONTEXT MANAGER ----------------
def _hash_token(token: str) -> str:
    """SHA-256 hash a Bearer token before storing/comparing in DB."""
    return hashlib.sha256(token.encode()).hexdigest()

_LOGIN_MAX_ATTEMPTS = 5
_LOGIN_LOCKOUT_MINUTES = 15

def _check_login_lockout(identifier: str, attempt_type: str = "admin"):
    """Return (is_locked, locked_until_str). Raises nothing."""
    try:
        with _db() as (cur, _):
            cur.execute(
                "SELECT locked_until FROM login_attempts WHERE identifier=%s AND attempt_type=%s",
                (identifier, attempt_type)
            )
            row = cur.fetchone()
        if row and row[0] and row[0] > datetime.datetime.now():
            return True, row[0].strftime("%H:%M")
    except Exception:
        pass
    return False, None

def _record_login_failure(identifier: str, attempt_type: str = "admin"):
    """Increment failure counter; lock account after _LOGIN_MAX_ATTEMPTS."""
    try:
        with _db() as (cur, conn):
            cur.execute(
                "INSERT INTO login_attempts (identifier, attempt_type, failed_count, last_attempt) "
                "VALUES (%s, %s, 1, NOW()) "
                "ON CONFLICT (identifier, attempt_type) DO UPDATE SET "
                "failed_count=login_attempts.failed_count+1, last_attempt=NOW()",
                (identifier, attempt_type)
            )
            conn.commit()
            cur.execute(
                "SELECT failed_count FROM login_attempts WHERE identifier=%s AND attempt_type=%s",
                (identifier, attempt_type)
            )
            row = cur.fetchone()
            if row and row[0] >= _LOGIN_MAX_ATTEMPTS:
                lockout_until = datetime.datetime.now() + datetime.timedelta(minutes=_LOGIN_LOCKOUT_MINUTES)
                cur.execute(
                    "UPDATE login_attempts SET locked_until=%s WHERE identifier=%s AND attempt_type=%s",
                    (lockout_until, identifier, attempt_type)
                )
                conn.commit()
    except Exception:
        pass

def _clear_login_failures(identifier: str, attempt_type: str = "admin"):
    """Reset failure counter on successful login."""
    try:
        with _db() as (cur, conn):
            cur.execute(
                "DELETE FROM login_attempts WHERE identifier=%s AND attempt_type=%s",
                (identifier, attempt_type)
            )
            conn.commit()
    except Exception:
        pass

@contextmanager
def _db():
    """Open a DB connection + buffered cursor; always close both on exit."""
    conn   = get_db_connection()
    cursor = conn.cursor(buffered=True)
    try:
        yield cursor, conn
    finally:
        try:  cursor.close()
        except Exception as _e: app_log.debug("cursor.close() failed: %s", _e)
        try:  conn.close()
        except Exception as _e: app_log.debug("conn.close() failed: %s", _e)

# ---------------- DB MIGRATION ----------------
# Trigger function backing every `... ON UPDATE CURRENT_TIMESTAMP`-style
# column from the old MySQL schema — Postgres has no column-level
# equivalent, so each such table gets a BEFORE UPDATE trigger calling this.
_UPDATED_AT_TRIGGER_FN = """
    CREATE OR REPLACE FUNCTION _set_updated_at() RETURNS TRIGGER AS $$
    BEGIN
        NEW.updated_at = CURRENT_TIMESTAMP;
        RETURN NEW;
    END;
    $$ LANGUAGE plpgsql;
"""

def _attach_updated_at_trigger(cursor, table):
    cursor.execute(f'DROP TRIGGER IF EXISTS trg_{table}_updated_at ON {table}')
    cursor.execute(f"""
        CREATE TRIGGER trg_{table}_updated_at BEFORE UPDATE ON {table}
        FOR EACH ROW EXECUTE FUNCTION _set_updated_at()
    """)

def init_db():
    db = get_db_connection()
    cursor = db.cursor(buffered=True)

    cursor.execute(_UPDATED_AT_TRIGGER_FN)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS employees (
            id SERIAL PRIMARY KEY,
            employee_id VARCHAR(50) UNIQUE NOT NULL,
            name VARCHAR(100) NOT NULL,
            email VARCHAR(150) DEFAULT NULL,
            face_image VARCHAR(255),
            qr_code VARCHAR(255)
        )
    """)
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS attendance (
            id SERIAL PRIMARY KEY,
            employee_id VARCHAR(50) NOT NULL,
            date DATE NOT NULL,
            login_time TIME DEFAULT NULL,
            logout_time TIME DEFAULT NULL,
            status VARCHAR(50) DEFAULT NULL,
            logout_status VARCHAR(50) DEFAULT NULL,
            attendance_type VARCHAR(50) DEFAULT NULL,
            UNIQUE (employee_id, date)
        )
    """)
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS holidays (
            id SERIAL PRIMARY KEY,
            date DATE UNIQUE NOT NULL,
            name VARCHAR(100) NOT NULL
        )
    """)
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS salary_config (
            id SERIAL PRIMARY KEY,
            employee_id VARCHAR(50) UNIQUE NOT NULL,
            salary_per_day DECIMAL(10,2) DEFAULT 0
        )
    """)
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS payroll_config (
            id SERIAL PRIMARY KEY,
            pf_employee_pct DECIMAL(5,2) DEFAULT 12.00,
            pf_employer_pct DECIMAL(5,2) DEFAULT 12.00,
            professional_tax DECIMAL(8,2) DEFAULT 200.00,
            tds_annual_pct DECIMAL(5,2) DEFAULT 0.00,
            pf_basic_cap DECIMAL(10,2) DEFAULT 15000.00
        )
    """)
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS admin_users (
            id SERIAL PRIMARY KEY,
            username VARCHAR(50) UNIQUE NOT NULL,
            password VARCHAR(255) NOT NULL,
            email VARCHAR(150) DEFAULT NULL,
            reset_token VARCHAR(64) DEFAULT NULL,
            reset_token_expiry TIMESTAMP DEFAULT NULL
        )
    """)
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS email_config (
            id SERIAL PRIMARY KEY,
            smtp_host VARCHAR(150) NOT NULL,
            smtp_port INT NOT NULL DEFAULT 587,
            smtp_user VARCHAR(150) NOT NULL,
            smtp_pass VARCHAR(255) NOT NULL,
            from_name VARCHAR(100) DEFAULT 'HR Department',
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    _attach_updated_at_trigger(cursor, "email_config")
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS leave_requests (
            id SERIAL PRIMARY KEY,
            employee_id VARCHAR(50) NOT NULL,
            leave_date DATE NOT NULL,
            reason VARCHAR(500) NOT NULL,
            status VARCHAR(20) DEFAULT 'Pending',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS notifications (
            id SERIAL PRIMARY KEY,
            recipient_type VARCHAR(20) NOT NULL CHECK (recipient_type IN ('admin', 'employee')),
            employee_id VARCHAR(50) NULL,
            title VARCHAR(255) NOT NULL,
            message TEXT NOT NULL,
            is_read BOOLEAN DEFAULT FALSE,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS resignation_requests (
            id SERIAL PRIMARY KEY,
            employee_id VARCHAR(50) NOT NULL,
            last_working_day DATE NOT NULL,
            reason TEXT NOT NULL,
            status VARCHAR(20) DEFAULT 'Pending',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS tickets (
            id SERIAL PRIMARY KEY,
            employee_id VARCHAR(50) NOT NULL,
            category VARCHAR(100) NOT NULL,
            subject VARCHAR(255) NOT NULL,
            description TEXT NOT NULL,
            priority VARCHAR(20) DEFAULT 'Medium',
            status VARCHAR(30) DEFAULT 'Open',
            admin_response TEXT DEFAULT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    _attach_updated_at_trigger(cursor, "tickets")
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS shifts (
            id SERIAL PRIMARY KEY,
            name VARCHAR(100) NOT NULL,
            start_time TIME NOT NULL,
            half_time  TIME NOT NULL,
            end_time   TIME NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS announcements (
            id SERIAL PRIMARY KEY,
            title VARCHAR(255) NOT NULL,
            content TEXT NOT NULL,
            priority VARCHAR(20) DEFAULT 'Normal' CHECK (priority IN ('Normal','Important','Urgent')),
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS break_config (
            id SERIAL PRIMARY KEY,
            break_name VARCHAR(100) NOT NULL,
            break_time TIME NOT NULL,
            duration_minutes INT NOT NULL DEFAULT 10,
            is_active SMALLINT DEFAULT 1
        )
    """)
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS incentive_goals (
            id SERIAL PRIMARY KEY,
            title VARCHAR(150) NOT NULL,
            description TEXT,
            incentive_amount DECIMAL(10,2) NOT NULL DEFAULT 0,
            is_active SMALLINT DEFAULT 1,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS employee_incentives (
            id SERIAL PRIMARY KEY,
            employee_id VARCHAR(50) NOT NULL,
            goal_id INT NOT NULL,
            month INT NOT NULL,
            year INT NOT NULL,
            amount DECIMAL(10,2) NOT NULL DEFAULT 0,
            notes TEXT,
            awarded_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS employee_experience (
            id SERIAL PRIMARY KEY,
            employee_id VARCHAR(50) NOT NULL,
            company VARCHAR(150) NOT NULL,
            designation VARCHAR(100) NOT NULL,
            from_year VARCHAR(10) NOT NULL,
            to_year VARCHAR(10) DEFAULT NULL,
            is_current SMALLINT DEFAULT 0,
            description TEXT DEFAULT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS employee_education (
            id SERIAL PRIMARY KEY,
            employee_id VARCHAR(50) NOT NULL,
            degree VARCHAR(150) NOT NULL,
            institution VARCHAR(200) NOT NULL,
            year_of_passing VARCHAR(10) DEFAULT NULL,
            percentage VARCHAR(20) DEFAULT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS leave_types (
            id SERIAL PRIMARY KEY,
            name VARCHAR(100) NOT NULL,
            annual_quota INT NOT NULL DEFAULT 12,
            is_paid SMALLINT DEFAULT 1,
            is_active SMALLINT DEFAULT 1,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS leave_balances (
            id SERIAL PRIMARY KEY,
            employee_id VARCHAR(50) NOT NULL,
            leave_type_id INT NOT NULL,
            year INT NOT NULL,
            total_days INT NOT NULL DEFAULT 0,
            used_days DECIMAL(4,1) NOT NULL DEFAULT 0,
            UNIQUE (employee_id, leave_type_id, year)
        )
    """)
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS employee_documents (
            id SERIAL PRIMARY KEY,
            employee_id VARCHAR(50) NOT NULL,
            doc_type VARCHAR(100) NOT NULL,
            original_name VARCHAR(255) NOT NULL,
            stored_name VARCHAR(255) NOT NULL,
            uploaded_by VARCHAR(20) DEFAULT 'admin',
            uploaded_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS performance_reviews (
            id SERIAL PRIMARY KEY,
            employee_id VARCHAR(50) NOT NULL,
            quarter SMALLINT NOT NULL,
            year INT NOT NULL,
            overall_rating DECIMAL(3,1) DEFAULT 0,
            reviewer_feedback TEXT,
            employee_comment TEXT,
            status VARCHAR(20) DEFAULT 'Draft',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            UNIQUE (employee_id, quarter, year)
        )
    """)
    _attach_updated_at_trigger(cursor, "performance_reviews")
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS performance_kpis (
            id SERIAL PRIMARY KEY,
            review_id INT NOT NULL,
            kpi_title VARCHAR(200) NOT NULL,
            description TEXT,
            target VARCHAR(200),
            achievement VARCHAR(200),
            weight INT DEFAULT 20,
            rating SMALLINT DEFAULT 0,
            comments TEXT
        )
    """)
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS hike_config (
            id SERIAL PRIMARY KEY,
            label VARCHAR(80) NOT NULL,
            min_rating DECIMAL(3,1) NOT NULL,
            max_rating DECIMAL(3,1) NOT NULL,
            hike_pct DECIMAL(5,2) DEFAULT 0,
            incentive_pct DECIMAL(5,2) DEFAULT 0,
            color VARCHAR(20) DEFAULT '#1e3a8a'
        )
    """)
    cursor.execute("SELECT COUNT(*) FROM hike_config")
    if cursor.fetchone()[0] == 0:
        for _lbl, _mn, _mx, _hp, _ip, _clr in [
            ("Exceptional",          4.5, 5.0, 20.00, 15.00, "#15803d"),
            ("Exceeds Expectations", 4.0, 4.4, 15.00, 10.00, "#2563eb"),
            ("Meets Expectations",   3.0, 3.9, 10.00,  5.00, "#7c3aed"),
            ("Needs Improvement",    2.0, 2.9,  5.00,  0.00, "#d97706"),
            ("Below Expectations",   0.0, 1.9,  0.00,  0.00, "#dc2626"),
        ]:
            cursor.execute(
                "INSERT INTO hike_config (label, min_rating, max_rating, hike_pct, incentive_pct, color) VALUES (%s,%s,%s,%s,%s,%s)",
                (_lbl, _mn, _mx, _hp, _ip, _clr)
            )
        db.commit()
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS overtime_records (
            id SERIAL PRIMARY KEY,
            employee_id VARCHAR(50) NOT NULL,
            date DATE NOT NULL,
            shift_end TIME NOT NULL,
            actual_logout TIME NOT NULL,
            ot_minutes INT NOT NULL DEFAULT 0,
            ot_pay DECIMAL(10,2) DEFAULT 0,
            status VARCHAR(20) DEFAULT 'Pending',
            notes TEXT DEFAULT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            UNIQUE (employee_id, date)
        )
    """)
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS onboarding_templates (
            id SERIAL PRIMARY KEY,
            name VARCHAR(200) NOT NULL,
            description TEXT,
            is_active SMALLINT DEFAULT 1,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS onboarding_template_tasks (
            id SERIAL PRIMARY KEY,
            template_id INT NOT NULL,
            task_title VARCHAR(300) NOT NULL,
            task_description TEXT,
            requires_document SMALLINT DEFAULT 0,
            due_days INT DEFAULT 7,
            sort_order INT DEFAULT 0
        )
    """)
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS employee_onboarding (
            id SERIAL PRIMARY KEY,
            employee_id VARCHAR(50) NOT NULL,
            template_id INT NOT NULL,
            assigned_date DATE NOT NULL,
            due_date DATE,
            status VARCHAR(20) DEFAULT 'In Progress',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS employee_onboarding_tasks (
            id SERIAL PRIMARY KEY,
            onboarding_id INT NOT NULL,
            template_task_id INT NOT NULL,
            employee_id VARCHAR(50) NOT NULL,
            task_title VARCHAR(300) NOT NULL,
            task_description TEXT,
            requires_document SMALLINT DEFAULT 0,
            due_days INT DEFAULT 7,
            status VARCHAR(20) DEFAULT 'Pending',
            completed_at TIMESTAMP NULL,
            document_path VARCHAR(500),
            admin_notes TEXT
        )
    """)
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS offer_letters (
            id SERIAL PRIMARY KEY,
            onboarding_id INT NOT NULL,
            employee_id VARCHAR(50) NOT NULL,
            designation VARCHAR(150),
            department VARCHAR(150),
            work_location VARCHAR(200),
            monthly_ctc DECIMAL(12,2) DEFAULT 0,
            joining_date DATE,
            offer_valid_until DATE,
            probation_months INT DEFAULT 6,
            reporting_to VARCHAR(150),
            additional_notes TEXT,
            generated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            sent_at TIMESTAMP DEFAULT NULL,
            status VARCHAR(20) DEFAULT 'draft',
            notice_period_days INT DEFAULT 30,
            candidate_address TEXT
        )
    """)
    for _col, _sql in [
        ("notice_period_days", "ALTER TABLE offer_letters ADD COLUMN IF NOT EXISTS notice_period_days INT DEFAULT 30"),
        ("candidate_address",  "ALTER TABLE offer_letters ADD COLUMN IF NOT EXISTS candidate_address TEXT"),
        ("response_token",         "ALTER TABLE offer_letters ADD COLUMN IF NOT EXISTS response_token VARCHAR(64) DEFAULT NULL"),
        ("candidate_response",     "ALTER TABLE offer_letters ADD COLUMN IF NOT EXISTS candidate_response VARCHAR(20) DEFAULT NULL"),
        ("responded_at",           "ALTER TABLE offer_letters ADD COLUMN IF NOT EXISTS responded_at TIMESTAMP DEFAULT NULL"),
        ("response_token_expiry",  "ALTER TABLE offer_letters ADD COLUMN IF NOT EXISTS response_token_expiry TIMESTAMP DEFAULT NULL"),
    ]:
        try:
            cursor.execute(_sql); db.commit()
        except psycopg2.Error:
            db.rollback()
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS audit_logs (
            id SERIAL PRIMARY KEY,
            actor VARCHAR(100) NOT NULL,
            actor_type VARCHAR(20) DEFAULT 'admin',
            action VARCHAR(150) NOT NULL,
            target_table VARCHAR(100),
            target_id VARCHAR(100),
            detail TEXT,
            ip_address VARCHAR(45),
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_actor ON audit_logs (actor)")
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_action ON audit_logs (action)")
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_created ON audit_logs (created_at)")
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS login_attempts (
            id SERIAL PRIMARY KEY,
            identifier VARCHAR(150) NOT NULL,
            attempt_type VARCHAR(20) DEFAULT 'admin',
            failed_count INT DEFAULT 0,
            last_attempt TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            locked_until TIMESTAMP DEFAULT NULL,
            UNIQUE (identifier, attempt_type)
        )
    """)
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS email_queue (
            id           SERIAL PRIMARY KEY,
            to_email     VARCHAR(255) NOT NULL,
            subject      VARCHAR(500) NOT NULL,
            html_body    TEXT   NOT NULL,
            attachment_b64 TEXT   DEFAULT NULL,
            attachment_filename VARCHAR(255) DEFAULT NULL,
            status       VARCHAR(20) DEFAULT 'pending' CHECK (status IN ('pending','sending','done','failed')),
            attempts     SMALLINT DEFAULT 0,
            last_error   TEXT DEFAULT NULL,
            created_at   TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            sent_at      TIMESTAMP DEFAULT NULL
        )
    """)
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_eq_status ON email_queue (status)")
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_eq_created ON email_queue (created_at)")
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS payroll_runs (
            id SERIAL PRIMARY KEY,
            year INT NOT NULL,
            month INT NOT NULL,
            processed_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            processed_by VARCHAR(100),
            email_count INT DEFAULT 0,
            UNIQUE (year, month)
        )
    """)
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS compoff_balance (
            id SERIAL PRIMARY KEY,
            employee_id VARCHAR(50) NOT NULL UNIQUE,
            earned_minutes INT DEFAULT 0,
            used_minutes INT DEFAULT 0,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    _attach_updated_at_trigger(cursor, "compoff_balance")
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS api_tokens (
            token VARCHAR(64) PRIMARY KEY,
            token_type VARCHAR(20) NOT NULL DEFAULT 'admin',
            identity VARCHAR(100) NOT NULL,
            created_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP,
            expires_at TIMESTAMP NOT NULL
        )
    """)
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS mobile_biometric_proofs (
            employee_id VARCHAR(50) PRIMARY KEY,
            nonce VARCHAR(64) DEFAULT NULL,
            nonce_expires_at TIMESTAMP DEFAULT NULL,
            verified_at TIMESTAMP DEFAULT NULL
        )
    """)
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS regularization_requests (
            id SERIAL PRIMARY KEY,
            employee_id VARCHAR(50) NOT NULL,
            request_date DATE NOT NULL,
            login_time TIME DEFAULT NULL,
            logout_time TIME DEFAULT NULL,
            reason TEXT NOT NULL,
            status VARCHAR(20) DEFAULT 'Pending',
            admin_note TEXT DEFAULT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            resolved_at TIMESTAMP DEFAULT NULL,
            UNIQUE (employee_id, request_date)
        )
    """)
    db.commit()
    # Seed default leave types if empty
    cursor.execute("SELECT COUNT(*) FROM leave_types")
    if cursor.fetchone()[0] == 0:
        cursor.executemany(
            "INSERT INTO leave_types (name, annual_quota, is_paid) VALUES (%s,%s,%s)",
            [
                ("Casual Leave",    12,  1),
                ("Sick Leave",      12,  1),
                ("Earned Leave",    15,  1),
                ("Maternity Leave", 90,  1),
                ("Paternity Leave",  5,  1),
                ("Comp-off",         0,  1),
            ]
        )
        db.commit()
    # Ensure Comp-off leave type exists
    cursor.execute("SELECT id FROM leave_types WHERE name='Comp-off' LIMIT 1")
    if not cursor.fetchone():
        cursor.execute("INSERT INTO leave_types (name, annual_quota, is_paid) VALUES ('Comp-off', 0, 1)")
        db.commit()
    # Seed default breaks if table is empty
    cursor.execute("SELECT COUNT(*) FROM break_config")
    if cursor.fetchone()[0] == 0:
        cursor.executemany(
            "INSERT INTO break_config (break_name, break_time, duration_minutes) VALUES (%s, %s, %s)",
            [
                ("Coffee Break 1", "11:00:00", 10),
                ("Lunch Break",    "13:00:00", 60),
                ("Coffee Break 2", "16:00:00", 10),
            ]
        )
        db.commit()
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS companies (
            id SERIAL PRIMARY KEY,
            name VARCHAR(200) NOT NULL,
            code VARCHAR(20) DEFAULT NULL,
            working_days VARCHAR(30) DEFAULT 'Mon,Tue,Wed,Thu,Fri',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    db.commit()
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS company_feature_settings (
            company_id INT PRIMARY KEY,
            face_auth_enabled  SMALLINT DEFAULT 1,
            qr_enabled         SMALLINT DEFAULT 1,
            fingerprint_enabled SMALLINT DEFAULT 0,
            geo_enabled        SMALLINT DEFAULT 0,
            geo_radius         INT DEFAULT 300,
            pin_enabled        SMALLINT DEFAULT 1,
            biometric_enabled  SMALLINT DEFAULT 0,
            notify_leave       SMALLINT DEFAULT 1,
            notify_payslip     SMALLINT DEFAULT 1,
            notify_resignation SMALLINT DEFAULT 1,
            notify_doc_expiry  SMALLINT DEFAULT 1,
            session_timeout    INT DEFAULT 30,
            late_deduction_pct DECIMAL(5,2) DEFAULT 10.00,
            half_day_deduction_pct DECIMAL(5,2) DEFAULT 50.00,
            grace_minutes      INT DEFAULT 15,
            holiday_pay        VARCHAR(20) DEFAULT 'paid' CHECK (holiday_pay IN ('paid','unpaid')),
            leave_pay          VARCHAR(20) DEFAULT 'exclude' CHECK (leave_pay IN ('exclude','absent')),
            shift_start        TIME DEFAULT '09:00:00',
            shift_half         TIME DEFAULT '13:00:00',
            shift_end          TIME DEFAULT '18:00:00',
            FOREIGN KEY (company_id) REFERENCES companies(id) ON DELETE CASCADE
        )
    """)
    db.commit()
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS shift_swap_requests (
            id SERIAL PRIMARY KEY,
            requester_id VARCHAR(50) NOT NULL,
            target_id VARCHAR(50) NOT NULL,
            requester_shift_id INT NOT NULL,
            target_shift_id INT NOT NULL,
            reason TEXT,
            status VARCHAR(20) DEFAULT 'Pending_Target' CHECK (status IN ('Pending_Target','Pending_Admin','Approved','Rejected','Rejected_Admin')),
            target_response TEXT,
            admin_response TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    _attach_updated_at_trigger(cursor, "shift_swap_requests")
    db.commit()

    # Create company_settings table (must precede the migration loop below,
    # which ALTERs this table — on a fresh install with nothing to migrate
    # from, an ALTER before the table exists silently no-ops instead of
    # erroring, so column order here isn't just cosmetic).
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS company_settings (
            id SERIAL PRIMARY KEY,
            company_name VARCHAR(200) DEFAULT 'My Company',
            company_tagline VARCHAR(300) DEFAULT 'Employee Attendance System',
            company_logo VARCHAR(255) DEFAULT NULL,
            currency_symbol VARCHAR(10) DEFAULT '₹',
            timezone VARCHAR(60) DEFAULT 'Asia/Kolkata',
            setup_done SMALLINT DEFAULT 0,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    _attach_updated_at_trigger(cursor, "company_settings")
    db.commit()
    # Add default shift columns if not present
    for col, default in [("shift_start","09:00:00"), ("shift_half","13:00:00"), ("shift_end","18:00:00")]:
        try:
            cursor.execute(f"ALTER TABLE company_settings ADD COLUMN IF NOT EXISTS {col} TIME DEFAULT '{default}'")
            db.commit()
        except Exception:
            pass

    # Migrations for existing installs
    for sql in [
        "ALTER TABLE attendance ADD COLUMN IF NOT EXISTS logout_status VARCHAR(50) DEFAULT NULL",
        "ALTER TABLE attendance ADD COLUMN IF NOT EXISTS attendance_type VARCHAR(50) DEFAULT NULL",
        "ALTER TABLE employees ADD COLUMN IF NOT EXISTS email VARCHAR(150) DEFAULT NULL",
        "ALTER TABLE employees ADD COLUMN IF NOT EXISTS role VARCHAR(100) DEFAULT NULL",
        "ALTER TABLE employees ADD COLUMN IF NOT EXISTS password VARCHAR(255) DEFAULT NULL",
        "ALTER TABLE employees ADD COLUMN IF NOT EXISTS shift_id INT DEFAULT NULL",
        "ALTER TABLE employees ADD COLUMN IF NOT EXISTS date_of_joining DATE DEFAULT NULL",
        "ALTER TABLE employees ADD COLUMN IF NOT EXISTS phone VARCHAR(20) DEFAULT NULL",
        "ALTER TABLE employees ADD COLUMN IF NOT EXISTS gender VARCHAR(20) DEFAULT NULL",
        "ALTER TABLE employees ADD COLUMN IF NOT EXISTS dob DATE DEFAULT NULL",
        "ALTER TABLE employees ADD COLUMN IF NOT EXISTS blood_group VARCHAR(10) DEFAULT NULL",
        "ALTER TABLE employees ADD COLUMN IF NOT EXISTS address TEXT DEFAULT NULL",
        "ALTER TABLE employees ADD COLUMN IF NOT EXISTS city VARCHAR(100) DEFAULT NULL",
        "ALTER TABLE employees ADD COLUMN IF NOT EXISTS state VARCHAR(100) DEFAULT NULL",
        "ALTER TABLE employees ADD COLUMN IF NOT EXISTS pincode VARCHAR(20) DEFAULT NULL",
        "ALTER TABLE employees ADD COLUMN IF NOT EXISTS emergency_contact_name VARCHAR(100) DEFAULT NULL",
        "ALTER TABLE employees ADD COLUMN IF NOT EXISTS emergency_contact_phone VARCHAR(20) DEFAULT NULL",
        "ALTER TABLE employees ADD COLUMN IF NOT EXISTS emergency_contact_relation VARCHAR(50) DEFAULT NULL",
        "ALTER TABLE employees ADD COLUMN IF NOT EXISTS aadhar_number VARCHAR(20) DEFAULT NULL",
        "ALTER TABLE employees ADD COLUMN IF NOT EXISTS pan_number VARCHAR(20) DEFAULT NULL",
        "ALTER TABLE employees ADD COLUMN IF NOT EXISTS bank_name VARCHAR(100) DEFAULT NULL",
        "ALTER TABLE employees ADD COLUMN IF NOT EXISTS bank_account VARCHAR(30) DEFAULT NULL",
        "ALTER TABLE employees ADD COLUMN IF NOT EXISTS bank_ifsc VARCHAR(20) DEFAULT NULL",
        "ALTER TABLE employees ADD COLUMN IF NOT EXISTS uan_number VARCHAR(30) DEFAULT NULL",
        "ALTER TABLE employees ADD COLUMN IF NOT EXISTS work_mode VARCHAR(20) DEFAULT 'office'",
        "ALTER TABLE employees ADD COLUMN IF NOT EXISTS work_lat DECIMAL(10,8) DEFAULT NULL",
        "ALTER TABLE employees ADD COLUMN IF NOT EXISTS work_lon DECIMAL(11,8) DEFAULT NULL",
        "ALTER TABLE salary_config ADD COLUMN IF NOT EXISTS last_revised DATE DEFAULT NULL",
        "ALTER TABLE admin_users ADD COLUMN IF NOT EXISTS email VARCHAR(150) DEFAULT NULL",
        "ALTER TABLE admin_users ADD COLUMN IF NOT EXISTS reset_token VARCHAR(64) DEFAULT NULL",
        "ALTER TABLE admin_users ADD COLUMN IF NOT EXISTS reset_token_expiry TIMESTAMP DEFAULT NULL",
        "ALTER TABLE email_config ADD COLUMN IF NOT EXISTS from_email VARCHAR(150) DEFAULT NULL",
        "ALTER TABLE employees ADD COLUMN IF NOT EXISTS about_me TEXT DEFAULT NULL",
        "ALTER TABLE employees ADD COLUMN IF NOT EXISTS manager_name VARCHAR(150) DEFAULT NULL",
        "ALTER TABLE employees ADD COLUMN IF NOT EXISTS manager_id VARCHAR(20) DEFAULT NULL",
        "ALTER TABLE employees ADD COLUMN IF NOT EXISTS department VARCHAR(100) DEFAULT NULL",
        "ALTER TABLE employees ADD COLUMN IF NOT EXISTS designation VARCHAR(150) DEFAULT NULL",
        "ALTER TABLE employees ADD COLUMN IF NOT EXISTS is_active SMALLINT DEFAULT 1",
        "ALTER TABLE leave_requests ADD COLUMN IF NOT EXISTS leave_type_id INT DEFAULT NULL",
        "ALTER TABLE leave_requests ADD COLUMN IF NOT EXISTS is_half_day SMALLINT DEFAULT 0",
        "ALTER TABLE leave_requests ADD COLUMN IF NOT EXISTS half_day_session VARCHAR(10) DEFAULT NULL",
        "ALTER TABLE company_settings ADD COLUMN IF NOT EXISTS company_code VARCHAR(10) DEFAULT NULL",
        "ALTER TABLE admin_users ADD COLUMN IF NOT EXISTS role VARCHAR(20) DEFAULT 'admin'",
        "ALTER TABLE attendance ADD COLUMN IF NOT EXISTS worked_minutes INT DEFAULT 0",
        "ALTER TABLE attendance ADD COLUMN IF NOT EXISTS last_relogin TIME DEFAULT NULL",
        "ALTER TABLE salary_config ADD COLUMN IF NOT EXISTS monthly_ctc DECIMAL(12,2) DEFAULT 0",
        "ALTER TABLE salary_config ADD COLUMN IF NOT EXISTS basic_pct INT DEFAULT 50",
        "ALTER TABLE company_settings ADD COLUMN IF NOT EXISTS compoff_min_ot_minutes INT DEFAULT 120",
        "ALTER TABLE company_settings ADD COLUMN IF NOT EXISTS compoff_minutes_per_day INT DEFAULT 480",
        "ALTER TABLE company_settings ADD COLUMN IF NOT EXISTS late_deduction_pct DECIMAL(5,2) DEFAULT 10.00",
        "ALTER TABLE company_settings ADD COLUMN IF NOT EXISTS half_day_deduction_pct DECIMAL(5,2) DEFAULT 50.00",
        "ALTER TABLE company_settings ADD COLUMN IF NOT EXISTS grace_minutes INT DEFAULT 15",
        "ALTER TABLE company_settings ADD COLUMN IF NOT EXISTS holiday_pay VARCHAR(20) DEFAULT 'paid' CHECK (holiday_pay IN ('paid','unpaid'))",
        "ALTER TABLE company_settings ADD COLUMN IF NOT EXISTS leave_pay VARCHAR(20) DEFAULT 'exclude' CHECK (leave_pay IN ('exclude','absent'))",
        "ALTER TABLE employees ADD COLUMN IF NOT EXISTS joining_date DATE DEFAULT NULL",
        "ALTER TABLE employees ADD COLUMN IF NOT EXISTS company_id INT DEFAULT NULL",
        "ALTER TABLE employee_documents ADD COLUMN IF NOT EXISTS expiry_date DATE DEFAULT NULL",
        "ALTER TABLE overtime_records ADD COLUMN IF NOT EXISTS requested_by_employee SMALLINT DEFAULT 0",
        "ALTER TABLE overtime_records ADD COLUMN IF NOT EXISTS employee_reason VARCHAR(500) DEFAULT NULL",
        "ALTER TABLE leave_requests ADD COLUMN IF NOT EXISTS cancelled_at TIMESTAMP DEFAULT NULL",
        "ALTER TABLE salary_config ADD COLUMN IF NOT EXISTS last_hike_quarter SMALLINT DEFAULT NULL",
        "ALTER TABLE salary_config ADD COLUMN IF NOT EXISTS last_hike_year INT DEFAULT NULL",
        "ALTER TABLE company_settings ADD COLUMN IF NOT EXISTS default_onboarding_template_id INT DEFAULT NULL",
        "ALTER TABLE employee_onboarding_tasks ADD COLUMN IF NOT EXISTS employee_note VARCHAR(500) DEFAULT NULL",
        "ALTER TABLE company_settings ADD COLUMN IF NOT EXISTS fingerprint_enabled SMALLINT DEFAULT 0",
        "ALTER TABLE company_settings ADD COLUMN IF NOT EXISTS qr_enabled SMALLINT DEFAULT 1",
        "ALTER TABLE company_settings ADD COLUMN IF NOT EXISTS face_enabled SMALLINT DEFAULT 1",
        "ALTER TABLE company_settings ADD COLUMN IF NOT EXISTS location_enabled SMALLINT DEFAULT 1",
        "ALTER TABLE company_settings ADD COLUMN IF NOT EXISTS employee_password_auth SMALLINT DEFAULT 1",
        "ALTER TABLE employees ADD COLUMN IF NOT EXISTS fingerprint_credential_id VARCHAR(512) DEFAULT NULL",
        "ALTER TABLE employees ADD COLUMN IF NOT EXISTS fingerprint_public_key TEXT DEFAULT NULL",
        "ALTER TABLE employees ADD COLUMN IF NOT EXISTS fingerprint_sign_count INT DEFAULT 0",
        "ALTER TABLE company_settings ADD COLUMN IF NOT EXISTS face_auth_enabled SMALLINT DEFAULT 0",
        "ALTER TABLE company_settings ADD COLUMN IF NOT EXISTS geo_enabled SMALLINT DEFAULT 0",
        "ALTER TABLE company_settings ADD COLUMN IF NOT EXISTS geo_radius INT DEFAULT 100",
        "ALTER TABLE company_settings ADD COLUMN IF NOT EXISTS pin_enabled SMALLINT DEFAULT 1",
        "ALTER TABLE company_settings ADD COLUMN IF NOT EXISTS biometric_enabled SMALLINT DEFAULT 0",
        "ALTER TABLE company_settings ADD COLUMN IF NOT EXISTS notify_leave SMALLINT DEFAULT 1",
        "ALTER TABLE company_settings ADD COLUMN IF NOT EXISTS notify_payslip SMALLINT DEFAULT 1",
        "ALTER TABLE company_settings ADD COLUMN IF NOT EXISTS notify_resignation SMALLINT DEFAULT 1",
        "ALTER TABLE company_settings ADD COLUMN IF NOT EXISTS notify_doc_expiry SMALLINT DEFAULT 1",
        "ALTER TABLE company_settings ADD COLUMN IF NOT EXISTS session_timeout INT DEFAULT 30",
        "ALTER TABLE company_settings ADD COLUMN IF NOT EXISTS working_days VARCHAR(30) DEFAULT 'Mon,Tue,Wed,Thu,Fri'",
        "ALTER TABLE break_config ADD COLUMN IF NOT EXISTS break_type VARCHAR(20) DEFAULT 'coffee' CHECK (break_type IN ('coffee','lunch','custom'))",
        "ALTER TABLE break_config ADD COLUMN IF NOT EXISTS shift_id INT DEFAULT NULL",
        "ALTER TABLE companies ADD COLUMN IF NOT EXISTS working_days VARCHAR(30) DEFAULT 'Mon,Tue,Wed,Thu,Fri'",
        "ALTER TABLE onboarding_templates ADD COLUMN IF NOT EXISTS role VARCHAR(100) DEFAULT NULL",
        "ALTER TABLE shifts ADD COLUMN IF NOT EXISTS company_id INT DEFAULT NULL",
        "ALTER TABLE companies ADD COLUMN IF NOT EXISTS pin VARCHAR(10) DEFAULT NULL",
        "ALTER TABLE break_config ADD COLUMN IF NOT EXISTS company_id INT DEFAULT NULL",
        "ALTER TABLE announcements ADD COLUMN IF NOT EXISTS visibility VARCHAR(20) DEFAULT 'public' CHECK (visibility IN ('public','private'))",
        "ALTER TABLE announcements ADD COLUMN IF NOT EXISTS target_employee_id VARCHAR(50) DEFAULT NULL",
        "ALTER TABLE employees ADD COLUMN IF NOT EXISTS reset_token VARCHAR(80) DEFAULT NULL",
        "ALTER TABLE employees ADD COLUMN IF NOT EXISTS reset_token_expiry TIMESTAMP DEFAULT NULL",
    ]:
        try:
            cursor.execute(sql)
            db.commit()
        except psycopg2.Error:
            db.rollback()

    # Back-fill password for existing employees that have none (default PIN = 1234)
    cursor.execute("SELECT employee_id FROM employees WHERE password IS NULL")
    for (eid,) in cursor.fetchall():
        cursor.execute(
            "UPDATE employees SET password=%s WHERE employee_id=%s",
            (generate_password_hash('1234'), eid)
        )
    db.commit()

    # One-time migration: reset ALL employees to default PIN 1234
    try:
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS _applied_migrations (
                name VARCHAR(100) PRIMARY KEY,
                applied_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        db.commit()
        cursor.execute("SELECT 1 FROM _applied_migrations WHERE name='default_pin_1234'")
        if not cursor.fetchone():
            cursor.execute("UPDATE employees SET password=%s", (generate_password_hash('1234'),))
            cursor.execute("INSERT INTO _applied_migrations (name) VALUES ('default_pin_1234')")
            db.commit()
    except Exception:
        pass

    # Migration: add force_pin_change column and flag employees on default PIN
    try:
        cursor.execute("ALTER TABLE employees ADD COLUMN IF NOT EXISTS force_pin_change SMALLINT DEFAULT 0")
        db.commit()
    except psycopg2.Error:
        db.rollback()
    try:
        cursor.execute("SELECT 1 FROM _applied_migrations WHERE name='force_pin_change_flag'")
        if not cursor.fetchone():
            default_hash = generate_password_hash('1234')
            cursor.execute("SELECT employee_id, password FROM employees")
            for eid, pwd_hash in cursor.fetchall():
                if pwd_hash and check_password_hash(pwd_hash, '1234'):
                    cursor.execute("UPDATE employees SET force_pin_change=1 WHERE employee_id=%s", (eid,))
            cursor.execute("INSERT INTO _applied_migrations (name) VALUES ('force_pin_change_flag')")
            db.commit()
    except Exception:
        pass

    # Performance indexes migration
    try:
        cursor.execute("SELECT 1 FROM _applied_migrations WHERE name='perf_indexes_v1'")
        if not cursor.fetchone():
            _idx_stmts = [
                "CREATE INDEX IF NOT EXISTS idx_leave_emp ON leave_requests(employee_id)",
                "CREATE INDEX IF NOT EXISTS idx_leave_status ON leave_requests(status)",
                "CREATE INDEX IF NOT EXISTS idx_tickets_emp ON tickets(employee_id)",
                "CREATE INDEX IF NOT EXISTS idx_tickets_status ON tickets(status)",
                "CREATE INDEX IF NOT EXISTS idx_resign_emp ON resignation_requests(employee_id)",
                "CREATE INDEX IF NOT EXISTS idx_notif_emp ON notifications(employee_id)",
                "CREATE INDEX IF NOT EXISTS idx_notif_read ON notifications(is_read)",
                "CREATE INDEX IF NOT EXISTS idx_onboard_emp ON employee_onboarding(employee_id)",
                "CREATE INDEX IF NOT EXISTS idx_onboard_status ON employee_onboarding(status)",
                "CREATE INDEX IF NOT EXISTS idx_payroll_emp ON payroll_runs(employee_id)",
            ]
            for stmt in _idx_stmts:
                try:
                    cursor.execute(stmt)
                    db.commit()
                except Exception:
                    db.rollback()
            cursor.execute("INSERT INTO _applied_migrations (name) VALUES ('perf_indexes_v1')")
            db.commit()
    except Exception:
        pass

    # Performance indexes v2 — high-traffic columns missing from v1
    try:
        cursor.execute("SELECT 1 FROM _applied_migrations WHERE name='perf_indexes_v2'")
        if not cursor.fetchone():
            _idx_stmts_v2 = [
                "CREATE INDEX IF NOT EXISTS idx_att_date ON attendance(date)",
                "CREATE INDEX IF NOT EXISTS idx_emp_active ON employees(is_active)",
                "CREATE INDEX IF NOT EXISTS idx_emp_company ON employees(company_id)",
                "CREATE INDEX IF NOT EXISTS idx_leave_date ON leave_requests(leave_date)",
            ]
            for stmt in _idx_stmts_v2:
                try:
                    cursor.execute(stmt)
                    db.commit()
                except Exception:
                    db.rollback()
            cursor.execute("INSERT INTO _applied_migrations (name) VALUES ('perf_indexes_v2')")
            db.commit()
    except Exception:
        pass

    cursor.execute("SELECT COUNT(*) FROM company_settings")
    if cursor.fetchone()[0] == 0:
        cursor.execute("INSERT INTO company_settings (setup_done) VALUES (0)")
        db.commit()

    # Seed admin from env — only if no admin exists yet
    _admin_user = os.environ.get("ADMIN_USERNAME", "admin").strip()
    _admin_pass = os.environ.get("ADMIN_PASSWORD", "").strip()
    cursor.execute("SELECT COUNT(*) FROM admin_users")
    admin_count = cursor.fetchone()[0]
    if admin_count == 0 and _admin_pass:
        cursor.execute(
            "INSERT INTO admin_users (username, password) VALUES (%s, %s)",
            (_admin_user, generate_password_hash(_admin_pass))
        )
        db.commit()
        app_log.info("Admin created: username=%s", _admin_user)
        admin_count = 1
    elif admin_count == 0 and not _admin_pass:
        app_log.warning("ADMIN_PASSWORD not set in .env — complete setup via /setup")

    # Auto-mark setup done for existing installs that already have an admin
    if admin_count > 0:
        cursor.execute("UPDATE company_settings SET setup_done=1 WHERE setup_done=0")
        db.commit()

    cursor.close()
    db.close()


def assign_leave_balances_for_employee(cursor, employee_id, year=None):
    """Auto-assign leave balances for all active leave types for a new/existing employee."""
    if year is None:
        year = datetime.date.today().year
    cursor.execute("SELECT id, annual_quota FROM leave_types WHERE is_active=1")
    for lt_id, quota in cursor.fetchall():
        cursor.execute("""
            INSERT INTO leave_balances (employee_id, leave_type_id, year, total_days, used_days)
            VALUES (%s, %s, %s, %s, 0)
            ON CONFLICT (employee_id, leave_type_id, year) DO UPDATE SET
                total_days = CASE WHEN leave_balances.used_days = 0
                                  THEN EXCLUDED.total_days ELSE leave_balances.total_days END
        """, (employee_id, lt_id, year, quota))


def init_master_db():
    """Create the att_master tenant-registry schema and its tenants table if
    they don't exist. Postgres has no mid-connection database switch, so this
    is a schema within the shared database now, not a separate physical
    database the way MySQL's att_master was."""
    try:
        # Schema must exist before get_master_db() can SET search_path to it,
        # so this first connection stays on the default (public) schema —
        # get_db_connection() now always resets search_path explicitly on
        # every borrow, so it's safe to use here without leaking att_master
        # onto whichever connection the pool hands out next.
        db = get_db_connection()
        cur = db.cursor()
        cur.execute('CREATE SCHEMA IF NOT EXISTS att_master')
        db.commit()
        cur.close(); db.close()

        from database import get_master_db
        db = get_master_db()
        cur = db.cursor()
        cur.execute("""
            CREATE TABLE IF NOT EXISTS tenants (
                id SERIAL PRIMARY KEY,
                company_name VARCHAR(200) NOT NULL,
                subdomain VARCHAR(100) UNIQUE NOT NULL,
                db_name VARCHAR(100) UNIQUE NOT NULL,
                admin_email VARCHAR(200) DEFAULT NULL,
                plan VARCHAR(50) DEFAULT 'starter',
                status VARCHAR(20) DEFAULT 'active',
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        db.commit()
        cur.close()
        db.close()
    except Exception as _e:
        app_log.warning("init_master_db failed (non-fatal for single-tenant mode): %s", _e)


def init_tenant_db(schema_name: str):
    """Initialize schema in a freshly created tenant schema."""
    from flask import g as _g
    _g.tenant_db = schema_name
    init_db()


# ---------------- NOTIFICATION HELPER ----------------
def _create_notification(recipient_type, title, message, employee_id=None):
    try:
        db = get_db_connection()
        cursor = db.cursor()
        cursor.execute(
            "INSERT INTO notifications (recipient_type, employee_id, title, message) VALUES (%s,%s,%s,%s)",
            (recipient_type, employee_id, title, message)
        )
        db.commit()
        cursor.close(); db.close()
    except Exception:
        pass


# ---------------- ADMIN GUARD ----------------
def admin_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        # If both admin and employee keys exist, drop the employee key and continue as admin
        if session.get("admin_logged_in") and session.get("employee_id"):
            session.pop("employee_id", None)
            session.pop("employee_name", None)
            session.pop("employee_role", None)
        if not session.get("admin_logged_in"):
            is_ajax = (
                request.headers.get("X-Requested-With") == "XMLHttpRequest"
                or request.headers.get("Accept", "").startswith("application/json")
                or request.headers.get("Content-Type", "").startswith("application/json")
                or request.is_json
            )
            if is_ajax:
                return jsonify({"ok": False, "msg": "Session expired. Please log in again.", "redirect": url_for("auth.admin_login")}), 401
            return redirect(url_for("auth.admin_login"))
        return f(*args, **kwargs)
    return wrapper

def employee_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        # If admin is also logged in, redirect to admin panel instead of clearing session
        if session.get("admin_logged_in") and session.get("employee_id"):
            session.pop("employee_id", None)
            session.pop("employee_name", None)
            session.pop("employee_role", None)
            return redirect("/admin")
        if not session.get("employee_id"):
            return redirect("/employee_login")
        if session.get("_fpc") and request.endpoint != "force_change_pin":
            return redirect("/force_change_pin")
        return f(*args, **kwargs)
    return wrapper

def manager_or_admin_required(f):
    """Allow access to admin users whose role is 'admin' or 'manager'.

    Managers can access leave, attendance, and employee views but NOT
    salary, settings, or user management (enforced by the views themselves).
    """
    @wraps(f)
    def wrapper(*args, **kwargs):
        if not session.get("admin_logged_in"):
            is_ajax = (
                request.headers.get("X-Requested-With") == "XMLHttpRequest"
                or request.headers.get("Accept", "").startswith("application/json")
                or request.is_json
            )
            if is_ajax:
                return jsonify({"ok": False, "msg": "Session expired. Please log in again.", "redirect": url_for("auth.admin_login")}), 401
            return redirect(url_for("auth.admin_login"))
        admin_role = session.get("admin_role", "admin")
        if admin_role not in ("admin", "manager"):
            return jsonify({"ok": False, "msg": "Insufficient permissions."}), 403
        return f(*args, **kwargs)
    return wrapper

# ---------------- ATTENDANCE HELPERS ----------------
def _td_to_time(val):
    """Convert a timedelta or datetime.time to datetime.time. psycopg2
    already returns TIME columns as datetime.time (returned as-is below);
    the timedelta path handles values computed via time arithmetic elsewhere
    in the app, and is what mysql-connector used to return directly."""
    if val is None:
        return None
    if isinstance(val, datetime.time):
        return val
    total = int(val.total_seconds())
    h, rem = divmod(total, 3600)
    m, s   = divmod(rem, 60)
    return datetime.time(h % 24, m, s)

def get_employee_shift(emp_id, cursor):
    """Return (shift_start, shift_half, shift_end, shift_name) for employee.
    Falls back to global defaults if no shift assigned."""
    cursor.execute(
        "SELECT s.start_time, s.half_time, s.end_time, s.name "
        "FROM employees e JOIN shifts s ON e.shift_id = s.id "
        "WHERE e.employee_id = %s",
        (emp_id,)
    )
    row = cursor.fetchone()
    if row:
        return _td_to_time(row[0]), _td_to_time(row[1]), _td_to_time(row[2]), row[3]
    return SHIFT_START, SHIFT_HALF, SHIFT_END, "Default"

def get_attendance_type(login_status, logout_status):
    if not login_status:
        return "Absent"
    if not logout_status:
        return "Half Day" if login_status == "Half Day Login" else "Present"
    if login_status == "Half Day Login":
        return "Half Day"
    if logout_status in ("Half Day Logout", "Early Logout"):
        return "Half Day"
    if login_status == "Late Login":
        return "Late - Full Day"
    return "Full Day"

def classify_by_worked_minutes(login_status, total_minutes, s_start, s_end):
    """Classify attendance based on cumulative worked minutes vs shift length."""
    today_d = datetime.date.today()
    shift_mins = max(1, int((
        datetime.datetime.combine(today_d, s_end) -
        datetime.datetime.combine(today_d, s_start)
    ).total_seconds() / 60))
    if total_minutes >= shift_mins * 0.75:
        return "Late - Full Day" if login_status == "Late Login" else "Full Day"
    return "Half Day"

def calculate_deduction(salary_per_day, attendance_type):
    spd = float(salary_per_day)
    if attendance_type == "Full Day":
        return 0.0
    if attendance_type == "Approved Leave":
        # If leave is configured as 'absent', deduct full day; else no deduction
        return spd if LEAVE_PAY == 'absent' else 0.0
    if attendance_type == "Holiday":
        # If holiday is unpaid, deduct full day
        return spd if HOLIDAY_PAY == 'unpaid' else 0.0
    if attendance_type == "Late - Full Day":
        return round(spd * LATE_DEDUCTION_RATE, 2)
    if attendance_type in ("Half Day", "Present"):
        return round(spd * HALF_DAY_RATE, 2)
    if attendance_type == "Absent":
        return spd
    return 0.0

def infer_type_legacy(status, login_time, logout_time):
    if not login_time:
        return "Absent"
    if not logout_time:
        return "Half Day" if status == "Half Day Login" else "Present"
    if status in ("Half Day Logout", "Early Logout"):
        return "Half Day"
    return "Full Day"

def detect_overtime(employee_id, date, logout_time):
    try:
        db = get_db_connection()
        cursor = db.cursor(buffered=True)
        cursor.execute(
            "SELECT s.end_time FROM employees e JOIN shifts s ON e.shift_id=s.id WHERE e.employee_id=%s",
            (employee_id,)
        )
        row = cursor.fetchone()
        shift_end = _td_to_time(row[0]) if row else SHIFT_END
        logout_t = _td_to_time(logout_time) if not isinstance(logout_time, datetime.time) else logout_time
        if logout_t is None or shift_end is None:
            cursor.close(); db.close(); return
        end_mins = shift_end.hour * 60 + shift_end.minute
        out_mins = logout_t.hour * 60 + logout_t.minute
        ot_minutes = out_mins - end_mins
        if ot_minutes < 30:
            cursor.close(); db.close(); return
        cursor.execute(
            "SELECT COALESCE(salary_per_day,0) FROM salary_config WHERE employee_id=%s",
            (employee_id,)
        )
        sc = cursor.fetchone()
        spd = float(sc[0]) if sc else 0.0
        ot_pay = round((spd / 8 / 60) * ot_minutes, 2)
        cursor.execute("""
            INSERT INTO overtime_records (employee_id, date, shift_end, actual_logout, ot_minutes, ot_pay, status)
            VALUES (%s,%s,%s,%s,%s,%s,'Pending')
            ON CONFLICT (employee_id, date) DO UPDATE SET
                actual_logout=EXCLUDED.actual_logout, ot_minutes=EXCLUDED.ot_minutes, ot_pay=EXCLUDED.ot_pay
        """, (employee_id, date, shift_end, logout_t, ot_minutes, ot_pay))
        db.commit()
        cursor.close(); db.close()
    except Exception:
        pass

def get_working_days(year, month):
    _, last_day = calendar.monthrange(year, month)
    return [
        datetime.date(year, month, d)
        for d in range(1, last_day + 1)
        if datetime.date(year, month, d).weekday() != 6
    ]

def fetch_holidays_set(year, month):
    _, last_day = calendar.monthrange(year, month)
    db = get_db_connection()
    cursor = db.cursor(buffered=True)
    cursor.execute(
        "SELECT date FROM holidays WHERE date BETWEEN %s AND %s",
        (datetime.date(year, month, 1), datetime.date(year, month, last_day))
    )
    holidays = {row[0] for row in cursor.fetchall()}
    cursor.close()
    db.close()
    return holidays

def get_billable_past_days(year, month):
    today = datetime.date.today()
    # Holidays are included — they count as paid working days
    return [d for d in get_working_days(year, month) if d <= today]

def fetch_leave_map(year, month):
    """Return {emp_id: set(leave_dates)} for approved leaves in the given month."""
    _, last_day = calendar.monthrange(year, month)
    db = get_db_connection()
    cursor = db.cursor(buffered=True)
    cursor.execute(
        "SELECT employee_id, leave_date FROM leave_requests "
        "WHERE status = 'Approved' AND leave_date BETWEEN %s AND %s",
        (datetime.date(year, month, 1), datetime.date(year, month, last_day))
    )
    leave_map = {}
    for eid, ld in cursor.fetchall():
        leave_map.setdefault(eid, set()).add(ld)
    cursor.close()
    db.close()
    return leave_map

# ---------------- EMAIL HELPERS ----------------
def get_email_config():
    db = get_db_connection()
    cursor = db.cursor(buffered=True)
    cursor.execute("SELECT smtp_host, smtp_port, smtp_user, smtp_pass, from_name, from_email FROM email_config ORDER BY id DESC LIMIT 1")
    row = cursor.fetchone()
    cursor.close()
    db.close()
    if row:
        return {
            "host": row[0], "port": row[1], "user": row[2], "password": decrypt_pii(row[3]),
            "from_name": row[4], "from_email": row[5] or row[2]
        }
    # Fall back to .env values so team members don't need to configure via UI
    smtp_host = os.environ.get("SMTP_HOST", "")
    smtp_user = os.environ.get("SMTP_USER", "")
    smtp_pass = os.environ.get("SMTP_PASS", "")
    if smtp_host and smtp_user and smtp_pass:
        return {
            "host": smtp_host,
            "port": int(os.environ.get("SMTP_PORT", 587)),
            "user": smtp_user,
            "password": smtp_pass,
            "from_name": os.environ.get("SMTP_FROM_NAME", "Attendance System"),
            "from_email": os.environ.get("SMTP_FROM_EMAIL", smtp_user),
        }
    return None

def get_admin_emails():
    """Return a list of all admin email addresses that have been set."""
    db = get_db_connection()
    cursor = db.cursor(buffered=True)
    cursor.execute("SELECT email FROM admin_users WHERE email IS NOT NULL AND email != ''")
    emails = [row[0] for row in cursor.fetchall()]
    cursor.close(); db.close()
    return emails

def build_salary_slip_html(emp_name, emp_id, emp_email, month_name, year, month, salary_data,
                           company_name="", emp_designation="", emp_dept="",
                           pan="", uan="", bank_account="", bank_name="",
                           payroll_cfg=None):
    e = salary_data
    pc = payroll_cfg or {}

    # ── Salary structure ──────────────────────────────────────────
    monthly_ctc  = float(e.get("monthly_ctc", 0))
    basic_pct    = int(e.get("basic_pct", 50))
    if monthly_ctc <= 0 and float(e.get("spd", 0)) > 0:
        monthly_ctc = round(float(e["spd"]) * 26, 2)

    basic        = round(monthly_ctc * basic_pct / 100, 2)
    hra          = round(monthly_ctc * 0.20, 2)
    # Cap conveyance so gross never exceeds CTC
    conveyance   = round(min(1600.0, max(0, monthly_ctc - basic - hra)), 2)
    special_all  = round(max(0, monthly_ctc - basic - hra - conveyance), 2)
    gross_salary = round(basic + hra + conveyance + special_all, 2)

    # ── LOP: standard 26-day denominator (Indian payroll norm) ───
    full_d  = int(e.get("full_days", 0))
    late_d  = int(e.get("late_days", 0))
    half_d  = int(e.get("half_days", 0))
    lop_days     = float(e.get("absent", 0))
    paid_days_display = full_d + late_d + half_d   # integer count for display
    lop_ded      = round(gross_salary / 26 * lop_days, 2)
    gross_earned = round(gross_salary - lop_ded, 2)

    # ── Statutory deductions ─────────────────────────────────────
    pf_pct        = float(pc.get("pf_employee_pct", 12))
    pf_er_pct     = float(pc.get("pf_employer_pct", 12))
    pf_cap_basic  = float(pc.get("pf_basic_cap", 15000))
    pt_monthly    = float(pc.get("professional_tax", 200))
    tds_ann_pct   = float(pc.get("tds_annual_pct", 0))

    # PF on capped basic; TDS = annual taxable (CTC×12) × rate ÷ 12
    pf_ded        = round(min(basic, pf_cap_basic) * pf_pct / 100, 2)
    pf_er_ded     = round(min(basic, pf_cap_basic) * pf_er_pct / 100, 2)
    annual_ctc    = monthly_ctc * 12
    tds_ded       = round(annual_ctc * tds_ann_pct / 100 / 12, 2)
    # Cap statutory deductions to gross earned (net cannot go below 0)
    stat_ded      = pf_ded + pt_monthly + tds_ded
    if stat_ded > gross_earned:
        ratio     = gross_earned / stat_ded if stat_ded > 0 else 0
        pf_ded    = round(pf_ded * ratio, 2)
        pt_monthly = round(pt_monthly * ratio, 2)
        tds_ded   = round(tds_ded * ratio, 2)
    total_ded     = round(lop_ded + pf_ded + pt_monthly + tds_ded, 2)
    net_pay       = max(0, round(gross_earned - pf_ded - pt_monthly - tds_ded, 2))

    emp_row_extra = ""
    if emp_designation: emp_row_extra += f"<tr><td>Designation</td><td>{_html.escape(str(emp_designation))}</td></tr>"
    if emp_dept:        emp_row_extra += f"<tr><td>Department</td><td>{_html.escape(str(emp_dept))}</td></tr>"
    if pan:             emp_row_extra += f"<tr><td>PAN</td><td>{_html.escape(str(pan))}</td></tr>"
    if uan:             emp_row_extra += f"<tr><td>UAN</td><td>{_html.escape(str(uan))}</td></tr>"
    if bank_account:
        masked = '*'*len(bank_account[:-4]) + bank_account[-4:]
        emp_row_extra += f"<tr><td>Bank A/C</td><td>{_html.escape(masked)}</td></tr>"
    if bank_name:       emp_row_extra += f"<tr><td>Bank</td><td>{_html.escape(str(bank_name))}</td></tr>"

    incentive_row = ""
    if e.get("incentive", 0) > 0:
        incentive_row = f'<tr><td>Incentive / Bonus</td><td class="green">+ Rs. {e["incentive"]:.2f}</td><td></td><td></td></tr>'

    return f"""<!DOCTYPE html>
<html>
<head>
<meta charset="UTF-8">
<style>
  *{{box-sizing:border-box;margin:0;padding:0}}
  body{{font-family:"Segoe UI",Arial,sans-serif;background:#f0f4ff;color:#1e293b}}
  .wrap{{max-width:800px;margin:20px auto;background:#fff;border-radius:16px;overflow:hidden;box-shadow:0 8px 40px rgba(0,0,0,.12)}}
  .hdr{{background:linear-gradient(135deg,#0f2460,#1e3a8a);padding:28px 32px;color:#fff;display:flex;justify-content:space-between;align-items:center}}
  .hdr-left h1{{font-size:20px;font-weight:800;margin-bottom:4px}}
  .hdr-left p{{font-size:13px;opacity:.75}}
  .hdr-right{{text-align:right}}
  .hdr-right .slip-num{{font-size:12px;opacity:.7;margin-bottom:4px}}
  .hdr-right .month{{font-size:18px;font-weight:700}}
  .emp-bar{{background:#dbeafe;padding:16px 32px;display:grid;grid-template-columns:1fr 1fr;gap:4px 40px;font-size:13px}}
  .emp-bar td:first-child{{font-weight:700;color:#1e3a8a;white-space:nowrap}}
  .emp-bar td{{padding:3px 6px;color:#1e293b}}
  .att-strip{{display:grid;grid-template-columns:repeat(6,1fr);gap:0;border-bottom:1px solid #e2e8f0}}
  .att-cell{{text-align:center;padding:14px 8px;border-right:1px solid #e2e8f0}}
  .att-cell:last-child{{border-right:none}}
  .att-cell .num{{font-size:22px;font-weight:800}}
  .att-cell .lbl{{font-size:10px;color:#64748b;margin-top:2px;text-transform:uppercase;letter-spacing:.3px}}
  .body{{padding:24px 32px}}
  .two-col{{display:grid;grid-template-columns:1fr 1fr;gap:24px;margin-bottom:16px}}
  .sec-title{{font-size:12px;font-weight:800;color:#1e3a8a;text-transform:uppercase;letter-spacing:.5px;padding-bottom:7px;border-bottom:2px solid #dbeafe;margin-bottom:10px}}
  table.pay-tbl{{width:100%;border-collapse:collapse;font-size:13px}}
  table.pay-tbl td{{padding:7px 10px;border-bottom:1px solid #f1f5f9;vertical-align:top}}
  table.pay-tbl td:last-child{{text-align:right;font-weight:600;white-space:nowrap}}
  table.pay-tbl tr.tot td{{background:#f8fafc;font-weight:800;border-top:2px solid #dbeafe;border-bottom:2px solid #dbeafe;font-size:14px}}
  .net-box{{background:linear-gradient(135deg,#0f2460,#1e3a8a);color:#fff;border-radius:12px;padding:18px 24px;display:flex;justify-content:space-between;align-items:center;margin-top:16px}}
  .net-box .lbl{{font-size:13px;opacity:.8}}
  .net-box .amt{{font-size:28px;font-weight:900}}
  .footer{{background:#f8fafc;padding:14px 32px;text-align:center;font-size:11px;color:#94a3b8;border-top:1px solid #e2e8f0;display:flex;justify-content:space-between;align-items:center}}
  .print-btn{{display:flex;gap:10px;margin:18px 32px 4px;justify-content:flex-end}}
  .btn{{padding:9px 20px;border:none;border-radius:9px;font-size:13px;font-weight:700;cursor:pointer}}
  .btn-print{{background:#0f2460;color:#fff}}
  .btn-back{{background:#f1f5f9;color:#64748b}}
  .green{{color:#16a34a}} .red{{color:#ef4444}} .yellow{{color:#f59e0b}}
  @media print{{
    body{{background:#fff}}
    .wrap{{box-shadow:none;margin:0;border-radius:0}}
    .print-btn{{display:none}}
    .btn{{display:none}}
  }}
</style>
</head>
<body>
<div class="wrap">
  <div class="print-btn">
    <button class="btn btn-back" onclick="history.back()">&#8592; Back</button>
    <button class="btn btn-print" onclick="window.print()">&#128438; Download / Print PDF</button>
  </div>

  <div class="hdr">
    <div class="hdr-left">
      <h1>{_html.escape(str(company_name)) if company_name else "Payslip"}</h1>
      <p>Salary Slip — {month_name}</p>
    </div>
    <div class="hdr-right">
      <div class="slip-num">Slip ID: {_html.escape(str(emp_id))}-{year}{month:02d}</div>
      <div class="month">{month_name}</div>
    </div>
  </div>

  <div class="emp-bar">
    <table>
      <tr><td>Employee Name</td><td>{_html.escape(str(emp_name))}</td></tr>
      <tr><td>Employee ID</td><td>{_html.escape(str(emp_id))}</td></tr>
      <tr><td>Email</td><td>{_html.escape(str(emp_email)) if emp_email else 'N/A'}</td></tr>
      {emp_row_extra}
    </table>
    <table>
      <tr><td>Pay Period</td><td>{month_name}</td></tr>
      <tr><td>Working Days (Standard)</td><td>26</td></tr>
      <tr><td>Days Present</td><td>{paid_days_display}</td></tr>
      <tr><td>LOP Days</td><td>{int(lop_days)}</td></tr>
      <tr><td>Monthly CTC</td><td>Rs. {monthly_ctc:,.2f}</td></tr>
    </table>
  </div>

  <div class="att-strip">
    <div class="att-cell"><div class="num green">{full_d}</div><div class="lbl">Full Days</div></div>
    <div class="att-cell"><div class="num yellow">{late_d}</div><div class="lbl">Late Days</div></div>
    <div class="att-cell"><div class="num yellow">{half_d}</div><div class="lbl">Half Days</div></div>
    <div class="att-cell"><div class="num red">{int(lop_days)}</div><div class="lbl">LOP / Absent</div></div>
    <div class="att-cell"><div class="num" style="color:#3b82f6">{e.get('holiday_days',0)}</div><div class="lbl">Holidays</div></div>
    <div class="att-cell"><div class="num" style="color:#9333ea">{e.get('leave_days',0)}</div><div class="lbl">Leave (Paid)</div></div>
  </div>

  <div class="body">
    <div class="two-col">
      <div>
        <div class="sec-title">Earnings (Monthly)</div>
        <table class="pay-tbl">
          <tr><td>Basic Salary ({basic_pct}% of CTC)</td><td>Rs. {basic:,.2f}</td></tr>
          <tr><td>House Rent Allowance (HRA)</td><td>Rs. {hra:,.2f}</td></tr>
          <tr><td>Conveyance Allowance</td><td>Rs. {conveyance:,.2f}</td></tr>
          <tr><td>Special Allowance</td><td>Rs. {special_all:,.2f}</td></tr>
          {incentive_row}
          <tr class="tot"><td>Gross Salary</td><td>Rs. {gross_salary:,.2f}</td></tr>
        </table>
        <div style="margin-top:10px;background:#f0fdf4;border:1px solid #bbf7d0;border-radius:8px;padding:10px 12px;font-size:12px;color:#15803d;">
          <b>PF — Employer Contribution ({pf_er_pct:.0f}%)</b>: Rs. {pf_er_ded:,.2f}
          <div style="color:#64748b;margin-top:2px;font-size:11px;">Company's share — not deducted from your pay</div>
        </div>
      </div>
      <div>
        <div class="sec-title">Deductions</div>
        <table class="pay-tbl">
          <tr><td>Loss of Pay — LOP ({int(lop_days)} days × Rs.{gross_salary/26:,.2f})</td><td class="red">Rs. {lop_ded:,.2f}</td></tr>
          <tr><td>PF — Employee Contribution ({pf_pct:.0f}% of Basic)</td><td class="red">Rs. {pf_ded:,.2f}</td></tr>
          <tr><td>Professional Tax</td><td class="red">Rs. {pt_monthly:,.2f}</td></tr>
          {"<tr><td>TDS — Income Tax (annual " + f"{tds_ann_pct:.1f}%" + ")</td><td class='red'>Rs. " + f"{tds_ded:,.2f}</td></tr>" if tds_ded > 0 else ""}
          <tr class="tot"><td>Total Deductions</td><td>Rs. {total_ded:,.2f}</td></tr>
        </table>
        <div style="margin-top:10px;background:#fff7ed;border:1px solid #fed7aa;border-radius:8px;padding:10px 12px;font-size:12px;color:#92400e;">
          <b>Gross Earned</b> (after LOP): Rs. {gross_earned:,.2f}
          <div style="color:#64748b;margin-top:2px;font-size:11px;">Gross Salary − LOP before other deductions</div>
        </div>
      </div>
    </div>

    <div class="net-box">
      <div>
        <div class="lbl">Net Take-Home Pay</div>
        <div style="font-size:11px;opacity:.65;margin-top:4px;">
          Rs.{gross_salary:,.2f} − LOP Rs.{lop_ded:,.2f} − PF Rs.{pf_ded:,.2f} − PT Rs.{pt_monthly:,.2f}{f" − TDS Rs.{tds_ded:,.2f}" if tds_ded > 0 else ""}
        </div>
      </div>
      <div class="amt">Rs. {net_pay:,.2f}</div>
    </div>
  </div>

  <div class="footer">
    <span>This is a system-generated payslip. Contact HR for any discrepancies.</span>
    <span>Generated on {datetime.date.today().strftime('%d %B %Y')}</span>
  </div>
</div>
</body>
</html>"""

def send_email_smtp(to_email, subject, html_body, config, attachment_bytes=None, attachment_filename=None):
    from_addr = config.get("from_email") or config["user"]

    msg = MIMEMultipart("mixed")
    msg["Subject"] = subject
    msg["From"]    = f"{config['from_name']} <{from_addr}>"
    msg["To"]      = to_email
    alt = MIMEMultipart("alternative")
    alt.attach(MIMEText(html_body, "html", "utf-8"))
    msg.attach(alt)

    if attachment_bytes and attachment_filename:
        part = MIMEBase("application", "octet-stream")
        part.set_payload(attachment_bytes)
        encoders.encode_base64(part)
        part.add_header("Content-Disposition", f'attachment; filename="{attachment_filename}"')
        msg.attach(part)

    context = ssl.create_default_context()
    port = int(config.get("port", 587))
    if port == 465:
        # Implicit SSL (port 465)
        with smtplib.SMTP_SSL(config["host"], port, context=context, timeout=20) as server:
            server.login(config["user"], config["password"])
            server.sendmail(from_addr, to_email, msg.as_string())
    else:
        # STARTTLS (port 587 / 25)
        with smtplib.SMTP(config["host"], port, timeout=20) as server:
            server.ehlo()
            server.starttls(context=context)
            server.ehlo()
            server.login(config["user"], config["password"])
            server.sendmail(from_addr, to_email, msg.as_string())

def send_email_async(to_email, subject, html_body, config,
                     attachment_bytes=None, attachment_filename=None, **_):
    """Enqueue an email for reliable delivery via the DB-backed email worker."""
    att_b64 = None
    if attachment_bytes:
        att_b64 = base64.b64encode(attachment_bytes).decode()
    try:
        db  = get_db_connection()
        cur = db.cursor()
        cur.execute(
            "INSERT INTO email_queue (to_email, subject, html_body, attachment_b64, attachment_filename) "
            "VALUES (%s,%s,%s,%s,%s)",
            (to_email, subject, html_body, att_b64, attachment_filename)
        )
        db.commit()
        cur.close(); db.close()
    except Exception as e:
        app_log.error("Failed to enqueue email to %s: %s", to_email, e)
        # Fall back to in-process send so the email is not silently dropped
        threading.Thread(
            target=lambda: send_email_smtp(to_email, subject, html_body, config,
                                           attachment_bytes=attachment_bytes,
                                           attachment_filename=attachment_filename),
            daemon=True
        ).start()


def _email_queue_worker():
    """Background thread: dequeues and sends emails, retries up to 3 times."""
    import time as _time
    while True:
        try:
            cfg = get_email_config()
            if not cfg:
                _time.sleep(30)
                continue
            db  = get_db_connection()
            cur = db.cursor(buffered=True)
            cur.execute(
                "SELECT id, to_email, subject, html_body, attachment_b64, attachment_filename "
                "FROM email_queue WHERE status='pending' AND attempts < 3 "
                "ORDER BY created_at LIMIT 10"
            )
            rows = cur.fetchall()
            for row in rows:
                eid, to_email, subject, html_body, att_b64, att_name = row
                cur.execute(
                    "UPDATE email_queue SET status='sending', attempts=attempts+1 WHERE id=%s", (eid,)
                )
                db.commit()
                try:
                    att_bytes = base64.b64decode(att_b64) if att_b64 else None
                    send_email_smtp(to_email, subject, html_body, cfg,
                                    attachment_bytes=att_bytes,
                                    attachment_filename=att_name)
                    cur.execute(
                        "UPDATE email_queue SET status='done', sent_at=NOW() WHERE id=%s", (eid,)
                    )
                except Exception as exc:
                    app_log.error("Email queue: send failed to %s: %s", to_email, exc)
                    cur.execute(
                        "UPDATE email_queue SET status='pending', last_error=%s WHERE id=%s",
                        (str(exc)[:500], eid)
                    )
                db.commit()
            cur.execute(
                "UPDATE email_queue SET status='failed' "
                "WHERE status='pending' AND attempts >= 3"
            )
            db.commit()
            cur.close(); db.close()
        except Exception as _we:
            app_log.error("Email queue worker error: %s", _we)
        _time.sleep(15)


# Start the email queue worker once (gunicorn spawns multiple workers; each gets its own thread)
threading.Thread(target=_email_queue_worker, daemon=True, name="email-queue-worker").start()

def build_attendance_email(employee_name, emp_id, action, status, time_str, today_str):
    color = "#16a34a" if action == "login" else "#2563eb"
    action_label = "Checked In" if action == "login" else "Checked Out"
    return f"""
<div style="font-family:Segoe UI,sans-serif;max-width:520px;margin:auto;background:#f8fafc;border-radius:16px;overflow:hidden;border:1px solid #dbeafe;">
  <div style="background:#1e3a8a;padding:24px 28px;color:white;">
    <div style="font-size:20px;font-weight:700;">&#127970; Employee Attendance System</div>
    <div style="font-size:13px;opacity:0.75;margin-top:4px;">Attendance Confirmation</div>
  </div>
  <div style="padding:28px;">
    <p style="font-size:15px;color:#1e293b;margin-bottom:20px;">Hi <strong>{employee_name}</strong>,</p>
    <div style="background:#ffffff;border:1px solid #dbeafe;border-radius:12px;padding:20px;margin-bottom:20px;">
      <div style="font-size:28px;font-weight:700;color:{color};text-align:center;margin-bottom:4px;">{action_label}</div>
      <div style="text-align:center;color:#64748b;font-size:13px;">{today_str}</div>
      <hr style="border:none;border-top:1px solid #e2e8f0;margin:16px 0;">
      <table style="width:100%;font-size:14px;color:#1e293b;">
        <tr><td style="color:#64748b;padding:4px 0;">Employee ID</td><td style="text-align:right;font-weight:600;">{emp_id}</td></tr>
        <tr><td style="color:#64748b;padding:4px 0;">Time</td><td style="text-align:right;font-weight:600;">{time_str}</td></tr>
        <tr><td style="color:#64748b;padding:4px 0;">Status</td><td style="text-align:right;font-weight:600;color:{color};">{status}</td></tr>
      </table>
    </div>
    <p style="font-size:12px;color:#94a3b8;text-align:center;">This is an automated message. Please do not reply.</p>
  </div>
</div>"""

def get_employee_incentive_total(cursor, emp_id, year, month):
    cursor.execute(
        "SELECT COALESCE(SUM(amount),0) FROM employee_incentives WHERE employee_id=%s AND year=%s AND month=%s",
        (emp_id, year, month)
    )
    return float(cursor.fetchone()[0])

def compute_salary_entry(emp_id, name, spd, att_map, billable_past,
                         holidays_set=None, leave_dates=None):
    if holidays_set is None:
        holidays_set = set()
    if leave_dates is None:
        leave_dates = set()

    emp_att = att_map.get(emp_id, {})
    full_days = half_days = late_days = absent_days = 0
    holiday_days = leave_days_count = 0

    for d in billable_past:
        if d in holidays_set:
            if HOLIDAY_PAY == 'paid':
                full_days += 1
            else:
                absent_days += 1   # unpaid holiday = counts as absent deduction
            holiday_days += 1
        elif d in leave_dates:
            if LEAVE_PAY == 'absent':
                absent_days += 1   # count approved leave as absent
            else:
                leave_days_count += 1  # exclude from working days, no pay/no deduction
        else:
            row = emp_att.get(d)
            if row:
                _, _, login_t, logout_t, status, _logout_status, att_type = row
                final = att_type if att_type else infer_type_legacy(status, login_t, logout_t)
                if final == "Full Day":
                    full_days += 1
                elif final == "Late - Full Day":
                    late_days += 1
                elif final in ("Half Day", "Present"):
                    half_days += 1
                else:
                    absent_days += 1
            else:
                absent_days += 1

    effective_billable = len(billable_past) - leave_days_count

    spd_f      = float(spd)
    full_earn  = round(full_days  * spd_f, 2)
    late_earn  = round(late_days  * spd_f * (1 - LATE_DEDUCTION_RATE), 2)
    half_earn  = round(half_days  * spd_f * (1 - HALF_DAY_RATE), 2)
    net        = round(full_earn + late_earn + half_earn, 2)
    gross      = round(spd_f * effective_billable, 2)
    deduction  = round(gross - net, 2)

    return {
        "emp_id":        emp_id,
        "name":          name,
        "spd":           round(spd_f, 2),
        "billable":      effective_billable,
        "holiday_days":  holiday_days,
        "leave_days":    leave_days_count,
        "full_days":     full_days,
        "half_days":     half_days,
        "late_days":     late_days,
        "absent":        absent_days,
        "full_earn":     full_earn,
        "late_earn":     late_earn,
        "half_earn":     half_earn,
        "gross":         gross,
        "absent_ded":    round(absent_days * spd_f, 2),
        "half_ded":      round(half_days   * spd_f * HALF_DAY_RATE, 2),
        "late_ded":      round(late_days   * spd_f * LATE_DEDUCTION_RATE, 2),
        "deduction":     deduction,
        "net":           net,
    }

# ---------------- ERROR HANDLERS ----------------
import traceback as _traceback

def _error_page(code, icon, title, subtitle, hint):
    back_admin = session.get("admin_logged_in")
    back_emp   = session.get("employee_id")
    back_link  = "/admin" if back_admin else ("/employee_portal" if back_emp else "/")
    back_label = "Go to Admin Dashboard" if back_admin else ("Go to My Portal" if back_emp else "Go to Home")
    return f"""<!doctype html>
<html lang="en"><head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>{code} – {title}</title>
<style>
  *{{margin:0;padding:0;box-sizing:border-box;font-family:"Segoe UI",sans-serif}}
  body{{min-height:100vh;background:#f1f5f9;display:flex;align-items:center;justify-content:center;}}
  .box{{background:#fff;border:1px solid #e2e8f0;border-radius:20px;padding:52px 44px;text-align:center;max-width:480px;width:90%;box-shadow:0 8px 32px rgba(0,0,0,0.08);}}
  .icon{{font-size:72px;margin-bottom:18px;}}
  .code{{font-size:80px;font-weight:900;line-height:1;color:#1e3a8a;margin-bottom:6px;}}
  .title{{font-size:22px;font-weight:700;color:#1e293b;margin-bottom:8px;}}
  .sub{{font-size:14px;color:#64748b;margin-bottom:6px;line-height:1.6;}}
  .hint{{font-size:12px;color:#94a3b8;margin-bottom:28px;}}
  a.btn{{display:inline-block;padding:12px 28px;background:#1e3a8a;color:#fff;border-radius:10px;font-size:14px;font-weight:700;text-decoration:none;transition:0.2s;margin:4px;}}
  a.btn:hover{{background:#1d4ed8;}}
  a.sec{{display:inline-block;padding:12px 20px;background:#f1f5f9;color:#374151;border-radius:10px;font-size:14px;font-weight:600;text-decoration:none;transition:0.2s;margin:4px;border:1px solid #e2e8f0;}}
  a.sec:hover{{background:#e2e8f0;}}
</style></head><body>
<div class="box">
  <div class="icon">{icon}</div>
  <div class="code">{code}</div>
  <div class="title">{title}</div>
  <div class="sub">{subtitle}</div>
  <div class="hint">{hint}</div>
  <a href="{back_link}" class="btn">{back_label}</a>
  <a href="javascript:history.back()" class="sec">← Go Back</a>
</div>
</body></html>""", code

# ---------------- ERROR ALERTING (malfunction detection) ----------------
# The catch-all exception handler below tells users "the team has been
# notified" — this is what actually makes that true. Emails admins on
# unhandled errors, deduped by error signature so one hot failing endpoint
# can't flood inboxes or the email_queue table.
_error_alert_cache = {}
_error_alert_lock  = threading.Lock()
_ERROR_ALERT_COOLDOWN = 900  # 15 min — same error signature won't re-alert sooner

def _alert_on_error(tb_text, context=""):
    """Best-effort: any failure here is swallowed so alerting can never
    mask or crash on top of the original error being reported."""
    try:
        # Dedup key = exception type + the line that actually raised it, not
        # the full traceback (which varies request-to-request — different
        # IDs, line numbers in called code, etc. would defeat deduping).
        last_line = tb_text.strip().splitlines()[-1] if tb_text.strip() else "unknown"
        sig = hashlib.sha256(last_line.encode()).hexdigest()[:16]
        now = time.time()
        with _error_alert_lock:
            last_sent = _error_alert_cache.get(sig)
            if last_sent and now - last_sent < _ERROR_ALERT_COOLDOWN:
                return
            _error_alert_cache[sig] = now
            if len(_error_alert_cache) > 500:  # cap unbounded growth
                _error_alert_cache.clear()

        cfg = get_email_config()
        if not cfg:
            return
        admins = get_admin_emails()
        if not admins:
            return

        body = (
            "<pre style='white-space:pre-wrap;font-family:monospace;font-size:13px'>"
            f"Time: {datetime.datetime.now().isoformat()}\n"
            f"Route: {_html.escape(context or 'unknown')}\n"
            f"Method: {_html.escape(request.method if request else 'n/a')}\n"
            f"Remote IP: {_html.escape(request.remote_addr or '') if request else 'n/a'}\n\n"
            f"{_html.escape(tb_text)}</pre>"
        )
        subject = f"⚠️ Application error — {context or 'unknown route'}"
        for admin_email in admins:
            send_email_async(admin_email, subject, body, cfg)
    except Exception as _alert_err:
        app_log.error("Failed to send error alert email: %s", _alert_err)


@app.errorhandler(404)
def not_found(e):
    return _error_page(404, "🔍", "Page Not Found",
        "The page you're looking for doesn't exist or has been moved.",
        "Check the URL or use one of the links below to get back on track.")

@app.errorhandler(403)
def forbidden(e):
    return _error_page(403, "🔒", "Access Denied",
        "You don't have permission to access this page.",
        "Please log in with the right account or contact your administrator.")

@app.errorhandler(500)
def internal_error(e):
    tb = _traceback.format_exc()
    app_log.error("500 error: %s", tb.replace('\n', '\\n'))
    _alert_on_error(tb, context=request.path if request else "")
    return _error_page(500, "⚙️", "Internal Server Error",
        "Something went wrong on our end. The error has been logged.",
        "Please try again in a moment or contact your administrator.")

@app.errorhandler(Exception)
def unhandled_exception(e):
    if isinstance(e, HTTPException):
        return _error_page(e.code, "⚠️", e.name,
            "The page you requested could not be processed.",
            "Use the buttons below to navigate back.")
    tb = _traceback.format_exc()
    app_log.error("Unhandled exception: %s", tb.replace('\n', '\\n'))
    _alert_on_error(tb, context=request.path if request else "")
    return _error_page(500, "⚙️", "Internal Server Error",
        "An unexpected error occurred. The team has been notified.",
        "Please try again or contact your administrator.")

# ---------------- HOME ----------------
@app.route("/")
def home():
    return render_template("index.html", auth_cfg=get_auth_config())

# ---------------- ADMIN LOGIN ----------------
@app.route("/admin")
@admin_required
def admin():
    db     = get_db_connection()
    cursor = db.cursor(buffered=True)
    today  = datetime.date.today()
    active_cid = session.get("active_company_id")
    _co_filter = "AND e.company_id=%s" if active_cid else ""
    _co_sub    = "AND employee_id IN (SELECT employee_id FROM employees WHERE company_id=%s)" if active_cid else ""
    _co_args   = (active_cid,) if active_cid else ()

    if active_cid:
        cursor.execute("SELECT COUNT(*) FROM employees WHERE company_id=%s", _co_args)
    else:
        cursor.execute("SELECT COUNT(*) FROM employees")
    total = cursor.fetchone()[0]

    cursor.execute(
        f"SELECT COUNT(DISTINCT employee_id) FROM attendance WHERE date=%s AND login_time IS NOT NULL {_co_sub}",
        (today,) + _co_args
    )
    present = cursor.fetchone()[0]

    cursor.execute(
        f"SELECT COUNT(DISTINCT employee_id) FROM attendance WHERE date=%s AND status='Late Login' {_co_sub}",
        (today,) + _co_args
    )
    late = cursor.fetchone()[0]

    cursor.execute(
        f"SELECT e.employee_id, e.name, a.login_time, a.logout_time, a.status, "
        f"       a.logout_status, a.attendance_type, e.role "
        f"FROM employees e "
        f"LEFT JOIN attendance a ON e.employee_id=a.employee_id AND a.date=%s "
        f"WHERE 1=1 {_co_filter} ORDER BY e.name",
        (today,) + _co_args
    )
    today_rows = cursor.fetchall()

    if active_cid:
        cursor.execute("SELECT employee_id, name FROM employees WHERE company_id=%s ORDER BY name", _co_args)
    else:
        cursor.execute("SELECT employee_id, name FROM employees ORDER BY name")
    all_employees = cursor.fetchall()

    cursor.execute(
        f"SELECT COUNT(*) FROM leave_requests WHERE status='Pending' {_co_sub}",
        _co_args
    )
    pending_leaves = cursor.fetchone()[0]

    cursor.execute(
        f"SELECT COUNT(*) FROM resignation_requests WHERE status='Pending' {_co_sub}",
        _co_args
    )
    pending_resignations = cursor.fetchone()[0]

    cursor.execute(
        f"SELECT COUNT(*) FROM tickets WHERE status IN ('Open','In Progress') {_co_sub}",
        _co_args
    )
    pending_tickets = cursor.fetchone()[0]

    try:
        cursor.execute("SELECT COUNT(*) FROM overtime_records WHERE status='Pending'")
        pending_ot = cursor.fetchone()[0]
    except Exception:
        pending_ot = 0

    cursor.execute("SELECT id, name, COALESCE(code,'') FROM companies ORDER BY name")
    companies_list = cursor.fetchall()

    # Onboarding summary for dashboard widget
    try:
        cursor.execute("""
            SELECT
              SUM(CASE WHEN status != 'Completed' THEN 1 ELSE 0 END),
              SUM(CASE WHEN status = 'Completed' THEN 1 ELSE 0 END),
              SUM(CASE WHEN status != 'Completed' AND due_date < %s THEN 1 ELSE 0 END)
            FROM employee_onboarding
        """, (today,))
        _ob = cursor.fetchone()
        ob_active    = int(_ob[0] or 0)
        ob_completed = int(_ob[1] or 0)
        ob_overdue   = int(_ob[2] or 0)
        cursor.execute("""
            SELECT eo.id, e.name, ot.name, eo.due_date
            FROM employee_onboarding eo
            JOIN employees e ON eo.employee_id = e.employee_id
            JOIN onboarding_templates ot ON eo.template_id = ot.id
            WHERE eo.status != 'Completed' AND eo.due_date < %s
            ORDER BY eo.due_date LIMIT 5
        """, (today,))
        ob_overdue_list = cursor.fetchall()
    except Exception:
        ob_active = ob_completed = ob_overdue = 0
        ob_overdue_list = []

    cursor.execute("SELECT id, break_name, break_time, duration_minutes, is_active FROM break_config ORDER BY break_time")
    break_rows = cursor.fetchall()
    breaks_display = []
    for b in break_rows:
        bt = b[2]
        if hasattr(bt, 'seconds'):
            h, m = divmod(bt.seconds // 60, 60)
        else:
            h, m = bt.hour, bt.minute
        ampm = "AM" if h < 12 else "PM"
        h12 = h % 12 or 12
        breaks_display.append({
            "id": b[0], "name": b[1],
            "time_str": "%02d:%02d %s" % (h12, m, ampm),
            "duration": b[3], "is_active": b[4]
        })

    cursor.close()
    db.close()

    return render_template("admin.html",
        total=total,
        present=present,
        absent=total - present,
        late=late,
        today=today.strftime("%d %b %Y"),
        active_nav="dashboard",
        today_rows=today_rows,
        all_employees=all_employees,
        shift_start=SHIFT_START.strftime("%I:%M %p"),
        shift_end=SHIFT_END.strftime("%I:%M %p"),
        pending_leaves=pending_leaves,
        pending_resignations=pending_resignations,
        pending_ot=pending_ot,
        pending_tickets=pending_tickets,
        now_month=today.month,
        now_year=today.year,
        breaks_display=breaks_display,
        companies_list=companies_list,
        ob_active=ob_active,
        ob_completed=ob_completed,
        ob_overdue=ob_overdue,
        ob_overdue_list=ob_overdue_list,
    )

# ---------------- LIVE DASHBOARD API ----------------
@app.route("/api/dashboard_live")
@admin_required
def dashboard_live():
    def fmt(t):
        if t is None:
            return None
        if hasattr(t, "strftime"):
            return t.strftime("%H:%M:%S")
        total = int(t.total_seconds())
        h, rem = divmod(total, 3600)
        m, s   = divmod(rem, 60)
        return f"{h:02d}:{m:02d}:{s:02d}"

    db     = get_db_connection()
    cursor = db.cursor(buffered=True)
    today  = datetime.date.today()
    active_cid = session.get("active_company_id")
    _co_filter = "AND e.company_id=%s" if active_cid else ""
    _co_sub    = "AND employee_id IN (SELECT employee_id FROM employees WHERE company_id=%s)" if active_cid else ""
    _co_args   = (active_cid,) if active_cid else ()

    if active_cid:
        cursor.execute("SELECT COUNT(*) FROM employees WHERE company_id=%s", _co_args)
    else:
        cursor.execute("SELECT COUNT(*) FROM employees")
    total = cursor.fetchone()[0]

    cursor.execute(
        f"SELECT COUNT(DISTINCT employee_id) FROM attendance WHERE date=%s AND login_time IS NOT NULL {_co_sub}",
        (today,) + _co_args
    )
    present = cursor.fetchone()[0]

    cursor.execute(
        f"SELECT COUNT(DISTINCT employee_id) FROM attendance WHERE date=%s AND status='Late Login' {_co_sub}",
        (today,) + _co_args
    )
    late = cursor.fetchone()[0]

    cursor.execute(
        f"SELECT e.employee_id, e.name, a.login_time, a.logout_time, "
        f"       a.status, a.logout_status, a.attendance_type, e.role "
        f"FROM employees e "
        f"LEFT JOIN attendance a ON e.employee_id=a.employee_id AND a.date=%s "
        f"WHERE 1=1 {_co_filter} ORDER BY e.name",
        (today,) + _co_args
    )
    rows = []
    for emp_id, name, login_t, logout_t, status, logout_s, att_type, role in cursor.fetchall():
        rows.append({
            "emp_id":   emp_id,
            "name":     name,
            "role":     role or "",
            "login_t":  fmt(login_t),
            "logout_t": fmt(logout_t),
            "status":   status or "",
            "logout_s": logout_s or "",
            "att_type": att_type or "",
        })

    cursor.execute(f"SELECT COUNT(*) FROM leave_requests WHERE status='Pending' {_co_sub}", _co_args)
    pending_leaves = cursor.fetchone()[0]

    cursor.execute(f"SELECT COUNT(*) FROM resignation_requests WHERE status='Pending' {_co_sub}", _co_args)
    pending_resignations = cursor.fetchone()[0]

    cursor.execute(f"SELECT COUNT(*) FROM tickets WHERE status IN ('Open','In Progress') {_co_sub}", _co_args)
    pending_tickets = cursor.fetchone()[0]

    cursor.close(); db.close()

    return jsonify({
        "total":   total,
        "present": present,
        "absent":  total - present,
        "late":    late,
        "rows":    rows,
        "pending_leaves":       pending_leaves,
        "pending_resignations": pending_resignations,
        "pending_tickets":      pending_tickets,
    })

# ---------------- CHART DATA API ----------------


# ---------------- TODAY FILTERED VIEWS ----------------
def _today_pending_counts(cursor):
    cursor.execute("SELECT COUNT(*) FROM leave_requests WHERE status='Pending'")
    pl = cursor.fetchone()[0]
    cursor.execute("SELECT COUNT(*) FROM resignation_requests WHERE status='Pending'")
    pr = cursor.fetchone()[0]
    cursor.execute("SELECT COUNT(*) FROM tickets WHERE status IN ('Open','In Progress')")
    pt = cursor.fetchone()[0]
    return pl, pr, pt

@app.route("/today_present")
@admin_required
def today_present():
    db = get_db_connection()
    cursor = db.cursor(buffered=True)
    today = datetime.date.today()
    active_cid = session.get("active_company_id")
    _co = "AND e.company_id=%s" if active_cid else ""
    _args = (today,) + ((active_cid,) if active_cid else ())
    cursor.execute(f"""
        SELECT e.employee_id, e.name, e.role, a.login_time, a.logout_time,
               a.status, a.logout_status, a.attendance_type
        FROM employees e
        JOIN attendance a ON e.employee_id = a.employee_id AND a.date = %s
        WHERE a.login_time IS NOT NULL {_co}
        ORDER BY a.login_time
    """, _args)
    rows = cursor.fetchall()
    pl, pr, pt = _today_pending_counts(cursor)
    cursor.close(); db.close()
    return render_template("today_attendance.html",
        filter_type="present", title="Present Today",
        rows=rows, today=today.strftime("%d %b %Y",
        active_nav="attendance",
    ),
        pending_leaves=pl, pending_resignations=pr, pending_tickets=pt)

@app.route("/today_absent")
@admin_required
def today_absent():
    db = get_db_connection()
    cursor = db.cursor(buffered=True)
    today = datetime.date.today()
    active_cid = session.get("active_company_id")
    _co = "AND e.company_id=%s" if active_cid else ""
    _args = (today,) + ((active_cid,) if active_cid else ())
    cursor.execute(f"""
        SELECT e.employee_id, e.name, e.role
        FROM employees e
        LEFT JOIN attendance a ON e.employee_id = a.employee_id AND a.date = %s
        WHERE a.employee_id IS NULL {_co}
        ORDER BY e.name
    """, _args)
    rows = cursor.fetchall()
    pl, pr, pt = _today_pending_counts(cursor)
    cursor.close(); db.close()
    return render_template("today_attendance.html",
        filter_type="absent", title="Absent Today",
        rows=rows, today=today.strftime("%d %b %Y",
        active_nav="attendance",
    ),
        pending_leaves=pl, pending_resignations=pr, pending_tickets=pt)

@app.route("/today_late")
@admin_required
def today_late():
    db = get_db_connection()
    cursor = db.cursor(buffered=True)
    today = datetime.date.today()
    active_cid = session.get("active_company_id")
    _co = "AND e.company_id=%s" if active_cid else ""
    _args = (today,) + ((active_cid,) if active_cid else ())
    cursor.execute(f"""
        SELECT e.employee_id, e.name, e.role, a.login_time, a.status
        FROM employees e
        JOIN attendance a ON e.employee_id = a.employee_id AND a.date = %s
        WHERE a.status IN ('Late Login', 'Half Day Login') {_co}
        ORDER BY a.login_time
    """, _args)
    rows = cursor.fetchall()
    pl, pr, pt = _today_pending_counts(cursor)
    cursor.close(); db.close()
    return render_template("today_attendance.html",
        filter_type="late", title="Late Logins Today",
        rows=rows, today=today.strftime("%d %b %Y",
        active_nav="attendance",
    ),
        pending_leaves=pl, pending_resignations=pr, pending_tickets=pt)

# ---------------- ADMIN ACTIONS ----------------
@app.route("/admin_action", methods=["POST"])
@admin_required
def admin_action():
    action = request.form.get("action")
    db     = get_db_connection()
    cursor = db.cursor(buffered=True)

    if action == "register":
        try:
            name            = request.form["name"]
            emp_id          = request.form["emp_id"].strip()
            email           = request.form.get("email", "").strip() or None
            role            = request.form.get("role", "").strip() or None
            date_of_joining = request.form.get("date_of_joining", "").strip() or None
            work_mode       = request.form.get("work_mode", "office").strip() or "office"
            work_lat_raw    = request.form.get("work_lat", "").strip()
            work_lon_raw    = request.form.get("work_lon", "").strip()
            work_lat        = float(work_lat_raw) if work_lat_raw else None
            work_lon        = float(work_lon_raw) if work_lon_raw else None
            company_id_raw  = request.form.get("company_id", "").strip()
            company_id      = int(company_id_raw) if company_id_raw.isdigit() else None
            # Extended fields
            department      = request.form.get("department", "").strip() or None
            phone           = request.form.get("phone", "").strip() or None
            manager_id      = request.form.get("manager_id", "").strip() or None
            manager_name    = request.form.get("manager_name", "").strip() or None
            salary_per_day_raw = request.form.get("salary_per_day", "").strip()
            salary_per_day  = float(salary_per_day_raw) if salary_per_day_raw else None
            gender          = request.form.get("gender", "").strip() or None
            dob_raw         = request.form.get("dob", "").strip()
            dob             = dob_raw if dob_raw else None
            blood_group     = request.form.get("blood_group", "").strip() or None
            address         = request.form.get("address", "").strip() or None
            city            = request.form.get("city", "").strip() or None
            state           = request.form.get("state", "").strip() or None
            pincode         = request.form.get("pincode", "").strip() or None
            ec_name         = request.form.get("emergency_contact_name", "").strip() or None
            ec_phone        = request.form.get("emergency_contact_phone", "").strip() or None
            ec_relation     = request.form.get("emergency_contact_relation", "").strip() or None
            aadhar          = encrypt_pii(request.form.get("aadhar_number", "").strip() or None)
            pan             = encrypt_pii(request.form.get("pan_number", "").strip().upper() or None)
            bank_name       = request.form.get("bank_name", "").strip() or None
            bank_account    = encrypt_pii(request.form.get("bank_account", "").strip() or None)
            bank_ifsc       = encrypt_pii(request.form.get("bank_ifsc", "").strip().upper() or None)
            uan             = encrypt_pii(request.form.get("uan_number", "").strip() or None)
            file            = request.files["face"]
        except (KeyError, ValueError) as _e:
            cursor.close(); db.close()
            flash(f"Missing or invalid field in registration form: {_e}", "error")
            return redirect("/admin")
        # Auto-increment emp_id if it's already taken
        cursor.execute("SELECT 1 FROM employees WHERE employee_id = %s", (emp_id,))
        if cursor.fetchone():
            prefix = ''.join(c for c in emp_id if not c.isdigit())
            if prefix:
                cursor.execute(
                    "SELECT employee_id FROM employees WHERE employee_id LIKE %s",
                    (prefix + "%",)
                )
                max_seq = 0
                for (eid,) in cursor.fetchall():
                    sfx = eid[len(prefix):]
                    if sfx.isdigit():
                        max_seq = max(max_seq, int(sfx))
                emp_id = f"{prefix}{max_seq + 1:03d}"
        _img_ok, _img_err = _validate_image_file(file)
        if not _img_ok:
            flash(_img_err, "error")
            cursor.close(); db.close()
            return redirect("/admin")
        filepath = os.path.join(app.config["UPLOAD_FOLDER"], emp_id + ".jpg")
        file.save(filepath)

        # Validate that the uploaded photo contains a detectable face
        if _face_recognition_available:
            test_img = face_recognition.load_image_file(filepath)
            if not face_recognition.face_encodings(test_img):
                os.remove(filepath)
                flash("No face detected in the uploaded photo. Please upload a clear, well-lit front-facing photo.", "error")
                cursor.close()
                db.close()
                return redirect("/admin")

        qr_path    = generate_qr(emp_id)
        auto_pass  = secrets.token_urlsafe(8)   # e.g. "aB3xQ7mR"
        hashed_pwd = generate_password_hash(auto_pass)
        try:
            cursor.execute(
                "INSERT INTO employees (name, employee_id, email, role, face_image, qr_code, password, "
                "date_of_joining, work_mode, work_lat, work_lon, company_id, "
                "department, phone, manager_id, manager_name, "
                "gender, dob, blood_group, "
                "address, city, state, pincode, "
                "emergency_contact_name, emergency_contact_phone, emergency_contact_relation, "
                "aadhar_number, pan_number, bank_name, bank_account, bank_ifsc, uan_number, "
                "force_pin_change) "
                "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,"
                "%s,%s,%s,%s,"
                "%s,%s,%s,"
                "%s,%s,%s,%s,"
                "%s,%s,%s,"
                "%s,%s,%s,%s,%s,%s,1)",
                (name, emp_id, email, role, filepath, qr_path, hashed_pwd,
                 date_of_joining, work_mode, work_lat, work_lon, company_id,
                 department, phone, manager_id, manager_name,
                 gender, dob, blood_group,
                 address, city, state, pincode,
                 ec_name, ec_phone, ec_relation,
                 aadhar, pan, bank_name, bank_account, bank_ifsc, uan)
            )
            db.commit()
            if salary_per_day is not None:
                cursor.execute(
                    "INSERT INTO salary_config (employee_id, salary_per_day) VALUES (%s,%s) "
                    "ON CONFLICT (employee_id) DO UPDATE SET salary_per_day=%s",
                    (emp_id, salary_per_day, salary_per_day)
                )
                db.commit()
            _enroll_fingerprint_from_form(emp_id, cursor, db)
            assign_leave_balances_for_employee(cursor, emp_id)
            db.commit()
            flash(f"✅ Employee '{name}' registered! ID: {emp_id} | Password: {auto_pass}", "success")
            # Send welcome email with credentials
            if not email:
                flash("⚠️ No email address provided — credentials email not sent. Share them manually.", "error")
            else:
                _ecfg = get_email_config()
                if not _ecfg:
                    flash("⚠️ SMTP not configured — credentials email not sent. Go to Email Settings to set it up.", "error")
                else:
                    _welcome_html = f"""
<div style="font-family:'Segoe UI',sans-serif;max-width:520px;margin:0 auto;background:#f8fafc;padding:32px 24px;border-radius:16px;">
  <div style="background:linear-gradient(135deg,#1e3a8a,#2563eb);border-radius:12px;padding:28px 24px;text-align:center;margin-bottom:24px;">
    <div style="font-size:36px;margin-bottom:8px;">👋</div>
    <h1 style="color:#fff;font-size:22px;margin:0;">Welcome to the Team!</h1>
    <p style="color:rgba(255,255,255,0.8);font-size:14px;margin:6px 0 0;">Your employee account has been created</p>
  </div>
  <p style="color:#1e293b;font-size:15px;margin-bottom:20px;">Hi <strong>{name}</strong>, here are your login credentials for the Attendance Portal:</p>
  <div style="background:#fff;border:1px solid #dbeafe;border-radius:12px;padding:20px 24px;margin-bottom:20px;">
    <table style="width:100%;font-size:14px;border-collapse:collapse;">
      <tr>
        <td style="color:#64748b;padding:8px 0;border-bottom:1px solid #f1f5f9;font-weight:600;width:40%;">Employee ID</td>
        <td style="color:#1e293b;padding:8px 0;border-bottom:1px solid #f1f5f9;font-weight:700;">{emp_id}</td>
      </tr>
      <tr>
        <td style="color:#64748b;padding:8px 0;font-weight:600;">Password</td>
        <td style="color:#1e293b;padding:8px 0;font-weight:700;font-family:monospace;font-size:15px;">{auto_pass}</td>
      </tr>
    </table>
  </div>
  <div style="background:#fffbeb;border:1px solid #fde68a;border-radius:10px;padding:12px 16px;font-size:13px;color:#92400e;margin-bottom:20px;">
    🔒 Please change your password after your first login for security.
  </div>
  <p style="color:#64748b;font-size:12px;text-align:center;margin:0;">This is an automated message — please do not reply.</p>
</div>"""
                    try:
                        send_email_smtp(email, f"Welcome {name} — Your Login Credentials", _welcome_html, _ecfg)
                        flash(f"📧 Credentials email sent to {email}", "success")
                    except Exception as _mail_err:
                        flash(f"⚠️ Email delivery failed: {_mail_err}. Share credentials manually.", "error")
        except psycopg2.IntegrityError:
            db.rollback()
            os.remove(filepath)
            flash(f"Employee ID '{emp_id}' already exists. Please use a different ID.", "error")
            cursor.close()
            db.close()
            return redirect("/admin")

    elif action == "update_face":
        emp_id   = request.form["emp_id"]
        file     = request.files["face"]
        cursor.execute("SELECT name FROM employees WHERE employee_id=%s", (emp_id,))
        row = cursor.fetchone()
        if not row:
            flash(f"Employee ID '{emp_id}' not found.", "error")
            cursor.close()
            db.close()
            return redirect("/admin")
        name = row[0]
        _img_ok, _img_err = _validate_image_file(file)
        if not _img_ok:
            flash(_img_err, "error")
            cursor.close(); db.close()
            return redirect("/admin")
        filepath = os.path.join(app.config["UPLOAD_FOLDER"], emp_id + ".jpg")
        file.save(filepath)
        if _face_recognition_available:
            test_img = face_recognition.load_image_file(filepath)
            if not face_recognition.face_encodings(test_img):
                os.remove(filepath)
                flash("No face detected in the uploaded photo. Please upload a clear, well-lit front-facing photo.", "error")
                cursor.close()
                db.close()
                return redirect("/admin")
        cursor.execute("UPDATE employees SET face_image=%s WHERE employee_id=%s", (filepath, emp_id))
        db.commit()
        flash(f"Face photo updated successfully for '{name}' (ID: {emp_id}).", "success")

    elif action == "reset_password":
        emp_id = request.form.get("emp_id", "").strip()
        cursor.execute("SELECT name FROM employees WHERE employee_id=%s", (emp_id,))
        row = cursor.fetchone()
        if not row:
            flash(f"Employee ID '{emp_id}' not found.", "error")
        else:
            cursor.execute(
                "UPDATE employees SET password=%s WHERE employee_id=%s",
                (generate_password_hash(emp_id), emp_id)
            )
            db.commit()
            flash(f"Password reset for '{row[0]}' ({emp_id}). They can now login using their Employee ID as the password.", "success")

    elif action == "holiday":
        cursor.execute(
            "INSERT INTO holidays (date, name) VALUES (%s,%s)",
            (request.form["date"], request.form["holiday_name"])
        )
        db.commit()

    elif action == "salary":
        emp_id = request.form["emp_id"]
        salary = request.form["salary"]
        cursor.execute("SELECT 1 FROM salary_config WHERE employee_id=%s", (emp_id,))
        if cursor.fetchone():
            cursor.execute(
                "UPDATE salary_config SET salary_per_day=%s WHERE employee_id=%s",
                (salary, emp_id)
            )
        else:
            cursor.execute(
                "INSERT INTO salary_config (employee_id, salary_per_day) VALUES (%s,%s)",
                (emp_id, salary)
            )
        db.commit()

    cursor.close()
    db.close()
    return redirect("/admin")

# ---------------- SETTINGS (unified) ----------------
@app.route("/settings")
@admin_required
def settings_page():
    tab    = request.args.get("tab", "company")
    db     = get_db_connection()
    cursor = db.cursor(buffered=True)

    # Email config
    cursor.execute("SELECT smtp_host, smtp_port, smtp_user, smtp_pass, from_name, from_email FROM email_config ORDER BY id DESC LIMIT 1")
    row = cursor.fetchone()
    email_config = {"host": row[0], "port": row[1], "user": row[2], "password": row[3], "from_name": row[4], "from_email": row[5] or row[2]} if row else None

    # Shifts (with company)
    cursor.execute("""
        SELECT s.id, s.name, s.start_time, s.half_time, s.end_time,
               COALESCE(s.company_id, 0), COALESCE(c.name, '')
        FROM shifts s
        LEFT JOIN companies c ON c.id = s.company_id
        ORDER BY c.name, s.start_time
    """)
    shift_rows = []
    for sid, sname, st, ht, et, scid, scname in cursor.fetchall():
        shift_rows.append({
            "id": sid, "name": sname,
            "start": _td_to_time(st).strftime("%H:%M") if st else "--",
            "half":  _td_to_time(ht).strftime("%H:%M") if ht else "--",
            "end":   _td_to_time(et).strftime("%H:%M") if et else "--",
            "company_id": scid, "company_name": scname,
        })
    cursor.execute("SELECT e.employee_id, e.name, e.role, s.name FROM employees e LEFT JOIN shifts s ON e.shift_id = s.id ORDER BY e.name")
    emp_list = [{"emp_id": r[0], "name": r[1], "role": r[2] or "", "shift": r[3] or "Default"} for r in cursor.fetchall()]

    # Company-specific shifts (company_id IS NOT NULL)
    cursor.execute("SELECT id, name, start_time, half_time, end_time, company_id FROM shifts WHERE company_id IS NOT NULL ORDER BY company_id, start_time")
    _co_shifts_raw = cursor.fetchall()
    company_shifts = {}
    for _csid, _csname, _csstart, _cshalf, _csend, _cscid in _co_shifts_raw:
        def _tdfmt(v):
            if v is None: return "--"
            if isinstance(v, datetime.timedelta):
                _s = int(v.total_seconds()); return "%02d:%02d" % (_s // 3600, (_s % 3600) // 60)
            if isinstance(v, datetime.time): return v.strftime("%H:%M")
            return str(v)[:5]
        company_shifts.setdefault(_cscid, []).append((_csid, _csname, _tdfmt(_csstart), _tdfmt(_cshalf), _tdfmt(_csend)))

    # Company-specific breaks (company_id IS NOT NULL), nested per shift
    cursor.execute("SELECT id, break_name, break_time, duration_minutes, is_active, company_id, COALESCE(shift_id,0) FROM break_config WHERE company_id IS NOT NULL ORDER BY company_id, shift_id, break_time")
    _co_breaks_raw = cursor.fetchall()
    company_breaks = {}
    for _cbid, _cbname, _cbt, _cbdur, _cbactive, _cbcid, _cbsid in _co_breaks_raw:
        if _cbt is None: _cbt_str = "--"
        elif isinstance(_cbt, datetime.timedelta):
            _s = int(_cbt.total_seconds()); _cbt_str = "%02d:%02d" % (_s // 3600, (_s % 3600) // 60)
        elif isinstance(_cbt, datetime.time): _cbt_str = _cbt.strftime("%H:%M")
        else: _cbt_str = str(_cbt)[:5]
        company_breaks.setdefault(_cbcid, {}).setdefault(_cbsid, []).append((_cbid, _cbname, _cbt_str, _cbdur, _cbactive))

    # Breaks (with shift_id) — pre-format break_time as HH:MM
    cursor.execute("SELECT id, break_name, break_time, duration_minutes, is_active, COALESCE(shift_id,0) FROM break_config WHERE company_id IS NULL ORDER BY shift_id, break_time")
    breaks = []
    for _bid, _bname, _bt, _bdur, _bactive, _bshift in cursor.fetchall():
        if _bt is None:
            _bt_str = "--"
        elif isinstance(_bt, datetime.timedelta):
            _s = int(_bt.total_seconds()); _bt_str = "%02d:%02d" % (_s // 3600, (_s % 3600) // 60)
        elif isinstance(_bt, datetime.time):
            _bt_str = _bt.strftime("%H:%M")
        else:
            _bt_str = str(_bt)[:5]
        breaks.append((_bid, _bname, _bt_str, _bdur, _bactive, _bshift))

    # Salary
    cursor.execute("""
        SELECT e.employee_id, e.name, COALESCE(s.salary_per_day, 0), e.role, s.last_revised,
               COALESCE(e.phone,''), COALESCE(e.email,'')
        FROM employees e
        LEFT JOIN salary_config s ON e.employee_id = s.employee_id
        ORDER BY e.name
    """)
    salaries = cursor.fetchall()

    # Announcements (admin sees all; include visibility and target employee name)
    cursor.execute("""
        SELECT a.id, a.title, a.content, a.priority, a.created_at,
               COALESCE(a.visibility,'public'), COALESCE(a.target_employee_id,''), COALESCE(e.name,'')
        FROM announcements a
        LEFT JOIN employees e ON e.employee_id = a.target_employee_id
        ORDER BY a.created_at DESC
    """)
    ann_list = cursor.fetchall()
    pub_anns  = [r for r in ann_list if r[5] == 'public']
    priv_anns = [r for r in ann_list if r[5] == 'private']

    # Employee list for private announcement targeting
    cursor.execute("SELECT employee_id, name FROM employees WHERE is_active=1 ORDER BY name")
    ann_emp_list = cursor.fetchall()

    # Pending counts
    cursor.execute("SELECT COUNT(*) FROM leave_requests WHERE status='Pending'")
    pending_leaves = cursor.fetchone()[0]
    cursor.execute("SELECT COUNT(*) FROM resignation_requests WHERE status='Pending'")
    pending_resignations = cursor.fetchone()[0]
    cursor.execute("SELECT COUNT(*) FROM tickets WHERE status='Open'")
    pending_tickets = cursor.fetchone()[0]

    cursor.execute("SELECT COALESCE(company_code,''), COALESCE(default_onboarding_template_id,0) FROM company_settings LIMIT 1")
    _cr = cursor.fetchone()
    company_code = _cr[0] if _cr else ""
    default_onboarding_tpl = int(_cr[1]) if _cr and _cr[1] else 0

    # Company stats
    cursor.execute("SELECT COUNT(*) FROM employees")
    total_employees = cursor.fetchone()[0]
    cursor.execute("""
        SELECT COUNT(*) FROM employees e
        WHERE NOT EXISTS (
            SELECT 1 FROM resignation_requests r
            WHERE r.employee_id = e.employee_id AND r.status = 'Approved'
        )
    """)
    active_employees = cursor.fetchone()[0]
    cursor.execute("SELECT COUNT(DISTINCT department) FROM employees WHERE department IS NOT NULL AND department != ''")
    total_departments = cursor.fetchone()[0]
    cursor.execute("SELECT COUNT(*) FROM shifts")
    total_shifts = cursor.fetchone()[0]
    cursor.execute("SELECT id, name FROM onboarding_templates WHERE is_active=1 ORDER BY name")
    onboarding_templates = cursor.fetchall()

    cursor.execute("""
        SELECT c.id, c.name, COALESCE(c.code,''), c.created_at,
               COUNT(e.id) AS emp_count,
               COALESCE(c.working_days,'Mon,Tue,Wed,Thu,Fri'),
               CASE WHEN c.pin IS NOT NULL AND c.pin != '' THEN 1 ELSE 0 END AS has_pin
        FROM companies c
        LEFT JOIN employees e ON e.company_id = c.id
        GROUP BY c.id, c.name, c.code, c.created_at, c.working_days, c.pin
        ORDER BY c.name
    """)
    companies = cursor.fetchall()

    # Feature flags — per-company when active, global otherwise
    _active_cid_settings = session.get("active_company_id")
    fr = get_co_features(_active_cid_settings)
    cursor.execute("SELECT COALESCE(working_days,'Mon,Tue,Wed,Thu,Fri'), COALESCE(company_name,''), COALESCE(timezone,'Asia/Kolkata') FROM company_settings LIMIT 1")
    _gset = cursor.fetchone()
    features = {
        "face_auth":    fr["face_auth_enabled"],
        "geo":          fr["geo_enabled"],
        "geo_radius":   fr["geo_radius"],
        "qr":           fr["qr_enabled"],
        "pin":          fr["pin_enabled"],
        "fingerprint":  fr["fingerprint_enabled"],
        "biometric":    fr["biometric_enabled"],
        "notify_leave": fr["notify_leave"],
        "notify_payslip": fr["notify_payslip"],
        "notify_resignation": fr["notify_resignation"],
        "notify_doc_expiry":  fr["notify_doc_expiry"],
        "session_timeout": fr["session_timeout"],
        "working_days": (_gset[0] if _gset else "Mon,Tue,Wed,Thu,Fri").split(","),
        "company_name": _gset[1] if _gset else "",
        "timezone":     _gset[2] if _gset else "Asia/Kolkata",
        # salary rules from company features
        "late_deduction_pct": fr["late_deduction_pct"],
        "half_day_deduction_pct": fr["half_day_deduction_pct"],
        "grace_minutes": fr["grace_minutes"],
        "holiday_pay": fr["holiday_pay"],
        "leave_pay": fr["leave_pay"],
        "shift_start": fr["shift_start"],
        "shift_half": fr["shift_half"],
        "shift_end": fr["shift_end"],
    }

    # Resolve salary/shift display values: company-specific overrides global
    def _td_str(v):
        if v is None: return None
        if isinstance(v, str): return v[:5]
        if isinstance(v, datetime.timedelta):
            t = int(v.total_seconds()); return "%02d:%02d" % (t//3600, (t%3600)//60)
        if isinstance(v, datetime.time): return v.strftime("%H:%M")
        return str(v)[:5]

    _co_shift_start = _td_str(fr.get("shift_start")) or SHIFT_START.strftime("%H:%M")
    _co_shift_half  = _td_str(fr.get("shift_half"))  or SHIFT_HALF.strftime("%H:%M")
    _co_shift_end   = _td_str(fr.get("shift_end"))   or SHIFT_END.strftime("%H:%M")

    cursor.close(); db.close()
    return render_template("settings.html",
        tab=tab,
        email_config=email_config,
        company_code=company_code,
        total_employees=total_employees,
        active_employees=active_employees,
        total_departments=total_departments,
        total_shifts=total_shifts,
        companies=companies,
        company_shifts=company_shifts,
        company_breaks=company_breaks,
        shifts=shift_rows,
        emp_list=emp_list,
        breaks=breaks,
        salaries=salaries,
        ann_list=ann_list,
        pending_leaves=pending_leaves,
        pending_resignations=pending_resignations,
        pending_tickets=pending_tickets,
        saved=request.args.get("saved") == "1",
        active_nav="settings",
        default_start=_co_shift_start,
        default_half=_co_shift_half,
        default_end=_co_shift_end,
        now_month=datetime.date.today().month,
        now_year=datetime.date.today().year,
        default_onboarding_tpl=default_onboarding_tpl,
        onboarding_templates=onboarding_templates,
        late_deduction_pct=round(fr["late_deduction_pct"], 1),
        half_day_deduction_pct=round(fr["half_day_deduction_pct"], 1),
        grace_minutes=fr["grace_minutes"],
        holiday_pay=fr["holiday_pay"],
        leave_pay=fr["leave_pay"],
        auth_config={
            "face_enabled":            fr["face_auth_enabled"],
            "qr_enabled":              fr["qr_enabled"],
            "fingerprint_enabled":     fr["fingerprint_enabled"],
            "location_enabled":        fr["geo_enabled"],
            "employee_password_auth":  True,
        },
        features=features,
    )

# ---------------- SAVE DEFAULT ONBOARDING TEMPLATE ----------------
@app.route("/save_default_onboarding_template", methods=["POST"])
@admin_required
def save_default_onboarding_template():
    tpl_id = request.form.get("default_onboarding_template_id") or None
    if tpl_id == "0" or tpl_id == "":
        tpl_id = None
    db = get_db_connection(); cursor = db.cursor(buffered=True)
    cursor.execute("UPDATE company_settings SET default_onboarding_template_id=%s", (tpl_id,))
    db.commit(); cursor.close(); db.close()
    flash("Default onboarding template saved.", "success")
    return redirect("/onboarding?tab=templates")

# ---------------- SAVE SALARY RULES ----------------

# ---------------- TOGGLE AUTH METHOD ----------------
_TOGGLE_COLUMN_MAP = {
    "fingerprint": "fingerprint_enabled",
    "qr":          "qr_enabled",
    "face":        "face_enabled",
    "location":    "location_enabled",
    "password":    "employee_password_auth",
}
_TOGGLE_LABEL_MAP = {
    "fingerprint": "Fingerprint / Biometric",
    "qr":          "QR Code",
    "face":        "Face Recognition",
    "location":    "Location Verification",
    "password":    "Password Login",
}

@app.route("/toggle_auth_method", methods=["POST"])
@admin_required
def toggle_auth_method():
    method  = request.form.get("method", "")
    enabled = request.form.get("enabled", "0") == "1"
    if method not in _TOGGLE_COLUMN_MAP:
        flash("Invalid authentication method.", "danger")
        return redirect("/settings?tab=attendance")
    column = _TOGGLE_COLUMN_MAP[method]
    label  = _TOGGLE_LABEL_MAP[method]
    active_cid = session.get("active_company_id")
    # Map old column names to company_feature_settings column names
    _cfs_map = {"face_enabled": "face_auth_enabled", "location_enabled": "geo_enabled",
                "employee_password_auth": None}  # password auth stays global
    cfs_col = _cfs_map.get(column, column)
    if active_cid and cfs_col:
        _upsert_co_feature(active_cid, cfs_col, 1 if enabled else 0)
    else:
        _VALID_CS_TOGGLE = frozenset(_TOGGLE_COLUMN_MAP.values())
        if column not in _VALID_CS_TOGGLE:
            flash("Invalid setting.", "danger")
            return redirect("/settings?tab=attendance")
        db = get_db_connection(); cursor = db.cursor(buffered=True)
        cursor.execute(f"UPDATE company_settings SET {column}=%s", (1 if enabled else 0,))
        db.commit(); cursor.close(); db.close()
    state = "enabled" if enabled else "disabled"
    flash(f"{label} {state}.", "success")
    return redirect("/settings?tab=attendance")

@app.route("/toggle_fingerprint", methods=["POST"])
@admin_required
def toggle_fingerprint():
    enabled = request.form.get("enabled", "0") == "1"
    active_cid = session.get("active_company_id")
    if active_cid:
        _upsert_co_feature(active_cid, "fingerprint_enabled", 1 if enabled else 0)
    else:
        db = get_db_connection(); cursor = db.cursor(buffered=True)
        cursor.execute("UPDATE company_settings SET fingerprint_enabled=%s", (1 if enabled else 0,))
        db.commit(); cursor.close(); db.close()
    state = "enabled" if enabled else "disabled"
    flash(f"Fingerprint authentication {state}.", "success")
    return redirect("/settings?tab=attendance")

# ---------------- SAVE COMPANY CODE ----------------
@app.route("/save_company_code", methods=["POST"])
@admin_required
def save_company_code():
    code = request.form.get("company_code", "").strip().upper()[:10]
    db = get_db_connection(); cursor = db.cursor(buffered=True)
    cursor.execute("UPDATE company_settings SET company_code=%s", (code,))
    db.commit(); cursor.close(); db.close()
    flash(f"Company code set to '{code}'.", "success")
    return redirect("/settings?tab=company")

# ---------------- SAVE COMPANY INFO ----------------
@app.route("/save_company_info", methods=["POST"])
@admin_required
def save_company_info():
    import pytz as _pytz
    _VALID_DAYS = {"Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"}
    name     = request.form.get("company_name", "").strip()[:200]
    code     = request.form.get("company_code", "").strip().upper()[:10]
    timezone = request.form.get("timezone", "Asia/Kolkata").strip()
    w_days_raw = request.form.getlist("working_days")
    # Validate timezone against pytz database
    if timezone not in _pytz.all_timezones_set:
        flash("Invalid timezone selected.", "danger")
        return redirect("/settings?tab=company")
    # Validate day names
    w_days_set = set(w_days_raw)
    if w_days_set and not w_days_set.issubset(_VALID_DAYS):
        flash("Invalid working days selected.", "danger")
        return redirect("/settings?tab=company")
    w_days = ",".join(d for d in w_days_raw if d in _VALID_DAYS)
    db = get_db_connection(); cursor = db.cursor(buffered=True)
    cursor.execute(
        "UPDATE company_settings SET company_name=%s, company_code=%s, timezone=%s, working_days=%s",
        (name, code, timezone, w_days or "Mon,Tue,Wed,Thu,Fri")
    )
    db.commit(); cursor.close(); db.close()
    flash("Company info saved.", "success")
    return redirect("/settings?tab=company")

# ---------------- TOGGLE FEATURE (AJAX) ----------------
@app.route("/toggle_feature", methods=["POST"])
@admin_required
def toggle_feature():
    from flask import jsonify
    allowed = {
        "face_auth_enabled","geo_enabled","qr_enabled","pin_enabled",
        "fingerprint_enabled","biometric_enabled",
        "notify_leave","notify_payslip","notify_resignation","notify_doc_expiry",
    }
    data    = request.get_json(force=True) or {}
    feature = data.get("feature", "")
    value   = 1 if data.get("value") else 0
    if feature not in allowed:
        return jsonify({"ok": False, "error": "unknown feature"}), 400
    active_cid = session.get("active_company_id")
    # Explicit allowlist maps feature name → exact DB column (no dynamic interpolation)
    _CS_COL_MAP = {
        "face_auth_enabled":  "face_auth_enabled",
        "geo_enabled":        "geo_enabled",
        "qr_enabled":         "qr_enabled",
        "pin_enabled":        "pin_enabled",
        "fingerprint_enabled":"fingerprint_enabled",
        "biometric_enabled":  "biometric_enabled",
        "notify_leave":       "notify_leave",
        "notify_payslip":     "notify_payslip",
        "notify_resignation": "notify_resignation",
        "notify_doc_expiry":  "notify_doc_expiry",
    }
    cs_col = _CS_COL_MAP.get(feature)
    if not cs_col:
        return jsonify({"ok": False, "error": "unknown feature"}), 400
    if active_cid:
        _upsert_co_feature(active_cid, cs_col, value)
    else:
        db = get_db_connection(); cursor = db.cursor(buffered=True)
        cursor.execute(f"UPDATE company_settings SET {cs_col}=%s", (value,))
        db.commit(); cursor.close(); db.close()
    return jsonify({"ok": True})

# ---------------- SAVE GEO RADIUS ----------------
@app.route("/save_geo_radius", methods=["POST"])
@admin_required
def save_geo_radius():
    try:
        radius = int(request.form.get("geo_radius", 100))
        if not (50 <= radius <= 5000):
            raise ValueError
    except (ValueError, TypeError):
        flash("Geo radius must be between 50 and 5000 metres.", "danger")
        return redirect("/settings?tab=attendance")
    active_cid = session.get("active_company_id")
    if active_cid:
        _upsert_co_feature(active_cid, "geo_radius", radius)
    else:
        db = get_db_connection(); cursor = db.cursor(buffered=True)
        cursor.execute("UPDATE company_settings SET geo_radius=%s", (radius,))
        db.commit(); cursor.close(); db.close()
    flash("Attendance settings saved.", "success")
    return redirect("/settings?tab=attendance")

# ---------------- SAVE SECURITY SETTINGS ----------------
@app.route("/save_security_settings", methods=["POST"])
@admin_required
def save_security_settings():
    try:
        timeout = int(request.form.get("session_timeout", 30))
        if not (5 <= timeout <= 1440):
            raise ValueError
    except (ValueError, TypeError):
        flash("Session timeout must be between 5 and 1440 minutes.", "danger")
        return redirect("/settings?tab=security")
    active_cid = session.get("active_company_id")
    if active_cid:
        _upsert_co_feature(active_cid, "session_timeout", timeout)
    else:
        db = get_db_connection(); cursor = db.cursor(buffered=True)
        cursor.execute("UPDATE company_settings SET session_timeout=%s", (timeout,))
        db.commit(); cursor.close(); db.close()
    flash("Security settings saved.", "success")
    return redirect("/settings?tab=security")


# ---------------- COMPANIES ----------------

@app.route("/switch_company", methods=["POST"])
@admin_required
def switch_company():
    cid  = request.form.get("company_id", "").strip()
    pin  = request.form.get("pin", "").strip()
    dest = _safe_redirect(request.form.get("next", ""), "/admin")
    if not cid:
        session.pop("active_company_id", None)
        flash("Switched to: All Companies", "success")
        return redirect(dest)
    try:
        cid = int(cid)
    except ValueError:
        return redirect(dest)
    db = get_db_connection(); cur = db.cursor(buffered=True)
    cur.execute("SELECT name, COALESCE(pin,'') FROM companies WHERE id=%s", (cid,))
    row = cur.fetchone()
    cur.close(); db.close()
    if not row:
        flash("Company not found.", "error")
        return redirect(dest)
    cname, stored_pin = row
    if stored_pin and stored_pin != pin:
        flash(f"Incorrect PIN for {cname}.", "error")
        return redirect(dest + ("&" if "?" in dest else "?") + "pin_error=1&pin_cid=" + str(cid))
    session["active_company_id"] = cid
    flash(f"Switched to: {cname}", "success")
    return redirect(dest)

@app.route("/clear_company", methods=["POST"])
@admin_required
def clear_company():
    session.pop("active_company_id", None)
    flash("Viewing all companies.", "success")
    return redirect(_safe_redirect(request.form.get("next", ""), "/admin"))

@app.route("/set_company_pin", methods=["POST"])
@admin_required
def set_company_pin():
    cid = request.form.get("company_id", "").strip()
    pin = request.form.get("pin", "").strip()
    if not cid:
        flash("Invalid request.", "error")
        return redirect("/settings?tab=company")
    db = get_db_connection(); cur = db.cursor(buffered=True)
    cur.execute("UPDATE companies SET pin=%s WHERE id=%s", (pin or None, int(cid)))
    db.commit(); cur.close(); db.close()
    flash("PIN " + ("set." if pin else "removed."), "success")
    return redirect("/settings?tab=company")

@app.route("/companies")
@admin_required
def view_companies():
    return redirect("/settings?tab=company")


@app.route("/companies/add", methods=["POST"])
@admin_required
def add_company():
    name        = request.form.get("name", "").strip()
    code        = request.form.get("code", "").strip().upper()[:20] or None
    redirect_to = request.form.get("redirect_to", "companies")
    dest        = "/settings?tab=company" if redirect_to == "settings" else "/companies"
    if not name:
        flash("Company name is required.", "error")
        return redirect(dest)
    w_days = ",".join(request.form.getlist("working_days")) or "Mon,Tue,Wed,Thu,Fri"
    db = get_db_connection(); cursor = db.cursor(buffered=True)
    cursor.execute("INSERT INTO companies (name, code, working_days) VALUES (%s, %s, %s) RETURNING id", (name, code, w_days))
    new_cid = cursor.fetchone()[0]
    db.commit()

    shift_names  = request.form.getlist("shift_name[]")
    shift_starts = request.form.getlist("shift_start[]")
    shift_halfs  = request.form.getlist("shift_half[]")
    shift_ends   = request.form.getlist("shift_end[]")
    for sname, sstart, shalf, send in zip(shift_names, shift_starts, shift_halfs, shift_ends):
        sname = sname.strip(); sstart = sstart.strip(); shalf = shalf.strip(); send = send.strip()
        if sname and sstart and shalf and send:
            cursor.execute(
                "INSERT INTO shifts (name, start_time, half_time, end_time, company_id) VALUES (%s,%s,%s,%s,%s)",
                (sname,
                 sstart + ":00" if len(sstart) == 5 else sstart,
                 shalf  + ":00" if len(shalf)  == 5 else shalf,
                 send   + ":00" if len(send)   == 5 else send,
                 new_cid)
            )
    db.commit()

    break_names = request.form.getlist("break_name[]")
    break_times = request.form.getlist("break_time[]")
    break_durs  = request.form.getlist("break_duration[]")
    for bname, btime, bdur in zip(break_names, break_times, break_durs):
        bname = bname.strip(); btime = btime.strip(); bdur = bdur.strip()
        if bname and btime and bdur.isdigit():
            cursor.execute(
                "INSERT INTO break_config (break_name, break_time, duration_minutes, company_id) VALUES (%s,%s,%s,%s)",
                (bname, btime + ":00" if len(btime) == 5 else btime, int(bdur), new_cid)
            )
    db.commit()
    cursor.close(); db.close()
    flash(f"Company '{name}' added.", "success")
    return redirect(dest)


@app.route("/companies/<int:cid>/edit", methods=["POST"])
@admin_required
def edit_company(cid):
    name        = request.form.get("name", "").strip()
    new_code    = (request.form.get("code", "").strip().upper()[:20]) or None
    redirect_to = request.form.get("redirect_to", "companies")
    dest        = "/settings?tab=company" if redirect_to == "settings" else "/companies"

    if not name:
        flash("Company name is required.", "error")
        return redirect(dest)

    db = get_db_connection(); cursor = db.cursor(buffered=True)

    w_days = ",".join(request.form.getlist("working_days")) or "Mon,Tue,Wed,Thu,Fri"

    cursor.execute("SELECT COALESCE(code,'') FROM companies WHERE id=%s", (cid,))
    row      = cursor.fetchone()
    old_code = (row[0] or "").strip().upper() if row else ""

    cursor.execute("UPDATE companies SET name=%s, code=%s, working_days=%s WHERE id=%s", (name, new_code, w_days, cid))
    db.commit()

    renamed_count = 0
    if old_code and new_code and old_code != new_code:
        cursor.execute(
            "SELECT employee_id FROM employees WHERE company_id=%s AND employee_id LIKE %s",
            (cid, old_code + "%")
        )
        to_rename = [
            (r[0], new_code + r[0][len(old_code):])
            for r in cursor.fetchall() if r[0].startswith(old_code)
        ]

        related_tables = [
            "attendance", "salary_config", "leave_requests", "notifications",
            "resignation_requests", "tickets", "employee_incentives",
            "employee_experience", "employee_education", "leave_balances",
            "employee_documents", "performance_reviews", "overtime_records",
            "regularization_requests", "compoff_balance", "employee_onboarding",
        ]

        for old_eid, new_eid in to_rename:
            for tbl in related_tables:
                try:
                    cursor.execute(
                        f"UPDATE {tbl} SET employee_id=%s WHERE employee_id=%s",
                        (new_eid, old_eid)
                    )
                except Exception:
                    pass

            old_img = os.path.join(app.config["UPLOAD_FOLDER"], old_eid + ".jpg")
            new_img = os.path.join(app.config["UPLOAD_FOLDER"], new_eid + ".jpg")
            old_qr  = os.path.join("static", "qrcodes", old_eid + ".png")
            new_qr  = os.path.join("static", "qrcodes", new_eid + ".png")

            cursor.execute(
                "UPDATE employees SET employee_id=%s, face_image=%s, qr_code=%s "
                "WHERE employee_id=%s AND company_id=%s",
                (new_eid, new_img, new_qr, old_eid, cid)
            )

            if os.path.exists(old_img):
                try: os.rename(old_img, new_img)
                except Exception: pass
            if os.path.exists(old_qr):
                try: os.rename(old_qr, new_qr)
                except Exception: pass

            renamed_count += 1

        db.commit()
        flash(
            f"Company updated. {renamed_count} employee ID(s) renamed: "
            f"{old_code}xxx → {new_code}xxx.",
            "success"
        )
    else:
        flash("Company updated.", "success")

    cursor.close(); db.close()
    return redirect(dest)


@app.route("/companies/<int:cid>/delete", methods=["POST"])
@admin_required
def delete_company(cid):
    redirect_to = request.form.get("redirect_to", "companies")
    dest        = "/settings?tab=company" if redirect_to == "settings" else "/companies"
    db = get_db_connection(); cursor = db.cursor(buffered=True)
    cursor.execute("SELECT COUNT(*) FROM employees WHERE company_id=%s", (cid,))
    count = cursor.fetchone()[0]
    if count > 0:
        cursor.close(); db.close()
        flash(f"Cannot delete: {count} employee(s) are assigned to this company.", "error")
        return redirect(dest)
    cursor.execute("DELETE FROM companies WHERE id=%s", (cid,))
    db.commit(); cursor.close(); db.close()
    flash("Company deleted.", "success")
    return redirect(dest)


# ---------------- ANNOUNCEMENTS ----------------
@app.route("/announcements", methods=["GET", "POST"])
@admin_required
def announcements_admin():
    db     = get_db_connection()
    cursor = db.cursor(buffered=True)
    if request.method == "POST":
        action = request.form.get("action")
        if action == "add":
            visibility = request.form.get("visibility", "public")
            target_emp = request.form.get("target_employee_id", "").strip() or None
            if visibility == "private" and not target_emp:
                flash("Please select an employee for a private announcement.", "error")
                cursor.close(); db.close()
                return redirect("/performance?tab=announcements")
            if visibility == "public":
                target_emp = None
            cursor.execute(
                "INSERT INTO announcements (title, content, priority, visibility, target_employee_id) VALUES (%s,%s,%s,%s,%s)",
                (request.form["title"], request.form["content"], request.form.get("priority","Normal"), visibility, target_emp)
            )
            db.commit()
            flash("Announcement posted.", "success")
        elif action == "delete":
            cursor.execute("DELETE FROM announcements WHERE id=%s", (request.form["ann_id"],))
            db.commit()
            flash("Announcement deleted.", "success")
        cursor.close(); db.close()
        return redirect("/performance?tab=announcements")
    cursor.close(); db.close()
    return redirect("/performance?tab=announcements")

# ---------------- INDIAN PUBLIC HOLIDAYS ----------------
def get_indian_holidays(year):
    """Returns sorted list of (date, name) for major Indian public holidays."""
    fixed = [
        (1,  1,  "New Year's Day"),
        (1,  26, "Republic Day"),
        (8,  15, "Independence Day"),
        (10, 2,  "Gandhi Jayanti"),
        (12, 25, "Christmas Day"),
    ]
    variable_by_year = {
        2025: [
            (1, 14, "Makar Sankranti / Pongal"),
            (2, 26, "Maha Shivaratri"),
            (3, 14, "Holi"),
            (3, 31, "Eid ul-Fitr"),
            (4, 14, "Dr. Ambedkar Jayanti"),
            (4, 18, "Good Friday"),
            (5,  1, "Maharashtra Day / Labour Day"),
            (6,  7, "Eid ul-Adha"),
            (8, 16, "Janmashtami"),
            (10, 2,  "Dussehra / Vijayadasami"),
            (10, 20, "Diwali (Lakshmi Puja)"),
            (11,  5, "Guru Nanak Jayanti"),
        ],
        2026: [
            (1, 14, "Makar Sankranti / Pongal"),
            (2, 15, "Maha Shivaratri"),
            (3,  5, "Holi"),
            (3, 20, "Eid ul-Fitr"),
            (4,  3, "Good Friday"),
            (4, 14, "Dr. Ambedkar Jayanti / Baisakhi"),
            (5,  1, "Maharashtra Day / Labour Day"),
            (5, 27, "Eid ul-Adha"),
            (8, 21, "Janmashtami"),
            (10, 21, "Dussehra / Vijayadasami"),
            (10, 30, "Diwali (Lakshmi Puja)"),
            (11, 25, "Guru Nanak Jayanti"),
        ],
        2027: [
            (1, 14, "Makar Sankranti / Pongal"),
            (3,  5, "Maha Shivaratri"),
            (3, 26, "Holi"),
            (4,  2, "Good Friday"),
            (4, 14, "Dr. Ambedkar Jayanti"),
            (5,  1, "Maharashtra Day / Labour Day"),
            (8, 15, "Independence Day"),
            (9,  4, "Janmashtami"),
            (10, 8,  "Dussehra / Vijayadasami"),
            (10, 17, "Diwali (Lakshmi Puja)"),
            (11, 14, "Guru Nanak Jayanti"),
        ],
    }
    result = []
    for m, d, name in fixed:
        try:
            result.append((datetime.date(year, m, d), name))
        except ValueError:
            pass
    for m, d, name in variable_by_year.get(year, []):
        try:
            result.append((datetime.date(year, m, d), name))
        except ValueError:
            pass
    return sorted(result, key=lambda x: x[0])

# ---------------- VIEW HOLIDAYS (legacy helper — no route; /view_holidays redirects to /leave_holidays) ----------------
def view_holidays():
    year = int(request.args.get("year", datetime.date.today().year))
    db     = get_db_connection()
    cursor = db.cursor(buffered=True)
    cursor.execute("SELECT * FROM holidays ORDER BY date")
    data = cursor.fetchall()
    cursor.close()
    db.close()

    # Build holiday map: date -> (id, name)
    holiday_map = {}
    for row in data:
        date_val = row[1]
        if isinstance(date_val, datetime.date):
            holiday_map[date_val] = (row[0], row[2])

    # Build calendar data, weeks starting Sunday (firstweekday=6)
    sun_cal = calendar.Calendar(firstweekday=6)
    today   = datetime.date.today()
    cal_data = []
    for month in range(1, 13):
        month_holidays = {}  # day_number -> (id, name)
        for date_obj, (hid, hname) in holiday_map.items():
            if date_obj.year == year and date_obj.month == month:
                month_holidays[date_obj.day] = (hid, hname)
        cal_data.append({
            'month_num':  month,
            'month_name': calendar.month_name[month],
            'weeks':      sun_cal.monthdayscalendar(year, month),
            'holidays':   month_holidays,
        })

    return render_template("holidays.html", holidays=data, cal_data=cal_data,
                           year=year, today=today,
        active_nav="leaves",
    )


# ---------------- EMPLOYEE DETAIL PAGE ----------------
# ---------------- ADD EMPLOYEE (from employees page) ----------------
# ---------------- UPDATE EMPLOYEE PHOTO ----------------
# ---------------- REGENERATE QR ----------------
# ---------------- LEAVE TYPES ADMIN ----------------


# ---------------- SHIFTS (redirect to settings) ----------------


# ──────────────────────── SHIFT SWAP REQUESTS ────────────────────────


# ---------------- AUTO GENERATE EMPLOYEE ID ----------------
# ---------------- BREAK CONFIG ----------------
@app.route("/api/breaks")
@limiter.limit("30 per minute")
def api_breaks():
    if not (session.get("admin_logged_in") or session.get("employee_id")):
        auth = request.headers.get("Authorization", "")
        if not auth.startswith("Bearer "):
            return jsonify({"ok": False, "msg": "Unauthorized"}), 401
    db     = get_db_connection()
    cursor = db.cursor(buffered=True)
    cursor.execute("SELECT id, break_name, break_time, duration_minutes FROM break_config WHERE is_active=1 ORDER BY break_time")
    rows = cursor.fetchall()
    cursor.close(); db.close()
    result = []
    for row in rows:
        bt = row[2]
        if hasattr(bt, 'seconds'):
            total = bt.seconds
            h, m = divmod(total // 60, 60)
        else:
            h, m = bt.hour, bt.minute
        result.append({"id": row[0], "name": row[1],
                        "hour": h, "minute": m,
                        "duration": row[3]})
    return jsonify(result)

@app.route("/break_config")
@admin_required
def view_break_config():
    return redirect("/settings?tab=shifts")

@app.route("/add_break", methods=["POST"])
@admin_required
def add_break():
    name     = request.form.get("break_name", "").strip()
    btime    = request.form.get("break_time", "")
    duration = int(request.form.get("duration_minutes", 10) or 10)
    dest     = _safe_redirect(request.form.get("redirect", ""), _safe_referrer_redirect(request.referrer or "", "/employees?tab=schedule"))
    cid_raw  = request.form.get("company_id", "").strip()
    company_id = int(cid_raw) if cid_raw.isdigit() else None
    sid_raw  = request.form.get("shift_id", "").strip()
    shift_id = int(sid_raw) if sid_raw.isdigit() else None
    db     = get_db_connection()
    cursor = db.cursor(buffered=True)
    if company_id:
        cursor.execute(
            "INSERT INTO break_config (break_name, break_time, duration_minutes, company_id, shift_id) VALUES (%s,%s,%s,%s,%s)",
            (name, btime, duration, company_id, shift_id)
        )
    else:
        cursor.execute("INSERT INTO break_config (break_name, break_time, duration_minutes, shift_id) VALUES (%s,%s,%s,%s)",
                       (name, btime, duration, shift_id))
    db.commit(); cursor.close(); db.close()
    flash("Break added successfully.", "success")
    return redirect(dest)

@app.route("/update_break", methods=["POST"])
@app.route("/update_break/<int:bid>", methods=["POST"])
@admin_required
def update_break(bid=None):
    if bid is None:
        try: bid = int(request.form.get("break_id", ""))
        except: return redirect("/employees?tab=schedule")
    name     = request.form.get("break_name", "").strip()
    btime    = request.form.get("break_time", "")
    duration = int(request.form.get("duration_minutes", 10) or 10)
    active   = 1 if request.form.get("is_active") else 0
    dest     = _safe_redirect(request.form.get("redirect", ""), _safe_referrer_redirect(request.referrer or "", "/employees?tab=schedule"))
    sid_raw  = request.form.get("shift_id", "").strip()
    db     = get_db_connection()
    cursor = db.cursor(buffered=True)
    if sid_raw.isdigit():
        cursor.execute(
            "UPDATE break_config SET break_name=%s, break_time=%s, duration_minutes=%s, is_active=%s, shift_id=%s WHERE id=%s",
            (name, btime, duration, active, int(sid_raw), bid)
        )
    else:
        cursor.execute(
            "UPDATE break_config SET break_name=%s, break_time=%s, duration_minutes=%s, is_active=%s WHERE id=%s",
            (name, btime, duration, active, bid)
        )
    db.commit(); cursor.close(); db.close()
    flash("Break updated.", "success")
    return redirect(dest)

@app.route("/delete_break", methods=["POST"])
@app.route("/delete_break/<int:bid>", methods=["POST"])
@admin_required
def delete_break(bid=None):
    if bid is None:
        try: bid = int(request.form.get("break_id", ""))
        except: return redirect("/employees?tab=schedule")
    dest = request.form.get("redirect") or "/employees?tab=schedule"
    db     = get_db_connection()
    cursor = db.cursor(buffered=True)
    cursor.execute("DELETE FROM break_config WHERE id=%s", (bid,))
    db.commit(); cursor.close(); db.close()
    flash("Break deleted.", "success")
    return redirect(dest)

# ---------------- VIEW SALARY CONFIG ----------------


# ---------------- MONTHLY ATTENDANCE REPORT ----------------

# ---------------- EMPLOYEE ATTENDANCE DETAIL ----------------

# ---------------- MANUAL ATTENDANCE CORRECTION ----------------


# ---------------- BULK MARK ATTENDANCE ----------------


# ---------------- MONTHLY REPORT EXCEL EXPORT ----------------

# ---------------- ABSENTEE REPORT EMAIL ----------------

# ---------------- SALARY REPORT ----------------


# ---------------- EMAIL CONFIG ----------------
@app.route("/email_config", methods=["GET", "POST"])
@admin_required
def email_config():
    db     = get_db_connection()
    cursor = db.cursor(buffered=True)

    if request.method == "POST":
        host       = request.form["smtp_host"].strip()
        port       = int(request.form["smtp_port"])
        user       = request.form["smtp_user"].strip()
        password   = request.form["smtp_pass"].strip()
        from_name  = request.form.get("from_name", "Attendance System").strip()
        from_email = request.form.get("from_email", "").strip() or user

        cursor.execute("DELETE FROM email_config")
        cursor.execute(
            "INSERT INTO email_config (smtp_host, smtp_port, smtp_user, smtp_pass, from_name, from_email) VALUES (%s,%s,%s,%s,%s,%s)",
            (host, port, user, encrypt_pii(password), from_name, from_email)
        )
        db.commit()
        cursor.close()
        db.close()
        return redirect("/settings?tab=email&saved=1")

    cursor.execute("SELECT smtp_host, smtp_port, smtp_user, smtp_pass, from_name, from_email FROM email_config ORDER BY id DESC LIMIT 1")
    row    = cursor.fetchone()
    config = {"host": row[0], "port": row[1], "user": row[2], "password": decrypt_pii(row[3]), "from_name": row[4], "from_email": row[5] or row[2]} if row else None
    cursor.close()
    db.close()

    return render_template("email_config.html",
        config=config,
        saved=request.args.get("saved",
        active_nav="salary",
    ) == "1",
    )

# ---------------- SEND SALARY EMAIL (single) ----------------

# ---------------- SEND ALL SALARY EMAILS ----------------

# ---------------- PAYROLL LOCK / UNLOCK ----------------


# ---------------- TEST EMAIL ----------------
@app.route("/test_email", methods=["POST"])
@admin_required
def test_email():
    to_email = request.form.get("test_to", "").strip()
    config   = get_email_config()
    if not config:
        return jsonify({"ok": False, "msg": "Email not configured yet."})
    if not to_email:
        return jsonify({"ok": False, "msg": "Enter a test recipient email."})
    try:
        send_email_smtp(
            to_email,
            "Test Email - Attendance System",
            "<h2>Test email from Employee Attendance System</h2><p>Email configuration is working correctly.</p>",
            config,
        )
        return jsonify({"ok": True, "msg": f"Test email sent to {to_email}"})
    except Exception:
        app_log.error("Test email send failed", exc_info=True)
        return jsonify({"ok": False, "msg": "Failed to send test email. Check email settings."})

# ---------------- LOCATION ----------------

# ---------------- DISTANCE CHECK ----------------
def is_within_range(user_lat, user_lon, office_lat, office_lon):
    R       = 6371000
    phi1    = math.radians(user_lat)
    phi2    = math.radians(office_lat)
    dphi    = math.radians(office_lat - user_lat)
    dlambda = math.radians(office_lon - user_lon)
    a = (
        math.sin(dphi / 2) ** 2
        + math.cos(phi1) * math.cos(phi2) * math.sin(dlambda / 2) ** 2
    )
    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))
    return (R * c) <= OFFICE_RADIUS_M

# ---------------- ATTENDANCE (LOGIN + LOGOUT) ----------------

# ================================================================
#  EMPLOYEE PORTAL
# ================================================================

@app.route("/update_my_profile", methods=["POST"])
@employee_required
def update_my_profile():
    emp_id = session["employee_id"]
    fields = {
        "phone":                      request.form.get("phone", "").strip() or None,
        "gender":                     request.form.get("gender", "").strip() or None,
        "dob":                        request.form.get("dob", "").strip() or None,
        "blood_group":                request.form.get("blood_group", "").strip() or None,
        "address":                    request.form.get("address", "").strip() or None,
        "city":                       request.form.get("city", "").strip() or None,
        "state":                      request.form.get("state", "").strip() or None,
        "pincode":                    request.form.get("pincode", "").strip() or None,
        "emergency_contact_name":     request.form.get("emergency_contact_name", "").strip() or None,
        "emergency_contact_phone":    request.form.get("emergency_contact_phone", "").strip() or None,
        "emergency_contact_relation": request.form.get("emergency_contact_relation", "").strip() or None,
        "about_me":                   request.form.get("about_me", "").strip() or None,
    }
    db = get_db_connection()
    cursor = db.cursor(buffered=True)
    cursor.execute("""
        UPDATE employees SET
            phone=%s, gender=%s, dob=%s, blood_group=%s,
            address=%s, city=%s, state=%s, pincode=%s,
            emergency_contact_name=%s, emergency_contact_phone=%s, emergency_contact_relation=%s,
            about_me=%s
        WHERE employee_id=%s
    """, (*fields.values(), emp_id))
    db.commit(); cursor.close(); db.close()
    return redirect("/employee_portal?profile_saved=1#my-profile")


@app.route("/update_my_bank_details", methods=["POST"])
@employee_required
def update_my_bank_details():
    emp_id = session["employee_id"]
    fields = {
        "aadhar_number": encrypt_pii(request.form.get("aadhar_number", "").strip() or None),
        "pan_number":    encrypt_pii(request.form.get("pan_number", "").upper().strip() or None),
        "bank_name":     request.form.get("bank_name", "").strip() or None,
        "bank_account":  encrypt_pii(request.form.get("bank_account", "").strip() or None),
        "bank_ifsc":     encrypt_pii(request.form.get("bank_ifsc", "").upper().strip() or None),
        "uan_number":    encrypt_pii(request.form.get("uan_number", "").strip() or None),
    }
    db = get_db_connection()
    cursor = db.cursor(buffered=True)
    cursor.execute("""
        UPDATE employees SET
            aadhar_number=%s, pan_number=%s, bank_name=%s,
            bank_account=%s, bank_ifsc=%s, uan_number=%s
        WHERE employee_id=%s
    """, (*fields.values(), emp_id))
    db.commit(); cursor.close(); db.close()
    return redirect("/employee_portal?bank_saved=1#my-profile")


@app.route("/add_experience", methods=["POST"])
@employee_required
def add_experience():
    emp_id = session["employee_id"]
    company     = request.form.get("company", "").strip()
    designation = request.form.get("designation", "").strip()
    from_year   = request.form.get("from_year", "").strip()
    to_year     = request.form.get("to_year", "").strip() or None
    is_current  = 1 if request.form.get("is_current") else 0
    description = request.form.get("description", "").strip() or None
    if not company or not designation or not from_year:
        return redirect("/employee_portal?exp_error=1#my-profile")
    db = get_db_connection()
    cursor = db.cursor(buffered=True)
    cursor.execute(
        "INSERT INTO employee_experience (employee_id, company, designation, from_year, to_year, is_current, description) "
        "VALUES (%s,%s,%s,%s,%s,%s,%s)",
        (emp_id, company, designation, from_year, to_year, is_current, description)
    )
    db.commit(); cursor.close(); db.close()
    return redirect("/employee_portal?exp_saved=1#my-profile")


@app.route("/delete_experience/<int:entry_id>", methods=["POST"])
@employee_required
def delete_experience(entry_id):
    emp_id = session["employee_id"]
    db = get_db_connection()
    cursor = db.cursor(buffered=True)
    cursor.execute("DELETE FROM employee_experience WHERE id=%s AND employee_id=%s", (entry_id, emp_id))
    db.commit(); cursor.close(); db.close()
    return redirect("/employee_portal#my-profile")


@app.route("/add_education_entry", methods=["POST"])
@employee_required
def add_education_entry():
    emp_id = session["employee_id"]
    degree          = request.form.get("degree", "").strip()
    institution     = request.form.get("institution", "").strip()
    year_of_passing = request.form.get("year_of_passing", "").strip() or None
    percentage      = request.form.get("percentage", "").strip() or None
    if not degree or not institution:
        return redirect("/employee_portal?edu_error=1#my-profile")
    db = get_db_connection()
    cursor = db.cursor(buffered=True)
    cursor.execute(
        "INSERT INTO employee_education (employee_id, degree, institution, year_of_passing, percentage) "
        "VALUES (%s,%s,%s,%s,%s)",
        (emp_id, degree, institution, year_of_passing, percentage)
    )
    db.commit(); cursor.close(); db.close()
    return redirect("/employee_portal?edu_saved=1#my-profile")


@app.route("/delete_education_entry/<int:entry_id>", methods=["POST"])
@employee_required
def delete_education_entry(entry_id):
    emp_id = session["employee_id"]
    db = get_db_connection()
    cursor = db.cursor(buffered=True)
    cursor.execute("DELETE FROM employee_education WHERE id=%s AND employee_id=%s", (entry_id, emp_id))
    db.commit(); cursor.close(); db.close()
    return redirect("/employee_portal#my-profile")


@app.route("/update_my_photo", methods=["POST"])
@employee_required
def update_my_photo():
    from flask import send_from_directory
    import numpy as np
    from PIL import Image
    import base64, io
    emp_id = session["employee_id"]
    file = request.files.get("photo")
    ok, err = _validate_image_file(file)
    if not ok:
        return redirect("/employee_portal?photo_error=bad_format#my-profile")
    try:
        img = Image.open(file.stream).convert("RGB")
        img_array = np.array(img)
        if _face_recognition_available:
            locs = face_recognition.face_locations(img_array)
            if not locs:
                return redirect("/employee_portal?photo_error=no_face#my-profile")
        save_path = os.path.join(app.config["UPLOAD_FOLDER"], emp_id + ".jpg")
        img.save(save_path, "JPEG", quality=90)
        db = get_db_connection()
        cursor = db.cursor(buffered=True)
        cursor.execute("UPDATE employees SET face_image=%s WHERE employee_id=%s", (emp_id + ".jpg", emp_id))
        db.commit(); cursor.close(); db.close()
        return redirect("/employee_portal?photo_saved=1#my-profile")
    except Exception:
        return redirect("/employee_portal?photo_error=failed#my-profile")


@app.route("/my_qr")
@employee_required
def my_qr():
    from flask import send_file
    emp_id = session["employee_id"]
    qr_path = os.path.join("static", "qrcodes", emp_id + ".png")
    if not os.path.exists(qr_path):
        # Auto-generate QR and save path to DB
        generated = generate_qr(emp_id)
        db = get_db_connection()
        cursor = db.cursor(buffered=True)
        cursor.execute("UPDATE employees SET qr_code=%s WHERE employee_id=%s", (generated, emp_id))
        db.commit(); cursor.close(); db.close()
        qr_path = generated
    return send_file(os.path.abspath(qr_path), as_attachment=True,
                     download_name=f"QR_{emp_id}.png", mimetype="image/png")


@app.route("/my_id_card")
@employee_required
def my_id_card():
    from PIL import Image, ImageDraw, ImageFont
    import io as _io2
    from flask import send_file

    emp_id = session["employee_id"]
    db = get_db_connection()
    cursor = db.cursor(buffered=True)
    cursor.execute("""
        SELECT e.employee_id, e.name, e.role, e.email, e.face_image, e.date_of_joining,
               sh.name AS shift_name, e.blood_group, e.phone
        FROM employees e
        LEFT JOIN shifts sh ON e.shift_id = sh.id
        WHERE e.employee_id = %s
    """, (emp_id,))
    row = cursor.fetchone()
    if not row:
        cursor.execute("""
            SELECT employee_id, name, role, email, face_image, date_of_joining,
                   NULL, blood_group, phone
            FROM employees WHERE employee_id=%s
        """, (emp_id,))
        row = cursor.fetchone()
    cursor.close(); db.close()

    # ── Colours ──────────────────────────────────────────
    DARK   = (15,  40, 100)
    BLUE   = (30,  58, 138)
    MID    = (37,  99, 235)
    LIGHT  = (59, 130, 246)
    PALE   = (219, 234, 254)
    WHITE  = (255, 255, 255)
    LGRAY  = (241, 245, 249)
    MGRAY  = (100, 116, 139)
    DGRAY  = (15,  23,  42)
    GOLD   = (251, 191,  36)
    RED    = (220,  38,  38)

    # ── Font loader ──────────────────────────────────────
    def fnt(size, bold=False):
        candidates = (
            ["C:/Windows/Fonts/arialbd.ttf",
             "C:/Windows/Fonts/calibrib.ttf",
             "/System/Library/Fonts/Supplemental/Arial Bold.ttf",
             "/System/Library/Fonts/Helvetica.ttc",
             "/Library/Fonts/Arial Bold.ttf",
             "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"]
            if bold else
            ["C:/Windows/Fonts/arial.ttf",
             "C:/Windows/Fonts/calibri.ttf",
             "/System/Library/Fonts/Supplemental/Arial.ttf",
             "/System/Library/Fonts/Helvetica.ttc",
             "/Library/Fonts/Arial.ttf",
             "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"]
        )
        for p in candidates:
            try: return ImageFont.truetype(p, size)
            except: pass
        return ImageFont.load_default()

    def _safe_text(text):
        try:
            text.encode('latin-1')
            return text
        except (UnicodeEncodeError, UnicodeDecodeError):
            return text.encode('ascii', 'replace').decode('ascii')

    def tw(draw, text, font):
        bb = draw.textbbox((0,0), _safe_text(text), font=font)
        return bb[2]-bb[0]

    def cx(draw, text, font, card_w, y, color):
        t = _safe_text(text)
        draw.text(((card_w - tw(draw, t, font))//2, y), t, font=font, fill=color)

    # ── Vertical card size (portrait) ────────────────────
    CW, CH = 500, 820

    # ════════════════════════════════════════════════════
    #  FRONT
    # ════════════════════════════════════════════════════
    front = Image.new("RGB", (CW, CH), WHITE)
    fd    = ImageDraw.Draw(front)

    # -- Top header --
    fd.rectangle([(0, 0), (CW, 110)], fill=BLUE)
    # Decorative circle top-right
    fd.ellipse([(CW-100, -60), (CW+60, 100)], fill=MID)
    cx(fd, "EMPLOYEE ID CARD", fnt(18, bold=True), CW, 18, WHITE)
    cx(fd, "Attendance Management System", fnt(11), CW, 52, PALE)
    # Thin gold accent line
    fd.rectangle([(0, 108), (CW, 113)], fill=GOLD)

    # -- Photo section --
    fd.rectangle([(0, 113), (CW, 370)], fill=LGRAY)
    PH_W  = 160
    PH_H  = 190
    PH_CX = CW // 2
    PH_X  = PH_CX - PH_W // 2
    PH_Y  = 128
    # Gold border box
    fd.rounded_rectangle([(PH_X-5, PH_Y-5), (PH_X+PH_W+5, PH_Y+PH_H+5)],
                         radius=8, fill=GOLD)
    # White inner border
    fd.rounded_rectangle([(PH_X-2, PH_Y-2), (PH_X+PH_W+2, PH_Y+PH_H+2)],
                         radius=6, fill=WHITE)
    # Photo
    photo_path = os.path.join("dataset", emp_id + ".jpg")
    try:
        ph = Image.open(photo_path).convert("RGB").resize((PH_W, PH_H), Image.LANCZOS)
        front.paste(ph, (PH_X, PH_Y))
    except Exception:
        fd.rounded_rectangle([(PH_X, PH_Y), (PH_X+PH_W, PH_Y+PH_H)], radius=4, fill=MID)
        ini = row[1][0].upper() if row and row[1] else "?"
        cx(fd, ini, fnt(56, bold=True), CW, PH_Y + PH_H//2 - 38, WHITE)

    # Name & role
    name_str = (row[1] or "Unknown")[:24]
    role_str  = (row[2] or "Employee")[:28]
    cx(fd, name_str,  fnt(18, bold=True), CW, 328, DGRAY)
    cx(fd, role_str,  fnt(12),            CW, 352, MGRAY)

    # Blue separator
    fd.rectangle([(40, 372), (CW-40, 374)], fill=PALE)

    # -- Info rows (centered) --
    info_rows = [
        ("Employee ID", row[0]  if row            else "-"),
        ("Email",       row[3]  if row and row[3] else "-"),
        ("Phone",       row[8]  if row and row[8] else "-"),
        ("Blood Group", row[7]  if row and row[7] else "-"),
    ]
    y = 390
    for i, (lbl, val) in enumerate(info_rows):
        if i % 2 == 0:
            fd.rectangle([(0, y-4), (CW, y+38)], fill=LGRAY)
        cx(fd, lbl,            fnt(10),            CW, y+2,  MGRAY)
        cx(fd, str(val)[:34],  fnt(13, bold=True), CW, y+17, DGRAY)
        y += 44

    # Blood group badge (prominent red pill)
    bg_val = row[7] if row and row[7] else None
    if bg_val:
        bw = tw(fd, bg_val, fnt(13, bold=True)) + 28
        bx = (CW - bw) // 2
        by = y + 8
        fd.rounded_rectangle([(bx, by), (bx+bw, by+32)], radius=16, fill=RED)
        cx(fd, bg_val, fnt(13, bold=True), CW, by+8, WHITE)

    # -- Footer --
    fd.rectangle([(0, CH-60), (CW, CH)], fill=BLUE)
    fd.rectangle([(0, CH-62), (CW, CH-60)], fill=GOLD)
    cx(fd, "Confidential  |  Not Transferable", fnt(10), CW, CH-44, PALE)
    cx(fd, "Property of the Organization",       fnt(10), CW, CH-26, (160,185,240))

    # ════════════════════════════════════════════════════
    #  BACK
    # ════════════════════════════════════════════════════
    back = Image.new("RGB", (CW, CH), LGRAY)
    bd   = ImageDraw.Draw(back)

    # Top header (same style)
    bd.rectangle([(0, 0), (CW, 110)], fill=BLUE)
    bd.ellipse([(CW-100, -60), (CW+60, 100)], fill=MID)
    cx(bd, "ATTENDANCE MANAGEMENT SYSTEM", fnt(14, bold=True), CW, 22, WHITE)
    cx(bd, "Employee Attendance Card", fnt(11), CW, 52, PALE)
    bd.rectangle([(0, 108), (CW, 113)], fill=GOLD)

    # QR code — large and centered
    qr_path = os.path.join("static", "qrcodes", emp_id + ".png")
    if not os.path.exists(qr_path):
        qr_path = generate_qr(emp_id)

    QS   = 240
    qr_x = (CW - QS) // 2
    qr_y = 148
    # White card behind QR
    bd.rounded_rectangle([(qr_x-16, qr_y-16), (qr_x+QS+16, qr_y+QS+16)],
                         radius=14, fill=WHITE)
    try:
        qr_img = Image.open(qr_path).convert("RGB").resize((QS, QS), Image.LANCZOS)
        back.paste(qr_img, (qr_x, qr_y))
    except Exception:
        cx(bd, "QR NOT AVAILABLE", fnt(13), CW, qr_y+QS//2, MGRAY)

    cx(bd, "Scan to Mark Attendance",      fnt(14, bold=True), CW, qr_y+QS+28, BLUE)
    cx(bd, row[0] if row else "",          fnt(12),            CW, qr_y+QS+52, MGRAY)

    # Divider
    bd.rectangle([(40, qr_y+QS+78), (CW-40, qr_y+QS+80)], fill=(203,213,225))

    # Info below QR
    sub_info = [
        ("Name",         (row[1] or "-")[:26] if row else "-"),
        ("Designation",  (row[2] or "-")[:26] if row else "-"),
        ("Blood Group",  (row[7] or "-")      if row else "-"),
    ]
    BP = 36
    sy = qr_y + QS + 94
    for lbl2, val2 in sub_info:
        cx(bd, lbl2, fnt(10),            CW, sy,    MGRAY)
        cx(bd, val2, fnt(12, bold=True), CW, sy+14, DGRAY)
        sy += 42

    # "If found" note
    bd.rectangle([(BP, sy+8), (CW-BP, sy+10)], fill=(203,213,225))
    cx(bd, "If found, please return to:", fnt(10),            CW, sy+18, MGRAY)
    cx(bd, "HR Department",               fnt(12, bold=True), CW, sy+34, BLUE)
    if row and row[3]:
        cx(bd, row[3][:34], fnt(10), CW, sy+54, MGRAY)

    # Magnetic stripe
    bd.rectangle([(0, CH-100), (CW, CH-68)], fill=DARK)

    # Footer
    bd.rectangle([(0, CH-60), (CW, CH)], fill=BLUE)
    bd.rectangle([(0, CH-62), (CW, CH-60)], fill=GOLD)
    cx(bd, "Authorized Personnel Only  |  Not Transferable", fnt(10), CW, CH-44, PALE)
    cx(bd, "Misuse is subject to disciplinary action",        fnt(10), CW, CH-26, (160,185,240))

    # ════════════════════════════════════════════════════
    #  COMBINE side by side  (front | gap | back)
    # ════════════════════════════════════════════════════
    GAP   = 40
    LBL_H = 24
    BGCOL = (215, 225, 240)
    total = Image.new("RGB", (CW*2 + GAP, CH + LBL_H), BGCOL)
    td    = ImageDraw.Draw(total)

    td.text((10,  4), "FRONT", font=fnt(13, bold=True), fill=BLUE)
    td.text((CW + GAP + 10, 4), "BACK", font=fnt(13, bold=True), fill=BLUE)

    total.paste(front, (0,       LBL_H))
    total.paste(back,  (CW+GAP,  LBL_H))

    buf = _io2.BytesIO()
    total.save(buf, format="PNG", dpi=(200, 200))
    buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name=f"IDCard_{emp_id}.png", mimetype="image/png")


def _build_id_card_buf(emp_id):
    """Generate the front+back ID card PNG and return a BytesIO buffer, or None if not found."""
    from PIL import Image, ImageDraw, ImageFont
    import io as _io2

    db = get_db_connection()
    cursor = db.cursor(buffered=True)
    cursor.execute("""
        SELECT e.employee_id, e.name, e.role, e.email, e.face_image, e.date_of_joining,
               sh.name AS shift_name, e.blood_group, e.phone
        FROM employees e
        LEFT JOIN shifts sh ON e.shift_id = sh.id
        WHERE e.employee_id = %s
    """, (emp_id,))
    row = cursor.fetchone()
    if not row:
        cursor.execute("""
            SELECT employee_id, name, role, email, face_image, date_of_joining,
                   NULL, blood_group, phone
            FROM employees WHERE employee_id=%s
        """, (emp_id,))
        row = cursor.fetchone()
    cursor.close(); db.close()

    if not row:
        return None

    DARK  = (15,  40, 100)
    BLUE  = (30,  58, 138)
    MID   = (37,  99, 235)
    LIGHT = (59, 130, 246)
    PALE  = (219, 234, 254)
    WHITE = (255, 255, 255)
    LGRAY = (241, 245, 249)
    MGRAY = (100, 116, 139)
    DGRAY = (15,  23,  42)
    GOLD  = (251, 191,  36)
    RED   = (220,  38,  38)

    def fnt(size, bold=False):
        candidates = (
            ["C:/Windows/Fonts/arialbd.ttf",
             "C:/Windows/Fonts/calibrib.ttf",
             "/System/Library/Fonts/Supplemental/Arial Bold.ttf",
             "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"]
            if bold else
            ["C:/Windows/Fonts/arial.ttf",
             "C:/Windows/Fonts/calibri.ttf",
             "/System/Library/Fonts/Supplemental/Arial.ttf",
             "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"]
        )
        for p in candidates:
            try: return ImageFont.truetype(p, size)
            except: pass
        return ImageFont.load_default()

    def _safe_text(text):
        try:
            text.encode('latin-1')
            return text
        except (UnicodeEncodeError, UnicodeDecodeError):
            return text.encode('ascii', 'replace').decode('ascii')

    def tw(draw, text, font):
        bb = draw.textbbox((0, 0), _safe_text(text), font=font)
        return bb[2] - bb[0]

    def cx(draw, text, font, card_w, y, color):
        t = _safe_text(text)
        draw.text(((card_w - tw(draw, t, font)) // 2, y), t, font=font, fill=color)

    CW, CH = 500, 820

    # ── FRONT ──────────────────────────────────────────────
    front = Image.new("RGB", (CW, CH), WHITE)
    fd    = ImageDraw.Draw(front)

    fd.rectangle([(0, 0), (CW, 110)], fill=BLUE)
    fd.ellipse([(CW-100, -60), (CW+60, 100)], fill=MID)
    cx(fd, "EMPLOYEE ID CARD", fnt(18, bold=True), CW, 18, WHITE)
    cx(fd, "Attendance Management System", fnt(11), CW, 52, PALE)
    fd.rectangle([(0, 108), (CW, 113)], fill=GOLD)

    fd.rectangle([(0, 113), (CW, 370)], fill=LGRAY)
    PH_W, PH_H = 160, 190
    PH_X = CW // 2 - PH_W // 2
    PH_Y = 128
    fd.rounded_rectangle([(PH_X-5, PH_Y-5), (PH_X+PH_W+5, PH_Y+PH_H+5)], radius=8, fill=GOLD)
    fd.rounded_rectangle([(PH_X-2, PH_Y-2), (PH_X+PH_W+2, PH_Y+PH_H+2)], radius=6, fill=WHITE)
    photo_path = os.path.join("dataset", emp_id + ".jpg")
    try:
        ph = Image.open(photo_path).convert("RGB").resize((PH_W, PH_H), Image.LANCZOS)
        front.paste(ph, (PH_X, PH_Y))
    except Exception:
        fd.rounded_rectangle([(PH_X, PH_Y), (PH_X+PH_W, PH_Y+PH_H)], radius=4, fill=MID)
        ini = row[1][0].upper() if row and row[1] else "?"
        cx(fd, ini, fnt(56, bold=True), CW, PH_Y + PH_H // 2 - 38, WHITE)

    cx(fd, (row[1] or "Unknown")[:24], fnt(18, bold=True), CW, 328, DGRAY)
    cx(fd, (row[2] or "Employee")[:28], fnt(12),            CW, 352, MGRAY)
    fd.rectangle([(40, 372), (CW-40, 374)], fill=PALE)

    info_rows = [
        ("Employee ID", row[0]  if row            else "-"),
        ("Email",       row[3]  if row and row[3] else "-"),
        ("Phone",       row[8]  if row and row[8] else "-"),
        ("Blood Group", row[7]  if row and row[7] else "-"),
    ]
    y = 390
    for i, (lbl, val) in enumerate(info_rows):
        if i % 2 == 0:
            fd.rectangle([(0, y-4), (CW, y+38)], fill=LGRAY)
        cx(fd, lbl,           fnt(10),            CW, y+2,  MGRAY)
        cx(fd, str(val)[:34], fnt(13, bold=True), CW, y+17, DGRAY)
        y += 44

    bg_val = row[7] if row and row[7] else None
    if bg_val:
        bw = tw(fd, bg_val, fnt(13, bold=True)) + 28
        bx = (CW - bw) // 2
        by = y + 8
        fd.rounded_rectangle([(bx, by), (bx+bw, by+32)], radius=16, fill=RED)
        cx(fd, bg_val, fnt(13, bold=True), CW, by+8, WHITE)

    fd.rectangle([(0, CH-60), (CW, CH)], fill=BLUE)
    fd.rectangle([(0, CH-62), (CW, CH-60)], fill=GOLD)
    cx(fd, "Confidential  |  Not Transferable", fnt(10), CW, CH-44, PALE)
    cx(fd, "Property of the Organization",       fnt(10), CW, CH-26, (160,185,240))

    # ── BACK ───────────────────────────────────────────────
    back = Image.new("RGB", (CW, CH), LGRAY)
    bd   = ImageDraw.Draw(back)

    bd.rectangle([(0, 0), (CW, 110)], fill=BLUE)
    bd.ellipse([(CW-100, -60), (CW+60, 100)], fill=MID)
    cx(bd, "ATTENDANCE MANAGEMENT SYSTEM", fnt(14, bold=True), CW, 22, WHITE)
    cx(bd, "Employee Attendance Card", fnt(11), CW, 52, PALE)
    bd.rectangle([(0, 108), (CW, 113)], fill=GOLD)

    qr_path = os.path.join("static", "qrcodes", emp_id + ".png")
    if not os.path.exists(qr_path):
        qr_path = generate_qr(emp_id)
    QS = 240
    qr_x = (CW - QS) // 2
    qr_y = 148
    bd.rounded_rectangle([(qr_x-16, qr_y-16), (qr_x+QS+16, qr_y+QS+16)], radius=14, fill=WHITE)
    try:
        qr_img = Image.open(qr_path).convert("RGB").resize((QS, QS), Image.LANCZOS)
        back.paste(qr_img, (qr_x, qr_y))
    except Exception:
        cx(bd, "QR NOT AVAILABLE", fnt(13), CW, qr_y+QS//2, MGRAY)

    cx(bd, "Scan to Mark Attendance", fnt(14, bold=True), CW, qr_y+QS+28, BLUE)
    cx(bd, row[0] if row else "",     fnt(12),            CW, qr_y+QS+52, MGRAY)
    bd.rectangle([(40, qr_y+QS+78), (CW-40, qr_y+QS+80)], fill=(203,213,225))

    sub_info = [
        ("Name",        (row[1] or "-")[:26] if row else "-"),
        ("Designation", (row[2] or "-")[:26] if row else "-"),
        ("Blood Group", (row[7] or "-")      if row else "-"),
    ]
    sy = qr_y + QS + 94
    for lbl2, val2 in sub_info:
        cx(bd, lbl2, fnt(10),            CW, sy,    MGRAY)
        cx(bd, val2, fnt(12, bold=True), CW, sy+14, DGRAY)
        sy += 42

    bd.rectangle([(36, sy+8), (CW-36, sy+10)], fill=(203,213,225))
    cx(bd, "If found, please return to:", fnt(10),            CW, sy+18, MGRAY)
    cx(bd, "HR Department",               fnt(12, bold=True), CW, sy+34, BLUE)
    if row and row[3]:
        cx(bd, row[3][:34], fnt(10), CW, sy+54, MGRAY)

    bd.rectangle([(0, CH-100), (CW, CH-68)], fill=DARK)
    bd.rectangle([(0, CH-60),  (CW, CH)],    fill=BLUE)
    bd.rectangle([(0, CH-62),  (CW, CH-60)], fill=GOLD)
    cx(bd, "Authorized Personnel Only  |  Not Transferable", fnt(10), CW, CH-44, PALE)
    cx(bd, "Misuse is subject to disciplinary action",        fnt(10), CW, CH-26, (160,185,240))

    # ── Combine front + back ───────────────────────────────
    GAP, LBL_H = 40, 24
    BGCOL = (215, 225, 240)
    total = Image.new("RGB", (CW*2 + GAP, CH + LBL_H), BGCOL)
    td = ImageDraw.Draw(total)
    td.text((10, 4),               "FRONT", font=fnt(13, bold=True), fill=BLUE)
    td.text((CW + GAP + 10, 4),   "BACK",  font=fnt(13, bold=True), fill=BLUE)
    total.paste(front, (0,      LBL_H))
    total.paste(back,  (CW+GAP, LBL_H))

    buf = _io2.BytesIO()
    total.save(buf, format="PNG", dpi=(200, 200))
    buf.seek(0)
    return buf


@app.route("/employee_portal")
@employee_required
def employee_portal():
    emp_id = session["employee_id"]
    db     = get_db_connection()
    cursor = db.cursor(buffered=True)

    cursor.execute("""
        SELECT e.employee_id, e.name, e.role, e.email, e.face_image,
               e.date_of_joining,
               COALESCE(sc.salary_per_day, 0) AS salary_per_day,
               sh.name AS shift_name, sh.start_time AS shift_start, sh.end_time AS shift_end,
               e.phone, e.gender, e.dob, e.blood_group,
               e.address, e.city, e.state, e.pincode,
               e.emergency_contact_name, e.emergency_contact_phone, e.emergency_contact_relation,
               e.aadhar_number, e.pan_number, e.bank_name, e.bank_account, e.bank_ifsc, e.uan_number,
               e.qr_code, e.work_mode, e.about_me, e.manager_name, e.department,
               e.fingerprint_credential_id
        FROM employees e
        LEFT JOIN salary_config sc ON e.employee_id = sc.employee_id
        LEFT JOIN shifts sh ON e.shift_id = sh.id
        WHERE e.employee_id = %s
    """, (emp_id,))
    emp = list(cursor.fetchone())
    # emp indices:
    # [0]=id [1]=name [2]=role [3]=email [4]=face_image [5]=date_of_joining
    # [6]=salary_per_day [7]=shift_name [8]=shift_start [9]=shift_end
    # [10]=phone [11]=gender [12]=dob [13]=blood_group
    # [14]=address [15]=city [16]=state [17]=pincode
    # [18]=emergency_contact_name [19]=emergency_contact_phone [20]=emergency_contact_relation
    # [21]=aadhar_number [22]=pan_number [23]=bank_name [24]=bank_account [25]=bank_ifsc [26]=uan_number
    # [27]=qr_code [28]=work_mode [29]=about_me [30]=manager_name [31]=department
    # [32]=fingerprint_credential_id
    fp_enrolled = bool(emp[32]) if len(emp) > 32 else False
    # Decrypt PII fields
    for _pii_idx in (21, 22, 24, 25, 26):
        if _pii_idx < len(emp):
            emp[_pii_idx] = decrypt_pii(emp[_pii_idx])

    today = datetime.date.today()
    cursor.execute(
        "SELECT login_time, logout_time, status, logout_status, attendance_type "
        "FROM attendance WHERE employee_id=%s AND date=%s",
        (emp_id, today)
    )
    today_att = cursor.fetchone()

    year  = int(request.args.get("year",  today.year))
    month = int(request.args.get("month", today.month))
    _, last_day = calendar.monthrange(year, month)
    cursor.execute("""
        SELECT date, login_time, logout_time, status, logout_status, attendance_type
        FROM attendance
        WHERE employee_id=%s AND date BETWEEN %s AND %s
        ORDER BY date DESC
    """, (emp_id, datetime.date(year, month, 1), datetime.date(year, month, last_day)))
    monthly_att = cursor.fetchall()

    holidays_set  = fetch_holidays_set(year, month)
    # Fetch holiday names for attendance calendar tooltips
    cursor.execute(
        "SELECT date, name FROM holidays WHERE date BETWEEN %s AND %s",
        (datetime.date(year, month, 1), datetime.date(year, month, calendar.monthrange(year, month)[1]))
    )
    att_hol_name_map = {row[0]: row[1] for row in cursor.fetchall()}
    billable_past = get_billable_past_days(year, month)
    att_by_date   = {r[0]: r for r in monthly_att}
    full_days = half_days = late_days = absent_days = 0
    total_seconds = 0
    for d in billable_past:
        row = att_by_date.get(d)
        if row:
            _, login_t, logout_t, status, _ls, att_type = row
            final = att_type if att_type else infer_type_legacy(status, login_t, logout_t)
            if   final in ("Full Day", "Approved Leave"): full_days   += 1
            elif final == "Late - Full Day":             late_days   += 1
            elif final in ("Half Day", "Present"):       half_days   += 1
            else:                                        absent_days += 1
            if login_t and logout_t:
                li = login_t.total_seconds()  if hasattr(login_t,  "total_seconds") else (login_t.hour*3600  + login_t.minute*60  + login_t.second)
                lo = logout_t.total_seconds() if hasattr(logout_t, "total_seconds") else (logout_t.hour*3600 + logout_t.minute*60 + logout_t.second)
                if lo > li:
                    total_seconds += int(lo - li)
        else:
            absent_days += 1

    total_hours_str = f"{total_seconds // 3600}h {(total_seconds % 3600) // 60}m"
    billable_count  = len(billable_past)
    present_equiv   = full_days + late_days + half_days * 0.5
    att_pct         = round(present_equiv / billable_count * 100, 1) if billable_count else 0

    # Calendar data for JS rendering
    cal_data = {}
    _, month_days = calendar.monthrange(year, month)
    for day in range(1, month_days + 1):
        d = datetime.date(year, month, day)
        if d in holidays_set:
            cal_data[day] = "holiday"
        elif d.weekday() == 6:
            cal_data[day] = "weekend"
        elif d > today:
            cal_data[day] = "future"
        else:
            row = att_by_date.get(d)
            if row:
                _, login_t, logout_t, status, _ls, att_type = row
                final = att_type if att_type else infer_type_legacy(status, login_t, logout_t)
                if   final == "Full Day":               cal_data[day] = "full"
                elif final == "Late - Full Day":        cal_data[day] = "late"
                elif final in ("Half Day", "Present"):  cal_data[day] = "half"
                else:                                   cal_data[day] = "absent"
            else:
                cal_data[day] = "absent"
    cal_hol_names = {d.day: n for d, n in att_hol_name_map.items()}
    cal_year      = year
    cal_month     = month
    cal_first_dow = datetime.date(year, month, 1).weekday()  # 0=Mon

    cursor.execute("""
        SELECT lr.leave_date, lr.reason, lr.status, lr.created_at,
               COALESCE(lt.name, '') AS leave_type_name, lr.id
        FROM leave_requests lr
        LEFT JOIN leave_types lt ON lr.leave_type_id = lt.id
        WHERE lr.employee_id=%s
        ORDER BY lr.created_at DESC LIMIT 20
    """, (emp_id,))
    my_leaves = cursor.fetchall()

    cursor.execute("""
        SELECT last_working_day, reason, status, created_at
        FROM resignation_requests WHERE employee_id=%s
        ORDER BY created_at DESC LIMIT 1
    """, (emp_id,))
    my_resignation = cursor.fetchone()

    cursor.execute("""
        SELECT id, category, subject, priority, status, admin_response, created_at
        FROM tickets WHERE employee_id=%s
        ORDER BY created_at DESC LIMIT 20
    """, (emp_id,))
    my_tickets = cursor.fetchall()

    # Leave types & per-type balances
    try:
        cursor.execute(
            "SELECT id, name, annual_quota, is_paid FROM leave_types WHERE is_active=1 ORDER BY id"
        )
        leave_types_list = cursor.fetchall()
        # Ensure balances exist for this employee
        assign_leave_balances_for_employee(cursor, emp_id, today.year)
        # Fetch from leave_balances table
        cursor.execute("""
            SELECT lt.id, lt.name, lt.annual_quota, lt.is_paid,
                   COALESCE(lb.total_days, lt.annual_quota) as total,
                   COALESCE(lb.used_days, 0) as used
            FROM leave_types lt
            LEFT JOIN leave_balances lb ON lb.employee_id=%s
                AND lb.leave_type_id=lt.id AND lb.year=%s
            WHERE lt.is_active=1 ORDER BY lt.id
        """, (emp_id, today.year))
        leave_type_balances = []
        annual_leave_quota = 0
        leaves_used = 0
        for lt_id, lt_name, lt_quota, lt_paid, total, used in cursor.fetchall():
            used = float(used or 0)
            total = int(total or lt_quota)
            remaining = max(0, total - used)
            leave_type_balances.append({
                "id": lt_id, "name": lt_name, "quota": total,
                "used": used, "balance": remaining, "is_paid": lt_paid
            })
            annual_leave_quota += total
            leaves_used += used
        leave_balance = max(0, annual_leave_quota - leaves_used)
    except Exception:
        leave_type_balances = []
        annual_leave_quota  = 12
        cursor.execute("""
            SELECT COUNT(*) FROM leave_requests
            WHERE employee_id=%s AND EXTRACT(YEAR FROM leave_date)=%s AND status IN ('Approved','Pending')
        """, (emp_id, today.year))
        leaves_used   = cursor.fetchone()[0] or 0
        leave_balance = max(0, annual_leave_quota - leaves_used)

    # Announcements for dashboard (public + private addressed to this employee)
    cursor.execute("""
        SELECT id, title, content, priority, created_at
        FROM announcements
        WHERE COALESCE(visibility,'public') = 'public'
           OR (visibility = 'private' AND target_employee_id = %s)
        ORDER BY created_at DESC LIMIT 10
    """, (emp_id,))
    announcements = cursor.fetchall()

    # Pending leave count for nav badge
    cursor.execute("SELECT COUNT(*) FROM leave_requests WHERE employee_id=%s AND status='Pending'", (emp_id,))
    pending_leaves_count = cursor.fetchone()[0] or 0

    # Open ticket count for nav badge
    cursor.execute("SELECT COUNT(*) FROM tickets WHERE employee_id=%s AND status='Open'", (emp_id,))
    open_tickets_count = cursor.fetchone()[0] or 0

    # Unread notification count for bell icon
    try:
        cursor.execute(
            "SELECT COUNT(*) FROM notifications WHERE recipient_type='employee' AND employee_id=%s AND is_read=FALSE",
            (emp_id,)
        )
        unread_notifications_web = cursor.fetchone()[0] or 0
    except Exception:
        unread_notifications_web = 0

    # Upcoming holidays (next 3 from today) for dashboard widget
    cursor.execute("""
        SELECT date, name FROM holidays WHERE date >= %s ORDER BY date LIMIT 3
    """, (today,))
    upcoming_holidays = cursor.fetchall()

    # Upcoming holidays for leave planning panel (rest of year, up to 15)
    cursor.execute("""
        SELECT date, name FROM holidays
        WHERE date >= %s AND EXTRACT(YEAR FROM date) = %s
        ORDER BY date LIMIT 15
    """, (today, today.year))
    leave_holidays = cursor.fetchall()

    # Holiday calendar data for employee view
    hol_year = int(request.args.get("hol_year", today.year))
    cursor.execute("SELECT id, date, name FROM holidays WHERE EXTRACT(YEAR FROM date)=%s ORDER BY date", (hol_year,))
    hol_rows = cursor.fetchall()
    hol_map = {}
    for row in hol_rows:
        date_val = row[1]
        if isinstance(date_val, datetime.date):
            hol_map[date_val] = (row[0], row[2])
    sun_cal_obj = calendar.Calendar(firstweekday=6)
    emp_hol_cal = []
    for _m in range(1, 13):
        m_hols = {}
        for _d, (_hid, _hname) in hol_map.items():
            if _d.month == _m:
                m_hols[_d.day] = (_hid, _hname)
        emp_hol_cal.append({
            'month_num':  _m,
            'month_name': calendar.month_name[_m],
            'weeks':      sun_cal_obj.monthdayscalendar(hol_year, _m),
            'holidays':   m_hols,
        })

    # Employee's own incentive history
    try:
        cursor.execute("""
            SELECT ig.title, ig.description, ei.month, ei.year, ei.amount, ei.notes, ei.awarded_at
            FROM employee_incentives ei
            JOIN incentive_goals ig ON ei.goal_id = ig.id
            WHERE ei.employee_id = %s
            ORDER BY ei.year DESC, ei.month DESC, ei.awarded_at DESC
        """, (emp_id,))
        my_incentives = cursor.fetchall()
        cursor.execute(
            "SELECT COALESCE(SUM(amount),0) FROM employee_incentives WHERE employee_id=%s AND year=%s",
            (emp_id, today.year)
        )
        total_incentive_year = float(cursor.fetchone()[0])
    except Exception:
        my_incentives = []
        total_incentive_year = 0.0

    # Employee work experience & education
    try:
        cursor.execute(
            "SELECT id, company, designation, from_year, to_year, is_current, description "
            "FROM employee_experience WHERE employee_id=%s ORDER BY is_current DESC, from_year DESC",
            (emp_id,)
        )
        my_experience = [
            {"id": r[0], "company": r[1], "designation": r[2], "from_year": r[3],
             "to_year": r[4], "is_current": r[5], "description": r[6]}
            for r in cursor.fetchall()
        ]
    except Exception:
        my_experience = []

    try:
        cursor.execute(
            "SELECT id, degree, institution, year_of_passing, percentage "
            "FROM employee_education WHERE employee_id=%s ORDER BY year_of_passing DESC",
            (emp_id,)
        )
        my_education = [
            {"id": r[0], "degree": r[1], "institution": r[2], "year_of_passing": r[3], "percentage": r[4]}
            for r in cursor.fetchall()
        ]
    except Exception:
        my_education = []

    try:
        cursor.execute(
            "SELECT id, doc_type, original_name, uploaded_by, uploaded_at FROM employee_documents WHERE employee_id=%s ORDER BY uploaded_at DESC",
            (emp_id,)
        )
        my_docs = cursor.fetchall()
    except Exception:
        my_docs = []

    try:
        cursor.execute(
            "SELECT date, shift_end, actual_logout, ot_minutes, ot_pay, status FROM overtime_records WHERE employee_id=%s AND EXTRACT(YEAR FROM date)=%s ORDER BY date DESC LIMIT 20",
            (emp_id, today.year)
        )
        my_overtime = cursor.fetchall()
    except Exception:
        my_overtime = []

    # Salary summary for Earnings tab
    salary_per_day = float(emp[6]) if emp[6] else 0.0
    gross_this_month = (full_days + late_days) * salary_per_day + half_days * salary_per_day * 0.5
    deduction_this_month = absent_days * salary_per_day + half_days * salary_per_day * 0.5
    try:
        cursor.execute(
            "SELECT COALESCE(SUM(amount),0) FROM employee_incentives WHERE employee_id=%s AND month=%s AND year=%s",
            (emp_id, today.month, today.year)
        )
        incentives_this_month = float(cursor.fetchone()[0])
    except Exception:
        incentives_this_month = 0.0
    try:
        cursor.execute(
            "SELECT COALESCE(SUM(ot_pay),0) FROM overtime_records WHERE employee_id=%s AND EXTRACT(MONTH FROM date)=%s AND EXTRACT(YEAR FROM date)=%s AND status='Approved'",
            (emp_id, today.month, today.year)
        )
        ot_pay_this_month = float(cursor.fetchone()[0] or 0)
    except Exception:
        ot_pay_this_month = 0.0
    net_this_month = gross_this_month + incentives_this_month + ot_pay_this_month

    # Comp-off balance
    try:
        cursor.execute("SELECT COALESCE(compoff_minutes_per_day,480) FROM company_settings LIMIT 1")
        mpd_row = cursor.fetchone()
        compoff_mpd = int(mpd_row[0]) if mpd_row else 480
        cursor.execute("SELECT COALESCE(earned_minutes,0), COALESCE(used_minutes,0) FROM compoff_balance WHERE employee_id=%s", (emp_id,))
        co_row = cursor.fetchone() or (0, 0)
        compoff_earned_days = round(co_row[0] / compoff_mpd, 1) if compoff_mpd else 0
        compoff_avail_days  = round(max(0, co_row[0] - co_row[1]) / compoff_mpd, 1) if compoff_mpd else 0
    except Exception:
        compoff_earned_days = 0
        compoff_avail_days  = 0

    # Last 3 months payslip summaries
    recent_payslips = []
    py2, pm2 = today.year, today.month
    for _ in range(3):
        pm2 -= 1
        if pm2 == 0:
            pm2 = 12; py2 -= 1
        _, ld = calendar.monthrange(py2, pm2)
        cursor.execute("""
            SELECT date, login_time, logout_time, status, logout_status, attendance_type
            FROM attendance WHERE employee_id=%s AND date BETWEEN %s AND %s
        """, (emp_id, datetime.date(py2, pm2, 1), datetime.date(py2, pm2, ld)))
        p_att = cursor.fetchall()
        p_billable = get_billable_past_days(py2, pm2)
        p_att_map  = {r[0]: r for r in p_att}
        p_full = p_late = p_half = p_absent = 0
        for d in p_billable:
            row = p_att_map.get(d)
            if row:
                _, lt, lot, st, _ls, at = row
                final = at if at else infer_type_legacy(st, lt, lot)
                if   final in ("Full Day", "Approved Leave"): p_full   += 1
                elif final == "Late - Full Day":              p_late   += 1
                elif final in ("Half Day", "Present"):        p_half   += 1
                else:                                         p_absent += 1
            else:
                p_absent += 1
        p_gross = (p_full + p_late) * salary_per_day + p_half * salary_per_day * 0.5
        try:
            cursor.execute("SELECT COALESCE(SUM(amount),0) FROM employee_incentives WHERE employee_id=%s AND month=%s AND year=%s", (emp_id, pm2, py2))
            p_inc = float(cursor.fetchone()[0])
        except Exception:
            p_inc = 0.0
        try:
            cursor.execute("SELECT COALESCE(SUM(ot_pay),0) FROM overtime_records WHERE employee_id=%s AND EXTRACT(MONTH FROM date)=%s AND EXTRACT(YEAR FROM date)=%s AND status='Approved'", (emp_id, pm2, py2))
            p_ot = float(cursor.fetchone()[0] or 0)
        except Exception:
            p_ot = 0.0
        recent_payslips.append({
            'month': calendar.month_name[pm2], 'year': py2,
            'gross': p_gross, 'incentives': p_inc, 'ot_pay': p_ot,
            'net': p_gross + p_inc + p_ot,
            'present': p_full + p_late + p_half, 'absent': p_absent,
        })

    # Shift swap data
    try:
        cursor.execute("""
            SELECT ssr.id, ssr.target_id, et.name, ts.name AS tgt_shift,
                   ssr.reason, ssr.status, ssr.created_at
            FROM shift_swap_requests ssr
            JOIN employees et ON et.employee_id = ssr.target_id
            JOIN shifts ts ON ts.id = ssr.target_shift_id
            WHERE ssr.requester_id=%s ORDER BY ssr.created_at DESC LIMIT 20
        """, (emp_id,))
        my_swap_requests = cursor.fetchall()
        cursor.execute("""
            SELECT ssr.id, ssr.requester_id, er.name, rs.name AS req_shift,
                   ssr.reason, ssr.status, ssr.created_at
            FROM shift_swap_requests ssr
            JOIN employees er ON er.employee_id = ssr.requester_id
            JOIN shifts rs ON rs.id = ssr.requester_shift_id
            WHERE ssr.target_id=%s AND ssr.status='Pending_Target' ORDER BY ssr.created_at DESC
        """, (emp_id,))
        incoming_swap_requests = cursor.fetchall()
        cursor.execute("""
            SELECT e.employee_id, e.name, COALESCE(s.shift_name,''),
                   COALESCE(TO_CHAR(s.start_time,'HH24:MI'),''),
                   COALESCE(TO_CHAR(s.end_time,'HH24:MI'),''),
                   COALESCE(e.department,''), COALESCE(e.designation,'')
            FROM employees e
            LEFT JOIN shifts s ON s.id = e.shift_id
            WHERE e.employee_id != %s AND e.is_active=1
            ORDER BY e.name
        """, (emp_id,))
        swap_eligible_employees = cursor.fetchall()
    except Exception:
        my_swap_requests = []
        incoming_swap_requests = []
        swap_eligible_employees = []

    cursor.close(); db.close()

    # Build last 12 months list for pay slips section
    payslip_months = []
    py, pm = today.year, today.month
    for _ in range(12):
        payslip_months.append((py, pm, calendar.month_name[pm]))
        pm -= 1
        if pm == 0:
            pm = 12; py -= 1

    return render_template("employee_portal.html",
        emp=emp,
        today_date=today,
        today=today.strftime("%d %b %Y"),
        today_long=today.strftime("%A, %d %B %Y"),
        today_att=today_att,
        monthly_att=monthly_att,
        full_days=full_days, late_days=late_days,
        half_days=half_days, absent_days=absent_days,
        billable=billable_count,
        my_leaves=my_leaves,
        my_resignation=my_resignation,
        my_tickets=my_tickets,
        leave_sent=request.args.get("leave_sent") == "1",
        resigned=request.args.get("resigned") == "1",
        ticket_sent=request.args.get("ticket_sent") == "1",
        month_name=datetime.date(year, month, 1).strftime("%B %Y"),
        selected_month=f"{year}-{month:02d}",
        att_pct=att_pct,
        total_hours=total_hours_str,
        cal_data=cal_data,
        cal_hol_names=cal_hol_names,
        cal_year=cal_year,
        cal_month=cal_month,
        cal_first_dow=cal_first_dow,
        sel_year=year,
        sel_month=month,
        years=list(range(today.year - 2, today.year + 1)),
        months=[(i, datetime.date(year, i, 1).strftime("%B")) for i in range(1, 13)],
        payslip_months=payslip_months,
        leave_balance=leave_balance,
        leaves_used=leaves_used,
        annual_leave_quota=annual_leave_quota,
        leave_type_balances=leave_type_balances,
        leave_types_for_form=[{"id": lt[0], "name": lt[1]} for lt in (leave_types_list if leave_types_list else [])],
        announcements=announcements,
        pending_leaves_count=pending_leaves_count,
        open_tickets_count=open_tickets_count,
        unread_notifications_web=unread_notifications_web,
        upcoming_holidays=upcoming_holidays,
        leave_holidays=leave_holidays,
        hol_year=hol_year,
        emp_hol_cal=emp_hol_cal,
        all_holidays_list=hol_rows,
        my_incentives=my_incentives,
        total_incentive_year=total_incentive_year,
        my_experience=my_experience,
        my_education=my_education,
        my_docs=my_docs,
        my_overtime=my_overtime,
        compoff_avail_days=compoff_avail_days,
        compoff_earned_days=compoff_earned_days,
        salary_per_day=salary_per_day,
        gross_this_month=gross_this_month,
        deduction_this_month=deduction_this_month,
        incentives_this_month=incentives_this_month,
        ot_pay_this_month=ot_pay_this_month,
        net_this_month=net_this_month,
        recent_payslips=recent_payslips,
        my_swap_requests=my_swap_requests,
        incoming_swap_requests=incoming_swap_requests,
        swap_eligible_employees=swap_eligible_employees,
        swap_sent=request.args.get("swap_sent") == "1",
        swap_responded=request.args.get("swap_responded") == "1",
        swap_error=request.args.get("swap_error", ""),
        fp_enrolled=fp_enrolled,
        fp_enabled=get_auth_config().get("fingerprint_enabled", False),
    )


# ─────────────────────────── PERFORMANCE MANAGEMENT ───────────────────────────

RATING_LABELS = {0: "Not Rated", 1: "Unsatisfactory", 2: "Needs Improvement",
                 3: "Meets Expectations", 4: "Exceeds Expectations", 5: "Outstanding"}

@app.route("/performance")
@admin_required
def performance():
    today  = datetime.date.today()
    q      = int(request.args.get("quarter", (today.month - 1) // 3 + 1))
    yr     = int(request.args.get("year", today.year))
    dept   = request.args.get("dept", "")
    active_tab = request.args.get("tab", "performance")

    db = get_db_connection()
    cursor = db.cursor(buffered=True)

    active_cid = session.get("active_company_id")
    dept_filter = "AND e.department=%s" if dept else ""
    co_filter   = "AND e.company_id=%s" if active_cid else ""
    params = [yr, q] + ([dept] if dept else []) + ([active_cid] if active_cid else [])
    cursor.execute(f"""
        SELECT e.employee_id, e.name, COALESCE(e.role,''), COALESCE(e.department,''),
               pr.id, COALESCE(pr.overall_rating,0), COALESCE(pr.status,'—'),
               (SELECT COUNT(*) FROM performance_kpis pk WHERE pk.review_id=pr.id) AS kpi_count
        FROM employees e
        LEFT JOIN performance_reviews pr
            ON pr.employee_id=e.employee_id AND pr.year=%s AND pr.quarter=%s
        WHERE e.is_active=1 {dept_filter} {co_filter}
        ORDER BY e.name
    """, params)
    employees = cursor.fetchall()

    if active_cid:
        cursor.execute("SELECT department FROM employees WHERE is_active=1 AND department IS NOT NULL AND department!='' AND company_id=%s GROUP BY department ORDER BY MIN(id) ASC", (active_cid,))
    else:
        cursor.execute("SELECT department FROM employees WHERE is_active=1 AND department IS NOT NULL AND department!='' GROUP BY department ORDER BY MIN(id) ASC")
    departments = [r[0] for r in cursor.fetchall()]

    # Announcements (admin sees all)
    cursor.execute("""
        SELECT a.id, a.title, a.content, a.priority, a.created_at,
               COALESCE(a.visibility,'public'), COALESCE(a.target_employee_id,''), COALESCE(e.name,'')
        FROM announcements a
        LEFT JOIN employees e ON e.employee_id = a.target_employee_id
        ORDER BY a.created_at DESC
    """)
    ann_list  = cursor.fetchall()
    pub_anns  = [r for r in ann_list if r[5] == 'public']
    priv_anns = [r for r in ann_list if r[5] == 'private']

    cursor.execute("SELECT employee_id, name FROM employees WHERE is_active=1 ORDER BY name")
    ann_emp_list = cursor.fetchall()

    cursor.execute("SELECT COUNT(*) FROM leave_requests WHERE status='Pending'")
    pending_leaves = cursor.fetchone()[0]
    cursor.execute("SELECT COUNT(*) FROM resignation_requests WHERE status='Pending'")
    pending_resignations = cursor.fetchone()[0]
    cursor.execute("SELECT COUNT(*) FROM tickets WHERE status IN ('Open','In Progress')")
    pending_tickets = cursor.fetchone()[0]
    cursor.execute("SELECT COALESCE(company_name,'') FROM company_settings LIMIT 1")
    co = cursor.fetchone()

    cursor.execute("SELECT id, label, min_rating, max_rating, hike_pct, incentive_pct, color FROM hike_config ORDER BY min_rating DESC")
    hike_bands = cursor.fetchall()

    hike_employees = []
    total_hike_cost = 0.0
    total_bonus_pool = 0.0
    hike_eligible_count = 0
    if active_tab == 'hike':
        _hike_co = "AND e.company_id=%s" if active_cid else ""
        _hike_params = (yr, q) + ((active_cid,) if active_cid else ())
        cursor.execute(f"""
            SELECT e.employee_id, e.name, COALESCE(e.role,''), COALESCE(e.department,''),
                   COALESCE(pr.overall_rating,0), COALESCE(pr.status,'—'),
                   COALESCE(sc.monthly_ctc,0)
            FROM employees e
            LEFT JOIN performance_reviews pr ON pr.employee_id=e.employee_id AND pr.year=%s AND pr.quarter=%s
            LEFT JOIN salary_config sc ON sc.employee_id=e.employee_id
            WHERE e.is_active=1 {_hike_co}
            ORDER BY e.name
        """, _hike_params)
        for (h_eid, h_name, h_role, h_dept, h_rating, h_status, h_ctc) in cursor.fetchall():
            h_rating = float(h_rating or 0)
            h_ctc    = float(h_ctc or 0)
            band_label, band_color, hike_pct, inc_pct = "Not Rated", "#94a3b8", 0.0, 0.0
            if h_rating > 0:
                for (_, blabel, bmin, bmax, bhike, binc, bcolor) in hike_bands:
                    if float(bmin) <= h_rating <= float(bmax):
                        band_label, band_color, hike_pct, inc_pct = blabel, bcolor, float(bhike), float(binc)
                        break
                hike_eligible_count += 1
            new_ctc = round(h_ctc * (1 + hike_pct / 100), 2) if h_ctc > 0 and hike_pct > 0 else h_ctc
            bonus   = round(h_ctc * inc_pct / 100, 2) if h_ctc > 0 and inc_pct > 0 else 0.0
            total_hike_cost  += max(0, new_ctc - h_ctc)
            total_bonus_pool += bonus
            hike_employees.append((h_eid, h_name, h_role, h_dept, h_rating, h_status,
                                   h_ctc, band_label, band_color, hike_pct, new_ctc, inc_pct, bonus))

    cursor.close(); db.close()

    return render_template("performance.html",
        employees=employees, departments=departments,
        quarter=q, year=yr, selected_dept=dept,
        rating_labels=RATING_LABELS,
        pending_leaves=pending_leaves,
        pending_resignations=pending_resignations,
        pending_tickets=pending_tickets, co=co,
        today=today,
        ann_list=ann_list,
        pub_anns=pub_anns,
        priv_anns=priv_anns,
        ann_emp_list=ann_emp_list,
        active_tab=active_tab,
        hike_bands=hike_bands,
        hike_employees=hike_employees,
        total_hike_cost=total_hike_cost,
        total_bonus_pool=total_bonus_pool,
        hike_eligible_count=hike_eligible_count,
    
        active_nav="performance",
    )


@app.route("/performance_review/<emp_id>", methods=["GET"])
@admin_required
def performance_review(emp_id):
    today = datetime.date.today()
    q     = int(request.args.get("quarter", (today.month - 1) // 3 + 1))
    yr    = int(request.args.get("year", today.year))

    db = get_db_connection()
    cursor = db.cursor(buffered=True)

    cursor.execute("""
        SELECT e.employee_id, e.name, COALESCE(e.role,''), COALESCE(e.department,''),
               COALESCE(e.email,''), COALESCE(e.phone,'')
        FROM employees e WHERE e.employee_id=%s
    """, (emp_id,))
    emp = cursor.fetchone()
    if not emp:
        cursor.close(); db.close()
        flash("Employee not found.", "error")
        return redirect("/performance")

    # Get or create review
    cursor.execute("""
        SELECT id, overall_rating, reviewer_feedback, employee_comment, status
        FROM performance_reviews WHERE employee_id=%s AND quarter=%s AND year=%s
    """, (emp_id, q, yr))
    review = cursor.fetchone()

    kpis = []
    if review:
        cursor.execute("""
            SELECT id, kpi_title, description, target, achievement, weight, rating, comments
            FROM performance_kpis WHERE review_id=%s ORDER BY id
        """, (review[0],))
        kpis = cursor.fetchall()

    # Past reviews for history tab
    cursor.execute("""
        SELECT id, quarter, year, overall_rating, status, created_at
        FROM performance_reviews WHERE employee_id=%s ORDER BY year DESC, quarter DESC LIMIT 8
    """, (emp_id,))
    history = cursor.fetchall()

    cursor.execute("SELECT COUNT(*) FROM leave_requests WHERE status='Pending'")
    pending_leaves = cursor.fetchone()[0]
    cursor.execute("SELECT COUNT(*) FROM resignation_requests WHERE status='Pending'")
    pending_resignations = cursor.fetchone()[0]
    cursor.execute("SELECT COUNT(*) FROM tickets WHERE status IN ('Open','In Progress')")
    pending_tickets = cursor.fetchone()[0]
    cursor.execute("SELECT COALESCE(company_name,'') FROM company_settings LIMIT 1")
    co = cursor.fetchone()
    cursor.close(); db.close()

    return render_template("performance_review.html",
        emp=emp, review=review, kpis=kpis, history=history,
        quarter=q, year=yr, rating_labels=RATING_LABELS,
        pending_leaves=pending_leaves,
        pending_resignations=pending_resignations,
        pending_tickets=pending_tickets, co=co
    ,
        active_nav="performance",
    )


@app.route("/performance_save_review", methods=["POST"])
@admin_required
def performance_save_review():
    emp_id   = request.form["employee_id"]
    q        = int(request.form["quarter"])
    yr       = int(request.form["year"])
    feedback = request.form.get("reviewer_feedback", "").strip()
    status   = request.form.get("status", "Draft")

    db = get_db_connection()
    cursor = db.cursor(buffered=True)
    cursor.execute("""
        INSERT INTO performance_reviews (employee_id, quarter, year, reviewer_feedback, status)
        VALUES (%s,%s,%s,%s,%s)
        ON CONFLICT (employee_id, quarter, year) DO UPDATE SET reviewer_feedback=%s, status=%s, updated_at=NOW()
    """, (emp_id, q, yr, feedback, status, feedback, status))
    db.commit()

    # Recalculate overall rating from KPIs
    cursor.execute("SELECT id FROM performance_reviews WHERE employee_id=%s AND quarter=%s AND year=%s", (emp_id, q, yr))
    rev = cursor.fetchone()
    if rev:
        cursor.execute("""
            SELECT weight, rating FROM performance_kpis WHERE review_id=%s AND rating > 0
        """, (rev[0],))
        kpi_rows = cursor.fetchall()
        if kpi_rows:
            total_weight = sum(r[0] for r in kpi_rows)
            weighted_sum = sum(r[0] * r[1] for r in kpi_rows)
            overall = round(weighted_sum / total_weight, 1) if total_weight > 0 else 0
            cursor.execute("UPDATE performance_reviews SET overall_rating=%s WHERE id=%s", (overall, rev[0]))
            db.commit()

    cursor.close(); db.close()
    flash("Review saved successfully.", "success")
    return redirect(f"/performance_review/{emp_id}?quarter={q}&year={yr}")


@app.route("/performance_add_kpi", methods=["POST"])
@admin_required
def performance_add_kpi():
    emp_id = request.form["employee_id"]
    q      = int(request.form["quarter"])
    yr     = int(request.form["year"])
    title  = request.form.get("kpi_title", "").strip()
    desc   = request.form.get("description", "").strip()
    target = request.form.get("target", "").strip()
    weight = int(request.form.get("weight", 20))

    if not title:
        flash("KPI title is required.", "error")
        return redirect(f"/performance_review/{emp_id}?quarter={q}&year={yr}")

    db = get_db_connection()
    cursor = db.cursor(buffered=True)

    # Ensure review exists
    cursor.execute("""
        INSERT INTO performance_reviews (employee_id, quarter, year, status)
        VALUES (%s,%s,%s,'Draft')
        ON CONFLICT (employee_id, quarter, year) DO UPDATE SET updated_at=NOW()
    """, (emp_id, q, yr))
    db.commit()

    cursor.execute("SELECT id FROM performance_reviews WHERE employee_id=%s AND quarter=%s AND year=%s", (emp_id, q, yr))
    rev_id = cursor.fetchone()[0]

    cursor.execute("""
        INSERT INTO performance_kpis (review_id, kpi_title, description, target, weight)
        VALUES (%s,%s,%s,%s,%s)
    """, (rev_id, title, desc, target, weight))
    db.commit()
    cursor.close(); db.close()
    flash("KPI added.", "success")
    return redirect(f"/performance_review/{emp_id}?quarter={q}&year={yr}")


@app.route("/performance_rate_kpi", methods=["POST"])
@admin_required
def performance_rate_kpi():
    kpi_id      = int(request.form["kpi_id"])
    emp_id      = request.form["employee_id"]
    q           = int(request.form["quarter"])
    yr          = int(request.form["year"])
    rating      = int(request.form.get("rating", 0))
    achievement = request.form.get("achievement", "").strip()
    comments    = request.form.get("comments", "").strip()

    db = get_db_connection()
    cursor = db.cursor(buffered=True)
    cursor.execute("""
        UPDATE performance_kpis SET rating=%s, achievement=%s, comments=%s WHERE id=%s
    """, (rating, achievement, comments, kpi_id))
    db.commit()

    # Recalculate overall rating
    cursor.execute("SELECT id FROM performance_reviews WHERE employee_id=%s AND quarter=%s AND year=%s", (emp_id, q, yr))
    rev = cursor.fetchone()
    if rev:
        cursor.execute("SELECT weight, rating FROM performance_kpis WHERE review_id=%s AND rating>0", (rev[0],))
        rows = cursor.fetchall()
        if rows:
            tw = sum(r[0] for r in rows); ws = sum(r[0]*r[1] for r in rows)
            cursor.execute("UPDATE performance_reviews SET overall_rating=%s WHERE id=%s",
                           (round(ws/tw, 1) if tw else 0, rev[0]))
            db.commit()

    cursor.close(); db.close()
    return redirect(f"/performance_review/{emp_id}?quarter={q}&year={yr}")


@app.route("/performance_delete_kpi", methods=["POST"])
@admin_required
def performance_delete_kpi():
    kpi_id = int(request.form["kpi_id"])
    emp_id = request.form["employee_id"]
    q      = int(request.form["quarter"])
    yr     = int(request.form["year"])
    db = get_db_connection()
    cursor = db.cursor(buffered=True)
    cursor.execute("DELETE FROM performance_kpis WHERE id=%s", (kpi_id,))
    db.commit()
    cursor.close(); db.close()
    return redirect(f"/performance_review/{emp_id}?quarter={q}&year={yr}")


@app.route("/my_performance")
@employee_required
def my_performance():
    emp_id = session["employee_id"]
    db = get_db_connection()
    cursor = db.cursor(buffered=True)

    cursor.execute("""
        SELECT pr.id, pr.quarter, pr.year, pr.overall_rating, pr.reviewer_feedback,
               pr.employee_comment, pr.status, pr.updated_at
        FROM performance_reviews pr
        WHERE pr.employee_id=%s ORDER BY pr.year DESC, pr.quarter DESC
    """, (emp_id,))
    reviews = cursor.fetchall()

    reviews_data = []
    for rev in reviews:
        cursor.execute("""
            SELECT kpi_title, target, achievement, weight, rating, comments
            FROM performance_kpis WHERE review_id=%s ORDER BY id
        """, (rev[0],))
        kpis = cursor.fetchall()
        reviews_data.append({"review": rev, "kpis": kpis})

    cursor.execute("SELECT name, COALESCE(role,''), COALESCE(department,''), face_image FROM employees WHERE employee_id=%s", (emp_id,))
    emp_info = cursor.fetchone()
    cursor.close(); db.close()

    return render_template("my_performance.html",
        reviews_data=reviews_data, emp_info=emp_info,
        emp_id=emp_id, rating_labels=RATING_LABELS
    )


@app.route("/performance_employee_comment", methods=["POST"])
@employee_required
def performance_employee_comment():
    rev_id  = int(request.form["review_id"])
    comment = request.form.get("comment", "").strip()
    emp_id  = session["employee_id"]
    db = get_db_connection()
    cursor = db.cursor(buffered=True)
    # Only allow comment on own review
    cursor.execute("UPDATE performance_reviews SET employee_comment=%s WHERE id=%s AND employee_id=%s",
                   (comment, rev_id, emp_id))
    db.commit()
    cursor.close(); db.close()
    flash("Comment submitted.", "success")
    return redirect("/my_performance")


@app.route("/performance_export")
@admin_required
def performance_export():
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    today = datetime.date.today()
    q  = int(request.args.get("quarter", (today.month - 1) // 3 + 1))
    yr = int(request.args.get("year", today.year))

    db = get_db_connection()
    cursor = db.cursor(buffered=True)

    cursor.execute("""
        SELECT e.employee_id, e.name, COALESCE(e.role,''), COALESCE(e.department,''),
               COALESCE(pr.overall_rating,0), COALESCE(pr.status,'Not Started'),
               COALESCE(pr.reviewer_feedback,''), COALESCE(pr.employee_comment,''),
               pr.id
        FROM employees e
        LEFT JOIN performance_reviews pr
            ON pr.employee_id=e.employee_id AND pr.year=%s AND pr.quarter=%s
        WHERE e.is_active=1
        ORDER BY e.name
    """, (yr, q))
    employees = cursor.fetchall()

    cursor.execute("""
        SELECT e.employee_id, e.name, pk.kpi_title, COALESCE(pk.description,''),
               COALESCE(pk.target,''), COALESCE(pk.achievement,''),
               pk.weight, COALESCE(pk.rating,0), COALESCE(pk.comments,'')
        FROM employees e
        JOIN performance_reviews pr ON pr.employee_id=e.employee_id AND pr.year=%s AND pr.quarter=%s
        JOIN performance_kpis pk ON pk.review_id=pr.id
        WHERE e.is_active=1
        ORDER BY e.name, pk.id
    """, (yr, q))
    kpis = cursor.fetchall()
    cursor.close(); db.close()

    wb = openpyxl.Workbook()

    # ── Styles ──
    hdr_font   = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill   = PatternFill("solid", fgColor="1E3A8A")
    hdr_align  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    alt_fill   = PatternFill("solid", fgColor="EFF6FF")
    thin       = Side(style="thin", color="BFDBFE")
    border     = Border(left=thin, right=thin, top=thin, bottom=thin)
    center     = Alignment(horizontal="center", vertical="center")

    def style_header(ws, cols):
        for col_idx, (title, width) in enumerate(cols, 1):
            cell = ws.cell(row=1, column=col_idx, value=title)
            cell.font      = hdr_font
            cell.fill      = hdr_fill
            cell.alignment = hdr_align
            cell.border    = border
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        ws.row_dimensions[1].height = 30

    def style_data_cell(cell, row_idx):
        cell.border    = border
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        if row_idx % 2 == 0:
            cell.fill = alt_fill

    # ── Sheet 1: Summary ──
    ws1 = wb.active
    ws1.title = f"Q{q} {yr} Summary"
    q_labels = {1:"Jan–Mar", 2:"Apr–Jun", 3:"Jul–Sep", 4:"Oct–Dec"}
    ws1.append([])
    ws1.merge_cells("A1:H1")
    title_cell = ws1["A1"]
    title_cell.value     = f"Performance Summary — Q{q} ({q_labels.get(q,'')}) {yr}"
    title_cell.font      = Font(bold=True, size=14, color="1E3A8A")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 36

    cols_s = [
        ("Employee ID", 16), ("Employee Name", 24), ("Role", 18), ("Department", 18),
        ("KPI Count", 12), ("Overall Rating (/ 5)", 20), ("Status", 16), ("Reviewer Feedback", 35),
    ]
    for col_idx, (title, width) in enumerate(cols_s, 1):
        cell = ws1.cell(row=2, column=col_idx, value=title)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = hdr_align
        cell.border    = border
        ws1.column_dimensions[get_column_letter(col_idx)].width = width
    ws1.row_dimensions[2].height = 28

    kpi_counts = {}
    for row in kpis:
        kpi_counts[row[0]] = kpi_counts.get(row[0], 0) + 1

    for r_idx, (emp_id, name, role, dept, rating, status, feedback, _, _rev_id) in enumerate(employees, 3):
        row_data = [emp_id, name, role, dept, kpi_counts.get(emp_id, 0),
                    rating if rating else "—", status, feedback]
        for c_idx, val in enumerate(row_data, 1):
            cell = ws1.cell(row=r_idx, column=c_idx, value=val)
            style_data_cell(cell, r_idx)
            if c_idx == 6 and isinstance(val, (int, float)) and val > 0:
                if val >= 4:   cell.font = Font(color="15803D", bold=True)
                elif val >= 3: cell.font = Font(color="1D4ED8", bold=True)
                else:           cell.font = Font(color="DC2626", bold=True)
        ws1.row_dimensions[r_idx].height = 22
    ws1.freeze_panes = "A3"

    # ── Sheet 2: KPI Details ──
    ws2 = wb.create_sheet(f"Q{q} {yr} KPI Details")
    cols_k = [
        ("Employee ID", 16), ("Employee Name", 22), ("KPI Title", 28),
        ("Description", 30), ("Target", 20), ("Achievement", 20),
        ("Weight (%)", 13), ("Rating (1–5)", 14), ("Comments", 30),
    ]
    for col_idx, (title, width) in enumerate(cols_k, 1):
        cell = ws2.cell(row=1, column=col_idx, value=title)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = hdr_align
        cell.border    = border
        ws2.column_dimensions[get_column_letter(col_idx)].width = width
    ws2.row_dimensions[1].height = 28

    for r_idx, (emp_id, name, title, desc, target, achievement, weight, rating, comments) in enumerate(kpis, 2):
        for c_idx, val in enumerate([emp_id, name, title, desc, target, achievement, weight, rating or "—", comments], 1):
            cell = ws2.cell(row=r_idx, column=c_idx, value=val)
            style_data_cell(cell, r_idx)
        ws2.row_dimensions[r_idx].height = 20
    ws2.freeze_panes = "A2"

    # ── Sheet 3: Import Template ──
    ws3 = wb.create_sheet("Import Template")
    note_fill = PatternFill("solid", fgColor="FFF9C4")
    note_font = Font(italic=True, size=10, color="92400E")
    ws3.merge_cells("A1:J1")
    n = ws3["A1"]
    n.value     = "Fill in the rows below and import this file. Required columns: employee_id, kpi_title, weight, rating. Quarter & Year are selected in the import dialog."
    n.font      = note_font
    n.fill      = note_fill
    n.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws3.row_dimensions[1].height = 40

    tpl_cols = [
        ("employee_id*", 18), ("kpi_title*", 28), ("description", 28),
        ("target", 20), ("achievement", 20), ("weight*", 12),
        ("rating* (1-5)", 14), ("comments", 28), ("status", 16), ("reviewer_feedback", 35),
    ]
    for col_idx, (title, width) in enumerate(tpl_cols, 1):
        cell = ws3.cell(row=2, column=col_idx, value=title)
        req_fill = PatternFill("solid", fgColor="1E3A8A") if title.endswith("*") else PatternFill("solid", fgColor="475569")
        cell.font      = Font(bold=True, color="FFFFFF", size=10)
        cell.fill      = req_fill
        cell.alignment = hdr_align
        cell.border    = border
        ws3.column_dimensions[get_column_letter(col_idx)].width = width
    ws3.row_dimensions[2].height = 28

    sample_rows = [
        ["EMP001", "Code Quality", "Maintain clean, tested code", "95% coverage", "92%", 30, 4, "Good progress", "Submitted", ""],
        ["EMP001", "Delivery Speed", "Complete tasks on time", "95% on-time", "90%", 30, 3, "", "", ""],
        ["EMP001", "Team Collaboration", "Cross-team work", "4 collabs/qtr", "", 20, 5, "Excellent", "", ""],
        ["EMP001", "Documentation", "Keep docs updated", "100% coverage", "80%", 20, 3, "", "", "Strong Q2 performance"],
        ["EMP002", "Sales Target", "Hit monthly targets", "₹5L / month", "₹4.8L", 50, 4, "Near target", "Draft", ""],
        ["EMP002", "Customer Satisfaction", "Maintain CSAT score", ">=4.5 / 5", "4.3", 50, 3, "Needs improvement", "", ""],
    ]
    for r_idx, row in enumerate(sample_rows, 3):
        for c_idx, val in enumerate(row, 1):
            cell = ws3.cell(row=r_idx, column=c_idx, value=val)
            cell.border    = border
            cell.alignment = Alignment(vertical="center")
            if r_idx % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="F0F9FF")
        ws3.row_dimensions[r_idx].height = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"performance_Q{q}_{yr}.xlsx"
    from flask import send_file
    return send_file(buf, as_attachment=True, download_name=filename,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/performance_import", methods=["POST"])
@admin_required
def performance_import():
    import io
    import openpyxl

    q_raw  = request.form.get("quarter", "").strip()
    yr_raw = request.form.get("year", "").strip()
    if not q_raw.isdigit() or not yr_raw.isdigit():
        flash("Invalid quarter or year.", "error")
        return redirect("/performance")
    q  = int(q_raw)
    yr = int(yr_raw)

    f = request.files.get("excel_file")
    if not f or not f.filename.endswith((".xlsx", ".xls")):
        flash("Please upload a valid Excel file (.xlsx or .xls).", "error")
        return redirect(f"/performance?tab=performance&quarter={q}&year={yr}")

    try:
        wb = openpyxl.load_workbook(io.BytesIO(f.read()), data_only=True)
    except Exception:
        flash("Could not read the Excel file. Make sure it is a valid .xlsx file.", "error")
        return redirect(f"/performance?tab=performance&quarter={q}&year={yr}")

    # Find the data sheet — use first sheet that isn't "Import Template"
    sheet = None
    for ws in wb.worksheets:
        if ws.title != "Import Template":
            sheet = ws
            break
    if sheet is None:
        sheet = wb.active

    rows = list(sheet.iter_rows(values_only=True))
    if len(rows) < 2:
        flash("The Excel file has no data rows.", "error")
        return redirect(f"/performance?tab=performance&quarter={q}&year={yr}")

    # Find header row (first row with 'employee_id' in it)
    header_row_idx = None
    headers = []
    for idx, row in enumerate(rows):
        cells = [str(c).strip().lower() if c is not None else "" for c in row]
        if "employee_id" in cells:
            header_row_idx = idx
            headers = cells
            break
    if header_row_idx is None:
        flash("Could not find header row. Make sure the file has an 'employee_id' column.", "error")
        return redirect(f"/performance?tab=performance&quarter={q}&year={yr}")

    def col(name):
        try: return headers.index(name)
        except ValueError: return None

    ci_emp      = col("employee_id")
    ci_title    = col("kpi_title")
    ci_desc     = col("description")
    ci_target   = col("target")
    ci_achieve  = col("achievement")
    ci_weight   = col("weight") or col("weight*")
    ci_rating   = next((col(x) for x in ["rating* (1-5)", "rating (1-5)", "rating"] if col(x) is not None), None)
    ci_comments = col("comments")
    ci_status   = col("status")
    ci_feedback = col("reviewer_feedback")

    if ci_emp is None or ci_title is None:
        flash("Missing required columns: 'employee_id' and 'kpi_title'.", "error")
        return redirect(f"/performance?tab=performance&quarter={q}&year={yr}")

    # Parse data rows
    data_rows = rows[header_row_idx + 1:]
    employees_data = {}  # emp_id → {feedback, status, kpis: [...]}
    skipped = 0
    for row in data_rows:
        if not any(c for c in row if c is not None):
            continue
        emp_id = str(row[ci_emp]).strip() if row[ci_emp] is not None else ""
        title  = str(row[ci_title]).strip() if ci_title is not None and row[ci_title] is not None else ""
        if not emp_id or not title:
            skipped += 1
            continue

        try:
            weight = int(row[ci_weight]) if ci_weight is not None and row[ci_weight] not in (None, "") else 20
        except (ValueError, TypeError):
            weight = 20
        weight = max(1, min(100, weight))

        try:
            rating = round(float(row[ci_rating]), 1) if ci_rating is not None and row[ci_rating] not in (None, "") else 0.0
        except (ValueError, TypeError):
            rating = 0.0
        rating = max(0.0, min(5.0, rating))

        desc     = str(row[ci_desc]).strip()     if ci_desc    is not None and row[ci_desc]    not in (None,"") else ""
        target   = str(row[ci_target]).strip()   if ci_target  is not None and row[ci_target]  not in (None,"") else ""
        achieve  = str(row[ci_achieve]).strip()  if ci_achieve is not None and row[ci_achieve] not in (None,"") else ""
        comments = str(row[ci_comments]).strip() if ci_comments is not None and row[ci_comments] not in (None,"") else ""
        status   = str(row[ci_status]).strip()   if ci_status  is not None and row[ci_status]   not in (None,"") else "Draft"
        feedback = str(row[ci_feedback]).strip() if ci_feedback is not None and row[ci_feedback] not in (None,"") else ""

        if status not in ("Draft", "Submitted", "Acknowledged"):
            status = "Draft"

        if emp_id not in employees_data:
            employees_data[emp_id] = {"status": status, "feedback": feedback, "kpis": []}
        if feedback:
            employees_data[emp_id]["feedback"] = feedback
        if status in ("Submitted", "Acknowledged"):
            employees_data[emp_id]["status"] = status

        employees_data[emp_id]["kpis"].append({
            "title": title, "description": desc, "target": target,
            "achievement": achieve, "weight": weight, "rating": rating, "comments": comments,
        })

    if not employees_data:
        flash("No valid data rows found in the file.", "error")
        return redirect(f"/performance?tab=performance&quarter={q}&year={yr}")

    db = get_db_connection()
    cursor = db.cursor(buffered=True)

    # Validate employee IDs exist
    cursor.execute("SELECT employee_id FROM employees WHERE is_active=1")
    valid_ids = {r[0] for r in cursor.fetchall()}

    imported = 0
    unknown  = []
    for emp_id, emp_data in employees_data.items():
        if emp_id not in valid_ids:
            unknown.append(emp_id)
            continue

        feedback = emp_data["feedback"]
        status   = emp_data["status"]
        kpis     = emp_data["kpis"]

        # Upsert review
        cursor.execute("""
            INSERT INTO performance_reviews (employee_id, quarter, year, reviewer_feedback, status)
            VALUES (%s,%s,%s,%s,%s)
            ON CONFLICT (employee_id, quarter, year) DO UPDATE SET
                reviewer_feedback=EXCLUDED.reviewer_feedback, status=EXCLUDED.status, updated_at=NOW()
        """, (emp_id, q, yr, feedback, status))
        db.commit()

        cursor.execute("SELECT id FROM performance_reviews WHERE employee_id=%s AND quarter=%s AND year=%s", (emp_id, q, yr))
        rev_id = cursor.fetchone()[0]

        # Replace KPIs
        cursor.execute("DELETE FROM performance_kpis WHERE review_id=%s", (rev_id,))
        for kpi in kpis:
            cursor.execute("""
                INSERT INTO performance_kpis (review_id, kpi_title, description, target, achievement, weight, rating, comments)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s)
            """, (rev_id, kpi["title"], kpi["description"], kpi["target"],
                  kpi["achievement"], kpi["weight"], kpi["rating"], kpi["comments"]))
        db.commit()

        # Recalculate overall rating
        cursor.execute("SELECT weight, rating FROM performance_kpis WHERE review_id=%s AND rating>0", (rev_id,))
        rated = cursor.fetchall()
        if rated:
            total_w   = sum(r[0] for r in rated)
            weighted  = sum(r[0] * r[1] for r in rated)
            overall   = round(weighted / total_w, 1) if total_w else 0
            cursor.execute("UPDATE performance_reviews SET overall_rating=%s WHERE id=%s", (overall, rev_id))
            db.commit()
        imported += 1

    cursor.close(); db.close()

    msg = f"✅ Imported {imported} employee(s) for Q{q} {yr}."
    if skipped:   msg += f" {skipped} row(s) skipped (missing ID or KPI title)."
    if unknown:   msg += f" Unknown employee IDs: {', '.join(unknown)}."
    flash(msg, "success")
    return redirect(f"/performance?tab=performance&quarter={q}&year={yr}")


# ================================================================
#  TICKETS  (web)
# ================================================================

@app.route("/raise_ticket", methods=["POST"])
@employee_required
def raise_ticket():
    emp_id      = session["employee_id"]
    category    = request.form.get("category", "").strip()
    subject     = request.form.get("subject", "").strip()
    description = request.form.get("description", "").strip()
    priority    = request.form.get("priority", "Medium").strip()
    if not category or not subject or not description:
        return redirect("/employee_portal#tickets")
    db     = get_db_connection()
    cursor = db.cursor(buffered=True)
    cursor.execute(
        "INSERT INTO tickets (employee_id, category, subject, description, priority) "
        "VALUES (%s,%s,%s,%s,%s)",
        (emp_id, category, subject, description, priority)
    )
    db.commit(); cursor.close(); db.close()
    return redirect("/employee_portal?ticket_sent=1#tickets")


@app.route("/tickets")
@admin_required
def tickets_view():
    db     = get_db_connection()
    cursor = db.cursor(buffered=True)
    cursor.execute("""
        SELECT t.id, t.employee_id, e.name, t.category, t.subject, t.description,
               t.priority, t.status, t.admin_response, t.created_at, t.updated_at
        FROM tickets t
        JOIN employees e ON t.employee_id = e.employee_id
        ORDER BY CASE WHEN t.status='Open' THEN 0 WHEN t.status='In Progress' THEN 1 WHEN t.status='Resolved' THEN 2 WHEN t.status='Closed' THEN 3 ELSE 4 END,
                 CASE WHEN t.priority='High' THEN 0 WHEN t.priority='Medium' THEN 1 WHEN t.priority='Low' THEN 2 ELSE 3 END, t.created_at DESC
    """)
    all_tickets = cursor.fetchall()
    cursor.execute("SELECT COUNT(*) FROM tickets WHERE status IN ('Open','In Progress')")
    pending_tickets = cursor.fetchone()[0]
    cursor.execute("SELECT COUNT(*) FROM leave_requests WHERE status='Pending'")
    pending_leaves = cursor.fetchone()[0]
    cursor.execute("SELECT COUNT(*) FROM resignation_requests WHERE status='Pending'")
    pending_resignations = cursor.fetchone()[0]
    cursor.close(); db.close()
    return render_template("tickets.html",
        all_tickets=all_tickets,
        pending_tickets=pending_tickets,
        pending_leaves=pending_leaves,
        pending_resignations=pending_resignations,
        today=datetime.date.today().strftime("%d %b %Y"),
        shift_start=SHIFT_START.strftime("%I:%M %p"),
        shift_end=SHIFT_END.strftime("%I:%M %p"),
        active_nav="leaves",
    )


@app.route("/ticket_action/<int:tid>", methods=["POST"])
@admin_required
def ticket_action(tid):
    new_status     = request.form.get("status", "").strip()
    admin_response = request.form.get("admin_response", "").strip()
    allowed = ("Open", "In Progress", "Resolved", "Closed")
    is_ajax = request.headers.get("X-Requested-With") == "XMLHttpRequest"
    if new_status not in allowed:
        return (jsonify({"ok": False, "msg": "Invalid status."}), 400) if is_ajax else redirect("/tickets")
    db     = get_db_connection()
    cursor = db.cursor(buffered=True)

    cursor.execute("""
        SELECT t.subject, t.category, t.priority, t.description,
               e.name, e.email
        FROM tickets t
        JOIN employees e ON t.employee_id = e.employee_id
        WHERE t.id = %s
    """, (tid,))
    row = cursor.fetchone()

    cursor.execute(
        "UPDATE tickets SET status=%s, admin_response=%s WHERE id=%s",
        (new_status, admin_response or None, tid)
    )
    db.commit(); cursor.close(); db.close()

    msg = ""
    msg_type = "success"
    if row and admin_response:
        subject_text, category, priority, description, emp_name, emp_email = row
        if emp_email:
            _ecfg = get_email_config()
            if _ecfg:
                status_color = {"Resolved": "#16a34a", "Closed": "#64748b",
                                "In Progress": "#d97706"}.get(new_status, "#2563eb")
                _html = f"""
<div style="font-family:'Segoe UI',sans-serif;max-width:560px;margin:0 auto;background:#f8fafc;border-radius:16px;overflow:hidden;border:1px solid #dbeafe;">
  <div style="background:linear-gradient(135deg,#1e3a8a,#2563eb);padding:24px 28px;color:white;">
    <div style="font-size:20px;font-weight:700;">🎫 Ticket Update</div>
    <div style="font-size:13px;opacity:0.75;margin-top:4px;">Employee Attendance System</div>
  </div>
  <div style="padding:28px;">
    <p style="font-size:15px;color:#1e293b;margin-bottom:20px;">Hi <strong>{emp_name}</strong>, your ticket has been updated.</p>
    <div style="background:#fff;border:1px solid #dbeafe;border-radius:12px;padding:18px 20px;margin-bottom:20px;">
      <div style="font-size:12px;color:#64748b;font-weight:600;text-transform:uppercase;letter-spacing:0.5px;margin-bottom:6px;">Ticket Subject</div>
      <div style="font-size:15px;color:#1e293b;font-weight:700;margin-bottom:14px;">{subject_text}</div>
      <div style="display:flex;gap:12px;flex-wrap:wrap;">
        <span style="background:#dbeafe;color:#1d4ed8;padding:3px 10px;border-radius:20px;font-size:12px;font-weight:600;">{category}</span>
        <span style="background:#fef9c3;color:#92400e;padding:3px 10px;border-radius:20px;font-size:12px;font-weight:600;">{priority} Priority</span>
        <span style="background:{status_color}22;color:{status_color};padding:3px 10px;border-radius:20px;font-size:12px;font-weight:700;">{new_status}</span>
      </div>
    </div>
    <div style="background:#f0fdf4;border:1px solid #86efac;border-radius:12px;padding:18px 20px;margin-bottom:20px;">
      <div style="font-size:12px;color:#15803d;font-weight:700;text-transform:uppercase;letter-spacing:0.5px;margin-bottom:8px;">Admin Response</div>
      <div style="font-size:14px;color:#1e293b;white-space:pre-line;">{admin_response}</div>
    </div>
    <p style="font-size:12px;color:#94a3b8;text-align:center;margin:0;">This is an automated message — please do not reply.</p>
  </div>
</div>"""
                send_email_async(emp_email, f"Ticket Update: {subject_text}", _html, _ecfg)
                msg = f"✅ Ticket updated — notification queued for {emp_email}"
            else:
                msg = "Ticket updated. SMTP not configured — email not sent."
                msg_type = "warning"
        else:
            msg = "Ticket updated. Employee has no email on record."
            msg_type = "warning"
    else:
        msg = "✅ Ticket status updated."

    if is_ajax:
        return jsonify({"ok": True, "msg": msg, "type": msg_type, "new_status": new_status})
    flash(msg, msg_type)
    return redirect("/tickets")


# ================================================================
#  REST API  (used by the Flutter mobile app)
# ================================================================

def api_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        auth = request.headers.get("Authorization", "")
        if not auth.startswith("Bearer "):
            return jsonify({"ok": False, "msg": "Unauthorized"}), 401
        token_hash = _hash_token(auth[7:])
        with _db() as (cursor, _conn):
            cursor.execute("DELETE FROM api_tokens WHERE expires_at < NOW()")
            _conn.commit()
            cursor.execute(
                "SELECT identity FROM api_tokens WHERE token=%s AND token_type='admin' AND expires_at > NOW()",
                (token_hash,)
            )
            row = cursor.fetchone()
        if not row:
            return jsonify({"ok": False, "msg": "Invalid or expired token"}), 401
        from flask import g as _g
        _g.api_user = row[0]
        return f(*args, **kwargs)
    return wrapper


@app.route("/api/login", methods=["POST"])
@limiter.limit("5 per minute")
@limiter.limit("20 per hour")
def api_login():
    data     = request.get_json() or {}
    username = data.get("username", "")
    password = data.get("password", "")
    if "\x00" in username or "\x00" in password:
        return jsonify({"ok": False, "msg": "Invalid credentials"}), 401
    with _db() as (cursor, conn):
        cursor.execute("SELECT password FROM admin_users WHERE username=%s", (username,))
        row = cursor.fetchone()
        if row and check_password_hash(row[0], password):
            token = secrets.token_hex(32)
            cursor.execute(
                "INSERT INTO api_tokens (token, token_type, identity, expires_at) "
                "VALUES (%s, 'admin', %s, NOW() + INTERVAL '24 hours')",
                (_hash_token(token), username)
            )
            conn.commit()
            return jsonify({"ok": True, "token": token, "username": username})
    return jsonify({"ok": False, "msg": "Invalid credentials"}), 401


@app.route("/api/logout", methods=["POST"])
def api_logout():
    auth = request.headers.get("Authorization", "")
    if auth.startswith("Bearer "):
        with _db() as (cursor, conn):
            cursor.execute("DELETE FROM api_tokens WHERE token=%s", (_hash_token(auth[7:]),))
            conn.commit()
    return jsonify({"ok": True})


@app.route("/api/dashboard", methods=["GET"])
@api_required
def api_dashboard():
    db     = get_db_connection()
    cursor = db.cursor(buffered=True)
    today  = datetime.date.today()

    cursor.execute("SELECT COUNT(*) FROM employees")
    total = cursor.fetchone()[0]
    cursor.execute(
        "SELECT COUNT(DISTINCT employee_id) FROM attendance WHERE date=%s AND login_time IS NOT NULL",
        (today,)
    )
    present = cursor.fetchone()[0]
    cursor.execute(
        "SELECT COUNT(DISTINCT employee_id) FROM attendance WHERE date=%s AND status='Late Login'",
        (today,)
    )
    late = cursor.fetchone()[0]
    cursor.execute("""
        SELECT e.employee_id, e.name, a.login_time, a.logout_time, a.status,
               a.logout_status, a.attendance_type
        FROM employees e
        LEFT JOIN attendance a ON e.employee_id=a.employee_id AND a.date=%s
        ORDER BY e.name
    """, (today,))
    rows = cursor.fetchall()
    today_rows = [
        {
            "employee_id": r[0], "name": r[1],
            "login_time":  str(r[2]) if r[2] else None,
            "logout_time": str(r[3]) if r[3] else None,
            "login_status": r[4], "logout_status": r[5], "attendance_type": r[6],
        }
        for r in rows
    ]
    cursor.execute("SELECT COUNT(*) FROM leave_requests WHERE status='Pending'")
    pending_leaves = cursor.fetchone()[0]
    cursor.execute("SELECT COUNT(*) FROM resignation_requests WHERE status='Pending'")
    pending_resignations = cursor.fetchone()[0]
    cursor.execute("SELECT COUNT(*) FROM tickets WHERE status IN ('Open','In Progress')")
    pending_tickets = cursor.fetchone()[0]
    cursor.execute("SELECT COUNT(*) FROM notifications WHERE recipient_type='admin' AND is_read=FALSE")
    unread_notifications = cursor.fetchone()[0]
    cursor.close(); db.close()

    return jsonify({
        "ok": True, "total": total, "present": present,
        "absent": total - present, "late": late,
        "today": today.strftime("%d %b %Y"), "today_rows": today_rows,
        "pending_leaves": pending_leaves, "pending_resignations": pending_resignations,
        "pending_tickets": pending_tickets, "unread_notifications": unread_notifications,
    })


@app.route("/api/email_config", methods=["GET"])
@api_required
def api_get_email_config():
    cfg = get_email_config()
    # Never return the SMTP password to clients — they only need to know config exists.
    safe_cfg = {k: v for k, v in cfg.items() if k != "password"}
    safe_cfg["password_set"] = bool(cfg.get("password"))
    return jsonify({"ok": True, "config": safe_cfg})


@app.route("/api/email_config", methods=["POST"])
@api_required
def api_save_email_config():
    data      = request.get_json() or {}
    host      = data.get("smtp_host", "").strip()
    port      = int(data.get("smtp_port", 587))
    user      = data.get("smtp_user", "").strip()
    password  = data.get("smtp_pass", "").strip()
    from_name = data.get("from_name", "HR Department").strip()
    if not host or not user or not password:
        return jsonify({"ok": False, "msg": "host, user and password required"}), 400
    db     = get_db_connection()
    cursor = db.cursor(buffered=True)
    cursor.execute("DELETE FROM email_config")
    cursor.execute(
        "INSERT INTO email_config (smtp_host, smtp_port, smtp_user, smtp_pass, from_name) VALUES (%s,%s,%s,%s,%s)",
        (host, port, user, encrypt_pii(password), from_name)
    )
    db.commit()
    cursor.close(); db.close()
    return jsonify({"ok": True})


# ---------------- API: LEAVE REQUESTS ----------------


# ---------------- API: RESIGNATION REQUESTS ----------------


def employee_api_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        auth = request.headers.get("Authorization", "")
        if not auth.startswith("Bearer "):
            return jsonify({"ok": False, "msg": "Unauthorized"}), 401
        token_hash = _hash_token(auth[7:])
        with _db() as (cursor, _conn):
            cursor.execute("DELETE FROM api_tokens WHERE expires_at < NOW()")
            _conn.commit()
            cursor.execute(
                "SELECT identity FROM api_tokens WHERE token=%s AND token_type='employee' AND expires_at > NOW()",
                (token_hash,)
            )
            row = cursor.fetchone()
        if not row:
            return jsonify({"ok": False, "msg": "Invalid or expired token"}), 401
        from flask import g as _g
        _g.api_emp_id = row[0]
        return f(*args, **kwargs)
    return wrapper


@app.route("/api/employee/login", methods=["POST"])
@limiter.limit("5 per minute")
@limiter.limit("20 per hour")
def api_employee_login():
    data   = request.get_json() or {}
    emp_id = data.get("employee_id", "").strip()
    password = data.get("password", "").strip()
    if not emp_id:
        return jsonify({"ok": False, "msg": "employee_id required"}), 400
    # Check lockout before hitting the DB with credentials
    locked, until = _check_login_lockout(emp_id, "employee")
    if locked:
        return jsonify({"ok": False, "msg": f"Account locked until {until}. Try again later."}), 429
    db     = get_db_connection()
    cursor = db.cursor(buffered=True)
    cursor.execute("SELECT name, email, password FROM employees WHERE employee_id=%s", (emp_id,))
    row = cursor.fetchone()
    cursor.close(); db.close()
    if not row:
        _record_login_failure(emp_id, "employee")
        return jsonify({"ok": False, "msg": "Invalid credentials"}), 401
    if not password:
        return jsonify({"ok": False, "msg": "Password required"}), 400
    if not row[2] or not check_password_hash(row[2], password):
        _record_login_failure(emp_id, "employee")
        return jsonify({"ok": False, "msg": "Invalid credentials"}), 401
    _clear_login_failures(emp_id, "employee")
    # Upgrade legacy hash to bcrypt transparently
    if row[2] and not row[2].startswith("$2"):
        with _db() as (_uc, _ud):
            _uc.execute("UPDATE employees SET password=%s WHERE employee_id=%s",
                        (generate_password_hash(password), emp_id))
            _ud.commit()
    token = secrets.token_hex(32)
    with _db() as (cursor, conn):
        cursor.execute(
            "INSERT INTO api_tokens (token, token_type, identity, expires_at) "
            "VALUES (%s, 'employee', %s, NOW() + INTERVAL '24 hours')",
            (_hash_token(token), emp_id)
        )
        conn.commit()
    return jsonify({"ok": True, "token": token, "employee_id": emp_id,
                    "name": row[0], "email": row[1]})


@app.route("/api/employee/logout", methods=["POST"])
def api_employee_logout():
    auth = request.headers.get("Authorization", "")
    if auth.startswith("Bearer "):
        with _db() as (cursor, conn):
            cursor.execute("DELETE FROM api_tokens WHERE token=%s", (_hash_token(auth[7:]),))
            conn.commit()
    return jsonify({"ok": True})


@app.route("/api/employee/change-password", methods=["POST"])
@employee_api_required
def api_employee_change_password():
    data = request.get_json() or {}
    current_password = data.get("current_password", "").strip()
    new_password     = data.get("new_password", "").strip()
    if not current_password or not new_password:
        return jsonify({"ok": False, "msg": "current_password and new_password required"}), 400
    if len(new_password) < 8:
        return jsonify({"ok": False, "msg": "New password must be at least 8 characters"}), 400
    from flask import g as _g
    emp_id = _g.api_emp_id
    with _db() as (cursor, conn):
        cursor.execute("SELECT password FROM employees WHERE employee_id=%s", (emp_id,))
        row = cursor.fetchone()
        if not row:
            return jsonify({"ok": False, "msg": "Employee not found"}), 404
        if not row[0] or not check_password_hash(row[0], current_password):
            return jsonify({"ok": False, "msg": "Current password is incorrect"}), 401
        cursor.execute(
            "UPDATE employees SET password=%s WHERE employee_id=%s",
            (generate_password_hash(new_password), emp_id)
        )
        conn.commit()
    return jsonify({"ok": True, "msg": "Password changed successfully"})


def _fmt_t(t):
    if t is None: return None
    if hasattr(t, 'strftime'): return t.strftime("%H:%M:%S")
    total = int(t.total_seconds())
    return "{:02d}:{:02d}:{:02d}".format(total // 3600, (total % 3600) // 60, total % 60)


@app.route("/api/employee/portal", methods=["GET"])
@employee_api_required
def api_employee_portal():
    from flask import g as _g
    emp_id = _g.api_emp_id
    db     = get_db_connection()
    cursor = db.cursor(buffered=True)
    today  = datetime.date.today()

    cursor.execute("""
        SELECT e.name, e.email, COALESCE(c.name, '') AS company_name
        FROM employees e
        LEFT JOIN companies c ON e.company_id = c.id
        WHERE e.employee_id=%s
    """, (emp_id,))
    emp = cursor.fetchone()

    cursor.execute(
        "SELECT login_time, logout_time, status, logout_status, attendance_type "
        "FROM attendance WHERE employee_id=%s AND date=%s", (emp_id, today)
    )
    att = cursor.fetchone()

    cursor.execute("""
        SELECT date, login_time, logout_time, status, logout_status, attendance_type
        FROM attendance WHERE employee_id=%s AND date >= %s
        ORDER BY date DESC LIMIT 10
    """, (emp_id, today - datetime.timedelta(days=30)))
    recent = cursor.fetchall()

    cursor.execute(
        "SELECT leave_date, reason, status, created_at FROM leave_requests "
        "WHERE employee_id=%s ORDER BY created_at DESC LIMIT 5", (emp_id,)
    )
    leaves = cursor.fetchall()

    cursor.execute(
        "SELECT last_working_day, reason, status, created_at FROM resignation_requests "
        "WHERE employee_id=%s ORDER BY created_at DESC LIMIT 1", (emp_id,)
    )
    resign = cursor.fetchone()
    cursor.execute(
        "SELECT COUNT(*) FROM notifications WHERE recipient_type='employee' AND employee_id=%s AND is_read=FALSE",
        (emp_id,)
    )
    unread_notifications = cursor.fetchone()[0]
    cursor.execute("""
        SELECT title, content, priority, created_at FROM announcements
        WHERE COALESCE(visibility,'public') = 'public'
           OR (visibility = 'private' AND target_employee_id = %s)
        ORDER BY created_at DESC LIMIT 5
    """, (emp_id,))
    ann_rows = cursor.fetchall()
    cursor.execute("SELECT role, department FROM employees WHERE employee_id=%s", (emp_id,))
    emp_extra = cursor.fetchone()
    cursor.close(); db.close()

    return jsonify({
        "ok": True,
        "employee_id": emp_id,
        "name": emp[0] if emp else emp_id,
        "email": emp[1] if emp else None,
        "company_name": emp[2] if emp else "",
        "today": today.strftime("%d %b %Y"),
        "today_attendance": {
            "login_time": _fmt_t(att[0]),
            "logout_time": _fmt_t(att[1]),
            "login_status": att[2],
            "logout_status": att[3],
            "attendance_type": att[4],
        } if att else None,
        "recent_attendance": [
            {"date": str(r[0]), "login_time": _fmt_t(r[1]), "logout_time": _fmt_t(r[2]),
             "login_status": r[3], "logout_status": r[4], "attendance_type": r[5]}
            for r in recent
        ],
        "recent_leaves": [
            {"leave_date": str(r[0]), "reason": r[1], "status": r[2],
             "requested_at": str(r[3])}
            for r in leaves
        ],
        "resignation": {
            "last_working_day": str(resign[0]),
            "reason": resign[1],
            "status": resign[2],
            "created_at": str(resign[3]),
        } if resign else None,
        "unread_notifications": unread_notifications,
        "role": emp_extra[0] if emp_extra else None,
        "department": emp_extra[1] if emp_extra else None,
        "announcements": [
            {"title": r[0], "content": r[1], "priority": r[2], "created_at": str(r[3])}
            for r in ann_rows
        ],
    })


@app.route("/api/employee/sync_punches", methods=["POST"])
@employee_api_required
def api_employee_sync_punches():
    """Batch-submit offline punches queued on the device when there was no connectivity."""
    from flask import g as _g
    emp_id  = _g.api_emp_id
    payload = request.get_json() or {}
    punches = payload.get("punches", [])
    if not punches:
        return jsonify({"ok": True, "results": []})

    db2  = get_db_connection()
    cur2 = db2.cursor(buffered=True)
    cur2.execute("SELECT name FROM employees WHERE employee_id=%s", (emp_id,))
    if not cur2.fetchone():
        cur2.close(); db2.close()
        return jsonify({"ok": False, "msg": "Employee not found"}), 404

    results = []
    for punch in punches:
        punched_at_str = punch.get("punched_at", "")
        lat = punch.get("lat")
        lon = punch.get("lon")
        try:
            _pt = datetime.datetime.fromisoformat(punched_at_str.replace("Z", "+00:00"))
            _pt = _pt.replace(tzinfo=None)
            _now = datetime.datetime.now()
            age = (_now - _pt).total_seconds()
            if age > 86400:
                results.append({"id": punch.get("id"), "ok": False, "msg": "Too old (>24 h)"})
                continue
            if _pt > _now + datetime.timedelta(minutes=5):
                results.append({"id": punch.get("id"), "ok": False, "msg": "Future timestamp rejected"})
                continue
        except (ValueError, TypeError):
            results.append({"id": punch.get("id"), "ok": False, "msg": "Invalid timestamp"})
            continue

        punch_date = _pt.date()
        punch_time = _pt.time()
        cur2.execute(
            "SELECT login_time, logout_time, status, worked_minutes, last_relogin "
            "FROM attendance WHERE employee_id=%s AND date=%s",
            (emp_id, punch_date)
        )
        rec = cur2.fetchone()
        login_time = rec[0] if rec else None
        logout_time = rec[1] if rec else None
        login_status = rec[2] if rec else None
        worked_mins = (rec[3] or 0) if rec else 0
        last_relogin = rec[4] if rec else None

        if not login_time:
            grace_time = (datetime.datetime.combine(punch_date, SHIFT_START) + datetime.timedelta(minutes=GRACE_MINUTES)).time()
            if punch_time <= grace_time:
                status = "Full Day Login"
            elif punch_time <= SHIFT_HALF:
                status = "Late Login"
            else:
                status = "Half Day Login"
            cur2.execute(
                "INSERT INTO attendance (employee_id, date, login_time, status) VALUES (%s,%s,%s,%s)",
                (emp_id, punch_date, punch_time, status)
            )
            db2.commit()
            results.append({"id": punch.get("id"), "ok": True, "action": "login", "status": status})
        elif not logout_time:
            session_start = last_relogin if last_relogin else login_time
            if not isinstance(session_start, datetime.time):
                session_start = _td_to_time(session_start)
            cur_dt    = datetime.datetime.combine(punch_date, punch_time)
            start_dt  = datetime.datetime.combine(punch_date, session_start)
            session_m = max(0, int((cur_dt - start_dt).total_seconds() / 60))
            total_m   = worked_mins + session_m
            if punch_time < SHIFT_HALF:
                out_status = "Half Day Logout"
            elif punch_time < SHIFT_END:
                out_status = "Early Logout"
            else:
                out_status = "Completed"
            att_type = classify_by_worked_minutes(login_status, total_m, SHIFT_START, SHIFT_END)
            cur2.execute(
                "UPDATE attendance SET logout_time=%s, logout_status=%s, attendance_type=%s, worked_minutes=%s "
                "WHERE employee_id=%s AND date=%s",
                (punch_time, out_status, att_type, total_m, emp_id, punch_date)
            )
            db2.commit()
            results.append({"id": punch.get("id"), "ok": True, "action": "logout", "status": out_status})
        else:
            results.append({"id": punch.get("id"), "ok": False, "msg": "Duplicate — day already complete"})

    cur2.close(); db2.close()
    _audit("sync_punches", "attendance", emp_id, f"Synced {len([r for r in results if r['ok']])} offline punches")
    return jsonify({"ok": True, "results": results})


@app.route("/api/employee/auth-config", methods=["GET"])
def api_employee_auth_config():
    """Return all authentication method flags (public, no token required)."""
    return jsonify({"ok": True, **get_auth_config()})


_IP_RE = re.compile(r'^\d{1,3}(\.\d{1,3}){3}$')

def _wa_rp_id():
    """WebAuthn Relying Party ID.
    Loopback: return the exact host so the browser's origin matches.
    Everything else: prefer ALLOWED_ORIGINS[0] (avoids Host-header injection and
    ensures a proper domain name is used even when accessed via LAN IP).
    Falls back to the raw host only when ALLOWED_ORIGINS is unconfigured or '*'."""
    host = request.host.split(":")[0]
    # Loopback: must match the browser origin exactly; return immediately
    if host in ("127.0.0.1", "::1", "localhost"):
        return host
    # Named hosts and LAN IPs: use the pinned production domain when available
    if _allowed_origins != "*" and _allowed_origins:
        canonical = urlparse(_allowed_origins[0]).hostname
        if canonical:
            return canonical
    # NOTE: returning a non-loopback IP here will be rejected by browsers as an RP ID
    return host

def _wa_check_rp_id(rp_id):
    """Return an error string if rp_id is a non-loopback IP (browsers reject these as RP IDs),
    or None if it looks usable."""
    if rp_id in ("127.0.0.1", "::1", "localhost"):
        return None
    if _IP_RE.match(rp_id):
        return (
            f"WebAuthn does not support IP addresses as RP IDs (got '{rp_id}'). "
            "Access the server via 'localhost' on the server machine, or configure a "
            "hostname (e.g. add an entry in your hosts file and set ALLOWED_ORIGINS)."
        )
    return None

def _wa_origins():
    """Return the set of acceptable WebAuthn origins for this host.
    Always accepts both 127.0.0.1 and localhost as equivalent loopback origins.
    For LAN IPs the only valid origin is the exact IP+port the browser used."""
    host   = request.host  # includes port if non-standard
    scheme = request.scheme
    origins = {f"{scheme}://{host}"}
    bare   = host.split(":")[0]
    if bare == "127.0.0.1":
        origins.add(f"{scheme}://{host.replace('127.0.0.1', 'localhost')}")
    elif bare == "localhost":
        origins.add(f"{scheme}://{host.replace('localhost', '127.0.0.1')}")
    # LAN IPs: the single origin already added above is correct
    return list(origins)

def _wa_b64url_decode(s):
    pad = "=" * (-len(s) % 4)
    return base64.urlsafe_b64decode(s + pad)

def _wa_b64url_encode(b):
    return base64.urlsafe_b64encode(b).rstrip(b"=").decode()

# How long a real, just-completed WebAuthn assertion stays usable to authorize
# a single subsequent check-in call, scoped to the specific employee_id that
# completed it. Prevents both replay across employees and indefinite reuse.
_WA_FP_VERIFY_WINDOW_SEC = 120

def _wa_fingerprint_recently_verified(emp_id):
    """One-time, employee-bound check: did this employee just complete a real
    WebAuthn signature verification in this session? Consumes the proof."""
    emp_id = (emp_id or "").strip().upper()
    verified_emp = session.pop("wa_fp_verified_emp_id", None)
    verified_at  = session.pop("wa_fp_verified_at", 0)
    return bool(emp_id) and verified_emp == emp_id and (time.time() - verified_at) <= _WA_FP_VERIFY_WINDOW_SEC


# ---- Mobile-app biometric attestation -------------------------------------
# The mobile app has no browser, so it can't do real WebAuthn (no native
# platform-authenticator API in React Native) and has no Flask session
# cookie to piggyback the proof above on. Instead it gets a weaker but still
# meaningfully-bound proof: a server-issued, single-use nonce minted only to
# the holder of a valid employee Bearer token, consumed by a second
# Bearer-authenticated call right after the device's local biometric/PIN
# check succeeds. This is NOT a cryptographic signature — it cannot detect a
# cloned/replayed device biometric — but unlike the old flow it cannot be
# satisfied without first proving possession of that exact employee's token.
_MOBILE_BIO_NONCE_TTL_SEC    = 60
_MOBILE_BIO_VERIFY_WINDOW_SEC = 120

def _mobile_biometric_issue_nonce(emp_id):
    """Mint a fresh single-use nonce for emp_id, replacing any prior one."""
    nonce = secrets.token_hex(16)
    with _db() as (cursor, conn):
        cursor.execute(
            "INSERT INTO mobile_biometric_proofs (employee_id, nonce, nonce_expires_at, verified_at) "
            "VALUES (%s, %s, NOW() + %s * INTERVAL '1 second', NULL) "
            "ON CONFLICT (employee_id) DO UPDATE SET "
            "nonce=EXCLUDED.nonce, nonce_expires_at=EXCLUDED.nonce_expires_at, verified_at=NULL",
            (emp_id, nonce, _MOBILE_BIO_NONCE_TTL_SEC)
        )
        conn.commit()
    return nonce

def _mobile_biometric_attest(emp_id, nonce):
    """Consume a nonce after the mobile app confirms a local biometric/device
    check for this exact authenticated employee. Returns (ok, err_msg)."""
    if not nonce:
        return False, "Missing nonce"
    with _db() as (cursor, conn):
        cursor.execute(
            "SELECT nonce, nonce_expires_at FROM mobile_biometric_proofs WHERE employee_id=%s",
            (emp_id,)
        )
        row = cursor.fetchone()
        if not row or row[0] != nonce or not row[1] or row[1] < datetime.datetime.now():
            return False, "Invalid or expired nonce"
        cursor.execute(
            "UPDATE mobile_biometric_proofs SET nonce=NULL, nonce_expires_at=NULL, verified_at=NOW() "
            "WHERE employee_id=%s",
            (emp_id,)
        )
        conn.commit()
    return True, None

def _mobile_biometric_recently_verified(emp_id):
    """One-time, employee-bound check mirroring _wa_fingerprint_recently_verified,
    but DB-backed (mobile has no Flask session) and gated by a real employee
    Bearer token at both the nonce-issue and attest steps above."""
    emp_id = (emp_id or "").strip().upper()
    if not emp_id:
        return False
    with _db() as (cursor, conn):
        cursor.execute(
            "SELECT verified_at FROM mobile_biometric_proofs WHERE employee_id=%s",
            (emp_id,)
        )
        row = cursor.fetchone()
        if not row or not row[0]:
            return False
        verified_at = row[0]
        cursor.execute(
            "UPDATE mobile_biometric_proofs SET verified_at=NULL WHERE employee_id=%s",
            (emp_id,)
        )
        conn.commit()
    return (datetime.datetime.now() - verified_at).total_seconds() <= _MOBILE_BIO_VERIFY_WINDOW_SEC


def _wa_verify_and_store_registration(emp_id, credential, challenge_b64, cursor, db):
    """Verify a WebAuthn registration response (real signature/attestation
    check) and persist the credential id + public key + sign count.
    `credential` may be a dict or a JSON string. Returns (ok, err_msg)."""
    if not _webauthn_available:
        return False, "Fingerprint enrollment is not available on this server."
    if not credential or not challenge_b64:
        return False, "Missing credential or challenge — please try enrolling again"
    # Rebuild the supported-alg list from session if available; fall back to the
    # same two algorithms we offer in generate_registration_options.
    _alg_ids = session.get("wa_reg_alg_ids") or [-7, -257]
    _supported_algs = [COSEAlgorithmIdentifier(v) for v in _alg_ids]
    _rp_id   = _wa_rp_id()
    _origins = _wa_origins()
    app_log.info("WebAuthn verify: emp=%s rp_id=%s origins=%s", emp_id, _rp_id, _origins)
    try:
        if isinstance(credential, str):
            credential = json.loads(credential)
        verified = webauthn.verify_registration_response(
            credential=credential,
            expected_challenge=_wa_b64url_decode(challenge_b64),
            expected_rp_id=_rp_id,
            expected_origin=_origins,
            supported_pub_key_algs=_supported_algs,
        )
    except Exception as exc:
        app_log.warning("WebAuthn registration failed: emp=%s rp_id=%s origins=%s err=%s",
                        emp_id, _rp_id, _origins, exc, exc_info=True)
        return False, f"Enrollment failed: {exc}"
    cred_id_b64 = _wa_b64url_encode(verified.credential_id)
    pubkey_b64  = base64.b64encode(verified.credential_public_key).decode()
    cursor.execute(
        "UPDATE employees SET fingerprint_credential_id=%s, fingerprint_public_key=%s, "
        "fingerprint_sign_count=%s WHERE employee_id=%s",
        (cred_id_b64, pubkey_b64, verified.sign_count, emp_id)
    )
    db.commit()
    return True, None


def _enroll_fingerprint_from_form(emp_id, cursor, db):
    """Shared by admin_action()/add_employee_page(): read the WebAuthn
    attestation posted by the registration form (if any), verify and store
    it, flashing a warning on failure. No-op if the field is empty."""
    fp_attestation = request.form.get("fingerprint_attestation", "").strip()
    if not fp_attestation:
        return
    _ok, _err = _wa_verify_and_store_registration(
        emp_id, fp_attestation, session.get("wa_reg_challenge"), cursor, db
    )
    session.pop("wa_reg_challenge", None)
    session.pop("wa_reg_alg_ids", None)
    if not _ok:
        flash(f"⚠️ Fingerprint enrollment failed verification: {_err}", "error")


@app.route("/api/admin/employee/<emp_id>/reset-fingerprint", methods=["POST"])
@admin_required
def admin_reset_employee_fingerprint(emp_id):
    """Admin: clear a specific employee's WebAuthn credential so they can re-enroll on a new device."""
    emp_id = emp_id.strip().upper()
    try:
        db = get_db_connection(); cursor = db.cursor(buffered=True)
        cursor.execute(
            "UPDATE employees SET fingerprint_credential_id=NULL, fingerprint_public_key=NULL, "
            "fingerprint_sign_count=0 WHERE employee_id=%s",
            (emp_id,)
        )
        db.commit()
        affected = cursor.rowcount
        cursor.close(); db.close()
        if affected == 0:
            return jsonify({"ok": False, "msg": "Employee not found"}), 404
        _audit("admin_reset_fingerprint", "employees", emp_id)
        return jsonify({"ok": True})
    except Exception:
        app_log.error("Failed to reset employee fingerprint", exc_info=True)
        return jsonify({"ok": False, "msg": "Failed to reset fingerprint. Please try again."}), 500


@app.route("/api/employee/mobile-biometric-nonce", methods=["POST"])
@limiter.limit("10 per minute")
@employee_api_required
def api_mobile_biometric_nonce():
    """Mobile app calls this (with its employee Bearer token) right before
    prompting the device's local biometric/PIN check. The returned nonce
    must be echoed back to /mobile-biometric-attest within 60s."""
    from flask import g as _g
    nonce = _mobile_biometric_issue_nonce(_g.api_emp_id)
    return jsonify({"ok": True, "nonce": nonce})


@app.route("/api/employee/mobile-biometric-attest", methods=["POST"])
@limiter.limit("10 per minute")
@employee_api_required
def api_mobile_biometric_attest():
    """Mobile app calls this immediately after a successful local
    LocalAuthentication.authenticateAsync(), turning that local-only signal
    into a server-side, employee-bound, single-use, time-boxed proof that
    /api/employee/qr-face-checkin will accept for fingerprint combos."""
    from flask import g as _g
    data  = request.get_json(force=True, silent=True) or {}
    nonce = (data.get("nonce") or "").strip()
    ok, err = _mobile_biometric_attest(_g.api_emp_id, nonce)
    if not ok:
        return jsonify({"ok": False, "msg": err}), 401
    return jsonify({"ok": True})


# ---------------- API: TICKETS (employee) ----------------

@app.route("/api/employee/tickets", methods=["GET"])
@employee_api_required
def api_employee_tickets():
    from flask import g as _g
    emp_id = _g.api_emp_id
    db     = get_db_connection()
    cursor = db.cursor(buffered=True)
    cursor.execute("""
        SELECT id, category, subject, description, priority, status, admin_response, created_at
        FROM tickets WHERE employee_id=%s ORDER BY created_at DESC LIMIT 30
    """, (emp_id,))
    rows = cursor.fetchall()
    cursor.close(); db.close()
    return jsonify({"ok": True, "tickets": [
        {"id": r[0], "category": r[1], "subject": r[2], "description": r[3],
         "priority": r[4], "status": r[5], "admin_response": r[6],
         "created_at": str(r[7])}
        for r in rows
    ]})


@app.route("/api/employee/raise_ticket", methods=["POST"])
@employee_api_required
def api_employee_raise_ticket():
    from flask import g as _g
    emp_id      = _g.api_emp_id
    data        = request.get_json() or {}
    category    = data.get("category", "").strip()
    subject     = data.get("subject", "").strip()
    description = data.get("description", "").strip()
    priority    = data.get("priority", "Medium").strip()
    if not category or not subject or not description:
        return jsonify({"ok": False, "msg": "category, subject and description required"}), 400
    db     = get_db_connection()
    cursor = db.cursor(buffered=True)
    cursor.execute(
        "INSERT INTO tickets (employee_id, category, subject, description, priority) VALUES (%s,%s,%s,%s,%s)",
        (emp_id, category, subject, description, priority)
    )
    db.commit(); cursor.close(); db.close()
    return jsonify({"ok": True, "msg": "Ticket raised successfully."})


# ---------------- API: EMPLOYEE — ATTENDANCE HISTORY ----------------


# ---------------- API: EMPLOYEE — LEAVE HISTORY + BALANCE ----------------


# ---------------- API: EMPLOYEE — CANCEL LEAVE ----------------


# ---------------- WEB: EMPLOYEE — CANCEL LEAVE ----------------


# ---------------- API: EMPLOYEE — REQUEST OVERTIME ----------------


# ---------------- API: ADMIN — DOCUMENT EXPIRY ALERTS ----------------

@app.route("/api/admin/expiring_documents", methods=["GET"])
@admin_required
def api_expiring_documents():
    days = int(request.args.get("days", 30))
    db   = get_db_connection()
    cursor = db.cursor(buffered=True)
    cursor.execute("""
        SELECT d.id, d.employee_id, e.name, d.doc_type, d.original_name, d.expiry_date,
               (d.expiry_date - CURRENT_DATE) AS days_left
        FROM employee_documents d
        JOIN employees e ON e.employee_id = d.employee_id
        WHERE d.expiry_date IS NOT NULL
          AND d.expiry_date >= CURRENT_DATE
          AND d.expiry_date <= CURRENT_DATE + (%s * INTERVAL '1 day')
        ORDER BY d.expiry_date ASC
    """, (days,))
    rows = cursor.fetchall()
    cursor.close(); db.close()
    return jsonify({
        "ok": True,
        "documents": [
            {"id": r[0], "employee_id": r[1], "employee_name": r[2],
             "doc_type": r[3], "filename": r[4],
             "expiry_date": str(r[5]), "days_left": r[6]}
            for r in rows
        ]
    })


# ---------------- API: EMPLOYEE — HOLIDAYS ----------------


# ---------------- API: EMPLOYEE — PROFILE ----------------

@app.route("/api/employee/profile", methods=["GET"])
@employee_api_required
def api_employee_profile():
    from flask import g as _g
    emp_id = _g.api_emp_id
    db     = get_db_connection()
    cursor = db.cursor(buffered=True)
    cursor.execute("""
        SELECT e.employee_id, e.name, e.email, e.role, e.department,
               e.phone, e.dob, e.gender, e.blood_group, e.address, e.city, e.state,
               e.pincode, e.about_me, e.emergency_contact_name, e.emergency_contact_phone,
               e.bank_name, e.bank_account, e.bank_ifsc, e.pan_number, e.aadhar_number,
               COALESCE(s.salary_per_day, 0), COALESCE(e.joining_date, e.date_of_joining),
               COALESCE(c.name, '')
        FROM employees e
        LEFT JOIN salary_config s ON e.employee_id = s.employee_id
        LEFT JOIN companies c ON e.company_id = c.id
        WHERE e.employee_id = %s
    """, (emp_id,))
    row = cursor.fetchone()
    cursor.close(); db.close()
    if not row:
        return jsonify({"ok": False, "msg": "Employee not found"}), 404
    return jsonify({
        "ok": True,
        "profile": {
            "employee_id": row[0], "name": row[1], "email": row[2],
            "role": row[3], "department": row[4],
            "phone": row[5],
            "dob": str(row[6]) if row[6] else None,
            "gender": row[7], "blood_group": row[8],
            "address": row[9], "city": row[10], "state": row[11], "pincode": row[12],
            "about_me": row[13],
            "emergency_contact_name": row[14], "emergency_contact_phone": row[15],
            "bank_name": row[16], "bank_account": decrypt_pii(row[17]), "bank_ifsc": decrypt_pii(row[18]),
            "pan_number": decrypt_pii(row[19]), "aadhar_number": decrypt_pii(row[20]),
            "salary_per_day": float(row[21]),
            "join_date": str(row[22]) if row[22] else None,
            "company_name": row[23],
            "photo_url": f"/dataset/{row[0]}.jpg",
        },
    })


@app.route("/api/employee/photo", methods=["POST"])
@employee_api_required
def api_employee_upload_photo():
    from flask import g as _g
    from PIL import Image
    emp_id = _g.api_emp_id
    file = request.files.get("photo")
    ok, err = _validate_image_file(file)
    if not ok:
        return jsonify({"ok": False, "msg": err}), 400
    try:
        img = Image.open(file.stream).convert("RGB")
        save_path = os.path.join(UPLOAD_FOLDER, emp_id + ".jpg")
        img.save(save_path, "JPEG", quality=85)
        return jsonify({"ok": True, "msg": "Photo uploaded successfully", "photo_url": f"/dataset/{emp_id}.jpg"})
    except Exception:
        return jsonify({"ok": False, "msg": "Failed to process image"}), 500


# ---------------- API: TICKETS (admin) ----------------

@app.route("/api/tickets", methods=["GET"])
@api_required
def api_tickets():
    db     = get_db_connection()
    cursor = db.cursor(buffered=True)
    cursor.execute("""
        SELECT t.id, t.employee_id, e.name, t.category, t.subject, t.description,
               t.priority, t.status, t.admin_response, t.created_at, t.updated_at
        FROM tickets t
        JOIN employees e ON t.employee_id = e.employee_id
        ORDER BY CASE WHEN t.status='Open' THEN 0 WHEN t.status='In Progress' THEN 1 WHEN t.status='Resolved' THEN 2 WHEN t.status='Closed' THEN 3 ELSE 4 END,
                 CASE WHEN t.priority='High' THEN 0 WHEN t.priority='Medium' THEN 1 WHEN t.priority='Low' THEN 2 ELSE 3 END, t.created_at DESC
    """)
    rows = cursor.fetchall()
    cursor.close(); db.close()
    return jsonify({"ok": True, "tickets": [
        {"id": r[0], "employee_id": r[1], "name": r[2], "category": r[3],
         "subject": r[4], "description": r[5], "priority": r[6],
         "status": r[7], "admin_response": r[8],
         "created_at": str(r[9]), "updated_at": str(r[10])}
        for r in rows
    ]})


@app.route("/api/tickets/<int:tid>/action", methods=["POST"])
@api_required
def api_ticket_action(tid):
    data           = request.get_json(silent=True) or {}
    new_status     = data.get("status", "").strip()
    admin_response = data.get("admin_response", "").strip()
    allowed = ("Open", "In Progress", "Resolved", "Closed")
    if new_status not in allowed:
        return jsonify({"ok": False, "msg": f"status must be one of {allowed}"}), 400
    db     = get_db_connection()
    cursor = db.cursor(buffered=True)
    cursor.execute(
        "UPDATE tickets SET status=%s, admin_response=%s WHERE id=%s",
        (new_status, admin_response or None, tid)
    )
    db.commit(); cursor.close(); db.close()
    return jsonify({"ok": True, "status": new_status})


# ---------------- PAY SLIPS ----------------


@app.route("/admin_payslips")
@admin_required
def admin_payslips():
    db = get_db_connection()
    cursor = db.cursor(buffered=True)

    cursor.execute("SELECT employee_id, name, role, COALESCE(phone,''), COALESCE(email,'') FROM employees ORDER BY name")
    employees = cursor.fetchall()

    cursor.execute("SELECT COUNT(*) FROM leave_requests WHERE status='Pending'")
    pending_leaves = cursor.fetchone()[0]
    cursor.execute("SELECT COUNT(*) FROM resignation_requests WHERE status='Pending'")
    pending_resignations = cursor.fetchone()[0]
    cursor.execute("SELECT COUNT(*) FROM tickets WHERE status IN ('Open','In Progress')")
    pending_tickets = cursor.fetchone()[0]

    cursor.close(); db.close()

    today = datetime.date.today()
    slip_months = []
    y, m = today.year, today.month
    for _ in range(12):
        slip_months.append((y, m, calendar.month_name[m]))
        m -= 1
        if m == 0:
            m = 12; y -= 1

    return render_template("admin_payslips.html",
        employees=employees,
        slip_months=slip_months,
        pending_leaves=pending_leaves,
        pending_resignations=pending_resignations,
        pending_tickets=pending_tickets
    ,
        active_nav="salary",
    )


@app.route("/payroll_settings", methods=["GET", "POST"])
@admin_required
def payroll_settings():
    db = get_db_connection()
    cursor = db.cursor(buffered=True)

    # Ensure at least one row exists
    cursor.execute("SELECT COUNT(*) FROM payroll_config")
    if cursor.fetchone()[0] == 0:
        cursor.execute("INSERT INTO payroll_config (pf_employee_pct, pf_employer_pct, professional_tax, tds_annual_pct, pf_basic_cap) VALUES (12,12,200,0,15000)")
        db.commit()

    if request.method == "POST":
        pf_emp  = float(request.form.get("pf_employee_pct", 12))
        pf_er   = float(request.form.get("pf_employer_pct", 12))
        pt      = float(request.form.get("professional_tax", 200))
        tds     = float(request.form.get("tds_annual_pct", 0))
        pf_cap  = float(request.form.get("pf_basic_cap", 15000))
        cursor.execute("""
            UPDATE payroll_config SET pf_employee_pct=%s, pf_employer_pct=%s,
            professional_tax=%s, tds_annual_pct=%s, pf_basic_cap=%s
        """, (pf_emp, pf_er, pt, tds, pf_cap))
        db.commit()

        # Update per-employee monthly CTC / basic_pct if submitted
        emp_ids = request.form.getlist("emp_id")
        for eid in emp_ids:
            ctc  = request.form.get(f"ctc_{eid}", "")
            bpct = request.form.get(f"bpct_{eid}", "50")
            if ctc:
                spd = round(float(ctc) / 26, 2)
                cursor.execute("""
                    INSERT INTO salary_config (employee_id, salary_per_day, monthly_ctc, basic_pct)
                    VALUES (%s, %s, %s, %s)
                    ON CONFLICT (employee_id) DO UPDATE SET
                        salary_per_day=%s, monthly_ctc=%s, basic_pct=%s
                """, (eid, spd, ctc, bpct, spd, ctc, bpct))
        db.commit()
        flash("Payroll settings saved.", "success")
        cursor.close(); db.close()
        return redirect("/payroll_settings")

    cursor.execute("SELECT pf_employee_pct, pf_employer_pct, professional_tax, tds_annual_pct, pf_basic_cap FROM payroll_config LIMIT 1")
    cfg = cursor.fetchone() or (12, 12, 200, 0, 15000)

    cursor.execute("""
        SELECT e.employee_id, e.name, e.role, e.department,
               COALESCE(s.monthly_ctc, 0), COALESCE(s.salary_per_day, 0), COALESCE(s.basic_pct, 50)
        FROM employees e
        LEFT JOIN salary_config s ON e.employee_id = s.employee_id
        ORDER BY e.name
    """)
    employees = cursor.fetchall()

    cursor.execute("SELECT COUNT(*) FROM leave_requests WHERE status='Pending'")
    pending_leaves = cursor.fetchone()[0]
    cursor.execute("SELECT COUNT(*) FROM resignation_requests WHERE status='Pending'")
    pending_resignations = cursor.fetchone()[0]
    cursor.execute("SELECT COUNT(*) FROM tickets WHERE status IN ('Open','In Progress')")
    pending_tickets = cursor.fetchone()[0]
    cursor.execute("SELECT COALESCE(company_name,'') FROM company_settings LIMIT 1")
    co = cursor.fetchone()
    cursor.close(); db.close()

    return render_template("payroll_settings.html",
        cfg=cfg, employees=employees,
        pending_leaves=pending_leaves,
        pending_resignations=pending_resignations,
        pending_tickets=pending_tickets,
        co=co
    ,
        active_nav="salary",
    )


# ---------------- API: SHIFTS (JSON) ----------------


# ================================================================
#  FEATURE 1: ANALYTICS
# ================================================================

@app.route("/analytics")
@admin_required
def analytics():
    db = get_db_connection()
    cursor = db.cursor(buffered=True)

    cursor.execute("SELECT company_name FROM company_settings LIMIT 1")
    row = cursor.fetchone()
    co = type('Co', (), {'company_name': row[0] if row else 'My Company'})()

    cursor.execute("SELECT COUNT(*) FROM leave_requests WHERE status='Pending'")
    pending_leaves = cursor.fetchone()[0]
    cursor.execute("SELECT COUNT(*) FROM resignation_requests WHERE status='Pending'")
    pending_resignations = cursor.fetchone()[0]
    cursor.execute("SELECT COUNT(*) FROM tickets WHERE status IN ('Open','In Progress')")
    pending_tickets = cursor.fetchone()[0]

    today = datetime.date.today()

    cursor.execute("SELECT COUNT(*) FROM employees")
    total_employees = cursor.fetchone()[0]

    _doj_start = today.replace(day=1)
    _doj_end   = datetime.date(today.year + 1, 1, 1) if today.month == 12 else today.replace(month=today.month + 1, day=1)
    cursor.execute(
        "SELECT COUNT(*) FROM employees WHERE date_of_joining >= %s AND date_of_joining < %s",
        (_doj_start, _doj_end)
    )
    new_this_month = cursor.fetchone()[0]

    cursor.execute(
        "SELECT COUNT(DISTINCT employee_id) FROM attendance WHERE date=%s AND login_time IS NOT NULL",
        (today,)
    )
    today_present = cursor.fetchone()[0]
    today_absent = max(0, total_employees - today_present)

    cursor.execute("SELECT date FROM holidays")
    all_holidays = {r[0] for r in cursor.fetchall()}

    def _working_days_in_month(y, m):
        _, last_day = calendar.monthrange(y, m)
        days = []
        for d in range(1, last_day + 1):
            dt = datetime.date(y, m, d)
            if dt.weekday() != 6 and dt not in all_holidays:
                days.append(dt)
        return days

    monthly_series = []
    for i in range(5, -1, -1):
        ref = today.replace(day=1) - datetime.timedelta(days=1) * (i * 28)
        ref = ref.replace(day=1)
        y, m = ref.year, ref.month
        working_days = _working_days_in_month(y, m)
        if not working_days:
            continue
        past_days = [d for d in working_days if d <= today]
        total_days = len(past_days)
        if total_days == 0:
            monthly_series.append({
                'month_label': datetime.date(y, m, 1).strftime("%b %Y"),
                'total_days': 0, 'present_days': 0, 'absent_days': 0, 'att_pct': 0
            })
            continue
        month_start = datetime.date(y, m, 1)
        if m == 12:
            month_end = datetime.date(y + 1, 1, 1)
        else:
            month_end = datetime.date(y, m + 1, 1)
        cursor.execute("""
            SELECT COUNT(DISTINCT employee_id) FROM attendance
            WHERE date >= %s AND date < %s AND login_time IS NOT NULL
        """, (month_start, month_end))
        present_records = cursor.fetchone()[0]
        expected = total_days * (total_employees or 1)
        present_pct = round(present_records / expected * 100, 1) if expected else 0
        monthly_series.append({
            'month_label': month_start.strftime("%b %Y"),
            'total_days': total_days,
            'present_days': present_records,
            'absent_days': max(0, expected - present_records),
            'att_pct': present_pct
        })

    if today.month >= 1:
        y, m = today.year, today.month
        working_days = _working_days_in_month(y, m)
        past_days = [d for d in working_days if d <= today]
        total_m = len(past_days)
        if total_m > 0:
            _ms = datetime.date(y, m, 1)
            _me = datetime.date(y + 1, 1, 1) if m == 12 else datetime.date(y, m + 1, 1)
            cursor.execute("""
                SELECT COUNT(DISTINCT employee_id) FROM attendance
                WHERE date >= %s AND date < %s AND login_time IS NOT NULL
            """, (_ms, _me))
            present_m = cursor.fetchone()[0]
            expected_m = total_m * (total_employees or 1)
            avg_attendance_pct = round(present_m / expected_m * 100, 1) if expected_m else 0
        else:
            avg_attendance_pct = 0
    else:
        avg_attendance_pct = 0

    cursor.execute("""
        SELECT department, COUNT(*) as cnt FROM employees
        WHERE department IS NOT NULL AND department != ''
        GROUP BY department ORDER BY cnt DESC
    """)
    dept_data = [{'department': r[0], 'count': r[1]} for r in cursor.fetchall()]

    cursor.execute("""
        SELECT lt.name, COUNT(*) as cnt
        FROM leave_requests lr
        JOIN leave_types lt ON lr.leave_type_id = lt.id
        WHERE lr.status='Approved' AND EXTRACT(YEAR FROM lr.leave_date)=%s
        GROUP BY lt.name ORDER BY cnt DESC
    """, (today.year,))
    leave_by_type = [{'name': r[0], 'count': r[1]} for r in cursor.fetchall()]

    cursor.execute("""
        SELECT e.employee_id, e.name,
               ROUND(COUNT(CASE WHEN a.login_time IS NOT NULL THEN 1 END)::NUMERIC /
                     GREATEST((LEAST((date_trunc('month', %s::date) + INTERVAL '1 month - 1 day')::date, %s::date) - %s::date) + 1, 1) * 100, 1) AS pct
        FROM employees e
        LEFT JOIN attendance a ON e.employee_id=a.employee_id AND EXTRACT(MONTH FROM a.date)=%s AND EXTRACT(YEAR FROM a.date)=%s
        GROUP BY e.employee_id, e.name
        ORDER BY pct DESC LIMIT 5
    """, (datetime.date(today.year, today.month, 1), today, datetime.date(today.year, today.month, 1), today.month, today.year))
    top_present = [{'name': r[1], 'employee_id': r[0], 'pct': float(r[2] or 0)} for r in cursor.fetchall()]

    cursor.execute("""
        SELECT gender, COUNT(*) as cnt FROM employees
        WHERE gender IS NOT NULL AND gender != ''
        GROUP BY gender
    """)
    gender_data = [{'gender': r[0], 'count': r[1]} for r in cursor.fetchall()]

    # Attendance heatmap — last 35 days (5 weeks) present count per day
    heatmap_start = today - datetime.timedelta(days=34)
    cursor.execute("""
        SELECT date, COUNT(DISTINCT employee_id) as cnt
        FROM attendance
        WHERE date BETWEEN %s AND %s AND login_time IS NOT NULL
        GROUP BY date
    """, (heatmap_start, today))
    heatmap_raw = {r[0]: r[1] for r in cursor.fetchall()}
    heatmap_data = []
    for i in range(35):
        d = heatmap_start + datetime.timedelta(days=i)
        heatmap_data.append({'date': d.strftime('%Y-%m-%d'), 'day': d.strftime('%a'), 'count': heatmap_raw.get(d, 0)})

    # Department-wise attendance rate this month
    cursor.execute("""
        SELECT e.department,
               COUNT(DISTINCT e.employee_id) as total_emp,
               COUNT(DISTINCT CASE WHEN a.login_time IS NOT NULL THEN a.employee_id END) as present_emp
        FROM employees e
        LEFT JOIN attendance a ON e.employee_id=a.employee_id AND EXTRACT(MONTH FROM a.date)=%s AND EXTRACT(YEAR FROM a.date)=%s
        WHERE e.department IS NOT NULL AND e.department != ''
        GROUP BY e.department
        ORDER BY present_emp DESC
    """, (today.month, today.year))
    dept_attendance = []
    for r in cursor.fetchall():
        dept, total, present = r[0], r[1], r[2]
        pct = round(present / total * 100, 1) if total else 0
        dept_attendance.append({'dept': dept, 'total': total, 'present': present, 'pct': pct})

    # Late arrival trend — last 14 days
    late_start = today - datetime.timedelta(days=13)
    cursor.execute("""
        SELECT date, COUNT(DISTINCT employee_id) as late_cnt
        FROM attendance
        WHERE date BETWEEN %s AND %s AND status='Late Login'
        GROUP BY date ORDER BY date ASC
    """, (late_start, today))
    late_raw = {r[0]: r[1] for r in cursor.fetchall()}
    late_trend = []
    for i in range(14):
        d = late_start + datetime.timedelta(days=i)
        late_trend.append({'date': d.strftime('%d %b'), 'count': late_raw.get(d, 0)})

    # Employee retention — tenure bands
    cursor.execute("SELECT date_of_joining FROM employees WHERE date_of_joining IS NOT NULL")
    retention = {'0-6m': 0, '6-12m': 0, '1-3y': 0, '3y+': 0}
    for (doj,) in cursor.fetchall():
        if isinstance(doj, str):
            try: doj = datetime.date.fromisoformat(doj)
            except Exception as _e:
                app_log.debug("Skipping bad date_of_joining value %r: %s", doj, _e)
                continue
        months = (today.year - doj.year) * 12 + (today.month - doj.month)
        if months < 6:       retention['0-6m'] += 1
        elif months < 12:    retention['6-12m'] += 1
        elif months < 36:    retention['1-3y'] += 1
        else:                retention['3y+'] += 1

    # Smart Alerts Panel
    smart_alerts = []

    # 1. Employees absent 3+ consecutive working days
    working_days_back = []
    for i in range(1, 15):
        d = today - datetime.timedelta(days=i)
        if d.weekday() != 6 and d not in all_holidays:
            working_days_back.append(d)
        if len(working_days_back) == 5:
            break
    last3 = working_days_back[:3]
    if len(last3) == 3:
        cursor.execute("""
            SELECT e.name, e.employee_id
            FROM employees e
            WHERE NOT EXISTS (
                SELECT 1 FROM attendance a
                WHERE a.employee_id = e.employee_id
                AND a.date IN (%s,%s,%s)
                AND a.login_time IS NOT NULL
            )
        """, (last3[0], last3[1], last3[2]))
        absent3 = cursor.fetchall()
        if absent3:
            names = ', '.join(r[1] for r in absent3[:3])
            extra = f' +{len(absent3)-3} more' if len(absent3) > 3 else ''
            smart_alerts.append({
                'level': 'danger',
                'icon': 'ti-user-off',
                'title': f'{len(absent3)} employee{"s" if len(absent3)>1 else ""} absent for 3+ consecutive days',
                'detail': names + extra,
                'link': '/monthly_report'
            })

    # 2. Leave requests spike this week vs last week
    week_start = today - datetime.timedelta(days=today.weekday())
    last_week_start = week_start - datetime.timedelta(days=7)
    cursor.execute("SELECT COUNT(*) FROM leave_requests WHERE leave_date >= %s", (week_start,))
    leaves_this_week = cursor.fetchone()[0]
    cursor.execute("SELECT COUNT(*) FROM leave_requests WHERE leave_date >= %s AND leave_date < %s", (last_week_start, week_start))
    leaves_last_week = cursor.fetchone()[0]
    if leaves_last_week > 0 and leaves_this_week > leaves_last_week * 1.4:
        pct_jump = round((leaves_this_week - leaves_last_week) / leaves_last_week * 100)
        smart_alerts.append({
            'level': 'warning',
            'icon': 'ti-calendar-up',
            'title': f'Leave requests spiked {pct_jump}% compared to last week',
            'detail': f'{leaves_this_week} requests this week vs {leaves_last_week} last week',
            'link': '/leave_requests'
        })

    # 3. Employees with attendance below 50% this month
    cursor.execute("""
        SELECT e.name, e.employee_id,
               COUNT(CASE WHEN a.login_time IS NOT NULL THEN 1 END) as present_days,
               COUNT(a.date) as total_days
        FROM employees e
        LEFT JOIN attendance a ON e.employee_id=a.employee_id
            AND EXTRACT(MONTH FROM a.date)=%s AND EXTRACT(YEAR FROM a.date)=%s
        GROUP BY e.employee_id, e.name
        HAVING COUNT(a.date) > 0
           AND (COUNT(CASE WHEN a.login_time IS NOT NULL THEN 1 END)::NUMERIC / COUNT(a.date)) < 0.5
    """, (today.month, today.year))
    low_att = cursor.fetchall()
    if low_att:
        names = ', '.join(r[1] for r in low_att[:3])
        extra = f' +{len(low_att)-3} more' if len(low_att) > 3 else ''
        smart_alerts.append({
            'level': 'warning',
            'icon': 'ti-chart-bar-off',
            'title': f'{len(low_att)} employee{"s" if len(low_att)>1 else ""} below 50% attendance this month',
            'detail': names + extra,
            'link': '/monthly_report'
        })

    # 4. High pending leave approvals
    if pending_leaves >= 5:
        smart_alerts.append({
            'level': 'warning',
            'icon': 'ti-clock-pause',
            'title': f'{pending_leaves} leave requests pending approval',
            'detail': 'Employees may be waiting — review and approve',
            'link': '/leave_requests'
        })

    # 5. New joiners who have never logged in
    cursor.execute("""
        SELECT e.name, e.employee_id FROM employees e
        WHERE e.date_of_joining >= %s
        AND NOT EXISTS (SELECT 1 FROM attendance a WHERE a.employee_id=e.employee_id AND a.login_time IS NOT NULL)
    """, (today - datetime.timedelta(days=30),))
    never_logged = cursor.fetchall()
    if never_logged:
        names = ', '.join(r[1] for r in never_logged[:3])
        extra = f' +{len(never_logged)-3} more' if len(never_logged) > 3 else ''
        smart_alerts.append({
            'level': 'info',
            'icon': 'ti-user-question',
            'title': f'{len(never_logged)} new joiner{"s" if len(never_logged)>1 else ""} {"have" if len(never_logged)>1 else "has"} never logged attendance',
            'detail': names + extra,
            'link': '/employees'
        })

    # 6. Pending overtime approvals
    cursor.execute("SELECT COUNT(*) FROM overtime_records WHERE status='Pending'")
    ot_pending_count = cursor.fetchone()[0]
    if ot_pending_count >= 3:
        smart_alerts.append({
            'level': 'info',
            'icon': 'ti-clock-bolt',
            'title': f'{ot_pending_count} overtime requests waiting for approval',
            'detail': 'Review pending OT requests from the dashboard',
            'link': '/overtime'
        })

    # 6. Documents expiring in next 30 days
    cursor.execute("""
        SELECT COUNT(*) FROM employee_documents
        WHERE expiry_date IS NOT NULL
          AND expiry_date >= CURRENT_DATE
          AND expiry_date <= CURRENT_DATE + INTERVAL '30 days'
    """)
    expiring_docs = cursor.fetchone()[0]
    if expiring_docs > 0:
        smart_alerts.append({
            'level': 'warning',
            'icon': 'ti-file-alert',
            'title': f'{expiring_docs} employee document{"s" if expiring_docs > 1 else ""} expiring within 30 days',
            'detail': 'Review and renew documents before they expire',
            'link': '/documents'
        })

    if not smart_alerts:
        smart_alerts.append({
            'level': 'success',
            'icon': 'ti-circle-check',
            'title': 'All systems healthy — no anomalies detected',
            'detail': 'Attendance, leaves and approvals are all on track',
            'link': ''
        })

    cursor.close(); db.close()

    return render_template("analytics.html",
        co=co,
        pending_leaves=pending_leaves,
        pending_resignations=pending_resignations,
        pending_tickets=pending_tickets,
        total_employees=total_employees,
        new_this_month=new_this_month,
        today_present=today_present,
        today_absent=today_absent,
        avg_attendance_pct=avg_attendance_pct,
        monthly_series=monthly_series,
        dept_data=dept_data,
        leave_by_type=leave_by_type,
        top_present=top_present,
        gender_data=gender_data,
        heatmap_data=heatmap_data,
        dept_attendance=dept_attendance,
        late_trend=late_trend,
        retention=retention,
        smart_alerts=smart_alerts,
    
        active_nav="analytics",
    )


# ================================================================
#  FEATURE 2: DOCUMENT MANAGEMENT
# ================================================================

_DOC_ALLOWED_EXT = {'pdf', 'jpg', 'jpeg', 'png', 'doc', 'docx', 'xls', 'xlsx'}

def _doc_admin_ctx(cursor):
    cursor.execute("SELECT company_name FROM company_settings LIMIT 1")
    row = cursor.fetchone()
    co = type('Co', (), {'company_name': row[0] if row else 'My Company'})()
    cursor.execute("SELECT COUNT(*) FROM leave_requests WHERE status='Pending'")
    pending_leaves = cursor.fetchone()[0]
    cursor.execute("SELECT COUNT(*) FROM resignation_requests WHERE status='Pending'")
    pending_resignations = cursor.fetchone()[0]
    cursor.execute("SELECT COUNT(*) FROM tickets WHERE status IN ('Open','In Progress')")
    pending_tickets = cursor.fetchone()[0]
    return co, pending_leaves, pending_resignations, pending_tickets


@app.route("/documents")
@admin_required
def documents():
    db = get_db_connection()
    cursor = db.cursor(buffered=True)
    co, pending_leaves, pending_resignations, pending_tickets = _doc_admin_ctx(cursor)

    cursor.execute("SELECT employee_id, name FROM employees ORDER BY name")
    employees = cursor.fetchall()

    sel_emp = request.args.get('emp_id', '')
    sel_emp_name = ''

    if sel_emp:
        cursor.execute("SELECT name FROM employees WHERE employee_id=%s", (sel_emp,))
        r = cursor.fetchone()
        sel_emp_name = r[0] if r else sel_emp
        cursor.execute("""
            SELECT d.id, d.employee_id, e.name, d.doc_type, d.original_name, d.stored_name,
                   d.uploaded_by, d.uploaded_at, d.expiry_date
            FROM employee_documents d JOIN employees e ON e.employee_id=d.employee_id
            WHERE d.employee_id=%s ORDER BY d.uploaded_at DESC
        """, (sel_emp,))
    else:
        cursor.execute("""
            SELECT d.id, d.employee_id, e.name, d.doc_type, d.original_name, d.stored_name,
                   d.uploaded_by, d.uploaded_at, d.expiry_date
            FROM employee_documents d JOIN employees e ON e.employee_id=d.employee_id
            ORDER BY d.uploaded_at DESC
        """)
    docs = cursor.fetchall()
    cursor.close(); db.close()

    return render_template("documents.html",
        co=co,
        pending_leaves=pending_leaves,
        pending_resignations=pending_resignations,
        pending_tickets=pending_tickets,
        employees=employees, docs=docs,
        sel_emp=sel_emp, sel_emp_name=sel_emp_name,
        today=datetime.date.today(),
    )


@app.route("/upload_document", methods=["POST"])
@admin_required
def upload_document():
    emp_id   = request.form.get('employee_id', '').strip()
    doc_type = request.form.get('doc_type', '').strip()
    f        = request.files.get('document')
    if not emp_id or not doc_type or not f or not f.filename:
        flash("All fields required.", "danger")
        return redirect('/documents')
    ok, err = _validate_upload(f, _DOC_ALLOWED_EXT)
    if not ok:
        flash(err, "danger")
        return redirect(f'/documents?emp_id={emp_id}')
    folder = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static', 'employee_docs', emp_id)
    os.makedirs(folder, exist_ok=True)
    orig_name   = f.filename
    stored_name = str(uuid.uuid4()) + '_' + secure_filename(orig_name)
    f.save(os.path.join(folder, stored_name))
    db = get_db_connection()
    cursor = db.cursor(buffered=True)
    expiry_raw  = request.form.get("expiry_date", "").strip()
    expiry_date = expiry_raw if expiry_raw else None
    cursor.execute(
        "INSERT INTO employee_documents (employee_id, doc_type, original_name, stored_name, uploaded_by, expiry_date) "
        "VALUES (%s,%s,%s,%s,'admin',%s)",
        (emp_id, doc_type, orig_name, stored_name, expiry_date)
    )
    db.commit(); cursor.close(); db.close()
    _audit("upload_document", "employee_documents", emp_id,
           f"doc_type={doc_type} file={orig_name} expiry={expiry_date or 'none'}")
    flash("Document uploaded successfully.", "success")
    raw_redirect = request.form.get('redirect_to') or f'/documents?emp_id={emp_id}'
    # Reject any redirect that leaves this origin (open-redirect prevention).
    # Only allow relative URLs (no scheme, no netloc).
    from urllib.parse import urlparse as _urlparse
    _p = _urlparse(raw_redirect)
    safe_redirect = raw_redirect if (not _p.scheme and not _p.netloc) else f'/documents?emp_id={emp_id}'
    return redirect(safe_redirect)


@app.route("/delete_document/<int:did>", methods=["POST"])
@admin_required
def delete_document(did):
    db = get_db_connection()
    cursor = db.cursor(buffered=True)
    cursor.execute("SELECT employee_id, stored_name FROM employee_documents WHERE id=%s", (did,))
    row = cursor.fetchone()
    if row:
        emp_id, stored_name = row
        fpath = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static', 'employee_docs', emp_id, stored_name)
        try:
            os.remove(fpath)
        except Exception:
            pass
        cursor.execute("DELETE FROM employee_documents WHERE id=%s", (did,))
        db.commit()
    cursor.close(); db.close()
    flash("Document deleted.", "success")
    return redirect(_safe_referrer_redirect(request.referrer or "", "/documents"))


@app.route("/download_document/<int:did>")
def download_document(did):
    is_admin = session.get("admin_logged_in")
    emp_session = session.get("employee_id")
    if not is_admin and not emp_session:
        return redirect("/employee_login")
    db = get_db_connection()
    cursor = db.cursor(buffered=True)
    cursor.execute("SELECT employee_id, original_name, stored_name FROM employee_documents WHERE id=%s", (did,))
    row = cursor.fetchone()
    cursor.close(); db.close()
    if not row:
        flash("Document not found.", "danger")
        return redirect('/documents')
    doc_emp_id, original_name, stored_name = row
    if not is_admin and emp_session != doc_emp_id:
        flash("Access denied.", "danger")
        return redirect('/employee_portal')
    folder = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static', 'employee_docs', doc_emp_id)
    return send_from_directory(folder, stored_name, as_attachment=True, download_name=original_name)


@app.route("/upload_my_document", methods=["POST"])
def upload_my_document():
    emp_id = session.get("employee_id")
    if not emp_id:
        return redirect("/employee_login")
    doc_type = request.form.get('doc_type', '').strip()
    f        = request.files.get('document')
    if not doc_type or not f or not f.filename:
        flash("All fields required.", "danger")
        return redirect('/employee_portal')
    ok, err = _validate_upload(f, _DOC_ALLOWED_EXT)
    if not ok:
        flash(err, "danger")
        return redirect('/employee_portal')
    folder = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static', 'employee_docs', emp_id)
    os.makedirs(folder, exist_ok=True)
    orig_name   = f.filename
    stored_name = str(uuid.uuid4()) + '_' + secure_filename(orig_name)
    f.save(os.path.join(folder, stored_name))
    db = get_db_connection()
    cursor = db.cursor(buffered=True)
    cursor.execute(
        "INSERT INTO employee_documents (employee_id, doc_type, original_name, stored_name, uploaded_by) VALUES (%s,%s,%s,%s,'employee')",
        (emp_id, doc_type, orig_name, stored_name)
    )
    db.commit(); cursor.close(); db.close()
    flash("Document uploaded successfully.", "success")
    return redirect('/employee_portal#documents')


@app.route("/delete_my_document/<int:did>", methods=["POST"])
def delete_my_document(did):
    emp_id = session.get("employee_id")
    if not emp_id:
        return redirect("/employee_login")
    db = get_db_connection()
    cursor = db.cursor(buffered=True)
    cursor.execute("SELECT employee_id, stored_name FROM employee_documents WHERE id=%s AND employee_id=%s", (did, emp_id))
    row = cursor.fetchone()
    if row:
        fpath = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static', 'employee_docs', emp_id, row[1])
        try:
            os.remove(fpath)
        except Exception:
            pass
        cursor.execute("DELETE FROM employee_documents WHERE id=%s AND employee_id=%s", (did, emp_id))
        db.commit()
    cursor.close(); db.close()
    flash("Document deleted.", "success")
    return redirect('/employee_portal#documents')


# ================================================================
#  FEATURE 3: OVERTIME TRACKING
# ================================================================


# ─────────────────────────── COMP-OFF MANAGEMENT ───────────────────────────


# Notification routes migrated to blueprints/notifications.py


# ── Tenant Provisioning ──────────────────────────────────────────────────────

_SUBDOMAIN_RE  = re.compile(r'^[a-z0-9\-]+$')
# Set SIGNUP_SECRET in .env to restrict who can create new organisations.
# Anyone who knows this token can provision a new tenant; keep it private.
_SIGNUP_SECRET = os.environ.get("SIGNUP_SECRET", "").strip()

@app.route("/create_org", methods=["GET"])
def create_org_page():
    if not _SIGNUP_SECRET:
        # Provisioning disabled: no SIGNUP_SECRET configured in .env
        return render_template("create_org.html", signup_disabled=True)
    return render_template("create_org.html", signup_disabled=False)


@app.route("/create_org", methods=["POST"])
def create_org():
    # Require a server-side secret token to prevent anonymous tenant creation.
    if not _SIGNUP_SECRET:
        flash("Organisation self-registration is disabled on this server.", "error")
        return redirect("/create_org")
    submitted_secret = request.form.get("signup_secret", "").strip()
    if not secrets.compare_digest(_SIGNUP_SECRET, submitted_secret):
        flash("Invalid signup code. Contact your administrator.", "error")
        return redirect("/create_org")

    company_name    = request.form.get("company_name", "").strip()
    subdomain       = request.form.get("subdomain", "").strip().lower()
    admin_username  = request.form.get("admin_username", "").strip()
    admin_password  = request.form.get("admin_password", "").strip()
    admin_email     = request.form.get("admin_email", "").strip() or None

    # Validate
    if not all([company_name, subdomain, admin_username, admin_password]):
        flash("All fields (company name, subdomain, admin username and password) are required.", "error")
        return redirect("/create_org")
    if not _SUBDOMAIN_RE.match(subdomain):
        flash("Subdomain may only contain lowercase letters, digits, and hyphens.", "error")
        return redirect("/create_org")
    if len(admin_password) < 8:
        flash("Admin password must be at least 8 characters.", "error")
        return redirect("/create_org")

    # Check subdomain not taken
    try:
        from database import get_master_db
        mconn = get_master_db()
        mcur  = mconn.cursor(buffered=True)
        mcur.execute("SELECT id FROM tenants WHERE subdomain=%s", (subdomain,))
        if mcur.fetchone():
            mcur.close(); mconn.close()
            flash(f"Subdomain '{subdomain}' is already taken. Choose another.", "error")
            return redirect("/create_org")
        mcur.close(); mconn.close()
    except Exception as exc:
        app_log.error("create_org subdomain check failed: %s", exc)
        flash("Could not check subdomain availability. Please try again.", "error")
        return redirect("/create_org")

    # Derive DB name
    db_name = "att_" + subdomain.replace("-", "_")

    try:
        from database import create_tenant_schema
        create_tenant_schema(db_name)
    except Exception as exc:
        app_log.error("create_org DB creation failed: %s", exc)
        flash("Failed to create organisation. Please contact support.", "error")
        return redirect("/create_org")

    try:
        from flask import g as _g
        _g.tenant_db = db_name
        init_tenant_db(db_name)
    except Exception as exc:
        app_log.error("create_org schema init failed: %s", exc)
        flash("Failed to initialise organisation schema. Please contact support.", "error")
        return redirect("/create_org")

    # Insert company settings and admin user into the new tenant DB
    try:
        from database import get_tenant_db
        tconn = get_tenant_db(db_name)
        tcur  = tconn.cursor()
        tcur.execute(
            "UPDATE company_settings SET company_name=%s, setup_done=1 WHERE id=1",
            (company_name,)
        )
        tcur.execute(
            "INSERT INTO admin_users (username, password, email) VALUES (%s, %s, %s)"
            " ON CONFLICT (username) DO UPDATE SET password=EXCLUDED.password",
            (admin_username, generate_password_hash(admin_password), admin_email)
        )
        tconn.commit()
        tcur.close(); tconn.close()
    except Exception as exc:
        flash(f"Failed to seed tenant data: {exc}", "error")
        return redirect("/create_org")

    # Register tenant in master DB
    try:
        from database import get_master_db
        mconn = get_master_db()
        mcur  = mconn.cursor()
        mcur.execute(
            "INSERT INTO tenants (company_name, subdomain, db_name, admin_email, status) "
            "VALUES (%s, %s, %s, %s, 'active')",
            (company_name, subdomain, db_name, admin_email)
        )
        mconn.commit()
        mcur.close(); mconn.close()
    except Exception as exc:
        flash(f"Tenant registered in DB but master registry failed: {exc}", "error")
        return redirect("/create_org")

    flash(f"Organisation '{company_name}' created! Subdomain: {subdomain}. You can now log in.", "success")
    return redirect("/admin_login")


# ─────────────────────────────────────────
#  ONBOARDING WORKFLOW
# ─────────────────────────────────────────

@app.route("/onboarding")
@admin_required
def onboarding():
    db = get_db_connection()
    cursor = db.cursor()
    active_tab = request.args.get("tab", "active")

    # Active onboardings with progress
    cursor.execute("""
        SELECT eo.id, e.employee_id, e.name, e.role, e.department,
               ot.name AS template_name, eo.assigned_date, eo.due_date, eo.status,
               COUNT(eot.id) AS total_tasks,
               SUM(CASE WHEN eot.status='Done' THEN 1 ELSE 0 END) AS done_tasks
        FROM employee_onboarding eo
        JOIN employees e ON e.employee_id = eo.employee_id
        JOIN onboarding_templates ot ON ot.id = eo.template_id
        LEFT JOIN employee_onboarding_tasks eot ON eot.onboarding_id = eo.id
        GROUP BY eo.id, e.employee_id, e.name, e.role, e.department,
                 ot.name, eo.assigned_date, eo.due_date, eo.status
        ORDER BY eo.assigned_date DESC
    """)
    active_onboardings = cursor.fetchall()

    # Templates with task count
    cursor.execute("""
        SELECT ot.id, ot.name, ot.description, ot.is_active,
               COUNT(tt.id) AS task_count, COALESCE(ot.role,'')
        FROM onboarding_templates ot
        LEFT JOIN onboarding_template_tasks tt ON tt.template_id = ot.id
        GROUP BY ot.id
        ORDER BY ot.created_at DESC
    """)
    templates = cursor.fetchall()

    # Employees list for assign dropdown
    cursor.execute("SELECT employee_id, name, role FROM employees WHERE is_active=1 ORDER BY name")
    emp_list = cursor.fetchall()

    # Active templates for assign dropdown (include role for JS filtering)
    cursor.execute("SELECT id, name, COALESCE(role,'') FROM onboarding_templates WHERE is_active=1 ORDER BY name")
    active_templates = cursor.fetchall()

    # Distinct employee roles for role filter dropdown
    cursor.execute("SELECT DISTINCT role FROM employees WHERE role IS NOT NULL AND role != '' ORDER BY role")
    employee_roles = [r[0] for r in cursor.fetchall()]

    today = datetime.date.today()
    total_active    = sum(1 for o in active_onboardings if o[8] != 'Completed')
    total_completed = sum(1 for o in active_onboardings if o[8] == 'Completed')
    total_overdue   = sum(1 for o in active_onboardings if o[7] and o[7] < today and o[8] != 'Completed')

    cursor.execute("SELECT COALESCE(default_onboarding_template_id, 0) FROM company_settings LIMIT 1")
    _dtpl = cursor.fetchone()
    default_onboarding_tpl = int(_dtpl[0]) if _dtpl and _dtpl[0] else 0

    co = get_company_settings()
    cursor.close(); db.close()
    return render_template("onboarding.html",
        active_onboardings=active_onboardings,
        templates=templates,
        emp_list=emp_list,
        active_templates=active_templates,
        employee_roles=employee_roles,
        active_tab=active_tab,
        co=co,
        today=today,
        total_active=total_active,
        total_completed=total_completed,
        total_overdue=total_overdue,
        default_onboarding_tpl=default_onboarding_tpl,
        pending_leaves=0, pending_resignations=0, pending_tickets=0
    ,
        active_nav="onboarding",
    )

@app.route("/onboarding_template_save", methods=["POST"])
@admin_required
def onboarding_template_save():
    db = get_db_connection(); cursor = db.cursor()
    tid    = request.form.get("template_id")
    name   = request.form.get("name", "").strip()
    desc   = request.form.get("description", "").strip()
    role   = request.form.get("role", "").strip() or None
    if not name:
        flash("Template name is required.", "error")
        return redirect("/onboarding?tab=templates")
    if tid:
        cursor.execute("UPDATE onboarding_templates SET name=%s, description=%s, role=%s WHERE id=%s", (name, desc, role, tid))
        flash("Template updated.", "success")
    else:
        cursor.execute("INSERT INTO onboarding_templates (name, description, role) VALUES (%s,%s,%s)", (name, desc, role))
        flash("Template created.", "success")
    db.commit(); cursor.close(); db.close()
    return redirect("/onboarding?tab=templates")

@app.route("/bulk_assign_onboarding", methods=["POST"])
@admin_required
def bulk_assign_onboarding():
    db = get_db_connection(); cursor = db.cursor()
    tid      = request.form.get("template_id")
    emp_ids  = request.form.getlist("employee_ids")
    today    = datetime.date.today()
    due_date = (today + datetime.timedelta(days=30)).isoformat()
    assigned = 0
    for emp_id in emp_ids:
        cursor.execute("SELECT id FROM employee_onboarding WHERE employee_id=%s AND template_id=%s AND status='In Progress'", (emp_id, tid))
        if cursor.fetchone():
            continue
        cursor.execute("INSERT INTO employee_onboarding (employee_id, template_id, assigned_date, due_date, status) VALUES (%s,%s,%s,%s,'In Progress') RETURNING id",
                       (emp_id, tid, today, due_date))
        ob_id = cursor.fetchone()[0]
        cursor.execute("SELECT id, task_title, task_description, requires_document, due_days FROM onboarding_template_tasks WHERE template_id=%s ORDER BY sort_order, id", (tid,))
        for tt in cursor.fetchall():
            cursor.execute("INSERT INTO employee_onboarding_tasks (onboarding_id, template_task_id, employee_id, task_title, task_description, requires_document, due_days, status) VALUES (%s,%s,%s,%s,%s,%s,%s,'Pending')",
                           (ob_id, tt[0], emp_id, tt[1], tt[2], tt[3], tt[4]))
        assigned += 1
        # Email notification
        try:
            cursor.execute("SELECT name, email FROM employees WHERE employee_id=%s", (emp_id,))
            _er = cursor.fetchone()
            cursor.execute("SELECT name FROM onboarding_templates WHERE id=%s", (tid,))
            _tr = cursor.fetchone()
            if _er and _er[1] and _tr:
                _ecfg = get_email_config()
                if _ecfg:
                    _html = (f"<p>Hi <strong>{_er[0]}</strong>,</p>"
                             f"<p>A new onboarding checklist <strong>'{_tr[0]}'</strong> has been assigned to you. Please complete all tasks by <strong>{due_date}</strong>.</p>")
                    send_email_async(_er[1], f"New Onboarding Checklist — {_tr[0]}", _html, _ecfg)
        except Exception:
            pass
    db.commit(); cursor.close(); db.close()
    flash(f"Onboarding assigned to {assigned} employee(s).", "success")
    return redirect("/employees")


@app.route("/export_onboarding_csv")
@admin_required
def export_onboarding_csv():
    import csv, io
    db = get_db_connection(); cursor = db.cursor()
    cursor.execute("""
        SELECT e.employee_id, e.name, e.department, ot.name,
               eo.assigned_date, eo.due_date, eo.status,
               COUNT(eot.id) AS total_tasks,
               SUM(CASE WHEN eot.status='Done' THEN 1 ELSE 0 END) AS done_tasks
        FROM employee_onboarding eo
        JOIN employees e ON eo.employee_id = e.employee_id
        JOIN onboarding_templates ot ON eo.template_id = ot.id
        LEFT JOIN employee_onboarding_tasks eot ON eot.onboarding_id = eo.id
        GROUP BY eo.id, e.employee_id, e.name, e.department, ot.name,
                 eo.assigned_date, eo.due_date, eo.status
        ORDER BY eo.assigned_date DESC
    """)
    rows = cursor.fetchall()
    cursor.close(); db.close()
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Employee ID","Name","Department","Template","Assigned Date","Due Date","Status","Total Tasks","Done Tasks","Progress %"])
    for r in rows:
        pct = round(int(r[8] or 0) / int(r[7] or 1) * 100) if r[7] else 0
        writer.writerow([r[0], r[1], r[2] or "", r[3], r[4], r[5], r[6], r[7], r[8] or 0, f"{pct}%"])
    output.seek(0)
    from flask import Response
    return Response(output.getvalue(), mimetype="text/csv",
                    headers={"Content-Disposition": f"attachment;filename=onboarding_export_{datetime.date.today()}.csv"})


@app.route("/onboarding_template_duplicate", methods=["POST"])
@admin_required
def onboarding_template_duplicate():
    db = get_db_connection(); cursor = db.cursor()
    tid = request.form.get("template_id")
    cursor.execute("SELECT name, description FROM onboarding_templates WHERE id=%s", (tid,))
    tpl = cursor.fetchone()
    if not tpl:
        flash("Template not found.", "error")
        cursor.close(); db.close()
        return redirect("/onboarding?tab=templates")
    cursor.execute(
        "INSERT INTO onboarding_templates (name, description, is_active) VALUES (%s, %s, 1) RETURNING id",
        (f"Copy of {tpl[0]}", tpl[1])
    )
    new_id = cursor.fetchone()[0]
    cursor.execute(
        "SELECT task_title, task_description, requires_document, due_days, sort_order "
        "FROM onboarding_template_tasks WHERE template_id=%s ORDER BY sort_order, id", (tid,)
    )
    for task in cursor.fetchall():
        cursor.execute(
            "INSERT INTO onboarding_template_tasks "
            "(template_id, task_title, task_description, requires_document, due_days, sort_order) "
            "VALUES (%s,%s,%s,%s,%s,%s)",
            (new_id, task[0], task[1], task[2], task[3], task[4])
        )
    db.commit(); cursor.close(); db.close()
    flash(f"Template duplicated as 'Copy of {tpl[0]}'.", "success")
    return redirect(f"/onboarding_template_detail/{new_id}")


@app.route("/onboarding_template_delete", methods=["POST"])
@admin_required
def onboarding_template_delete():
    db = get_db_connection(); cursor = db.cursor()
    tid = request.form.get("template_id")
    cursor.execute("DELETE FROM onboarding_template_tasks WHERE template_id=%s", (tid,))
    cursor.execute("DELETE FROM onboarding_templates WHERE id=%s", (tid,))
    db.commit(); cursor.close(); db.close()
    flash("Template deleted.", "success")
    return redirect("/onboarding?tab=templates")

@app.route("/onboarding_task_save", methods=["POST"])
@admin_required
def onboarding_task_save():
    db = get_db_connection(); cursor = db.cursor()
    task_id   = request.form.get("task_id")
    tid       = request.form.get("template_id")
    title     = request.form.get("task_title", "").strip()
    desc      = request.form.get("task_description", "").strip()
    req_doc   = 1 if request.form.get("requires_document") else 0
    due_days  = int(request.form.get("due_days", 7))
    sort_order= int(request.form.get("sort_order", 0))
    if not title:
        flash("Task title is required.", "error")
        return redirect(f"/onboarding_template_detail/{tid}")
    if task_id:
        cursor.execute("""UPDATE onboarding_template_tasks
                          SET task_title=%s, task_description=%s, requires_document=%s,
                              due_days=%s, sort_order=%s
                          WHERE id=%s""", (title, desc, req_doc, due_days, sort_order, task_id))
        flash("Task updated.", "success")
    else:
        cursor.execute("""INSERT INTO onboarding_template_tasks
                          (template_id, task_title, task_description, requires_document, due_days, sort_order)
                          VALUES (%s,%s,%s,%s,%s,%s)""", (tid, title, desc, req_doc, due_days, sort_order))
        flash("Task added.", "success")
    db.commit(); cursor.close(); db.close()
    return redirect(f"/onboarding_template_detail/{tid}")

@app.route("/onboarding_task_delete", methods=["POST"])
@admin_required
def onboarding_task_delete():
    db = get_db_connection(); cursor = db.cursor()
    task_id = request.form.get("task_id")
    cursor.execute("SELECT template_id FROM onboarding_template_tasks WHERE id=%s", (task_id,))
    row = cursor.fetchone()
    tid = row[0] if row else None
    cursor.execute("DELETE FROM onboarding_template_tasks WHERE id=%s", (task_id,))
    db.commit(); cursor.close(); db.close()
    flash("Task deleted.", "success")
    return redirect(f"/onboarding_template_detail/{tid}")

@app.route("/onboarding_template_detail/<int:tid>")
@admin_required
def onboarding_template_detail(tid):
    db = get_db_connection(); cursor = db.cursor()
    cursor.execute("SELECT id, name, description, is_active FROM onboarding_templates WHERE id=%s", (tid,))
    template = cursor.fetchone()
    cursor.execute("""SELECT id, task_title, task_description, requires_document, due_days, sort_order
                      FROM onboarding_template_tasks WHERE template_id=%s ORDER BY sort_order, id""", (tid,))
    tasks = cursor.fetchall()
    co = get_company_settings()
    cursor.close(); db.close()
    return render_template("onboarding_template_detail.html",
        template=template, tasks=tasks, co=co,
        pending_leaves=0, pending_resignations=0, pending_tickets=0
    ,
        active_nav="onboarding",
    )

@app.route("/onboarding_assign", methods=["POST"])
@admin_required
def onboarding_assign():
    db = get_db_connection(); cursor = db.cursor()
    emp_id   = request.form.get("employee_id")
    tid      = request.form.get("template_id")
    due_date = request.form.get("due_date") or None
    today    = datetime.date.today()

    # Check not already assigned same template
    cursor.execute("SELECT id FROM employee_onboarding WHERE employee_id=%s AND template_id=%s AND status='In Progress'",
                   (emp_id, tid))
    if cursor.fetchone():
        flash("This employee already has this onboarding in progress.", "error")
        cursor.close(); db.close()
        return redirect("/onboarding?tab=active")

    cursor.execute("INSERT INTO employee_onboarding (employee_id, template_id, assigned_date, due_date) VALUES (%s,%s,%s,%s) RETURNING id",
                   (emp_id, tid, today, due_date))
    ob_id = cursor.fetchone()[0]

    # Copy tasks from template
    cursor.execute("""SELECT id, task_title, task_description, requires_document, due_days
                      FROM onboarding_template_tasks WHERE template_id=%s ORDER BY sort_order, id""", (tid,))
    for task in cursor.fetchall():
        cursor.execute("""INSERT INTO employee_onboarding_tasks
                          (onboarding_id, template_task_id, employee_id, task_title, task_description, requires_document, due_days)
                          VALUES (%s,%s,%s,%s,%s,%s,%s)""",
                       (ob_id, task[0], emp_id, task[1], task[2], task[3], task[4]))
    db.commit()

    # Notification to employee
    try:
        cursor.execute("SELECT name FROM onboarding_templates WHERE id=%s", (tid,))
        tname = cursor.fetchone()[0]
        cursor.execute("""INSERT INTO employee_notifications (employee_id, title, message, notif_type)
                          VALUES (%s, 'Onboarding Started', %s, 'info')""",
                       (emp_id, f"Your onboarding checklist '{tname}' has been assigned. Please complete all tasks."))
        db.commit()
    except Exception:
        pass

    cursor.execute("SELECT name, email FROM employees WHERE employee_id=%s", (emp_id,))
    _er = cursor.fetchone(); emp_name = _er[0]; emp_email = _er[1] if _er else None
    # Email employee about new onboarding assignment
    if emp_email:
        _ecfg = get_email_config()
        if _ecfg:
            try:
                _safe_name = _html.escape(emp_name or emp_id)
                _safe_tname = _html.escape(tname or "")
                _ob_html = (f"<p>Hi <strong>{_safe_name}</strong>,</p>"
                            f"<p>A new onboarding checklist <strong>'{_safe_tname}'</strong> has been assigned to you.</p>"
                            f"<p>Due date: <strong>{due_date or 'Not set'}</strong></p>"
                            f"<p>Please log in to your employee portal and complete all tasks on time.</p>")
                send_email_async(emp_email, f"New Onboarding Checklist Assigned — {tname}", _ob_html, _ecfg)
            except Exception:
                pass
    cursor.close(); db.close()
    flash(f"Onboarding assigned to {emp_name}.", "success")
    return redirect("/onboarding?tab=active")

@app.route("/onboarding_detail/<int:ob_id>")
@admin_required
def onboarding_detail(ob_id):
    db = get_db_connection(); cursor = db.cursor()
    cursor.execute("""
        SELECT eo.id, e.employee_id, e.name, e.role, e.department,
               ot.name AS tname, eo.assigned_date, eo.due_date, eo.status
        FROM employee_onboarding eo
        JOIN employees e ON e.employee_id=eo.employee_id
        JOIN onboarding_templates ot ON ot.id=eo.template_id
        WHERE eo.id=%s
    """, (ob_id,))
    ob = cursor.fetchone()
    cursor.execute("""
        SELECT id, task_title, task_description, requires_document, due_days,
               status, completed_at, document_path, admin_notes, employee_note
        FROM employee_onboarding_tasks WHERE onboarding_id=%s ORDER BY id
    """, (ob_id,))
    tasks = cursor.fetchall()
    co = get_company_settings()
    cursor.close(); db.close()
    return render_template("onboarding_detail.html",
        ob=ob, tasks=tasks, co=co,
        today=datetime.date.today(),
        pending_leaves=0, pending_resignations=0, pending_tickets=0,
        active_nav="onboarding",
    )

@app.route("/onboarding_admin_task_update", methods=["POST"])
@admin_required
def onboarding_admin_task_update():
    db = get_db_connection(); cursor = db.cursor()
    task_id    = request.form.get("task_id")
    new_status = request.form.get("status")
    notes      = request.form.get("admin_notes", "")
    ob_id      = request.form.get("ob_id")
    completed  = datetime.datetime.now() if new_status == "Done" else None
    cursor.execute("""UPDATE employee_onboarding_tasks
                      SET status=%s, completed_at=%s, admin_notes=%s WHERE id=%s""",
                   (new_status, completed, notes, task_id))
    # Auto-complete onboarding if all tasks done
    cursor.execute("SELECT COUNT(*) FROM employee_onboarding_tasks WHERE onboarding_id=%s AND status!='Done'", (ob_id,))
    remaining = cursor.fetchone()[0]
    if remaining == 0:
        cursor.execute("UPDATE employee_onboarding SET status='Completed' WHERE id=%s", (ob_id,))
    db.commit(); cursor.close(); db.close()
    flash("Task updated.", "success")
    return redirect(f"/onboarding_detail/{ob_id}")

@app.route("/onboarding_close", methods=["POST"])
@admin_required
def onboarding_close():
    db = get_db_connection(); cursor = db.cursor()
    ob_id = request.form.get("ob_id")
    cursor.execute("UPDATE employee_onboarding SET status='Completed' WHERE id=%s", (ob_id,))
    db.commit(); cursor.close(); db.close()
    flash("Onboarding marked as completed.", "success")
    return redirect("/onboarding?tab=active")

# ── OFFER LETTER ──────────────────────────────────────────────────────────────
@app.route("/offer_letter/<int:ob_id>")
@admin_required
def offer_letter(ob_id):
    db = get_db_connection(); cursor = db.cursor()
    cursor.execute("""
        SELECT eo.id, e.employee_id, e.name, e.role, e.department, e.email,
               eo.assigned_date, e.date_of_joining
        FROM employee_onboarding eo
        JOIN employees e ON e.employee_id = eo.employee_id
        WHERE eo.id = %s
    """, (ob_id,))
    ob = cursor.fetchone()
    cursor.execute("SELECT COALESCE(monthly_ctc,0), COALESCE(salary_per_day,0) FROM salary_config WHERE employee_id=%s", (ob[1],))
    sal = cursor.fetchone() or (0, 0)
    monthly_ctc = float(sal[0]) or round(float(sal[1]) * 26, 2)
    cursor.execute("SELECT * FROM offer_letters WHERE onboarding_id=%s ORDER BY id DESC LIMIT 1", (ob_id,))
    existing = cursor.fetchone()
    co = get_company_settings()
    cursor.close(); db.close()
    return render_template("offer_letter.html", ob=ob, monthly_ctc=monthly_ctc,
                           existing=existing, co=co,
                           pending_leaves=0, pending_resignations=0, pending_tickets=0,
        active_nav="onboarding",
    )

@app.route("/offer_letter_save", methods=["POST"])
@admin_required
def offer_letter_save():
    ob_id         = request.form.get("ob_id")
    employee_id   = request.form.get("employee_id")
    designation   = request.form.get("designation","")
    department    = request.form.get("department","")
    work_location = request.form.get("work_location","")
    monthly_ctc   = request.form.get("monthly_ctc", 0) or 0
    joining_date  = request.form.get("joining_date") or None
    valid_until   = request.form.get("offer_valid_until") or None
    probation     = int(request.form.get("probation_months", 6))
    reporting_to  = request.form.get("reporting_to","")
    notes         = request.form.get("additional_notes","")
    notice_days   = int(request.form.get("notice_period_days", 30))
    candidate_addr= request.form.get("candidate_address","")
    db = get_db_connection(); cursor = db.cursor()
    # add new columns if they don't exist yet (migration)
    try:
        cursor.execute("ALTER TABLE offer_letters ADD COLUMN IF NOT EXISTS notice_period_days INT DEFAULT 30")
        db.commit()
    except psycopg2.Error:
        db.rollback()
    try:
        cursor.execute("ALTER TABLE offer_letters ADD COLUMN IF NOT EXISTS candidate_address TEXT")
        db.commit()
    except psycopg2.Error:
        db.rollback()
    cursor.execute("SELECT id FROM offer_letters WHERE onboarding_id=%s", (ob_id,))
    existing = cursor.fetchone()
    if existing:
        cursor.execute("""UPDATE offer_letters SET designation=%s,department=%s,work_location=%s,
            monthly_ctc=%s,joining_date=%s,offer_valid_until=%s,probation_months=%s,
            reporting_to=%s,additional_notes=%s,notice_period_days=%s,candidate_address=%s,
            generated_at=NOW(),status='draft',sent_at=NULL
            WHERE id=%s""",
            (designation,department,work_location,monthly_ctc,joining_date,valid_until,
             probation,reporting_to,notes,notice_days,candidate_addr,existing[0]))
        letter_id = existing[0]
    else:
        cursor.execute("""INSERT INTO offer_letters (onboarding_id,employee_id,designation,department,
            work_location,monthly_ctc,joining_date,offer_valid_until,probation_months,
            reporting_to,additional_notes,notice_period_days,candidate_address)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id""",
            (ob_id,employee_id,designation,department,work_location,monthly_ctc,
             joining_date,valid_until,probation,reporting_to,notes,notice_days,candidate_addr))
        letter_id = cursor.fetchone()[0]
    db.commit(); cursor.close(); db.close()
    return redirect(f"/offer_letter_view/{letter_id}")

@app.route("/offer_letter_view/<int:letter_id>")
@admin_required
def offer_letter_view(letter_id):
    db = get_db_connection(); cursor = db.cursor()
    cursor.execute("""
        SELECT ol.id,ol.onboarding_id,ol.employee_id,ol.designation,ol.department,
               ol.work_location,ol.monthly_ctc,ol.joining_date,ol.offer_valid_until,
               ol.probation_months,ol.reporting_to,ol.additional_notes,ol.generated_at,
               ol.sent_at,ol.status,
               COALESCE(ol.notice_period_days,30),COALESCE(ol.candidate_address,''),
               e.name, e.email
        FROM offer_letters ol
        JOIN employees e ON e.employee_id = ol.employee_id
        WHERE ol.id = %s
    """, (letter_id,))
    letter = cursor.fetchone()
    co = get_company_settings()
    cursor.close(); db.close()
    if not letter:
        flash("Offer letter not found.", "error")
        return redirect("/onboarding")
    return render_template("offer_letter_view.html", letter=letter, co=co)

def _generate_offer_letter_pdf(letter, co):
    """Build offer letter PDF with ReportLab and return bytes."""
    from io import BytesIO
    from reportlab.lib import colors as rl_colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Table,
                                    TableStyle, Spacer, HRFlowable)
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

    BLUE  = rl_colors.HexColor("#1d4ed8")
    DARK  = rl_colors.HexColor("#111827")
    GRAY  = rl_colors.HexColor("#6b7280")
    LIGHT = rl_colors.HexColor("#f3f4f6")

    emp_name      = letter[17]
    designation   = letter[3] or "the offered position"
    department    = letter[4] or ""
    work_location = letter[5] or ""
    monthly_ctc   = float(letter[6]) if letter[6] else 0
    joining_date  = letter[7].strftime("%d %B %Y") if letter[7] else "—"
    valid_until   = letter[8].strftime("%d %B %Y") if letter[8] else "7 days from date of issue"
    probation     = letter[9] or 6
    reporting_to  = letter[10] or "the Department Head"
    notes         = letter[11] or ""
    gen_date      = letter[12].strftime("%d %B %Y") if letter[12] else ""
    notice_days   = letter[15] or 30
    ref_num       = f"OL/{letter[2].upper()}/{letter[12].strftime('%Y') if letter[12] else ''}/{letter[0]:04d}"
    company       = co.get("company_name", "Company")
    co_address    = co.get("address", "")
    co_email_val  = co.get("email", "")

    def ps(name, **kw):
        base = dict(fontName="Helvetica", fontSize=10, leading=14, textColor=DARK)
        base.update(kw)
        return ParagraphStyle(name, **base)

    sNormal  = ps("normal")
    sBold    = ps("bold",   fontName="Helvetica-Bold")
    sSmall   = ps("small",  fontSize=8,  textColor=GRAY)
    sLabel   = ps("label",  fontSize=8,  fontName="Helvetica-Bold", textColor=BLUE, spaceAfter=4)
    sCenter  = ps("center", alignment=TA_CENTER)
    sRight   = ps("right",  alignment=TA_RIGHT)

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=20*mm, rightMargin=20*mm,
                            topMargin=14*mm, bottomMargin=16*mm)
    story = []

    # ── Blue top rule ──────────────────────────────────────────────────────
    story.append(HRFlowable(width="100%", thickness=4, color=BLUE, spaceAfter=8))

    # ── Letterhead row ─────────────────────────────────────────────────────
    addr_line = co_address
    if co_email_val:
        addr_line += f"  ·  {co_email_val}" if addr_line else co_email_val
    lh_data = [[
        [Paragraph(f"<b>{company}</b>", ps("co", fontSize=14, fontName="Helvetica-Bold")),
         Paragraph(addr_line, sSmall)],
        [Paragraph(f"<b>Date:</b> {gen_date}", sRight),
         Paragraph(f"<b>Ref:</b> {ref_num}", sRight)],
    ]]
    lh_tbl = Table(lh_data, colWidths=["55%", "45%"])
    lh_tbl.setStyle(TableStyle([
        ("VALIGN",  (0,0), (-1,-1), "TOP"),
        ("ALIGN",   (1,0), (1,-1),  "RIGHT"),
        ("BOTTOMPADDING", (0,0), (-1,-1), 6),
    ]))
    story.append(lh_tbl)
    story.append(HRFlowable(width="100%", thickness=1, color=rl_colors.HexColor("#e5e7eb"), spaceAfter=10))

    # ── To block ───────────────────────────────────────────────────────────
    story.append(Paragraph("<b>To,</b>", sNormal))
    story.append(Paragraph(emp_name, sNormal))
    story.append(Paragraph(f"Employee ID: {letter[2]}", sNormal))
    story.append(Spacer(1, 8))

    # ── Subject ────────────────────────────────────────────────────────────
    story.append(Paragraph(f"<u><b>Sub: Offer of Employment — {designation}</b></u>", sNormal))
    story.append(Spacer(1, 10))

    # ── Salutation ─────────────────────────────────────────────────────────
    story.append(Paragraph(f"Dear <b>{emp_name}</b>,", sNormal))
    story.append(Spacer(1, 8))

    # ── Opening paragraphs ─────────────────────────────────────────────────
    dept_txt  = f" in the <b>{department}</b> department" if department else ""
    loc_txt   = f", located at <b>{work_location}</b>" if work_location else ""
    story.append(Paragraph(
        f"We are pleased to offer you the position of <b>{designation}</b>{dept_txt} "
        f"at <b>{company}</b>{loc_txt}. You will be reporting to <b>{reporting_to}</b>.",
        sNormal))
    story.append(Spacer(1, 6))
    story.append(Paragraph(
        f"Your date of joining will be <b>{joining_date}</b>. Please report to the HR Department "
        f"on the joining date with your original documents for verification.",
        sNormal))
    story.append(Spacer(1, 6))
    story.append(Paragraph(
        f"At <b>{company}</b>, we believe in fostering a collaborative, growth-oriented environment "
        f"where every team member is empowered to make an impact. As a <b>{designation}</b>, "
        f"you will play a key role in driving our mission forward. We look forward to the "
        f"valuable perspective and expertise you will bring to the team.",
        sNormal))
    story.append(Spacer(1, 12))

    # ── Compensation ───────────────────────────────────────────────────────
    if monthly_ctc > 0:
        story.append(Paragraph("COMPENSATION DETAILS", sLabel))
        basic = round(monthly_ctc * 0.40, 2)
        hra   = round(monthly_ctc * 0.20, 2)
        sa    = round(monthly_ctc * 0.33, 2)
        pf    = round(monthly_ctc * 0.04, 2)
        gr    = round(monthly_ctc * 0.03, 2)
        def fmt(n): return f"₹{n:,.2f}"
        ctc_data = [
            ["Salary Component", "Monthly", "Annual"],
            ["Basic Salary",            fmt(basic),       fmt(basic*12)],
            ["House Rent Allowance",     fmt(hra),         fmt(hra*12)],
            ["Special Allowance",        fmt(sa),          fmt(sa*12)],
            ["PF — Employer (12%)",      fmt(pf),          fmt(pf*12)],
            ["Gratuity (4.81%)",         fmt(gr),          fmt(gr*12)],
            ["GROSS CTC",                fmt(monthly_ctc), fmt(monthly_ctc*12)],
        ]
        ctc_tbl = Table(ctc_data, colWidths=["50%", "25%", "25%"])
        ctc_tbl.setStyle(TableStyle([
            ("BACKGROUND",   (0, 0), (-1, 0),  LIGHT),
            ("BACKGROUND",   (0, -1), (-1, -1), DARK),
            ("TEXTCOLOR",    (0, 0), (-1, 0),  GRAY),
            ("TEXTCOLOR",    (0, -1), (-1, -1), rl_colors.white),
            ("FONTNAME",     (0, 0), (-1, 0),  "Helvetica-Bold"),
            ("FONTNAME",     (0, -1), (-1, -1), "Helvetica-Bold"),
            ("FONTSIZE",     (0, 0), (-1, -1), 9),
            ("ALIGN",        (1, 0), (-1, -1),  "RIGHT"),
            ("ROWBACKGROUNDS",(0,1), (-1,-2),  [rl_colors.white, rl_colors.HexColor("#f9fafb")]),
            ("GRID",         (0, 0), (-1, -2),  0.3, rl_colors.HexColor("#e5e7eb")),
            ("TOPPADDING",   (0, 0), (-1, -1),  6),
            ("BOTTOMPADDING",(0, 0), (-1, -1),  6),
            ("LEFTPADDING",  (0, 0), (-1, -1),  8),
            ("RIGHTPADDING", (0, 0), (-1, -1),  8),
        ]))
        story.append(ctc_tbl)
        story.append(Spacer(1, 10))

    # ── Notes ──────────────────────────────────────────────────────────────
    if notes:
        note_tbl = Table([[Paragraph(f"<b>Note:</b> {notes}", ps("note", fontSize=9, textColor=rl_colors.HexColor("#1e40af")))]],
                         colWidths=["100%"])
        note_tbl.setStyle(TableStyle([
            ("BACKGROUND",  (0,0), (-1,-1), rl_colors.HexColor("#eff6ff")),
            ("LEFTPADDING", (0,0), (-1,-1), 10),
            ("TOPPADDING",  (0,0), (-1,-1), 8),
            ("BOTTOMPADDING",(0,0),(-1,-1), 8),
        ]))
        story.append(note_tbl)
        story.append(Spacer(1, 10))

    # ── Terms & Conditions ─────────────────────────────────────────────────
    story.append(Paragraph("TERMS &amp; CONDITIONS", sLabel))
    tc_items = [
        "This offer is subject to satisfactory verification of your educational qualifications, credentials, and prior employment history.",
        f"You will serve a probationary period of <b>{probation} months</b> from the date of joining. Confirmation is subject to satisfactory performance.",
        f"Post-confirmation, either party may terminate employment by providing <b>{notice_days} days'</b> written notice or salary in lieu thereof. During probation, 7 days' notice applies.",
        "All compensation is subject to applicable statutory deductions (TDS, PF, ESI, Professional Tax) as per prevailing Indian law.",
        f"This offer is valid until <b>{valid_until}</b>. Non-acceptance by this date shall render this offer null and void.",
        "You shall maintain strict confidentiality of all proprietary and sensitive information of the Company during and after your employment.",
        "You will abide by the Company's HR policies, Code of Conduct, and all applicable rules as amended from time to time.",
        "A formal Appointment Letter will be issued upon joining. This offer letter does not constitute a contract of employment.",
    ]
    tc_data = [
        [Paragraph(f"{i+1}.&nbsp;&nbsp;{item}", ps(f"tc{i}", fontSize=9, leading=13, textColor=rl_colors.HexColor("#4b5563")))]
        for i, item in enumerate(tc_items)
    ]
    tc_tbl = Table(tc_data, colWidths=["100%"])
    tc_tbl.setStyle(TableStyle([
        ("TOPPADDING",    (0,0), (-1,-1), 3),
        ("BOTTOMPADDING", (0,0), (-1,-1), 3),
    ]))
    story.append(tc_tbl)
    story.append(Spacer(1, 12))

    story.append(Paragraph(
        f"We look forward to welcoming you to <b>{company}</b>. "
        "Please sign and return one copy of this letter to confirm your acceptance.",
        sNormal))
    story.append(Spacer(1, 6))
    story.append(Paragraph("Warm regards,", sNormal))
    story.append(Spacer(1, 24))

    # ── Signature row ──────────────────────────────────────────────────────
    sig_data = [[
        [HRFlowable(width="80%", thickness=1, color=DARK),
         Paragraph("<b>Authorised Signatory</b>", ps("sig", fontSize=9)),
         Paragraph(company, ps("sigt", fontSize=8, textColor=GRAY)),
         Paragraph("Human Resources", ps("sigt2", fontSize=8, textColor=GRAY))],
        [Paragraph("I hereby accept this offer and agree to all terms stated above.", ps("accnote", fontSize=8, textColor=GRAY)),
         HRFlowable(width="80%", thickness=1, color=DARK),
         Paragraph(f"<b>{emp_name}</b>", ps("csig", fontSize=9)),
         Paragraph("Candidate Signature", ps("csigt", fontSize=8, textColor=GRAY)),
         Paragraph("Date: _______________", ps("cdate", fontSize=8, textColor=GRAY))],
    ]]
    sig_tbl = Table(sig_data, colWidths=["48%", "52%"])
    sig_tbl.setStyle(TableStyle([("VALIGN", (0,0), (-1,-1), "TOP"), ("TOPPADDING", (0,0), (-1,-1), 0)]))
    story.append(sig_tbl)

    # ── Footer rule ────────────────────────────────────────────────────────
    story.append(Spacer(1, 16))
    story.append(HRFlowable(width="100%", thickness=1, color=rl_colors.HexColor("#e5e7eb")))
    foot_txt = company
    if co_address:
        foot_txt += f"  ·  {co_address}"
    story.append(Paragraph(f'<font size="8" color="#9ca3af">{foot_txt}&nbsp;&nbsp;&nbsp;Confidential — For addressee only</font>', sCenter))
    story.append(HRFlowable(width="100%", thickness=4, color=DARK, spaceBefore=6))

    doc.build(story)
    return buf.getvalue()


@app.route("/offer_letter_send/<int:letter_id>", methods=["POST"])
@admin_required
def offer_letter_send(letter_id):
    db = get_db_connection(); cursor = db.cursor()
    cursor.execute("""
        SELECT ol.id,ol.onboarding_id,ol.employee_id,ol.designation,ol.department,
               ol.work_location,ol.monthly_ctc,ol.joining_date,ol.offer_valid_until,
               ol.probation_months,ol.reporting_to,ol.additional_notes,ol.generated_at,
               ol.sent_at,ol.status,
               COALESCE(ol.notice_period_days,30),COALESCE(ol.candidate_address,''),
               e.name, e.email
        FROM offer_letters ol
        JOIN employees e ON e.employee_id = ol.employee_id
        WHERE ol.id = %s
    """, (letter_id,))
    letter = cursor.fetchone()
    co = get_company_settings()
    if not letter or not letter[18]:
        flash("Employee email not found.", "error")
        cursor.close(); db.close()
        return redirect(f"/offer_letter_view/{letter_id}")
    cfg = get_email_config()
    if not cfg:
        flash("Email not configured. Go to Settings → Email.", "error")
        cursor.close(); db.close()
        return redirect(f"/offer_letter_view/{letter_id}")
    try:
        emp_name      = letter[17]
        emp_email     = letter[18]
        designation   = letter[3] or "the offered position"
        department    = letter[4] or ""
        work_location = letter[5] or ""
        monthly_ctc   = float(letter[6]) if letter[6] else 0
        joining_date  = letter[7].strftime("%d %B %Y") if letter[7] else "—"
        valid_until   = letter[8].strftime("%d %B %Y") if letter[8] else "7 days from date of issue"
        probation     = letter[9] or 6
        reporting_to  = letter[10] or "the Department Head"
        notes         = letter[11] or ""
        gen_date      = letter[12].strftime("%d %B %Y") if letter[12] else ""
        notice_days   = letter[15] or 30
        ref_num       = f"OL/{letter[2].upper()}/{letter[12].strftime('%Y') if letter[12] else ''}/{letter[0]:04d}"
        company       = co.get("company_name", "Company")
        co_address    = co.get("address", "")
        co_email      = co.get("email", "")

        # Secure one-time token (shared by accept/reject AND pdf view)
        token = secrets.token_urlsafe(32)
        token_hash = hashlib.sha256(token.encode()).hexdigest()
        base_url    = _safe_app_url()
        accept_url  = f"{base_url}/offer_letter_respond/{token}/accept"
        reject_url  = f"{base_url}/offer_letter_respond/{token}/reject"
        pdf_view_url = f"{base_url}/offer_letter_pdf/{token}"
        pdf_dl_url   = f"{base_url}/offer_letter_pdf/{token}?dl=1"

        # ── Salary breakdown helper ────────────────────────────────────────
        def fmt(n): return f"{n:,.2f}"
        ctc_section = ""
        if monthly_ctc > 0:
            basic = round(monthly_ctc * 0.40, 2)
            hra   = round(monthly_ctc * 0.20, 2)
            sa    = round(monthly_ctc * 0.33, 2)
            pf    = round(monthly_ctc * 0.04, 2)
            gr    = round(monthly_ctc * 0.03, 2)
            ctc_section = f"""
            <p style="font-size:10px;font-weight:700;letter-spacing:1.5px;text-transform:uppercase;color:#1d4ed8;margin:20px 0 8px;">Compensation Details</p>
            <table style="width:100%;border-collapse:collapse;font-size:12.5px;margin-bottom:20px;">
              <thead><tr>
                <th style="background:#f3f4f6;color:#6b7280;font-size:10px;font-weight:700;text-transform:uppercase;padding:9px 12px;text-align:left;border-bottom:1px solid #e5e7eb;">Salary Component</th>
                <th style="background:#f3f4f6;color:#6b7280;font-size:10px;font-weight:700;text-transform:uppercase;padding:9px 12px;text-align:right;border-bottom:1px solid #e5e7eb;">Monthly (&#8377;)</th>
                <th style="background:#f3f4f6;color:#6b7280;font-size:10px;font-weight:700;text-transform:uppercase;padding:9px 12px;text-align:right;border-bottom:1px solid #e5e7eb;">Annual (&#8377;)</th>
              </tr></thead>
              <tbody>
                <tr><td style="padding:9px 12px;border-bottom:1px solid #f3f4f6;">Basic Salary</td><td style="padding:9px 12px;text-align:right;font-weight:600;border-bottom:1px solid #f3f4f6;">{fmt(basic)}</td><td style="padding:9px 12px;text-align:right;font-weight:600;border-bottom:1px solid #f3f4f6;">{fmt(basic*12)}</td></tr>
                <tr><td style="padding:9px 12px;border-bottom:1px solid #f3f4f6;">House Rent Allowance (HRA)</td><td style="padding:9px 12px;text-align:right;font-weight:600;border-bottom:1px solid #f3f4f6;">{fmt(hra)}</td><td style="padding:9px 12px;text-align:right;font-weight:600;border-bottom:1px solid #f3f4f6;">{fmt(hra*12)}</td></tr>
                <tr><td style="padding:9px 12px;border-bottom:1px solid #f3f4f6;">Special Allowance</td><td style="padding:9px 12px;text-align:right;font-weight:600;border-bottom:1px solid #f3f4f6;">{fmt(sa)}</td><td style="padding:9px 12px;text-align:right;font-weight:600;border-bottom:1px solid #f3f4f6;">{fmt(sa*12)}</td></tr>
                <tr><td style="padding:9px 12px;border-bottom:1px solid #f3f4f6;">PF — Employer Contribution</td><td style="padding:9px 12px;text-align:right;font-weight:600;border-bottom:1px solid #f3f4f6;">{fmt(pf)}</td><td style="padding:9px 12px;text-align:right;font-weight:600;border-bottom:1px solid #f3f4f6;">{fmt(pf*12)}</td></tr>
                <tr><td style="padding:9px 12px;">Gratuity (4.81% of Basic)</td><td style="padding:9px 12px;text-align:right;font-weight:600;">{fmt(gr)}</td><td style="padding:9px 12px;text-align:right;font-weight:600;">{fmt(gr*12)}</td></tr>
              </tbody>
              <tfoot><tr>
                <td style="padding:10px 12px;font-weight:800;background:#111827;color:#fff;">Gross CTC</td>
                <td style="padding:10px 12px;text-align:right;font-weight:800;background:#111827;color:#fff;">&#8377;{fmt(monthly_ctc)}</td>
                <td style="padding:10px 12px;text-align:right;font-weight:800;background:#111827;color:#fff;">&#8377;{fmt(monthly_ctc*12)}</td>
              </tr></tfoot>
            </table>"""

        notes_section = ""
        if notes:
            notes_section = f"""<div style="background:#eff6ff;border-left:3px solid #1d4ed8;padding:11px 16px;font-size:12.5px;color:#1e40af;border-radius:0 6px 6px 0;margin-bottom:16px;line-height:1.7;">
              <strong>Note:</strong> {notes}</div>"""

        dept_html = f' in the <strong>{department}</strong> department' if department else ''
        loc_html  = f', located at <strong>{work_location}</strong>' if work_location else ''

        html_body = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8"/>
<style>
  @keyframes burst {{
    0%   {{ transform:translate(var(--tx,0),0) rotate(0deg) scale(1); opacity:1; }}
    100% {{ transform:translate(var(--tx,0),var(--ty,-70px)) rotate(var(--rot,360deg)) scale(0); opacity:0; }}
  }}
  .cw {{ position:relative; display:inline-block; cursor:default; }}
  .cw:hover .cp {{ animation:burst .75s ease-out forwards; }}
  .cp {{ position:absolute; width:7px; height:7px; border-radius:2px;
         top:0; left:50%; opacity:0; pointer-events:none; }}
</style>
</head>
<body style="margin:0;padding:0;background:#e5e7eb;font-family:'Segoe UI',Arial,sans-serif;">
<div style="max-width:680px;margin:32px auto;background:#fff;border-radius:6px;overflow:hidden;box-shadow:0 4px 24px rgba(0,0,0,0.13);">

  <!-- Top accent -->
  <div style="height:4px;background:#1d4ed8;"></div>

  <!-- Hero banner -->
  <div style="background:linear-gradient(135deg,#1e3a8a 0%,#2563eb 100%);padding:40px 48px 32px;text-align:center;">
    <div class="cw">
      <span style="font-size:28px;font-weight:800;color:#fff;letter-spacing:-.5px;">
        &#127881; Congratulations, {emp_name}!
      </span>
      <!-- confetti pieces -->
      <span class="cp" style="background:#ff6b6b;--tx:-38px;--ty:-75px;--rot:240deg;animation-delay:.00s;"></span>
      <span class="cp" style="background:#ffd93d;--tx:-18px;--ty:-82px;--rot:180deg;animation-delay:.05s;"></span>
      <span class="cp" style="background:#6bcb77;--tx:  5px;--ty:-78px;--rot:300deg;animation-delay:.10s;"></span>
      <span class="cp" style="background:#4d96ff;--tx: 24px;--ty:-70px;--rot:120deg;animation-delay:.05s;"></span>
      <span class="cp" style="background:#ff6b6b;--tx: 42px;--ty:-80px;--rot: 60deg;animation-delay:.00s;"></span>
      <span class="cp" style="background:#c77dff;--tx:-50px;--ty:-60px;--rot:200deg;animation-delay:.12s;"></span>
      <span class="cp" style="background:#ffd93d;--tx: 55px;--ty:-65px;--rot:160deg;animation-delay:.08s;"></span>
      <span class="cp" style="background:#6bcb77;--tx:-25px;--ty:-90px;--rot:280deg;animation-delay:.03s;"></span>
      <span class="cp" style="background:#4d96ff;--tx: 30px;--ty:-88px;--rot:330deg;animation-delay:.15s;"></span>
      <span class="cp" style="background:#ff9f1c;--tx:  0px;--ty:-95px;--rot:  0deg;animation-delay:.07s;"></span>
    </div>
    <p style="color:#bfdbfe;font-size:14px;margin-top:10px;margin-bottom:0;">
      We are thrilled to welcome you to the <strong style="color:#fff;">{company}</strong> family!
    </p>
  </div>

  <!-- Letterhead meta -->
  <div style="padding:20px 48px 0;display:table;width:100%;box-sizing:border-box;">
    <div style="display:table-cell;vertical-align:top;">
      <div style="font-size:16px;font-weight:800;color:#111827;">{company}</div>
      <div style="font-size:11px;color:#9ca3af;margin-top:3px;">{co_address}{(' &nbsp;·&nbsp; ' + co_email) if co_email else ''}</div>
    </div>
    <div style="display:table-cell;vertical-align:top;text-align:right;font-size:11px;color:#6b7280;line-height:1.8;">
      <div><strong style="color:#111827;">Date:</strong> {gen_date}</div>
      <div><strong style="color:#111827;">Ref:</strong> {ref_num}</div>
    </div>
  </div>
  <hr style="border:none;border-top:1.5px solid #e5e7eb;margin:14px 48px 0;"/>

  <!-- Address + Subject -->
  <div style="padding:18px 48px 0;font-size:13px;color:#374151;line-height:1.8;">
    <div style="font-weight:700;">To,</div>
    <div>{emp_name}</div>
    <div>Employee ID: {letter[2]}</div>
  </div>
  <div style="padding:12px 48px 0;font-size:13px;font-weight:700;color:#111827;text-decoration:underline;">
    Sub: Offer of Employment — {designation}
  </div>

  <!-- Letter body -->
  <div style="padding:18px 48px 32px;">

    <p style="font-size:13px;color:#374151;line-height:1.9;margin-bottom:14px;">
      Dear <strong>{emp_name}</strong>,
    </p>

    <p style="font-size:13px;color:#374151;line-height:1.9;margin-bottom:14px;">
      We are pleased to offer you the position of <strong>{designation}</strong>{dept_html}
      at <strong>{company}</strong>{loc_html}.
      You will be reporting to <strong>{reporting_to}</strong>.
    </p>

    <p style="font-size:13px;color:#374151;line-height:1.9;margin-bottom:14px;">
      Your date of joining will be <strong>{joining_date}</strong>. Please report to the HR Department
      on the joining date with your original documents for verification.
    </p>

    <!-- About the role -->
    <div style="background:#f8fafc;border-radius:8px;padding:18px 20px;margin-bottom:18px;">
      <p style="font-size:10px;font-weight:700;letter-spacing:1.2px;text-transform:uppercase;color:#1d4ed8;margin:0 0 10px;">About Your Role</p>
      <p style="font-size:13px;color:#374151;line-height:1.85;margin:0;">
        As a <strong>{designation}</strong>{dept_html} at <strong>{company}</strong>, you will be entrusted with
        responsibilities that directly contribute to our organisational goals. You will collaborate with
        cross-functional teams, lead initiatives within your domain, and contribute to building a high-performance
        culture. We expect you to bring creativity, ownership, and a commitment to excellence to every task.
      </p>
    </div>

    <!-- What we offer -->
    <div style="background:#f0fdf4;border-radius:8px;padding:18px 20px;margin-bottom:18px;">
      <p style="font-size:10px;font-weight:700;letter-spacing:1.2px;text-transform:uppercase;color:#15803d;margin:0 0 10px;">What We Offer</p>
      <table style="width:100%;border-collapse:collapse;font-size:12.5px;">
        <tr>
          <td style="padding:5px 8px;width:50%;vertical-align:top;">&#127775; Competitive CTC &amp; annual reviews</td>
          <td style="padding:5px 8px;width:50%;vertical-align:top;">&#128218; Learning &amp; development budget</td>
        </tr>
        <tr>
          <td style="padding:5px 8px;vertical-align:top;">&#127968; Flexible work environment</td>
          <td style="padding:5px 8px;vertical-align:top;">&#129303; Inclusive &amp; collaborative culture</td>
        </tr>
        <tr>
          <td style="padding:5px 8px;vertical-align:top;">&#127775; Performance bonuses &amp; incentives</td>
          <td style="padding:5px 8px;vertical-align:top;">&#128200; Clear growth &amp; promotion path</td>
        </tr>
      </table>
    </div>

    {ctc_section}
    {notes_section}

    <!-- Terms & Conditions -->
    <p style="font-size:10px;font-weight:700;letter-spacing:1.5px;text-transform:uppercase;color:#1d4ed8;margin:20px 0 8px;">Terms &amp; Conditions</p>
    <ol style="padding-left:18px;font-size:12.5px;color:#4b5563;line-height:1.85;margin-bottom:20px;">
      <li style="margin-bottom:7px;">This offer is subject to satisfactory verification of your educational qualifications, credentials, and prior employment history.</li>
      <li style="margin-bottom:7px;">You will serve a probationary period of <strong>{probation} months</strong> from the date of joining. Confirmation is subject to satisfactory performance.</li>
      <li style="margin-bottom:7px;">Post-confirmation, either party may terminate employment by providing <strong>{notice_days} days'</strong> written notice or salary in lieu thereof. During probation, 7 days' notice applies.</li>
      <li style="margin-bottom:7px;">All compensation is subject to applicable statutory deductions (TDS, PF, ESI, Professional Tax) as per prevailing Indian law.</li>
      <li style="margin-bottom:7px;">This offer is valid until <strong>{valid_until}</strong>. Non-acceptance by this date shall render this offer null and void.</li>
      <li style="margin-bottom:7px;">You shall maintain strict confidentiality of all proprietary and sensitive information of the Company during and after your employment.</li>
      <li style="margin-bottom:7px;">You will abide by the Company's HR policies, Code of Conduct, and all applicable rules as amended from time to time.</li>
      <li style="margin-bottom:7px;">A formal Appointment Letter will be issued upon joining. This offer letter does not constitute a contract of employment.</li>
    </ol>

    <p style="font-size:13px;color:#374151;line-height:1.9;margin-bottom:20px;">
      We look forward to welcoming you to <strong>{company}</strong>.
      Please review your complete offer letter PDF below and respond using the buttons at the bottom.
    </p>

    <!-- PDF section -->
    <div style="border:2px solid #e5e7eb;border-radius:10px;padding:20px 24px;margin-bottom:24px;background:#fafafa;">
      <div style="display:table;width:100%;">
        <div style="display:table-cell;vertical-align:middle;">
          <div style="font-size:32px;display:inline-block;vertical-align:middle;margin-right:12px;">&#128196;</div>
          <div style="display:inline-block;vertical-align:middle;">
            <div style="font-size:13px;font-weight:700;color:#111827;">Offer Letter — {emp_name}.pdf</div>
            <div style="font-size:11px;color:#9ca3af;margin-top:2px;">Complete offer letter with salary breakdown &amp; terms</div>
          </div>
        </div>
      </div>
      <div style="margin-top:14px;display:flex;gap:10px;">
        <a href="{pdf_view_url}"
           style="display:inline-block;padding:10px 22px;background:#1d4ed8;color:#fff;font-size:12px;font-weight:700;text-decoration:none;border-radius:7px;">
          &#128065; &nbsp;View PDF
        </a>
        <a href="{pdf_dl_url}"
           style="display:inline-block;padding:10px 22px;background:#fff;color:#111827;font-size:12px;font-weight:700;text-decoration:none;border-radius:7px;border:1.5px solid #d1d5db;margin-left:10px;">
          &#8681; &nbsp;Download PDF
        </a>
      </div>
    </div>

    <!-- Accept / Reject -->
    <div style="margin:0 0 16px;text-align:center;">
      <a href="{accept_url}"
         style="display:inline-block;padding:14px 40px;background:#16a34a;color:#fff;font-size:14px;font-weight:700;text-decoration:none;border-radius:8px;margin-right:14px;letter-spacing:.3px;">
        &#10003;&nbsp; Accept Offer
      </a>
      <a href="{reject_url}"
         style="display:inline-block;padding:14px 40px;background:#dc2626;color:#fff;font-size:14px;font-weight:700;text-decoration:none;border-radius:8px;letter-spacing:.3px;">
        &#10005;&nbsp; Decline Offer
      </a>
    </div>
    <p style="font-size:11px;color:#9ca3af;text-align:center;margin-bottom:24px;">
      Each response button can be used only once. Contact HR to change your response.
    </p>

    <p style="font-size:13px;color:#374151;line-height:1.9;">Warm regards,</p>
    <p style="font-size:13px;color:#374151;font-weight:700;margin-top:4px;">{company} HR Team</p>
  </div>

  <!-- Footer -->
  <div style="border-top:1px solid #e5e7eb;padding:10px 48px;font-size:10px;color:#9ca3af;display:table;width:100%;box-sizing:border-box;">
    <span style="display:table-cell;">{company}{(' · ' + co_address) if co_address else ''}</span>
    <span style="display:table-cell;text-align:right;">Confidential — For addressee only</span>
  </div>
  <div style="height:4px;background:#111827;"></div>
</div>
</body></html>"""

        # Generate PDF
        pdf_bytes = _generate_offer_letter_pdf(letter, co)
        safe_name = emp_name.replace(" ", "_")
        send_email_smtp(
            emp_email,
            f"Offer Letter — {company}",
            html_body,
            cfg,
            attachment_bytes=pdf_bytes,
            attachment_filename=f"Offer_Letter_{safe_name}.pdf",
        )

        token_expiry = datetime.datetime.utcnow() + datetime.timedelta(days=30)
        cursor.execute(
            "UPDATE offer_letters SET sent_at=NOW(), status='sent', response_token=%s, "
            "response_token_expiry=%s, candidate_response=NULL, responded_at=NULL WHERE id=%s",
            (token_hash, token_expiry, letter_id)
        )
        db.commit()
        flash(f"Offer letter emailed to {emp_email}.", "success")
    except Exception as ex:
        flash(f"Email failed: {ex}", "error")
    cursor.close(); db.close()
    return redirect(f"/offer_letter_view/{letter_id}")


@app.route("/offer_letter_pdf/<token>")
def offer_letter_pdf(token):
    """Serve the offer letter PDF to the candidate (view or download) using their email token."""
    db = get_db_connection(); cursor = db.cursor(buffered=True)
    cursor.execute("""
        SELECT ol.id,ol.onboarding_id,ol.employee_id,ol.designation,ol.department,
               ol.work_location,ol.monthly_ctc,ol.joining_date,ol.offer_valid_until,
               ol.probation_months,ol.reporting_to,ol.additional_notes,ol.generated_at,
               ol.sent_at,ol.status,
               COALESCE(ol.notice_period_days,30),COALESCE(ol.candidate_address,''),
               e.name, e.email
        FROM offer_letters ol
        JOIN employees e ON e.employee_id = ol.employee_id
        WHERE ol.response_token = %s
          AND (ol.response_token_expiry IS NULL OR ol.response_token_expiry > NOW())
    """, (hashlib.sha256(token.encode()).hexdigest(),))
    letter = cursor.fetchone()
    cursor.close(); db.close()
    if not letter:
        return "<html><body style='font-family:Segoe UI,sans-serif;padding:60px;text-align:center;'>" \
               "<h2 style='color:#dc2626;'>Invalid or expired link.</h2>" \
               "<p>Please contact HR for a copy of your offer letter.</p></body></html>", 404
    co = get_company_settings()
    pdf_bytes = _generate_offer_letter_pdf(letter, co)
    emp_name  = letter[17]
    safe_name = secure_filename(emp_name.replace(" ", "_")) or "Employee"
    dl = request.args.get("dl", "0")
    disposition = "attachment" if dl == "1" else "inline"
    from flask import Response
    resp = Response(pdf_bytes, mimetype="application/pdf")
    resp.headers["Content-Disposition"] = f'{disposition}; filename="Offer_Letter_{safe_name}.pdf"'
    return resp


@app.route("/offer_letter_respond/<token>/<action>")
def offer_letter_respond(token, action):
    if action not in ("accept", "reject"):
        return "Invalid action.", 400
    db = get_db_connection(); cursor = db.cursor(buffered=True)
    cursor.execute(
        "SELECT id, employee_id, candidate_response, status FROM offer_letters "
        "WHERE response_token=%s AND (response_token_expiry IS NULL OR response_token_expiry > NOW())",
        (hashlib.sha256(token.encode()).hexdigest(),)
    )
    row = cursor.fetchone()
    if not row:
        cursor.close(); db.close()
        return """<html><body style="font-family:Segoe UI,sans-serif;text-align:center;padding:60px;color:#374151;">
          <h2 style="color:#dc2626;">Invalid or expired link.</h2>
          <p>This offer letter link is not valid. Please contact HR.</p></body></html>""", 404
    letter_id, emp_id, existing_response, status = row
    if existing_response:
        label = "accepted" if existing_response == "accept" else "declined"
        color = "#16a34a" if existing_response == "accept" else "#dc2626"
        cursor.close(); db.close()
        return f"""<html><body style="font-family:Segoe UI,sans-serif;text-align:center;padding:60px;color:#374151;">
          <h2 style="color:{color};">You have already {label} this offer.</h2>
          <p>Please contact HR if you wish to change your response.</p></body></html>"""
    cursor.execute(
        "UPDATE offer_letters SET candidate_response=%s, responded_at=NOW(), status=%s WHERE id=%s",
        (action, "accepted" if action == "accept" else "rejected", letter_id)
    )
    db.commit()
    cursor.close(); db.close()
    if action == "accept":
        return """<html><body style="font-family:Segoe UI,sans-serif;text-align:center;padding:60px;color:#374151;">
          <div style="font-size:56px;">&#127881;</div>
          <h2 style="color:#16a34a;margin-top:16px;">Offer Accepted!</h2>
          <p style="font-size:15px;margin-top:8px;">Thank you for accepting the offer. HR will reach out to you with next steps.</p>
          <p style="margin-top:24px;font-size:13px;color:#9ca3af;">You may close this window.</p></body></html>"""
    else:
        return """<html><body style="font-family:Segoe UI,sans-serif;text-align:center;padding:60px;color:#374151;">
          <div style="font-size:56px;">&#128533;</div>
          <h2 style="color:#dc2626;margin-top:16px;">Offer Declined</h2>
          <p style="font-size:15px;margin-top:8px;">We have noted your decision. Thank you for considering us. We wish you the best.</p>
          <p style="margin-top:24px;font-size:13px;color:#9ca3af;">You may close this window.</p></body></html>"""

# Employee portal onboarding
@app.route("/my_onboarding")
@employee_required
def my_onboarding():
    emp_id = session.get("employee_id")
    db = get_db_connection(); cursor = db.cursor()
    cursor.execute("""
        SELECT eo.id, ot.name, eo.assigned_date, eo.due_date, eo.status,
               COUNT(eot.id) AS total, SUM(CASE WHEN eot.status='Done' THEN 1 ELSE 0 END) AS done
        FROM employee_onboarding eo
        JOIN onboarding_templates ot ON ot.id=eo.template_id
        LEFT JOIN employee_onboarding_tasks eot ON eot.onboarding_id=eo.id
        WHERE eo.employee_id=%s
        GROUP BY eo.id, ot.name, eo.assigned_date, eo.due_date, eo.status ORDER BY eo.assigned_date DESC
    """, (emp_id,))
    onboardings = cursor.fetchall()

    selected_ob_id = request.args.get("ob_id")
    tasks = []
    selected_ob = None
    if not selected_ob_id and onboardings:
        selected_ob_id = onboardings[0][0]
    if selected_ob_id:
        cursor.execute("""SELECT id, task_title, task_description, requires_document,
                                 due_days, status, completed_at, document_path
                          FROM employee_onboarding_tasks
                          WHERE onboarding_id=%s AND employee_id=%s ORDER BY id""",
                       (selected_ob_id, emp_id))
        tasks = cursor.fetchall()
        for ob in onboardings:
            if ob[0] == int(selected_ob_id):
                selected_ob = ob
                break

    cursor.execute("SELECT employee_id, name, role, department, face_image FROM employees WHERE employee_id=%s", (emp_id,))
    emp = cursor.fetchone()
    cursor.close(); db.close()
    return render_template("my_onboarding.html",
        emp=emp, emp_id=emp_id, onboardings=onboardings, tasks=tasks,
        selected_ob=selected_ob, selected_ob_id=int(selected_ob_id) if selected_ob_id else None,
        today=datetime.date.today(),
    )

@app.route("/my_onboarding_task_done", methods=["POST"])
@employee_required
def my_onboarding_task_done():
    emp_id = session.get("employee_id")
    db = get_db_connection(); cursor = db.cursor()
    task_id      = request.form.get("task_id")
    ob_id        = request.form.get("ob_id")
    employee_note = request.form.get("employee_note", "").strip()[:500]

    cursor.execute("SELECT employee_id, requires_document FROM employee_onboarding_tasks WHERE id=%s", (task_id,))
    row = cursor.fetchone()
    if not row or row[0] != emp_id:
        flash("Not authorised.", "error")
        cursor.close(); db.close()
        return redirect("/my_onboarding")

    doc_path = None
    if 'document' in request.files:
        f = request.files['document']
        if f and f.filename:
            import os as _os
            upload_dir = _os.path.join("static", "onboarding_docs")
            _os.makedirs(upload_dir, exist_ok=True)
            safe_name = f"{emp_id}_{task_id}_{f.filename.replace(' ','_')}"
            f.save(_os.path.join(upload_dir, safe_name))
            doc_path = safe_name

    update_args = [datetime.datetime.now(), task_id]
    if doc_path:
        cursor.execute("UPDATE employee_onboarding_tasks SET status='Done', completed_at=%s, document_path=%s, employee_note=%s WHERE id=%s",
                       (datetime.datetime.now(), doc_path, employee_note or None, task_id))
    else:
        cursor.execute("UPDATE employee_onboarding_tasks SET status='Done', completed_at=%s, employee_note=%s WHERE id=%s",
                       (datetime.datetime.now(), employee_note or None, task_id))

    # Auto-complete if all done
    cursor.execute("SELECT COUNT(*) FROM employee_onboarding_tasks WHERE onboarding_id=%s AND status!='Done'", (ob_id,))
    remaining = cursor.fetchone()[0]
    if remaining == 0:
        cursor.execute("UPDATE employee_onboarding SET status='Completed' WHERE id=%s", (ob_id,))
    db.commit()

    # Email admin about task completion
    try:
        cursor.execute("SELECT task_title FROM employee_onboarding_tasks WHERE id=%s", (task_id,))
        _tt = cursor.fetchone()
        task_title = _tt[0] if _tt else "Task"
        cursor.execute("SELECT name FROM employees WHERE employee_id=%s", (emp_id,))
        _en = cursor.fetchone(); emp_name_ob = _en[0] if _en else emp_id
        _ecfg = get_email_config()
        admin_email = _ecfg.get("from_email") if _ecfg else None
        if admin_email and _ecfg:
            _msg = (f"<p><strong>{emp_name_ob}</strong> has completed the onboarding task:</p>"
                    f"<p style='background:#f0fdf4;padding:10px;border-radius:8px;'><strong>{task_title}</strong></p>")
            if remaining == 0:
                _msg += "<p style='color:#16a34a;font-weight:700;'>🎉 All tasks completed — onboarding marked as Complete!</p>"
            else:
                _msg += f"<p>{remaining} task(s) remaining.</p>"
            send_email_async(admin_email, f"Onboarding Task Done — {emp_name_ob}", _msg, _ecfg)
    except Exception:
        pass

    cursor.close(); db.close()
    flash("Task marked as done!", "success")
    return redirect(f"/my_onboarding?ob_id={ob_id}")


# ---------------- ADMIN TOOLS (Org Chart + Audit Logs combined) ----------------
@app.route("/org_chart")
@admin_required
def org_chart_page():
    db = get_db_connection(); cursor = db.cursor(buffered=True)
    active_cid = session.get("active_company_id")
    _co_sub = "AND employee_id IN (SELECT employee_id FROM employees WHERE company_id=%s)" if active_cid else ""
    _co_args = (active_cid,) if active_cid else ()
    cursor.execute(f"SELECT COUNT(*) FROM leave_requests WHERE status='Pending' {_co_sub}", _co_args)
    pending_leaves = cursor.fetchone()[0]
    cursor.execute(f"SELECT COUNT(*) FROM resignation_requests WHERE status='Pending' {_co_sub}", _co_args)
    pending_resignations = cursor.fetchone()[0]
    cursor.execute(f"SELECT COUNT(*) FROM tickets WHERE status='Open' {_co_sub}", _co_args)
    pending_tickets = cursor.fetchone()[0]
    if active_cid:
        cursor.execute("SELECT DISTINCT department FROM employees WHERE department IS NOT NULL AND department != '' AND company_id=%s ORDER BY department", (active_cid,))
    else:
        cursor.execute("SELECT DISTINCT department FROM employees WHERE department IS NOT NULL AND department != '' ORDER BY department")
    departments = [r[0] for r in cursor.fetchall()]
    co = get_company_settings()
    cursor.close(); db.close()
    return render_template("org_chart.html",
        co=co, departments=departments,
        pending_leaves=pending_leaves,
        pending_resignations=pending_resignations,
        pending_tickets=pending_tickets,
    
        active_nav="admin_tools",
    )

@app.route("/audit_logs")
def audit_logs_redirect():
    return redirect("/admin_tools?tab=audit_logs")

@app.route("/admin_tools")
@admin_required
def admin_tools():
    tab = request.args.get("tab", "org_chart")
    db = get_db_connection(); cursor = db.cursor(buffered=True)

    active_cid = session.get("active_company_id")
    _co_sub    = "AND employee_id IN (SELECT employee_id FROM employees WHERE company_id=%s)" if active_cid else ""
    _co_args   = (active_cid,) if active_cid else ()
    _co_emp    = "AND company_id=%s" if active_cid else ""

    cursor.execute(f"SELECT COUNT(*) FROM leave_requests WHERE status='Pending' {_co_sub}", _co_args)
    pending_leaves = cursor.fetchone()[0]
    cursor.execute(f"SELECT COUNT(*) FROM resignation_requests WHERE status='Pending' {_co_sub}", _co_args)
    pending_resignations = cursor.fetchone()[0]
    cursor.execute(f"SELECT COUNT(*) FROM tickets WHERE status='Open' {_co_sub}", _co_args)
    pending_tickets = cursor.fetchone()[0]

    if active_cid:
        cursor.execute("SELECT DISTINCT department FROM employees WHERE department IS NOT NULL AND department != '' AND company_id=%s ORDER BY department", (active_cid,))
    else:
        cursor.execute("SELECT DISTINCT department FROM employees WHERE department IS NOT NULL AND department != '' ORDER BY department")
    departments = [r[0] for r in cursor.fetchall()]

    # Audit logs — filter by employees of the active company when set
    actor_f  = request.args.get("actor", "").strip()
    action_f = request.args.get("action", "").strip()
    date_f   = request.args.get("date", "").strip()
    page     = max(1, int(request.args.get("page", 1)))
    per_page = 50
    conditions, params = [], []
    if actor_f:  conditions.append("actor LIKE %s"); params.append(f"%{actor_f}%")
    if action_f: conditions.append("action LIKE %s"); params.append(f"%{action_f}%")
    if date_f:   conditions.append("DATE(created_at) = %s"); params.append(date_f)
    if active_cid:
        # Show logs where the target_id is an employee of the active company,
        # OR the actor is an employee of the active company, OR it's an admin action
        conditions.append(
            "(target_id IN (SELECT employee_id FROM employees WHERE company_id=%s) "
            "OR actor IN (SELECT employee_id FROM employees WHERE company_id=%s) "
            "OR actor_type='admin')"
        )
        params += [active_cid, active_cid]
    where = ("WHERE " + " AND ".join(conditions)) if conditions else ""
    cursor.execute(f"SELECT COUNT(*) FROM audit_logs {where}", params)
    total = cursor.fetchone()[0]
    total_pages = max(1, (total + per_page - 1) // per_page)
    offset = (page - 1) * per_page
    cursor.execute(
        f"""SELECT id, actor, actor_type, action, target_table, target_id,
                   detail, ip_address, created_at
            FROM audit_logs {where} ORDER BY created_at DESC LIMIT %s OFFSET %s""",
        params + [per_page, offset]
    )
    logs = cursor.fetchall()
    if active_cid:
        cursor.execute(
            "SELECT DISTINCT actor FROM audit_logs WHERE actor IN "
            "(SELECT employee_id FROM employees WHERE company_id=%s) OR actor_type='admin' ORDER BY actor LIMIT 200",
            (active_cid,)
        )
    else:
        cursor.execute("SELECT DISTINCT actor FROM audit_logs ORDER BY actor LIMIT 200")
    actors = [r[0] for r in cursor.fetchall()]

    co = get_company_settings()
    cursor.close(); db.close()
    return render_template("admin_tools.html",
        co=co, tab=tab, departments=departments,
        logs=logs, total=total, page=page, total_pages=total_pages,
        actor_f=actor_f, action_f=action_f, date_f=date_f, actors=actors,
        pending_leaves=pending_leaves, pending_resignations=pending_resignations,
        pending_tickets=pending_tickets,
    
        active_nav="admin_tools",
    )


# old standalone routes kept for API


@app.route("/api/org_chart_data")
@admin_required
def api_org_chart_data():
    dept_filter = request.args.get("dept", "")
    active_cid  = session.get("active_company_id")
    db = get_db_connection(); cursor = db.cursor()
    query = """
        SELECT e.employee_id, e.name, e.role, e.department,
               e.manager_id, e.face_image,
               COALESCE(e.manager_name, '') as manager_name
        FROM employees e
        WHERE COALESCE(e.is_active, 1) = 1
    """
    params = []
    if active_cid:
        query += " AND e.company_id = %s"
        params.append(active_cid)
    if dept_filter:
        query += " AND e.department = %s"
        params.append(dept_filter)
    query += " ORDER BY e.name"
    cursor.execute(query, params)
    rows = cursor.fetchall()
    cursor.close(); db.close()

    emp_map = {}
    for r in rows:
        emp_map[r[0]] = {
            "id":         r[0],
            "name":       r[1],
            "role":       r[2] or "Employee",
            "department": r[3] or "",
            "manager_id": r[4],
            "has_photo":  bool(r[5] and os.path.exists(r[5])),
            "children":   []
        }

    roots = []
    for emp in emp_map.values():
        mid = emp["manager_id"]
        if mid and mid in emp_map and mid != emp["id"]:
            emp_map[mid]["children"].append(emp)
        else:
            roots.append(emp)

    # Sort children alphabetically
    def sort_tree(node):
        node["children"].sort(key=lambda x: x["name"])
        for child in node["children"]:
            sort_tree(child)
        return node

    roots.sort(key=lambda x: x["name"])
    tree = [sort_tree(r) for r in roots]
    return jsonify({"ok": True, "tree": tree, "total": len(emp_map)})


# ── /api/v1/ aliases ──────────────────────────────────────────────────────────
# Register every /api/<path> route also under /api/v1/<path>.  Existing mobile
# clients keep using /api/ with no changes; new integrations can start on v1.
# The view functions (and their decorators: @limiter, @api_required, etc.) are
# shared, so rate-limits and auth are identical on both prefixes.
def _register_api_v1_aliases():
    _seen = set()
    for _rule in list(app.url_map.iter_rules()):
        if not _rule.rule.startswith("/api/") or _rule.rule.startswith("/api/v"):
            continue
        _v1_rule = "/api/v1" + _rule.rule[4:]
        _vf      = app.view_functions.get(_rule.endpoint)
        if _vf is None:
            continue
        _ep_v1 = "v1_" + _rule.endpoint
        if _ep_v1 in _seen:
            continue
        _seen.add(_ep_v1)
        app.add_url_rule(
            _v1_rule,
            endpoint=_ep_v1,
            view_func=_vf,
            methods=_rule.methods,
        )

_register_api_v1_aliases()

# ---------------- RUN ----------------
if __name__ == "__main__":
    init_master_db()
    init_db()
    load_default_shift()
    load_salary_rules()
    import os as _os
    _cert = _os.environ.get("SSL_CERT_PATH") or _os.path.join(_os.path.dirname(__file__), "cert.pem")
    _key  = _os.environ.get("SSL_KEY_PATH") or _os.path.join(_os.path.dirname(__file__), "key.pem")
    if _os.path.exists(_cert) and _os.path.exists(_key):
        print("🔒  SSL cert found — starting on https://0.0.0.0:5000")
        app.run(host='0.0.0.0', port=5000, debug=False, use_reloader=False,
                ssl_context=(_cert, _key))
    else:
        print("⚠   No cert.pem / key.pem — starting on http://0.0.0.0:5000")
        print("    Fingerprint / WebAuthn requires HTTPS. Run: python generate_cert.py")
        app.run(host='0.0.0.0', port=5000, debug=False, use_reloader=False)
