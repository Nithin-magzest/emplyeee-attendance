"""Performance blueprint — reviews, KPIs. Hike/bonus live in blueprints/payroll.py."""
import datetime
from flask import Blueprint, request, session, redirect, render_template, flash
from database import get_db_connection
from utils.auth import admin_required, employee_required

performance_bp = Blueprint("performance", __name__)

RATING_LABELS = {0: "Not Rated", 1: "Unsatisfactory", 2: "Needs Improvement",
                 3: "Meets Expectations", 4: "Exceeds Expectations", 5: "Outstanding"}

@performance_bp.route("/performance")
@admin_required
def performance():
    today  = datetime.date.today()
    q      = int(request.args.get("quarter", (today.month - 1) // 3 + 1))
    yr     = int(request.args.get("year", today.year))
    dept   = request.args.get("dept", "")
    active_tab = request.args.get("tab", "performance")

    db = get_db_connection()
    cursor = db.cursor(buffered=True)

    active_cid = session.get("active_company_id")
    dept_filter = "AND e.department=%s" if dept else ""
    co_filter   = "AND e.company_id=%s" if active_cid else ""
    params = [yr, q] + ([dept] if dept else []) + ([active_cid] if active_cid else [])
    cursor.execute(f"""
        SELECT e.employee_id, e.name, COALESCE(e.role,''), COALESCE(e.department,''),
               pr.id, COALESCE(pr.overall_rating,0), COALESCE(pr.status,'—'),
               (SELECT COUNT(*) FROM performance_kpis pk WHERE pk.review_id=pr.id) AS kpi_count
        FROM employees e
        LEFT JOIN performance_reviews pr
            ON pr.employee_id=e.employee_id AND pr.year=%s AND pr.quarter=%s
        WHERE e.is_active=1 {dept_filter} {co_filter}
        ORDER BY e.name
    """, params)
    employees = cursor.fetchall()

    if active_cid:
        cursor.execute("SELECT department FROM employees WHERE is_active=1 AND department IS NOT NULL AND department!='' AND company_id=%s GROUP BY department ORDER BY MIN(id) ASC", (active_cid,))
    else:
        cursor.execute("SELECT department FROM employees WHERE is_active=1 AND department IS NOT NULL AND department!='' GROUP BY department ORDER BY MIN(id) ASC")
    departments = [r[0] for r in cursor.fetchall()]

    # Announcements (admin sees all)
    cursor.execute("""
        SELECT a.id, a.title, a.content, a.priority, a.created_at,
               COALESCE(a.visibility,'public'), COALESCE(a.target_employee_id,''), COALESCE(e.name,'')
        FROM announcements a
        LEFT JOIN employees e ON e.employee_id = a.target_employee_id
        ORDER BY a.created_at DESC
    """)
    ann_list  = cursor.fetchall()
    pub_anns  = [r for r in ann_list if r[5] == 'public']
    priv_anns = [r for r in ann_list if r[5] == 'private']

    cursor.execute("SELECT employee_id, name FROM employees WHERE is_active=1 ORDER BY name")
    ann_emp_list = cursor.fetchall()

    cursor.execute("SELECT COUNT(*) FROM leave_requests WHERE status='Pending'")
    pending_leaves = cursor.fetchone()[0]
    cursor.execute("SELECT COUNT(*) FROM resignation_requests WHERE status='Pending'")
    pending_resignations = cursor.fetchone()[0]
    cursor.execute("SELECT COUNT(*) FROM tickets WHERE status IN ('Open','In Progress')")
    pending_tickets = cursor.fetchone()[0]
    cursor.execute("SELECT COALESCE(company_name,'') FROM company_settings LIMIT 1")
    co = cursor.fetchone()

    cursor.execute("SELECT id, label, min_rating, max_rating, hike_pct, incentive_pct, color FROM hike_config ORDER BY min_rating DESC")
    hike_bands = cursor.fetchall()

    hike_employees = []
    total_hike_cost = 0.0
    total_bonus_pool = 0.0
    hike_eligible_count = 0
    if active_tab == 'hike':
        _hike_co = "AND e.company_id=%s" if active_cid else ""
        _hike_params = (yr, q) + ((active_cid,) if active_cid else ())
        cursor.execute(f"""
            SELECT e.employee_id, e.name, COALESCE(e.role,''), COALESCE(e.department,''),
                   COALESCE(pr.overall_rating,0), COALESCE(pr.status,'—'),
                   COALESCE(sc.monthly_ctc,0)
            FROM employees e
            LEFT JOIN performance_reviews pr ON pr.employee_id=e.employee_id AND pr.year=%s AND pr.quarter=%s
            LEFT JOIN salary_config sc ON sc.employee_id=e.employee_id
            WHERE e.is_active=1 {_hike_co}
            ORDER BY e.name
        """, _hike_params)
        for (h_eid, h_name, h_role, h_dept, h_rating, h_status, h_ctc) in cursor.fetchall():
            h_rating = float(h_rating or 0)
            h_ctc    = float(h_ctc or 0)
            band_label, band_color, hike_pct, inc_pct = "Not Rated", "#94a3b8", 0.0, 0.0
            if h_rating > 0:
                for (_, blabel, bmin, bmax, bhike, binc, bcolor) in hike_bands:
                    if float(bmin) <= h_rating <= float(bmax):
                        band_label, band_color, hike_pct, inc_pct = blabel, bcolor, float(bhike), float(binc)
                        break
                hike_eligible_count += 1
            new_ctc = round(h_ctc * (1 + hike_pct / 100), 2) if h_ctc > 0 and hike_pct > 0 else h_ctc
            bonus   = round(h_ctc * inc_pct / 100, 2) if h_ctc > 0 and inc_pct > 0 else 0.0
            total_hike_cost  += max(0, new_ctc - h_ctc)
            total_bonus_pool += bonus
            hike_employees.append((h_eid, h_name, h_role, h_dept, h_rating, h_status,
                                   h_ctc, band_label, band_color, hike_pct, new_ctc, inc_pct, bonus))

    cursor.close(); db.close()

    return render_template("performance.html",
        employees=employees, departments=departments,
        quarter=q, year=yr, selected_dept=dept,
        rating_labels=RATING_LABELS,
        pending_leaves=pending_leaves,
        pending_resignations=pending_resignations,
        pending_tickets=pending_tickets, co=co,
        today=today,
        ann_list=ann_list,
        pub_anns=pub_anns,
        priv_anns=priv_anns,
        ann_emp_list=ann_emp_list,
        active_tab=active_tab,
        hike_bands=hike_bands,
        hike_employees=hike_employees,
        total_hike_cost=total_hike_cost,
        total_bonus_pool=total_bonus_pool,
        hike_eligible_count=hike_eligible_count,
    )

@performance_bp.route("/performance_review/<emp_id>", methods=["GET"])
@admin_required
def performance_review(emp_id):
    today = datetime.date.today()
    q     = int(request.args.get("quarter", (today.month - 1) // 3 + 1))
    yr    = int(request.args.get("year", today.year))

    db = get_db_connection()
    cursor = db.cursor(buffered=True)

    cursor.execute("""
        SELECT e.employee_id, e.name, COALESCE(e.role,''), COALESCE(e.department,''),
               COALESCE(e.email,''), COALESCE(e.phone,'')
        FROM employees e WHERE e.employee_id=%s
    """, (emp_id,))
    emp = cursor.fetchone()
    if not emp:
        cursor.close(); db.close()
        flash("Employee not found.", "error")
        return redirect("/performance")

    # Get or create review
    cursor.execute("""
        SELECT id, overall_rating, reviewer_feedback, employee_comment, status
        FROM performance_reviews WHERE employee_id=%s AND quarter=%s AND year=%s
    """, (emp_id, q, yr))
    review = cursor.fetchone()

    kpis = []
    if review:
        cursor.execute("""
            SELECT id, kpi_title, description, target, achievement, weight, rating, comments
            FROM performance_kpis WHERE review_id=%s ORDER BY id
        """, (review[0],))
        kpis = cursor.fetchall()

    # Past reviews for history tab
    cursor.execute("""
        SELECT id, quarter, year, overall_rating, status, created_at
        FROM performance_reviews WHERE employee_id=%s ORDER BY year DESC, quarter DESC LIMIT 8
    """, (emp_id,))
    history = cursor.fetchall()

    cursor.execute("SELECT COUNT(*) FROM leave_requests WHERE status='Pending'")
    pending_leaves = cursor.fetchone()[0]
    cursor.execute("SELECT COUNT(*) FROM resignation_requests WHERE status='Pending'")
    pending_resignations = cursor.fetchone()[0]
    cursor.execute("SELECT COUNT(*) FROM tickets WHERE status IN ('Open','In Progress')")
    pending_tickets = cursor.fetchone()[0]
    cursor.execute("SELECT COALESCE(company_name,'') FROM company_settings LIMIT 1")
    co = cursor.fetchone()
    cursor.close(); db.close()

    return render_template("performance_review.html",
        emp=emp, review=review, kpis=kpis, history=history,
        quarter=q, year=yr, rating_labels=RATING_LABELS,
        pending_leaves=pending_leaves,
        pending_resignations=pending_resignations,
        pending_tickets=pending_tickets, co=co
    )

@performance_bp.route("/performance_save_review", methods=["POST"])
@admin_required
def performance_save_review():
    emp_id   = request.form["employee_id"]
    q        = int(request.form["quarter"])
    yr       = int(request.form["year"])
    feedback = request.form.get("reviewer_feedback", "").strip()
    status   = request.form.get("status", "Draft")

    db = get_db_connection()
    cursor = db.cursor(buffered=True)
    cursor.execute("""
        INSERT INTO performance_reviews (employee_id, quarter, year, reviewer_feedback, status)
        VALUES (%s,%s,%s,%s,%s)
        ON CONFLICT (employee_id, quarter, year) DO UPDATE SET reviewer_feedback=%s, status=%s, updated_at=NOW()
    """, (emp_id, q, yr, feedback, status, feedback, status))
    db.commit()

    # Recalculate overall rating from KPIs
    cursor.execute("SELECT id FROM performance_reviews WHERE employee_id=%s AND quarter=%s AND year=%s", (emp_id, q, yr))
    rev = cursor.fetchone()
    if rev:
        cursor.execute("""
            SELECT weight, rating FROM performance_kpis WHERE review_id=%s AND rating > 0
        """, (rev[0],))
        kpi_rows = cursor.fetchall()
        if kpi_rows:
            total_weight = sum(r[0] for r in kpi_rows)
            weighted_sum = sum(r[0] * r[1] for r in kpi_rows)
            overall = round(weighted_sum / total_weight, 1) if total_weight > 0 else 0
            cursor.execute("UPDATE performance_reviews SET overall_rating=%s WHERE id=%s", (overall, rev[0]))
            db.commit()

    cursor.close(); db.close()
    flash("Review saved successfully.", "success")
    return redirect(f"/performance_review/{emp_id}?quarter={q}&year={yr}")

@performance_bp.route("/performance_add_kpi", methods=["POST"])
@admin_required
def performance_add_kpi():
    emp_id = request.form["employee_id"]
    q      = int(request.form["quarter"])
    yr     = int(request.form["year"])
    title  = request.form.get("kpi_title", "").strip()
    desc   = request.form.get("description", "").strip()
    target = request.form.get("target", "").strip()
    weight = int(request.form.get("weight", 20))

    if not title:
        flash("KPI title is required.", "error")
        return redirect(f"/performance_review/{emp_id}?quarter={q}&year={yr}")

    db = get_db_connection()
    cursor = db.cursor(buffered=True)

    # Ensure review exists
    cursor.execute("""
        INSERT INTO performance_reviews (employee_id, quarter, year, status)
        VALUES (%s,%s,%s,'Draft')
        ON CONFLICT (employee_id, quarter, year) DO UPDATE SET updated_at=NOW()
    """, (emp_id, q, yr))
    db.commit()

    cursor.execute("SELECT id FROM performance_reviews WHERE employee_id=%s AND quarter=%s AND year=%s", (emp_id, q, yr))
    rev_id = cursor.fetchone()[0]

    cursor.execute("""
        INSERT INTO performance_kpis (review_id, kpi_title, description, target, weight)
        VALUES (%s,%s,%s,%s,%s)
    """, (rev_id, title, desc, target, weight))
    db.commit()
    cursor.close(); db.close()
    flash("KPI added.", "success")
    return redirect(f"/performance_review/{emp_id}?quarter={q}&year={yr}")

@performance_bp.route("/performance_rate_kpi", methods=["POST"])
@admin_required
def performance_rate_kpi():
    kpi_id      = int(request.form["kpi_id"])
    emp_id      = request.form["employee_id"]
    q           = int(request.form["quarter"])
    yr          = int(request.form["year"])
    rating      = int(request.form.get("rating", 0))
    achievement = request.form.get("achievement", "").strip()
    comments    = request.form.get("comments", "").strip()

    db = get_db_connection()
    cursor = db.cursor(buffered=True)
    cursor.execute("""
        UPDATE performance_kpis SET rating=%s, achievement=%s, comments=%s WHERE id=%s
    """, (rating, achievement, comments, kpi_id))
    db.commit()

    # Recalculate overall rating
    cursor.execute("SELECT id FROM performance_reviews WHERE employee_id=%s AND quarter=%s AND year=%s", (emp_id, q, yr))
    rev = cursor.fetchone()
    if rev:
        cursor.execute("SELECT weight, rating FROM performance_kpis WHERE review_id=%s AND rating>0", (rev[0],))
        rows = cursor.fetchall()
        if rows:
            tw = sum(r[0] for r in rows); ws = sum(r[0]*r[1] for r in rows)
            cursor.execute("UPDATE performance_reviews SET overall_rating=%s WHERE id=%s",
                           (round(ws/tw, 1) if tw else 0, rev[0]))
            db.commit()

    cursor.close(); db.close()
    return redirect(f"/performance_review/{emp_id}?quarter={q}&year={yr}")

@performance_bp.route("/performance_delete_kpi", methods=["POST"])
@admin_required
def performance_delete_kpi():
    kpi_id = int(request.form["kpi_id"])
    emp_id = request.form["employee_id"]
    q      = int(request.form["quarter"])
    yr     = int(request.form["year"])
    db = get_db_connection()
    cursor = db.cursor(buffered=True)
    cursor.execute("DELETE FROM performance_kpis WHERE id=%s", (kpi_id,))
    db.commit()
    cursor.close(); db.close()
    return redirect(f"/performance_review/{emp_id}?quarter={q}&year={yr}")

@performance_bp.route("/my_performance")
@employee_required
def my_performance():
    emp_id = session["employee_id"]
    db = get_db_connection()
    cursor = db.cursor(buffered=True)

    cursor.execute("""
        SELECT pr.id, pr.quarter, pr.year, pr.overall_rating, pr.reviewer_feedback,
               pr.employee_comment, pr.status, pr.updated_at
        FROM performance_reviews pr
        WHERE pr.employee_id=%s ORDER BY pr.year DESC, pr.quarter DESC
    """, (emp_id,))
    reviews = cursor.fetchall()

    reviews_data = []
    for rev in reviews:
        cursor.execute("""
            SELECT kpi_title, target, achievement, weight, rating, comments
            FROM performance_kpis WHERE review_id=%s ORDER BY id
        """, (rev[0],))
        kpis = cursor.fetchall()
        reviews_data.append({"review": rev, "kpis": kpis})

    cursor.execute("SELECT name, COALESCE(role,''), COALESCE(department,''), face_image FROM employees WHERE employee_id=%s", (emp_id,))
    emp_info = cursor.fetchone()
    cursor.close(); db.close()

    return render_template("my_performance.html",
        reviews_data=reviews_data, emp_info=emp_info,
        emp_id=emp_id, rating_labels=RATING_LABELS
    )

@performance_bp.route("/performance_employee_comment", methods=["POST"])
@employee_required
def performance_employee_comment():
    rev_id  = int(request.form["review_id"])
    comment = request.form.get("comment", "").strip()
    emp_id  = session["employee_id"]
    db = get_db_connection()
    cursor = db.cursor(buffered=True)
    # Only allow comment on own review
    cursor.execute("UPDATE performance_reviews SET employee_comment=%s WHERE id=%s AND employee_id=%s",
                   (comment, rev_id, emp_id))
    db.commit()
    cursor.close(); db.close()
    flash("Comment submitted.", "success")
    return redirect("/my_performance")

@performance_bp.route("/performance_export")
@admin_required
def performance_export():
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    today = datetime.date.today()
    q  = int(request.args.get("quarter", (today.month - 1) // 3 + 1))
    yr = int(request.args.get("year", today.year))

    db = get_db_connection()
    cursor = db.cursor(buffered=True)

    cursor.execute("""
        SELECT e.employee_id, e.name, COALESCE(e.role,''), COALESCE(e.department,''),
               COALESCE(pr.overall_rating,0), COALESCE(pr.status,'Not Started'),
               COALESCE(pr.reviewer_feedback,''), COALESCE(pr.employee_comment,''),
               pr.id
        FROM employees e
        LEFT JOIN performance_reviews pr
            ON pr.employee_id=e.employee_id AND pr.year=%s AND pr.quarter=%s
        WHERE e.is_active=1
        ORDER BY e.name
    """, (yr, q))
    employees = cursor.fetchall()

    cursor.execute("""
        SELECT e.employee_id, e.name, pk.kpi_title, COALESCE(pk.description,''),
               COALESCE(pk.target,''), COALESCE(pk.achievement,''),
               pk.weight, COALESCE(pk.rating,0), COALESCE(pk.comments,'')
        FROM employees e
        JOIN performance_reviews pr ON pr.employee_id=e.employee_id AND pr.year=%s AND pr.quarter=%s
        JOIN performance_kpis pk ON pk.review_id=pr.id
        WHERE e.is_active=1
        ORDER BY e.name, pk.id
    """, (yr, q))
    kpis = cursor.fetchall()
    cursor.close(); db.close()

    wb = openpyxl.Workbook()

    # ── Styles ──
    hdr_font   = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill   = PatternFill("solid", fgColor="1E3A8A")
    hdr_align  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    alt_fill   = PatternFill("solid", fgColor="EFF6FF")
    thin       = Side(style="thin", color="BFDBFE")
    border     = Border(left=thin, right=thin, top=thin, bottom=thin)
    center     = Alignment(horizontal="center", vertical="center")

    def style_header(ws, cols):
        for col_idx, (title, width) in enumerate(cols, 1):
            cell = ws.cell(row=1, column=col_idx, value=title)
            cell.font      = hdr_font
            cell.fill      = hdr_fill
            cell.alignment = hdr_align
            cell.border    = border
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        ws.row_dimensions[1].height = 30

    def style_data_cell(cell, row_idx):
        cell.border    = border
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        if row_idx % 2 == 0:
            cell.fill = alt_fill

    # ── Sheet 1: Summary ──
    ws1 = wb.active
    ws1.title = f"Q{q} {yr} Summary"
    q_labels = {1:"Jan–Mar", 2:"Apr–Jun", 3:"Jul–Sep", 4:"Oct–Dec"}
    ws1.append([])
    ws1.merge_cells("A1:H1")
    title_cell = ws1["A1"]
    title_cell.value     = f"Performance Summary — Q{q} ({q_labels.get(q,'')}) {yr}"
    title_cell.font      = Font(bold=True, size=14, color="1E3A8A")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 36

    cols_s = [
        ("Employee ID", 16), ("Employee Name", 24), ("Role", 18), ("Department", 18),
        ("KPI Count", 12), ("Overall Rating (/ 5)", 20), ("Status", 16), ("Reviewer Feedback", 35),
    ]
    for col_idx, (title, width) in enumerate(cols_s, 1):
        cell = ws1.cell(row=2, column=col_idx, value=title)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = hdr_align
        cell.border    = border
        ws1.column_dimensions[get_column_letter(col_idx)].width = width
    ws1.row_dimensions[2].height = 28

    kpi_counts = {}
    for row in kpis:
        kpi_counts[row[0]] = kpi_counts.get(row[0], 0) + 1

    for r_idx, (emp_id, name, role, dept, rating, status, feedback, _, _rev_id) in enumerate(employees, 3):
        row_data = [emp_id, name, role, dept, kpi_counts.get(emp_id, 0),
                    rating if rating else "—", status, feedback]
        for c_idx, val in enumerate(row_data, 1):
            cell = ws1.cell(row=r_idx, column=c_idx, value=val)
            style_data_cell(cell, r_idx)
            if c_idx == 6 and isinstance(val, (int, float)) and val > 0:
                if val >= 4:   cell.font = Font(color="15803D", bold=True)
                elif val >= 3: cell.font = Font(color="1D4ED8", bold=True)
                else:           cell.font = Font(color="DC2626", bold=True)
        ws1.row_dimensions[r_idx].height = 22
    ws1.freeze_panes = "A3"

    # ── Sheet 2: KPI Details ──
    ws2 = wb.create_sheet(f"Q{q} {yr} KPI Details")
    cols_k = [
        ("Employee ID", 16), ("Employee Name", 22), ("KPI Title", 28),
        ("Description", 30), ("Target", 20), ("Achievement", 20),
        ("Weight (%)", 13), ("Rating (1–5)", 14), ("Comments", 30),
    ]
    for col_idx, (title, width) in enumerate(cols_k, 1):
        cell = ws2.cell(row=1, column=col_idx, value=title)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = hdr_align
        cell.border    = border
        ws2.column_dimensions[get_column_letter(col_idx)].width = width
    ws2.row_dimensions[1].height = 28

    for r_idx, (emp_id, name, title, desc, target, achievement, weight, rating, comments) in enumerate(kpis, 2):
        for c_idx, val in enumerate([emp_id, name, title, desc, target, achievement, weight, rating or "—", comments], 1):
            cell = ws2.cell(row=r_idx, column=c_idx, value=val)
            style_data_cell(cell, r_idx)
        ws2.row_dimensions[r_idx].height = 20
    ws2.freeze_panes = "A2"

    # ── Sheet 3: Import Template ──
    ws3 = wb.create_sheet("Import Template")
    note_fill = PatternFill("solid", fgColor="FFF9C4")
    note_font = Font(italic=True, size=10, color="92400E")
    ws3.merge_cells("A1:J1")
    n = ws3["A1"]
    n.value     = "Fill in the rows below and import this file. Required columns: employee_id, kpi_title, weight, rating. Quarter & Year are selected in the import dialog."
    n.font      = note_font
    n.fill      = note_fill
    n.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws3.row_dimensions[1].height = 40

    tpl_cols = [
        ("employee_id*", 18), ("kpi_title*", 28), ("description", 28),
        ("target", 20), ("achievement", 20), ("weight*", 12),
        ("rating* (1-5)", 14), ("comments", 28), ("status", 16), ("reviewer_feedback", 35),
    ]
    for col_idx, (title, width) in enumerate(tpl_cols, 1):
        cell = ws3.cell(row=2, column=col_idx, value=title)
        req_fill = PatternFill("solid", fgColor="1E3A8A") if title.endswith("*") else PatternFill("solid", fgColor="475569")
        cell.font      = Font(bold=True, color="FFFFFF", size=10)
        cell.fill      = req_fill
        cell.alignment = hdr_align
        cell.border    = border
        ws3.column_dimensions[get_column_letter(col_idx)].width = width
    ws3.row_dimensions[2].height = 28

    sample_rows = [
        ["EMP001", "Code Quality", "Maintain clean, tested code", "95% coverage", "92%", 30, 4, "Good progress", "Submitted", ""],
        ["EMP001", "Delivery Speed", "Complete tasks on time", "95% on-time", "90%", 30, 3, "", "", ""],
        ["EMP001", "Team Collaboration", "Cross-team work", "4 collabs/qtr", "", 20, 5, "Excellent", "", ""],
        ["EMP001", "Documentation", "Keep docs updated", "100% coverage", "80%", 20, 3, "", "", "Strong Q2 performance"],
        ["EMP002", "Sales Target", "Hit monthly targets", "₹5L / month", "₹4.8L", 50, 4, "Near target", "Draft", ""],
        ["EMP002", "Customer Satisfaction", "Maintain CSAT score", ">=4.5 / 5", "4.3", 50, 3, "Needs improvement", "", ""],
    ]
    for r_idx, row in enumerate(sample_rows, 3):
        for c_idx, val in enumerate(row, 1):
            cell = ws3.cell(row=r_idx, column=c_idx, value=val)
            cell.border    = border
            cell.alignment = Alignment(vertical="center")
            if r_idx % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="F0F9FF")
        ws3.row_dimensions[r_idx].height = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"performance_Q{q}_{yr}.xlsx"
    from flask import send_file
    return send_file(buf, as_attachment=True, download_name=filename,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@performance_bp.route("/performance_import", methods=["POST"])
@admin_required
def performance_import():
    import io
    import openpyxl

    q_raw  = request.form.get("quarter", "").strip()
    yr_raw = request.form.get("year", "").strip()
    if not q_raw.isdigit() or not yr_raw.isdigit():
        flash("Invalid quarter or year.", "error")
        return redirect("/performance")
    q  = int(q_raw)
    yr = int(yr_raw)

    f = request.files.get("excel_file")
    if not f or not f.filename.endswith((".xlsx", ".xls")):
        flash("Please upload a valid Excel file (.xlsx or .xls).", "error")
        return redirect(f"/performance?tab=performance&quarter={q}&year={yr}")

    try:
        wb = openpyxl.load_workbook(io.BytesIO(f.read()), data_only=True)
    except Exception:
        flash("Could not read the Excel file. Make sure it is a valid .xlsx file.", "error")
        return redirect(f"/performance?tab=performance&quarter={q}&year={yr}")

    # Find the data sheet — use first sheet that isn't "Import Template"
    sheet = None
    for ws in wb.worksheets:
        if ws.title != "Import Template":
            sheet = ws
            break
    if sheet is None:
        sheet = wb.active

    rows = list(sheet.iter_rows(values_only=True))
    if len(rows) < 2:
        flash("The Excel file has no data rows.", "error")
        return redirect(f"/performance?tab=performance&quarter={q}&year={yr}")

    # Find header row (first row with 'employee_id' in it)
    header_row_idx = None
    headers = []
    for idx, row in enumerate(rows):
        cells = [str(c).strip().lower() if c is not None else "" for c in row]
        if "employee_id" in cells:
            header_row_idx = idx
            headers = cells
            break
    if header_row_idx is None:
        flash("Could not find header row. Make sure the file has an 'employee_id' column.", "error")
        return redirect(f"/performance?tab=performance&quarter={q}&year={yr}")

    def col(name):
        try: return headers.index(name)
        except ValueError: return None

    ci_emp      = col("employee_id")
    ci_title    = col("kpi_title")
    ci_desc     = col("description")
    ci_target   = col("target")
    ci_achieve  = col("achievement")
    ci_weight   = col("weight") or col("weight*")
    ci_rating   = next((col(x) for x in ["rating* (1-5)", "rating (1-5)", "rating"] if col(x) is not None), None)
    ci_comments = col("comments")
    ci_status   = col("status")
    ci_feedback = col("reviewer_feedback")

    if ci_emp is None or ci_title is None:
        flash("Missing required columns: 'employee_id' and 'kpi_title'.", "error")
        return redirect(f"/performance?tab=performance&quarter={q}&year={yr}")

    # Parse data rows
    data_rows = rows[header_row_idx + 1:]
    employees_data = {}  # emp_id → {feedback, status, kpis: [...]}
    skipped = 0
    for row in data_rows:
        if not any(c for c in row if c is not None):
            continue
        emp_id = str(row[ci_emp]).strip() if row[ci_emp] is not None else ""
        title  = str(row[ci_title]).strip() if ci_title is not None and row[ci_title] is not None else ""
        if not emp_id or not title:
            skipped += 1
            continue

        try:
            weight = int(row[ci_weight]) if ci_weight is not None and row[ci_weight] not in (None, "") else 20
        except (ValueError, TypeError):
            weight = 20
        weight = max(1, min(100, weight))

        try:
            rating = round(float(row[ci_rating]), 1) if ci_rating is not None and row[ci_rating] not in (None, "") else 0.0
        except (ValueError, TypeError):
            rating = 0.0
        rating = max(0.0, min(5.0, rating))

        desc     = str(row[ci_desc]).strip()     if ci_desc    is not None and row[ci_desc]    not in (None,"") else ""
        target   = str(row[ci_target]).strip()   if ci_target  is not None and row[ci_target]  not in (None,"") else ""
        achieve  = str(row[ci_achieve]).strip()  if ci_achieve is not None and row[ci_achieve] not in (None,"") else ""
        comments = str(row[ci_comments]).strip() if ci_comments is not None and row[ci_comments] not in (None,"") else ""
        status   = str(row[ci_status]).strip()   if ci_status  is not None and row[ci_status]   not in (None,"") else "Draft"
        feedback = str(row[ci_feedback]).strip() if ci_feedback is not None and row[ci_feedback] not in (None,"") else ""

        if status not in ("Draft", "Submitted", "Acknowledged"):
            status = "Draft"

        if emp_id not in employees_data:
            employees_data[emp_id] = {"status": status, "feedback": feedback, "kpis": []}
        if feedback:
            employees_data[emp_id]["feedback"] = feedback
        if status in ("Submitted", "Acknowledged"):
            employees_data[emp_id]["status"] = status

        employees_data[emp_id]["kpis"].append({
            "title": title, "description": desc, "target": target,
            "achievement": achieve, "weight": weight, "rating": rating, "comments": comments,
        })

    if not employees_data:
        flash("No valid data rows found in the file.", "error")
        return redirect(f"/performance?tab=performance&quarter={q}&year={yr}")

    db = get_db_connection()
    cursor = db.cursor(buffered=True)

    # Validate employee IDs exist
    cursor.execute("SELECT employee_id FROM employees WHERE is_active=1")
    valid_ids = {r[0] for r in cursor.fetchall()}

    imported = 0
    unknown  = []
    for emp_id, emp_data in employees_data.items():
        if emp_id not in valid_ids:
            unknown.append(emp_id)
            continue

        feedback = emp_data["feedback"]
        status   = emp_data["status"]
        kpis     = emp_data["kpis"]

        # Upsert review
        cursor.execute("""
            INSERT INTO performance_reviews (employee_id, quarter, year, reviewer_feedback, status)
            VALUES (%s,%s,%s,%s,%s)
            ON CONFLICT (employee_id, quarter, year) DO UPDATE SET
                reviewer_feedback=EXCLUDED.reviewer_feedback, status=EXCLUDED.status, updated_at=NOW()
        """, (emp_id, q, yr, feedback, status))
        db.commit()

        cursor.execute("SELECT id FROM performance_reviews WHERE employee_id=%s AND quarter=%s AND year=%s", (emp_id, q, yr))
        rev_id = cursor.fetchone()[0]

        # Replace KPIs
        cursor.execute("DELETE FROM performance_kpis WHERE review_id=%s", (rev_id,))
        for kpi in kpis:
            cursor.execute("""
                INSERT INTO performance_kpis (review_id, kpi_title, description, target, achievement, weight, rating, comments)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s)
            """, (rev_id, kpi["title"], kpi["description"], kpi["target"],
                  kpi["achievement"], kpi["weight"], kpi["rating"], kpi["comments"]))
        db.commit()

        # Recalculate overall rating
        cursor.execute("SELECT weight, rating FROM performance_kpis WHERE review_id=%s AND rating>0", (rev_id,))
        rated = cursor.fetchall()
        if rated:
            total_w   = sum(r[0] for r in rated)
            weighted  = sum(r[0] * r[1] for r in rated)
            overall   = round(weighted / total_w, 1) if total_w else 0
            cursor.execute("UPDATE performance_reviews SET overall_rating=%s WHERE id=%s", (overall, rev_id))
            db.commit()
        imported += 1

    cursor.close(); db.close()

    msg = f"✅ Imported {imported} employee(s) for Q{q} {yr}."
    if skipped:   msg += f" {skipped} row(s) skipped (missing ID or KPI title)."
    if unknown:   msg += f" Unknown employee IDs: {', '.join(unknown)}."
    flash(msg, "success")
    return redirect(f"/performance?tab=performance&quarter={q}&year={yr}")

