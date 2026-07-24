"""Payroll blueprint — salary, payslips, payroll settings, reports.

Migrated from app.py (25 routes: the 23 the manifest listed, plus two
companion POST handlers the original manifest comment collapsed into their
GET counterparts — /api/salary_config POST and /api/email_config POST are
separate functions, not the same one handling both methods). Extracted
build_salary_slip_html/compute_salary_entry into utils/salary_utils.py as
part of this move rather than importing them back from app.py, which would
have created a real circular import between this blueprint and the module
that registers it.
"""
import datetime
import calendar
import math
import io as _io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from flask import (
    Blueprint, request, session, redirect, jsonify, render_template,
    flash,
)

from database import get_db_connection
from extensions import app_log, limiter, log_security_event
from utils.auth import admin_required, employee_required, api_required, enforce_ownership, role_required, api_role_required
from utils.helpers import _audit, decrypt_pii, encrypt_pii
from utils.email_utils import get_email_config, send_email_async, send_email_smtp
from utils.attendance_utils import (
    get_working_days, fetch_holidays_set, get_billable_past_days, infer_type_legacy,
)
from utils.salary_utils import build_salary_slip_html, compute_salary_entry
import utils.config as cfg

payroll_bp = Blueprint("payroll", __name__)


@payroll_bp.route("/view_salary")
@role_required("admin")
@limiter.limit("10 per minute")
def view_salary():
    db = get_db_connection()
    cursor = db.cursor(buffered=True)
    active_cid = session.get("active_company_id")
    if active_cid:
        cursor.execute("""
            SELECT e.employee_id, e.name, COALESCE(s.salary_per_day, 0), e.role, s.last_revised,
                   COALESCE(e.phone,''), COALESCE(e.email,'')
            FROM employees e
            LEFT JOIN salary_config s ON e.employee_id = s.employee_id
            WHERE e.company_id = %s
            ORDER BY e.name
        """, (active_cid,))
    else:
        cursor.execute("""
            SELECT e.employee_id, e.name, COALESCE(s.salary_per_day, 0), e.role, s.last_revised,
                   COALESCE(e.phone,''), COALESCE(e.email,'')
            FROM employees e
            LEFT JOIN salary_config s ON e.employee_id = s.employee_id
            ORDER BY e.name
        """)
    data = cursor.fetchall()
    cursor.close()
    db.close()
    return render_template("salary.html", salaries=data, active_nav="salary")


@payroll_bp.route("/update_salary", methods=["POST"])
@role_required("admin")
@limiter.limit("20 per minute")
def update_salary():
    emp_id = request.form["emp_id"]
    salary = request.form["salary"]
    hike_date = request.form.get("hike_date") or None
    db = get_db_connection()
    cursor = db.cursor(buffered=True)
    cursor.execute("SELECT 1 FROM salary_config WHERE employee_id=%s", (emp_id,))
    if cursor.fetchone():
        cursor.execute(
            "UPDATE salary_config SET salary_per_day=%s, last_revised=%s WHERE employee_id=%s",
            (salary, hike_date or datetime.date.today(), emp_id)
        )
    else:
        cursor.execute(
            "INSERT INTO salary_config (employee_id, salary_per_day, last_revised) VALUES (%s,%s,%s)",
            (emp_id, salary, hike_date or datetime.date.today())
        )
    db.commit()
    cursor.close()
    db.close()
    _audit("update_salary", "salary_config", emp_id, f"salary_per_day set to {salary}")
    return redirect("/settings?tab=salary")

# ---------------- MONTHLY ATTENDANCE REPORT ----------------


@payroll_bp.route("/salary_report")
@role_required("admin")
@limiter.limit("10 per minute")
def salary_report():
    year = int(request.args.get("year", datetime.date.today().year))
    month = int(request.args.get("month", datetime.date.today().month))

    db = get_db_connection()
    cursor = db.cursor(buffered=True)

    active_cid = session.get("active_company_id")
    if active_cid:
        cursor.execute("""
            SELECT e.employee_id, e.name, e.email, COALESCE(s.salary_per_day, 0),
                   COALESCE(e.role,''), COALESCE(e.phone,'')
            FROM employees e
            LEFT JOIN salary_config s ON e.employee_id = s.employee_id
            WHERE e.company_id = %s
            ORDER BY e.name
        """, (active_cid,))
    else:
        cursor.execute("""
            SELECT e.employee_id, e.name, e.email, COALESCE(s.salary_per_day, 0),
                   COALESCE(e.role,''), COALESCE(e.phone,'')
            FROM employees e
            LEFT JOIN salary_config s ON e.employee_id = s.employee_id
            ORDER BY e.name
        """)
    employees = cursor.fetchall()

    _, last_day = calendar.monthrange(year, month)
    cursor.execute("""
        SELECT employee_id, date, login_time, logout_time, status, logout_status, attendance_type
        FROM attendance
        WHERE date BETWEEN %s AND %s
    """, (datetime.date(year, month, 1), datetime.date(year, month, last_day)))

    att_map = {}
    for row in cursor.fetchall():
        att_map.setdefault(row[0], {})[row[1]] = row

    cursor.execute("""
        SELECT employee_id, leave_date FROM leave_requests
        WHERE status = 'Approved' AND leave_date BETWEEN %s AND %s
    """, (datetime.date(year, month, 1), datetime.date(year, month, last_day)))
    leave_map = {}
    for eid, ld in cursor.fetchall():
        leave_map.setdefault(eid, set()).add(ld)

    cursor.execute(
        "SELECT processed_at, processed_by, email_count FROM payroll_runs WHERE year=%s AND month=%s",
        (year, month)
    )
    lock_row = cursor.fetchone()
    is_locked = lock_row is not None
    lock_info = {"at": lock_row[0], "by": lock_row[1], "count": lock_row[2]} if lock_row else None

    cursor.close()
    db.close()

    holidays_set = fetch_holidays_set(year, month)
    billable_past = get_billable_past_days(year, month)

    # Fetch all incentives for this month in one query
    try:
        db2 = get_db_connection()
        cursor2 = db2.cursor(buffered=True)
        cursor2.execute(
            "SELECT employee_id, COALESCE(SUM(amount),0) FROM employee_incentives WHERE year=%s AND month=%s GROUP BY employee_id",
            (year, month)
        )
        incentive_map = {r[0]: float(r[1]) for r in cursor2.fetchall()}
        cursor2.close()
        db2.close()
    except Exception:
        incentive_map = {}

    salary_data = []
    for emp_id, name, email, spd, role, phone in employees:
        entry = compute_salary_entry(emp_id, name, spd, att_map, billable_past,
                                     holidays_set=holidays_set,
                                     leave_dates=leave_map.get(emp_id, set()))
        inc = incentive_map.get(emp_id, 0.0)
        entry["incentive"] = inc
        entry["net"] = round(entry["net"] + inc, 2)
        entry["email"] = email
        entry["role"] = role
        entry["phone"] = phone
        salary_data.append(entry)

    months = [(i, datetime.date(year, i, 1).strftime("%B")) for i in range(1, 13)]
    years = list(range(datetime.date.today().year - 2, datetime.date.today().year + 1))

    email_cfg = get_email_config()

    return render_template("salary_report.html",
                           salary_data=salary_data,
                           month_name=datetime.date(year, month, 1).strftime("%B %Y"),
                           year=year, month=month,
                           months=months, years=years,
                           late_rate=int(cfg.LATE_DEDUCTION_RATE * 100),
                           half_rate=int(cfg.HALF_DAY_RATE * 100),
                           email_configured=email_cfg is not None,
                           is_locked=is_locked,
                           lock_info=lock_info,
                           active_nav="salary",
                           )


@payroll_bp.route("/salary_report_export")
@role_required("admin")
@limiter.limit("10 per minute")
def salary_report_export():
    from flask import send_file
    year = int(request.args.get("year", datetime.date.today().year))
    month = int(request.args.get("month", datetime.date.today().month))

    db = get_db_connection()
    cursor = db.cursor(buffered=True)
    active_cid = session.get("active_company_id")
    if active_cid:
        cursor.execute("""
            SELECT e.employee_id, e.name, e.email, COALESCE(s.salary_per_day, 0),
                   COALESCE(e.role,''), COALESCE(e.department,'')
            FROM employees e
            LEFT JOIN salary_config s ON e.employee_id = s.employee_id
            WHERE e.company_id = %s
            ORDER BY e.name
        """, (active_cid,))
    else:
        cursor.execute("""
            SELECT e.employee_id, e.name, e.email, COALESCE(s.salary_per_day, 0),
                   COALESCE(e.role,''), COALESCE(e.department,'')
            FROM employees e
            LEFT JOIN salary_config s ON e.employee_id = s.employee_id
            ORDER BY e.name
        """)
    employees = cursor.fetchall()

    _, last_day = calendar.monthrange(year, month)
    cursor.execute("""
        SELECT employee_id, date, login_time, logout_time, status, logout_status, attendance_type
        FROM attendance WHERE date BETWEEN %s AND %s
    """, (datetime.date(year, month, 1), datetime.date(year, month, last_day)))
    att_map = {}
    for row in cursor.fetchall():
        att_map.setdefault(row[0], {})[row[1]] = row

    cursor.execute("""
        SELECT employee_id, leave_date FROM leave_requests
        WHERE status='Approved' AND leave_date BETWEEN %s AND %s
    """, (datetime.date(year, month, 1), datetime.date(year, month, last_day)))
    leave_map = {}
    for eid, ld in cursor.fetchall():
        leave_map.setdefault(eid, set()).add(ld)

    cursor2 = db.cursor(buffered=True)
    cursor2.execute(
        "SELECT employee_id, COALESCE(SUM(amount),0) FROM employee_incentives WHERE year=%s AND month=%s GROUP BY employee_id",
        (year, month)
    )
    incentive_map = {r[0]: float(r[1]) for r in cursor2.fetchall()}
    cursor.close()
    cursor2.close()
    db.close()

    holidays_set = fetch_holidays_set(year, month)
    billable_past = get_billable_past_days(year, month)

    wb = openpyxl.Workbook()
    ws = wb.active
    month_name = datetime.date(year, month, 1).strftime("%B %Y")
    ws.title = f"Salary {month_name}"

    hdr_fill = PatternFill("solid", fgColor="1E3A8A")
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    alt_fill = PatternFill("solid", fgColor="EFF6FF")
    center = Alignment(horizontal="center", vertical="center")
    thin_side = Side(style="thin", color="CCCCCC")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    headers = ["#", "Employee ID", "Name", "Role", "Department",
               "Salary/Day", "Billable Days", "Present", "Absent",
               "Deduction", "Incentive", "Net Salary"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = center
        cell.border = thin_border

    col_widths = [5, 16, 22, 16, 16, 12, 14, 10, 10, 12, 12, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    for idx, (emp_id, name, email, spd, role, dept) in enumerate(employees, 1):
        entry = compute_salary_entry(emp_id, name, spd, att_map, billable_past,
                                     holidays_set=holidays_set,
                                     leave_dates=leave_map.get(emp_id, set()))
        inc = incentive_map.get(emp_id, 0.0)
        net = round(entry["net"] + inc, 2)
        row_data = [
            idx, emp_id, name, role, dept,
            float(spd), entry["billable"], entry["present"],
            entry["absent"], entry["deduction"], inc, net
        ]
        fill = alt_fill if idx % 2 == 0 else None
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=idx + 1, column=col, value=val)
            cell.border = thin_border
            cell.alignment = center
            if fill:
                cell.fill = fill

    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"Salary_Report_{month_name.replace(' ', '_')}.xlsx"
    return send_file(buf, as_attachment=True,
                     download_name=filename,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ---------------- EMAIL CONFIG ----------------

@payroll_bp.route("/email_config", methods=["GET", "POST"])
@role_required("admin")
@limiter.limit("10 per minute")
def email_config():
    # GET used to render this standalone page with the SMTP password
    # decrypted straight into the form's HTML and no step-up gate at all —
    # worse than the ciphertext-leak bug on the /settings copy of this form,
    # since here it was the real plaintext. Retired in favor of the
    # 2FA-gated Email tab; redirecting (rather than deleting the route)
    # keeps the "Setup Email First" link on the salary report page working.
    if request.method == "GET":
        return redirect("/settings?tab=email")

    db = get_db_connection()
    cursor = db.cursor(buffered=True)
    host = request.form["smtp_host"].strip()
    port = int(request.form["smtp_port"])
    user = request.form["smtp_user"].strip()
    password = request.form.get("smtp_pass", "").strip()
    from_name = request.form.get("from_name", "Attendance System").strip()
    from_email = request.form.get("from_email", "").strip() or user

    # A blank or masked password means "leave the stored one unchanged" —
    # previously any save re-encrypted whatever was in the field, and since
    # GET used to prefill that field with ciphertext, an untouched save
    # would silently corrupt the real password into unusable garbage.
    if password and password != "********":
        encrypted_password = encrypt_pii(password)
    else:
        cursor.execute("SELECT smtp_pass FROM email_config ORDER BY id DESC LIMIT 1")
        prev = cursor.fetchone()
        encrypted_password = prev[0] if prev else ""

    cursor.execute("DELETE FROM email_config")
    cursor.execute(
        "INSERT INTO email_config (smtp_host, smtp_port, smtp_user, smtp_pass, from_name, from_email) VALUES (%s,%s,%s,%s,%s,%s)",
        (host, port, user, encrypted_password, from_name, from_email)
    )
    db.commit()
    cursor.close()
    db.close()
    return redirect("/settings?tab=email&saved=1")

# ---------------- SEND SALARY EMAIL (single) ----------------


@payroll_bp.route("/send_salary_email", methods=["POST"])
@admin_required
def send_salary_email():
    emp_id = request.form["emp_id"]
    year = int(request.form["year"])
    month = int(request.form["month"])

    config = get_email_config()
    if not config:
        return jsonify({"ok": False, "msg": "Email not configured. Go to Email Settings first."})

    db = get_db_connection()
    cursor = db.cursor(buffered=True)

    cursor.execute("""
        SELECT e.name, e.email, COALESCE(s.salary_per_day,0),
               COALESCE(s.monthly_ctc,0), COALESCE(s.basic_pct,50),
               COALESCE(e.role,''), COALESCE(e.department,'')
        FROM employees e LEFT JOIN salary_config s ON e.employee_id=s.employee_id
        WHERE e.employee_id=%s
    """, (emp_id,))
    emp = cursor.fetchone()
    if not emp:
        cursor.close()
        db.close()
        return jsonify({"ok": False, "msg": "Employee not found."})

    name, email, spd, monthly_ctc, basic_pct, designation, dept = emp
    if not email:
        cursor.close()
        db.close()
        return jsonify({"ok": False, "msg": f"No email address set for {name}."})

    _, last_day = calendar.monthrange(year, month)
    cursor.execute("""
        SELECT employee_id, date, login_time, logout_time, status, logout_status, attendance_type
        FROM attendance WHERE employee_id=%s AND date BETWEEN %s AND %s
    """, (emp_id, datetime.date(year, month, 1), datetime.date(year, month, last_day)))

    att_map = {}
    for row in cursor.fetchall():
        att_map.setdefault(row[0], {})[row[1]] = row

    _, last_day2 = calendar.monthrange(year, month)
    cursor.execute(
        "SELECT leave_date FROM leave_requests "
        "WHERE status='Approved' AND employee_id=%s AND leave_date BETWEEN %s AND %s",
        (emp_id, datetime.date(year, month, 1), datetime.date(year, month, last_day2))
    )
    leave_dates = {row[0] for row in cursor.fetchall()}

    # Fetch payroll config
    db2 = get_db_connection()
    cur2 = db2.cursor(buffered=True)
    cur2.execute("SELECT pf_employee_pct,pf_employer_pct,professional_tax,tds_annual_pct,pf_basic_cap FROM payroll_config LIMIT 1")
    pc_row = cur2.fetchone()
    payroll_cfg = {"pf_employee_pct": float(pc_row[0] or 12), "pf_employer_pct": float(pc_row[1] or 12),
                   "professional_tax": float(pc_row[2] or 200), "tds_annual_pct": float(pc_row[3] or 0),
                   "pf_basic_cap": float(pc_row[4] or 15000)} if pc_row else {}
    cur2.close()
    db2.close()
    cursor.close()
    db.close()

    holidays_set = fetch_holidays_set(year, month)
    billable_past = get_billable_past_days(year, month)
    entry = compute_salary_entry(emp_id, name, spd, att_map, billable_past,
                                 holidays_set=holidays_set, leave_dates=leave_dates)
    entry["monthly_ctc"] = float(monthly_ctc) if float(monthly_ctc) > 0 else float(spd) * 26
    entry["basic_pct"] = int(basic_pct)
    month_name = datetime.date(year, month, 1).strftime("%B %Y")
    html_body = build_salary_slip_html(name, emp_id, email, month_name, year, month, entry,
                                       emp_designation=designation, emp_dept=dept,
                                       payroll_cfg=payroll_cfg)

    try:
        send_email_smtp(email, f"Salary Slip - {month_name}", html_body, config)
        return jsonify({"ok": True, "msg": f"Salary slip sent to {email}"})
    except Exception:
        app_log.error("Failed to send salary slip email to %s", email, exc_info=True)
        return jsonify({"ok": False, "msg": "Failed to send email. Check email settings."})

# ---------------- SEND ALL SALARY EMAILS ----------------


@payroll_bp.route("/send_all_salary_emails", methods=["POST"])
@role_required("admin")
@limiter.limit("5 per minute")
def send_all_salary_emails():
    year = int(request.form["year"])
    month = int(request.form["month"])

    config = get_email_config()
    if not config:
        return jsonify({"ok": False, "msg": "Email not configured. Go to Email Settings first."})

    db = get_db_connection()
    cursor = db.cursor(buffered=True)

    cursor.execute("""
        SELECT e.employee_id, e.name, e.email, COALESCE(s.salary_per_day, 0)
        FROM employees e
        LEFT JOIN salary_config s ON e.employee_id = s.employee_id
        ORDER BY e.name
    """)
    employees = cursor.fetchall()

    _, last_day = calendar.monthrange(year, month)
    cursor.execute("""
        SELECT employee_id, date, login_time, logout_time, status, logout_status, attendance_type
        FROM attendance WHERE date BETWEEN %s AND %s
    """, (datetime.date(year, month, 1), datetime.date(year, month, last_day)))

    att_map = {}
    for row in cursor.fetchall():
        att_map.setdefault(row[0], {})[row[1]] = row

    _, last_day_all = calendar.monthrange(year, month)
    cursor.execute(
        "SELECT employee_id, leave_date FROM leave_requests "
        "WHERE status='Approved' AND leave_date BETWEEN %s AND %s",
        (datetime.date(year, month, 1), datetime.date(year, month, last_day_all))
    )
    leave_map_all = {}
    for eid, ld in cursor.fetchall():
        leave_map_all.setdefault(eid, set()).add(ld)

    cursor.close()
    db.close()

    holidays_set = fetch_holidays_set(year, month)
    billable_past = get_billable_past_days(year, month)
    month_name = datetime.date(year, month, 1).strftime("%B %Y")

    sent = skipped = failed = 0
    errors = []

    for emp_id, name, email, spd in employees:
        if not email:
            skipped += 1
            continue
        entry = compute_salary_entry(emp_id, name, spd, att_map, billable_past,
                                     holidays_set=holidays_set,
                                     leave_dates=leave_map_all.get(emp_id, set()))
        html_body = build_salary_slip_html(name, emp_id, email, month_name, year, month, entry)
        try:
            send_email_smtp(email, f"Salary Slip - {month_name}", html_body, config)
            sent += 1
        except Exception:
            app_log.error("Failed to send salary slip to %s", name, exc_info=True)
            failed += 1
            errors.append(f"{name}: email delivery failed")

    if sent > 0:
        try:
            db2 = get_db_connection()
            cur2 = db2.cursor()
            actor = session.get("admin_username", "admin")
            cur2.execute("""
                INSERT INTO payroll_runs (year, month, processed_by, email_count)
                VALUES (%s, %s, %s, %s)
                ON CONFLICT (year, month) DO UPDATE SET processed_at=NOW(), processed_by=%s, email_count=%s
            """, (year, month, actor, sent, actor, sent))
            db2.commit()
            cur2.close()
            db2.close()
            _audit("lock_payroll", "payroll_runs", f"{year}-{month:02d}",
                   f"Payroll locked for {year}-{month:02d} after sending {sent} payslips")
        except Exception as _le:
            app_log.warning("Could not record payroll lock for %d-%02d: %s", year, month, _le)

    msg = f"Sent: {sent}, Skipped (no email): {skipped}, Failed: {failed}"
    if errors:
        msg += " | " + "; ".join(errors[:3])
    return jsonify({"ok": failed == 0, "msg": msg, "locked": sent > 0})

# ---------------- PAYROLL LOCK / UNLOCK ----------------


@payroll_bp.route("/lock_payroll", methods=["POST"])
@admin_required
def lock_payroll():
    year = int(request.form["year"])
    month = int(request.form["month"])
    actor = session.get("admin_username", "admin")
    db = get_db_connection()
    cursor = db.cursor()
    cursor.execute("""
        INSERT INTO payroll_runs (year, month, processed_by, email_count)
        VALUES (%s, %s, %s, 0)
        ON CONFLICT (year, month) DO UPDATE SET processed_at=NOW(), processed_by=%s
    """, (year, month, actor, actor))
    db.commit()
    cursor.close()
    db.close()
    _audit("lock_payroll", "payroll_runs", f"{year}-{month:02d}", f"Manually locked by {actor}")
    return jsonify({"ok": True, "msg": f"Payroll for {year}-{month:02d} locked."})


@payroll_bp.route("/unlock_payroll", methods=["POST"])
@admin_required
def unlock_payroll():
    year = int(request.form["year"])
    month = int(request.form["month"])
    actor = session.get("admin_username", "admin")
    db = get_db_connection()
    cursor = db.cursor()
    cursor.execute("DELETE FROM payroll_runs WHERE year=%s AND month=%s", (year, month))
    db.commit()
    cursor.close()
    db.close()
    _audit("unlock_payroll", "payroll_runs", f"{year}-{month:02d}", f"Unlocked by {actor}")
    return jsonify({"ok": True, "msg": f"Payroll for {year}-{month:02d} unlocked."})


# ---------------- TEST EMAIL ----------------

@payroll_bp.route("/my_payslip_summary/<int:year>/<int:month>")
@employee_required
def my_payslip_summary(year, month):
    import json as _json
    emp_id = session["employee_id"]
    db = get_db_connection()
    cursor = db.cursor(buffered=True)
    cursor.execute("SELECT COALESCE(salary_per_day,0) FROM salary_config WHERE employee_id=%s", (emp_id,))
    row = cursor.fetchone()
    spd = float(row[0]) if row else 0.0

    _, last_day = calendar.monthrange(year, month)
    cursor.execute("""
        SELECT date, login_time, logout_time, status, logout_status, attendance_type
        FROM attendance WHERE employee_id=%s AND date BETWEEN %s AND %s
    """, (emp_id, datetime.date(year, month, 1), datetime.date(year, month, last_day)))
    att_rows = cursor.fetchall()

    # Incentive total for this employee/month
    try:
        cursor.execute(
            "SELECT COALESCE(SUM(amount),0) FROM employee_incentives WHERE employee_id=%s AND year=%s AND month=%s",
            (emp_id, year, month)
        )
        incentive = float(cursor.fetchone()[0])
        cursor.execute("""
            SELECT ig.title, ei.amount, ei.notes
            FROM employee_incentives ei
            JOIN incentive_goals ig ON ei.goal_id = ig.id
            WHERE ei.employee_id=%s AND ei.year=%s AND ei.month=%s
        """, (emp_id, year, month))
        incentive_details = [{"title": r[0], "amount": float(r[1]), "notes": r[2] or ""} for r in cursor.fetchall()]
    except Exception:
        incentive = 0.0
        incentive_details = []

    try:
        cursor.execute(
            "SELECT COALESCE(SUM(ot_pay),0) FROM overtime_records WHERE employee_id=%s AND EXTRACT(MONTH FROM date)=%s AND EXTRACT(YEAR FROM date)=%s AND status='Approved'",
            (emp_id, month, year)
        )
        ot_pay = float(cursor.fetchone()[0])
    except Exception:
        ot_pay = 0.0

    cursor.close()
    db.close()

    att_map = {r[0]: r for r in att_rows}
    billable = get_billable_past_days(year, month)
    full = late = half = 0
    for d in billable:
        r = att_map.get(d)
        if r:
            final = r[5] if r[5] else infer_type_legacy(r[3], r[1], r[2])
            if final == "Full Day":
                full += 1
            elif final == "Late - Full Day":
                late += 1
            elif final in ("Half Day", "Present"):
                half += 1

    full_earn = round(full * spd, 2)
    late_earn = round(late * spd, 2)
    half_earn = round(half * spd * 0.5, 2)
    gross = full_earn + late_earn + half_earn
    pf = round(gross * 0.12, 2)
    net = round(gross - pf + incentive + ot_pay, 2)

    return _json.dumps({
        "salary_per_day": spd,
        "full_days": full, "late_days": late, "half_days": half,
        "full_earn": full_earn, "late_earn": late_earn, "half_earn": half_earn,
        "gross": gross, "pf": pf, "incentive": incentive,
        "incentive_details": incentive_details, "ot_pay": ot_pay, "net": net
    }), 200, {"Content-Type": "application/json"}


@payroll_bp.route("/my_attendance_pdf")
@employee_required
def my_attendance_pdf():
    emp_id = session["employee_id"]
    year = int(request.args.get("year", datetime.date.today().year))
    month = int(request.args.get("month", datetime.date.today().month))

    db = get_db_connection()
    cursor = db.cursor(buffered=True)
    cursor.execute("SELECT employee_id, name, role, email FROM employees WHERE employee_id=%s", (emp_id,))
    emp = cursor.fetchone()

    _, last_day = calendar.monthrange(year, month)
    cursor.execute("""
        SELECT date, login_time, logout_time, status, logout_status, attendance_type
        FROM attendance WHERE employee_id=%s AND date BETWEEN %s AND %s ORDER BY date
    """, (emp_id, datetime.date(year, month, 1), datetime.date(year, month, last_day)))
    monthly_att = cursor.fetchall()
    cursor.close()
    db.close()

    billable_past = get_billable_past_days(year, month)
    att_by_date = {r[0]: r for r in monthly_att}
    full_days = half_days = late_days = absent_days = total_sec = 0
    for d in billable_past:
        row = att_by_date.get(d)
        if row:
            _, login_t, logout_t, status, _ls, att_type = row
            final = att_type if att_type else infer_type_legacy(status, login_t, logout_t)
            if final == "Full Day":
                full_days += 1
            elif final == "Late - Full Day":
                late_days += 1
            elif final in ("Half Day", "Present"):
                half_days += 1
            else:
                absent_days += 1
            if login_t and logout_t:
                li = login_t.total_seconds() if hasattr(login_t, "total_seconds") else login_t.hour * \
                    3600 + login_t.minute * 60 + login_t.second
                lo = logout_t.total_seconds() if hasattr(logout_t, "total_seconds") else logout_t.hour * \
                    3600 + logout_t.minute * 60 + logout_t.second
                if lo > li:
                    total_sec += int(lo - li)
        else:
            absent_days += 1

    def fmt(t):
        if t is None:
            return "--"
        if hasattr(t, "strftime"):
            return t.strftime("%H:%M")
        s = int(t.total_seconds())
        return f"{s//3600:02d}:{(s%3600)//60:02d}"

    rows_html = ""
    for d in sorted(att_by_date.keys()):
        row = att_by_date[d]
        _, lt, lot, ls, _lo, at = row
        final = at if at else infer_type_legacy(ls, lt, lot)
        color = {"Full Day": "#16a34a", "Late - Full Day": "#d97706",
                 "Half Day": "#dc2626", "Present": "#d97706"}.get(final, "#6b7280")
        rows_html += f"<tr><td>{d.strftime('%d %b %Y')}</td><td>{d.strftime('%A')}</td><td>{fmt(lt)}</td><td>{fmt(lot)}</td><td style='color:{color};font-weight:600;'>{final or 'Absent'}</td></tr>"

    billable = len(billable_past)
    pct = round((full_days + late_days + half_days * 0.5) / billable * 100, 1) if billable else 0
    month_name = datetime.date(year, month, 1).strftime("%B %Y")
    total_h = f"{total_sec//3600}h {(total_sec%3600)//60}m"

    html = f"""<!doctype html><html><head><meta charset="UTF-8">
<title>Attendance Report — {emp[1]} — {month_name}</title>
<style>
  body {{ font-family: "Segoe UI", sans-serif; margin: 0; padding: 32px; color: #1e293b; background: white; }}
  .header {{ display: flex; justify-content: space-between; align-items: flex-start; margin-bottom: 28px; border-bottom: 2px solid #1e3a8a; padding-bottom: 18px; }}
  .title {{ font-size: 22px; font-weight: 700; color: #1e3a8a; }}
  .sub {{ font-size: 13px; color: #64748b; margin-top: 4px; }}
  .meta {{ text-align: right; font-size: 13px; color: #64748b; }}
  .stats {{ display: grid; grid-template-columns: repeat(6, 1fr); gap: 12px; margin-bottom: 24px; }}
  .stat {{ background: #f8fafc; border: 1px solid #dbeafe; border-radius: 10px; padding: 12px; text-align: center; }}
  .stat .n {{ font-size: 24px; font-weight: 700; }}
  .stat .l {{ font-size: 11px; color: #64748b; margin-top: 3px; }}
  .c-green {{ color: #16a34a; }} .c-yellow {{ color: #d97706; }} .c-red {{ color: #dc2626; }} .c-blue {{ color: #2563eb; }}
  table {{ width: 100%; border-collapse: collapse; font-size: 13px; }}
  th {{ background: #1e3a8a; color: white; padding: 10px 12px; text-align: left; }}
  td {{ padding: 9px 12px; border-bottom: 1px solid #e2e8f0; }}
  tr:last-child td {{ border-bottom: none; }}
  .footer {{ margin-top: 24px; font-size: 11px; color: #94a3b8; text-align: center; }}
  @media print {{ body {{ padding: 16px; }} button {{ display: none; }} }}
</style></head><body>
<div class="header">
  <div><div class="title">Attendance Report</div>
    <div class="sub">{emp[1]} &nbsp;·&nbsp; {emp[0]} &nbsp;·&nbsp; {emp[2] or 'Employee'}</div>
    <div class="sub">{month_name}</div></div>
  <div class="meta">Generated: {datetime.date.today().strftime('%d %b %Y')}<br>
    <button onclick="window.print()" style="margin-top:8px;padding:8px 16px;background:#1e3a8a;color:white;border:none;border-radius:8px;cursor:pointer;font-size:13px;">🖨️ Print / Save PDF</button>
  </div>
</div>
<div class="stats">
  <div class="stat"><div class="n c-green">{full_days}</div><div class="l">Full Days</div></div>
  <div class="stat"><div class="n c-yellow">{late_days}</div><div class="l">Late Days</div></div>
  <div class="stat"><div class="n c-yellow">{half_days}</div><div class="l">Half Days</div></div>
  <div class="stat"><div class="n c-red">{absent_days}</div><div class="l">Absent</div></div>
  <div class="stat"><div class="n c-blue">{pct}%</div><div class="l">Attendance</div></div>
  <div class="stat"><div class="n c-blue">{total_h}</div><div class="l">Total Hours</div></div>
</div>
<table><thead><tr><th>Date</th><th>Day</th><th>Login</th><th>Logout</th><th>Status</th></tr></thead>
<tbody>{rows_html}</tbody></table>
<div class="footer">Employee Attendance System &nbsp;·&nbsp; {emp[1]} &nbsp;·&nbsp; {month_name}</div>
</body></html>"""
    return html


@payroll_bp.route("/apply_hike", methods=["POST"])
@role_required("admin")
@limiter.limit("10 per minute")
def apply_hike():
    q = int(request.form.get("quarter", 1))
    yr = int(request.form.get("year", datetime.date.today().year))
    emp_ids = request.form.getlist("emp_ids")
    if not emp_ids:
        flash("No employees selected.", "error")
        return redirect(f"/performance?tab=hike&quarter={q}&year={yr}")

    db = get_db_connection()
    cursor = db.cursor(buffered=True)
    cursor.execute("SELECT min_rating, max_rating, hike_pct FROM hike_config ORDER BY min_rating DESC")
    bands = cursor.fetchall()

    # Batch-fetch instead of 2 SELECTs per employee — this is frequently a
    # whole-company action (dozens to hundreds of employees selected at
    # once), so the per-employee round-trips were the dominant cost here.
    cursor.execute(
        "SELECT employee_id, COALESCE(overall_rating,0) FROM performance_reviews "
        "WHERE employee_id = ANY(%s) AND quarter=%s AND year=%s",
        (emp_ids, q, yr)
    )
    ratings = {row[0]: float(row[1]) for row in cursor.fetchall()}

    cursor.execute(
        "SELECT employee_id, COALESCE(monthly_ctc,0), last_hike_quarter, last_hike_year "
        "FROM salary_config WHERE employee_id = ANY(%s)",
        (emp_ids,)
    )
    salaries = {row[0]: (float(row[1]), row[2], row[3]) for row in cursor.fetchall()}

    today = datetime.date.today()
    updated = 0
    for emp_id in emp_ids:
        rating = ratings.get(emp_id, 0.0)
        if rating == 0:
            continue
        hike_pct = 0.0
        for (mn, mx, hp) in bands:
            if float(mn) <= rating <= float(mx):
                hike_pct = float(hp)
                break
        if hike_pct <= 0:
            continue
        sc = salaries.get(emp_id)
        if not sc or sc[0] == 0:
            continue
        current_ctc, last_hike_q, last_hike_yr = sc
        # Idempotency: skip if this quarter's hike was already applied
        if last_hike_q == q and last_hike_yr == yr:
            continue
        new_ctc = round(current_ctc * (1 + hike_pct / 100), 2)
        new_spd = round(new_ctc / 26, 2)
        # The idempotency guard is repeated in this UPDATE's WHERE clause
        # (not just the Python-level skip above) so it's checked-and-set
        # atomically in one statement — closes the race where two concurrent
        # apply_hike submissions both read last_hike_quarter != q before
        # either had committed, which would otherwise double-apply the hike.
        cursor.execute(
            "UPDATE salary_config SET monthly_ctc=%s, salary_per_day=%s, last_revised=%s, "
            "last_hike_quarter=%s, last_hike_year=%s "
            "WHERE employee_id=%s AND (last_hike_quarter IS DISTINCT FROM %s OR last_hike_year IS DISTINCT FROM %s)",
            (new_ctc, new_spd, today, q, yr, emp_id, q, yr)
        )
        if cursor.rowcount:
            updated += 1

    db.commit()
    cursor.close()
    db.close()
    flash(f"Hike applied to {updated} employee(s) successfully.", "success")
    return redirect(f"/performance?tab=hike&quarter={q}&year={yr}")


@payroll_bp.route("/award_performance_bonus", methods=["POST"])
@role_required("admin")
@limiter.limit("10 per minute")
def award_performance_bonus():
    q = int(request.form.get("quarter", 1))
    yr = int(request.form.get("year", datetime.date.today().year))
    emp_ids = request.form.getlist("emp_ids")
    if not emp_ids:
        flash("No employees selected.", "error")
        return redirect(f"/performance?tab=hike&quarter={q}&year={yr}")

    db = get_db_connection()
    cursor = db.cursor(buffered=True)

    cursor.execute("SELECT id FROM incentive_goals WHERE title='Performance Bonus' LIMIT 1")
    row = cursor.fetchone()
    if row:
        goal_id = row[0]
    else:
        cursor.execute(
            "INSERT INTO incentive_goals (title, description, incentive_amount, is_active) VALUES (%s,%s,%s,%s) RETURNING id",
            ("Performance Bonus", "Quarterly performance-based incentive", 0, 1)
        )
        goal_id = cursor.fetchone()[0]
        db.commit()

    cursor.execute("SELECT min_rating, max_rating, incentive_pct FROM hike_config ORDER BY min_rating DESC")
    bands = cursor.fetchall()

    bonus_month = {1: 3, 2: 6, 3: 9, 4: 12}.get(q, q * 3)

    # Same batching approach as apply_hike above — 3 SELECTs per employee
    # collapsed into 3 SELECTs total, regardless of how many are selected.
    cursor.execute(
        "SELECT employee_id, COALESCE(overall_rating,0) FROM performance_reviews "
        "WHERE employee_id = ANY(%s) AND quarter=%s AND year=%s",
        (emp_ids, q, yr)
    )
    ratings = {row[0]: float(row[1]) for row in cursor.fetchall()}

    cursor.execute(
        "SELECT employee_id, COALESCE(monthly_ctc,0) FROM salary_config WHERE employee_id = ANY(%s)",
        (emp_ids,)
    )
    ctcs = {row[0]: float(row[1]) for row in cursor.fetchall()}

    cursor.execute(
        "SELECT DISTINCT employee_id FROM employee_incentives "
        "WHERE employee_id = ANY(%s) AND goal_id=%s AND month=%s AND year=%s",
        (emp_ids, goal_id, bonus_month, yr)
    )
    already_awarded = {row[0] for row in cursor.fetchall()}

    awarded = 0
    for emp_id in emp_ids:
        rating = ratings.get(emp_id, 0.0)
        if rating == 0:
            continue
        inc_pct = 0.0
        for (mn, mx, ip) in bands:
            if float(mn) <= rating <= float(mx):
                inc_pct = float(ip)
                break
        if inc_pct <= 0:
            continue
        ctc = ctcs.get(emp_id, 0.0)
        if ctc == 0:
            continue
        bonus_amount = round(ctc * inc_pct / 100, 2)
        if bonus_amount <= 0:
            continue
        # Skip if this bonus was already awarded for this employee/quarter/year
        # (Python-level fast path). ON CONFLICT DO NOTHING below is the real
        # guard against the race where two concurrent award requests both
        # read "not yet awarded" before either had committed — it relies on
        # the unique index on (employee_id, goal_id, month, year) created by
        # the incentives_unique_v1 migration in app.py's init_db().
        if emp_id in already_awarded:
            continue
        cursor.execute(
            "INSERT INTO employee_incentives (employee_id, goal_id, month, year, amount, notes) "
            "VALUES (%s,%s,%s,%s,%s,%s) "
            "ON CONFLICT (employee_id, goal_id, month, year) DO NOTHING",
            (emp_id, goal_id, bonus_month, yr, bonus_amount, f"Performance bonus Q{q} {yr} — Rating: {rating}/5")
        )
        if cursor.rowcount:
            awarded += 1

    db.commit()
    cursor.close()
    db.close()
    flash(f"Performance bonus awarded to {awarded} employee(s).", "success")
    return redirect(f"/performance?tab=hike&quarter={q}&year={yr}")


@payroll_bp.route("/save_hike_config", methods=["POST"])
@admin_required
def save_hike_config():
    q = request.form.get("quarter", "1")
    yr = request.form.get("year", str(datetime.date.today().year))
    ids = request.form.getlist("band_id")
    labels = request.form.getlist("band_label")
    min_rats = request.form.getlist("band_min")
    max_rats = request.form.getlist("band_max")
    hike_pcts = request.form.getlist("band_hike")
    inc_pcts = request.form.getlist("band_inc")

    n = min(len(ids), len(labels), len(min_rats), len(max_rats), len(hike_pcts), len(inc_pcts))
    if n == 0:
        flash("No band data received.", "error")
        return redirect(f"/performance?tab=hike&quarter={q}&year={yr}")

    db = get_db_connection()
    cursor = db.cursor(buffered=True)
    for i in range(n):
        try:
            cursor.execute(
                "UPDATE hike_config SET label=%s, min_rating=%s, max_rating=%s, hike_pct=%s, incentive_pct=%s WHERE id=%s",
                (labels[i], float(min_rats[i]), float(max_rats[i]),
                 float(hike_pcts[i]), float(inc_pcts[i]), int(ids[i]))
            )
        except (ValueError, TypeError):
            continue
    db.commit()
    cursor.close()
    db.close()
    flash("Hike band configuration saved.", "success")
    return redirect(f"/performance?tab=hike&quarter={q}&year={yr}")


@payroll_bp.route("/api/salary_config", methods=["GET"])
@api_required
@api_role_required("admin")
@limiter.limit("10 per minute")
def api_salary_config_get():
    db = get_db_connection()
    cursor = db.cursor(buffered=True)
    cursor.execute("""
        SELECT e.employee_id, e.name, COALESCE(s.salary_per_day, 0)
        FROM employees e
        LEFT JOIN salary_config s ON e.employee_id = s.employee_id
        ORDER BY e.name
    """)
    rows = cursor.fetchall()
    cursor.close()
    db.close()
    return jsonify({"ok": True, "salaries": [
        {"employee_id": r[0], "name": r[1], "salary_per_day": float(r[2])} for r in rows
    ]})


@payroll_bp.route("/api/salary_config", methods=["POST"])
@api_required
def api_salary_config_post():
    data = request.get_json() or {}
    emp_id = data.get("employee_id")
    salary = data.get("salary_per_day")
    if not emp_id or salary is None:
        return jsonify({"ok": False, "msg": "employee_id and salary_per_day required"}), 400
    db = get_db_connection()
    cursor = db.cursor(buffered=True)
    cursor.execute("SELECT 1 FROM salary_config WHERE employee_id=%s", (emp_id,))
    if cursor.fetchone():
        cursor.execute("UPDATE salary_config SET salary_per_day=%s WHERE employee_id=%s", (salary, emp_id))
    else:
        cursor.execute("INSERT INTO salary_config (employee_id, salary_per_day) VALUES (%s,%s)", (emp_id, salary))
    db.commit()
    cursor.close()
    db.close()
    return jsonify({"ok": True})


@payroll_bp.route("/api/monthly_report", methods=["GET"])
@api_required
def api_monthly_report():
    year = int(request.args.get("year", datetime.date.today().year))
    month = int(request.args.get("month", datetime.date.today().month))
    db = get_db_connection()
    cursor = db.cursor(buffered=True)
    cursor.execute("SELECT employee_id, name FROM employees ORDER BY name")
    employees = cursor.fetchall()
    _, last_day = calendar.monthrange(year, month)
    cursor.execute("""
        SELECT employee_id, date, login_time, logout_time, status, logout_status, attendance_type
        FROM attendance WHERE date BETWEEN %s AND %s
    """, (datetime.date(year, month, 1), datetime.date(year, month, last_day)))
    att_map = {}
    for row in cursor.fetchall():
        att_map.setdefault(row[0], {})[row[1]] = row
    cursor.close()
    db.close()
    holidays = fetch_holidays_set(year, month)
    working_days = get_working_days(year, month)
    today = datetime.date.today()
    report = []
    for emp_id, name in employees:
        emp_att = att_map.get(emp_id, {})
        full_days = half_days = late_days = absent = 0
        for d in working_days:
            if d > today or d in holidays:
                continue
            row = emp_att.get(d)
            if row:
                _, _, login_t, logout_t, status, _logout_status, att_type = row
                final = att_type if att_type else infer_type_legacy(status, login_t, logout_t)
                if final in ("Full Day", "Approved Leave"):
                    full_days += 1
                elif final == "Late - Full Day":
                    late_days += 1
                elif final in ("Half Day", "Present"):
                    half_days += 1
                else:
                    absent += 1
            else:
                absent += 1
        billable = len([d for d in working_days if d <= today and d not in holidays])
        present_equiv = full_days + late_days + half_days * 0.5
        pct = round(present_equiv / billable * 100, 1) if billable > 0 else 0
        report.append({"employee_id": emp_id, "name": name, "full_days": full_days,
                       "half_days": half_days, "late_days": late_days, "absent": absent,
                       "billable": billable, "pct": pct})
    return jsonify({"ok": True, "report": report,
                    "month_name": datetime.date(year, month, 1).strftime("%B %Y"),
                    "year": year, "month": month,
                    "holiday_count": len(holidays),
                    "total_working": len([d for d in working_days if d <= today and d not in holidays])})


@payroll_bp.route("/api/salary_report", methods=["GET"])
@api_required
@api_role_required("admin")
@limiter.limit("10 per minute")
def api_salary_report():
    year = int(request.args.get("year", datetime.date.today().year))
    month = int(request.args.get("month", datetime.date.today().month))
    db = get_db_connection()
    cursor = db.cursor(buffered=True)
    cursor.execute("""
        SELECT e.employee_id, e.name, e.email, COALESCE(s.salary_per_day, 0)
        FROM employees e
        LEFT JOIN salary_config s ON e.employee_id = s.employee_id
        ORDER BY e.name
    """)
    employees = cursor.fetchall()
    _, last_day = calendar.monthrange(year, month)
    cursor.execute("""
        SELECT employee_id, date, login_time, logout_time, status, logout_status, attendance_type
        FROM attendance WHERE date BETWEEN %s AND %s
    """, (datetime.date(year, month, 1), datetime.date(year, month, last_day)))
    att_map = {}
    for row in cursor.fetchall():
        att_map.setdefault(row[0], {})[row[1]] = row
    cursor.close()
    db.close()
    billable_past = get_billable_past_days(year, month)
    salary_data = []
    for emp_id, name, email, spd in employees:
        entry = compute_salary_entry(emp_id, name, spd, att_map, billable_past)
        entry["email"] = email
        salary_data.append(entry)
    return jsonify({"ok": True, "salary_data": salary_data,
                    "month_name": datetime.date(year, month, 1).strftime("%B %Y"),
                    "year": year, "month": month})


@payroll_bp.route("/api/email_config", methods=["GET"])
@api_required
def api_get_email_config():
    cfg = get_email_config()
    # Never return the SMTP password to clients — they only need to know config exists.
    safe_cfg = {k: v for k, v in cfg.items() if k != "password"}
    safe_cfg["password_set"] = bool(cfg.get("password"))
    return jsonify({"ok": True, "config": safe_cfg})


@payroll_bp.route("/api/email_config", methods=["POST"])
@api_required
def api_save_email_config():
    data = request.get_json() or {}
    host = data.get("smtp_host", "").strip()
    port = int(data.get("smtp_port", 587))
    user = data.get("smtp_user", "").strip()
    password = data.get("smtp_pass", "").strip()
    from_name = data.get("from_name", "HR Department").strip()
    if not host or not user or not password:
        return jsonify({"ok": False, "msg": "host, user and password required"}), 400
    db = get_db_connection()
    cursor = db.cursor(buffered=True)
    cursor.execute("DELETE FROM email_config")
    cursor.execute(
        "INSERT INTO email_config (smtp_host, smtp_port, smtp_user, smtp_pass, from_name) VALUES (%s,%s,%s,%s,%s)",
        (host, port, user, encrypt_pii(password), from_name)
    )
    db.commit()
    cursor.close()
    db.close()
    return jsonify({"ok": True})


@payroll_bp.route("/api/send_salary_email", methods=["POST"])
@api_required
def api_send_salary_email():
    data = request.get_json() or {}
    emp_id = data.get("emp_id")
    year = int(data.get("year", datetime.date.today().year))
    month = int(data.get("month", datetime.date.today().month))
    config = get_email_config()
    if not config:
        return jsonify({"ok": False, "msg": "Email not configured."})
    db = get_db_connection()
    cursor = db.cursor(buffered=True)
    cursor.execute(
        "SELECT name, email, COALESCE(s.salary_per_day,0) FROM employees e "
        "LEFT JOIN salary_config s ON e.employee_id=s.employee_id WHERE e.employee_id=%s",
        (emp_id,)
    )
    emp = cursor.fetchone()
    if not emp:
        cursor.close()
        db.close()
        return jsonify({"ok": False, "msg": "Employee not found."})
    name, email, spd = emp
    if not email:
        cursor.close()
        db.close()
        return jsonify({"ok": False, "msg": f"No email for {name}."})
    _, last_day = calendar.monthrange(year, month)
    cursor.execute("""
        SELECT employee_id, date, login_time, logout_time, status, logout_status, attendance_type
        FROM attendance WHERE employee_id=%s AND date BETWEEN %s AND %s
    """, (emp_id, datetime.date(year, month, 1), datetime.date(year, month, last_day)))
    att_map = {}
    for row in cursor.fetchall():
        att_map.setdefault(row[0], {})[row[1]] = row
    cursor.close()
    db.close()
    billable_past = get_billable_past_days(year, month)
    entry = compute_salary_entry(emp_id, name, spd, att_map, billable_past)
    month_name = datetime.date(year, month, 1).strftime("%B %Y")
    html_body = build_salary_slip_html(name, emp_id, email, month_name, year, month, entry)
    send_email_async(email, f"Salary Slip - {month_name}", html_body, config)
    return jsonify({"ok": True, "msg": f"Queued for {email}"})


@payroll_bp.route("/view_payslip/<emp_id>/<int:year>/<int:month>")
def view_payslip(emp_id, year, month):
    # emp_id is a raw URL parameter — this is exactly the BOLA/IDOR shape
    # (a valid session trying to view a DIFFERENT employee's salary data by
    # editing the URL). enforce_ownership() logs any denial at ERROR, which
    # feeds the alerting webhook (utils/alerts.py), not just the log stream.
    if not enforce_ownership(emp_id, "payslip", f"{year}-{month:02d}"):
        return redirect("/employee_login")
    # enforce_ownership() grants any admin-side session a bypass regardless
    # of role — too broad for this route specifically, since it renders
    # plaintext PAN/UAN/bank account details. Restrict the non-owner
    # (i.e. admin) path to the "admin" role; the employee's own session
    # viewing their own payslip already passed the ownership check above
    # and is unaffected by this.
    if session.get("admin_logged_in") and session.get("admin_role", "admin") != "admin":
        log_security_event(
            "access.denied", "Non-admin role attempted to view an employee payslip",
            level="ERROR", identifier=session.get("admin_username"),
            resource_type="payslip", resource_id=f"{emp_id}:{year}-{month:02d}",
        )
        return redirect("/employee_login")

    db = get_db_connection()
    cursor = db.cursor(buffered=True)
    cursor.execute("""
        SELECT e.name, e.email, COALESCE(s.salary_per_day, 0),
               COALESCE(s.monthly_ctc, 0), COALESCE(s.basic_pct, 50),
               COALESCE(e.role,''), COALESCE(e.department,''),
               COALESCE(e.pan_number,''), COALESCE(e.uan_number,''),
               COALESCE(e.bank_account,''), COALESCE(e.bank_name,'')
        FROM employees e
        LEFT JOIN salary_config s ON e.employee_id = s.employee_id
        WHERE e.employee_id = %s
    """, (emp_id,))
    row = cursor.fetchone()
    if not row:
        cursor.close()
        db.close()
        return "Employee not found", 404
    name, email, spd, monthly_ctc, basic_pct, designation, dept, pan, uan, bank_acct, bank_nm = row
    pan = decrypt_pii(pan)
    uan = decrypt_pii(uan)
    bank_acct = decrypt_pii(bank_acct)
    bank_nm = decrypt_pii(bank_nm)

    # Payroll config
    cursor.execute(
        "SELECT pf_employee_pct, pf_employer_pct, professional_tax, tds_annual_pct, pf_basic_cap FROM payroll_config LIMIT 1")
    pc_row = cursor.fetchone()
    payroll_cfg = {}
    if pc_row:
        payroll_cfg = {
            "pf_employee_pct": float(pc_row[0] or 12),
            "pf_employer_pct": float(pc_row[1] or 12),
            "professional_tax": float(pc_row[2] or 200),
            "tds_annual_pct": float(pc_row[3] or 0),
            "pf_basic_cap": float(pc_row[4] or 15000),
        }

    cursor.execute("SELECT COALESCE(company_name,'') FROM company_settings LIMIT 1")
    co_row = cursor.fetchone()
    company_name = co_row[0] if co_row else ""

    _, last_day = calendar.monthrange(year, month)
    cursor.execute("""
        SELECT date, employee_id, login_time, logout_time, status, logout_status, attendance_type
        FROM attendance
        WHERE employee_id=%s AND date BETWEEN %s AND %s
    """, (emp_id, datetime.date(year, month, 1), datetime.date(year, month, last_day)))
    att_map = {}
    for r in cursor.fetchall():
        att_map.setdefault(r[1], {})[r[0]] = r

    cursor.execute(
        "SELECT leave_date FROM leave_requests "
        "WHERE status='Approved' AND employee_id=%s AND leave_date BETWEEN %s AND %s",
        (emp_id, datetime.date(year, month, 1), datetime.date(year, month, last_day))
    )
    leave_dates = {r[0] for r in cursor.fetchall()}
    cursor.close()
    db.close()

    holidays_set = fetch_holidays_set(year, month)
    billable_past = get_billable_past_days(year, month)
    entry = compute_salary_entry(emp_id, name, spd, att_map, billable_past,
                                 holidays_set=holidays_set, leave_dates=leave_dates)
    entry["monthly_ctc"] = float(monthly_ctc) if float(monthly_ctc) > 0 else float(spd) * 26
    entry["basic_pct"] = int(basic_pct)

    month_name = calendar.month_name[month] + f" {year}"
    return build_salary_slip_html(
        name, emp_id, email, month_name, year, month, entry,
        company_name=company_name,
        emp_designation=designation, emp_dept=dept,
        pan=pan, uan=uan, bank_account=bank_acct, bank_name=bank_nm,
        payroll_cfg=payroll_cfg
    )


@payroll_bp.route("/download_payslip/<emp_id>/<int:year>/<int:month>")
def download_payslip(emp_id, year, month):
    html = view_payslip(emp_id, year, month)
    if not isinstance(html, str):
        return html  # redirect or error response
    from flask import Response
    filename = f"payslip_{emp_id}_{year}_{month:02d}.html"
    return Response(
        html,
        mimetype="text/html",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


@payroll_bp.route("/admin_payslips")
@admin_required
@limiter.limit("10 per minute")
def admin_payslips():
    db = get_db_connection()
    cursor = db.cursor(buffered=True)

    cursor.execute("SELECT employee_id, name, role, COALESCE(phone,''), COALESCE(email,'') FROM employees ORDER BY name")
    employees = cursor.fetchall()

    cursor.execute("SELECT COUNT(*) FROM leave_requests WHERE status='Pending'")
    pending_leaves = cursor.fetchone()[0]
    cursor.execute("SELECT COUNT(*) FROM resignation_requests WHERE status='Pending'")
    pending_resignations = cursor.fetchone()[0]
    cursor.execute("SELECT COUNT(*) FROM tickets WHERE status IN ('Open','In Progress')")
    pending_tickets = cursor.fetchone()[0]

    cursor.close()
    db.close()

    today = datetime.date.today()
    slip_months = []
    y, m = today.year, today.month
    for _ in range(12):
        slip_months.append((y, m, calendar.month_name[m]))
        m -= 1
        if m == 0:
            m = 12
            y -= 1

    return render_template("admin_payslips.html",
                           employees=employees,
                           slip_months=slip_months,
                           pending_leaves=pending_leaves,
                           pending_resignations=pending_resignations,
                           pending_tickets=pending_tickets
                           )


@payroll_bp.route("/payroll_settings", methods=["GET", "POST"])
@role_required("admin")
@limiter.limit("20 per minute")
def payroll_settings():
    db = get_db_connection()
    cursor = db.cursor(buffered=True)

    # Ensure at least one row exists
    cursor.execute("SELECT COUNT(*) FROM payroll_config")
    if cursor.fetchone()[0] == 0:
        cursor.execute(
            "INSERT INTO payroll_config (pf_employee_pct, pf_employer_pct, professional_tax, tds_annual_pct, pf_basic_cap) VALUES (12,12,200,0,15000)")
        db.commit()

    if request.method == "POST":
        try:
            # nan-injection findings below: float() accepts the literal
            # string "nan" without raising, so the try/except here doesn't
            # catch it -- the math.isnan() guard a few lines down is what
            # actually rejects it, not this try block.
            pf_emp = float(request.form.get("pf_employee_pct", 12))  # nosemgrep: python.flask.security.injection.nan-injection.nan-injection
            pf_er = float(request.form.get("pf_employer_pct", 12))  # nosemgrep: python.flask.security.injection.nan-injection.nan-injection
            pt = float(request.form.get("professional_tax", 200))  # nosemgrep: python.flask.security.injection.nan-injection.nan-injection
            tds = float(request.form.get("tds_annual_pct", 0))  # nosemgrep: python.flask.security.injection.nan-injection.nan-injection
            pf_cap = float(request.form.get("pf_basic_cap", 15000))  # nosemgrep: python.flask.security.injection.nan-injection.nan-injection
        except (ValueError, TypeError):
            flash("Invalid values.", "error")
            cursor.close()
            db.close()
            return redirect("/payroll_settings")
        # NaN would silently corrupt every downstream payroll calculation
        # that compares against these values (NaN comparisons are always
        # False) -- reject it explicitly since float() alone won't.
        if any(math.isnan(v) for v in (pf_emp, pf_er, pt, tds, pf_cap)):
            flash("Invalid values.", "error")
            cursor.close()
            db.close()
            return redirect("/payroll_settings")
        cursor.execute("""
            UPDATE payroll_config SET pf_employee_pct=%s, pf_employer_pct=%s,
            professional_tax=%s, tds_annual_pct=%s, pf_basic_cap=%s
        """, (pf_emp, pf_er, pt, tds, pf_cap))
        db.commit()

        # Update per-employee monthly CTC / basic_pct if submitted — batched
        # into a single multi-row upsert instead of one INSERT round trip
        # per employee; EXCLUDED.* lets Postgres reference each row's own
        # proposed values instead of repeating each param twice.
        emp_ids = request.form.getlist("emp_id")
        rows = []
        for eid in emp_ids:
            ctc = request.form.get(f"ctc_{eid}", "")
            bpct = request.form.get(f"bpct_{eid}", "50")
            if ctc:
                spd = round(float(ctc) / 26, 2)
                rows.append((eid, spd, ctc, bpct))
        if rows:
            placeholders = ",".join(["(%s,%s,%s,%s)"] * len(rows))
            params = [v for row in rows for v in row]
            cursor.execute(
                "INSERT INTO salary_config (employee_id, salary_per_day, monthly_ctc, basic_pct) "  # nosec B608
                f"VALUES {placeholders} "
                "ON CONFLICT (employee_id) DO UPDATE SET "
                "salary_per_day=EXCLUDED.salary_per_day, monthly_ctc=EXCLUDED.monthly_ctc, "
                "basic_pct=EXCLUDED.basic_pct",
                params
            )
        db.commit()
        flash("Payroll settings saved.", "success")
        cursor.close()
        db.close()
        return redirect("/payroll_settings")

    cursor.execute(
        "SELECT pf_employee_pct, pf_employer_pct, professional_tax, tds_annual_pct, pf_basic_cap FROM payroll_config LIMIT 1")
    cfg = cursor.fetchone() or (12, 12, 200, 0, 15000)

    cursor.execute("""
        SELECT e.employee_id, e.name, e.role, e.department,
               COALESCE(s.monthly_ctc, 0), COALESCE(s.salary_per_day, 0), COALESCE(s.basic_pct, 50)
        FROM employees e
        LEFT JOIN salary_config s ON e.employee_id = s.employee_id
        ORDER BY e.name
    """)
    employees = cursor.fetchall()

    cursor.execute("SELECT COUNT(*) FROM leave_requests WHERE status='Pending'")
    pending_leaves = cursor.fetchone()[0]
    cursor.execute("SELECT COUNT(*) FROM resignation_requests WHERE status='Pending'")
    pending_resignations = cursor.fetchone()[0]
    cursor.execute("SELECT COUNT(*) FROM tickets WHERE status IN ('Open','In Progress')")
    pending_tickets = cursor.fetchone()[0]
    cursor.execute("SELECT COALESCE(company_name,'') FROM company_settings LIMIT 1")
    co = cursor.fetchone()
    cursor.close()
    db.close()

    return render_template("payroll_settings.html",
                           cfg=cfg, employees=employees,
                           pending_leaves=pending_leaves,
                           pending_resignations=pending_resignations,
                           pending_tickets=pending_tickets,
                           co=co
                           )


# ---------------- API: SHIFTS (JSON) ----------------
