"""Attendance blueprint — check-in/out, shifts, breaks, reports.

Bandit B608 audit note: the nosec-marked queries below interpolate `_co`/
`_args`, a company-scoping fragment that's always a hardcoded literal chosen
by a bool (`"AND e.company_id=%s" if active_cid else ""`) — never user
input. Actual values are always %s-bound params.
"""
import os
import datetime
import calendar
import io as _io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from flask import (
    Blueprint, request, session, redirect, jsonify, render_template, flash,
)
from extensions import limiter, app_log
from database import get_db_connection
from utils.auth import admin_required, employee_required, api_required
from utils.helpers import get_auth_config, get_company_settings, _safe_redirect, _safe_referrer_redirect, co_scope_column, decrypt_pii
from utils.email_utils import get_email_config, send_email_smtp
from utils.attendance_utils import (
    classify_by_worked_minutes, detect_overtime, get_working_days,
    fetch_holidays_set, get_employee_shift, _td_to_time, infer_type_legacy,
    is_within_range,
)
from utils.face_utils import face_recognition, _face_recognition_available, _get_known_face_encoding
from utils.webauthn_utils import _wa_fingerprint_recently_verified
import utils.config as cfg

attendance_bp = Blueprint("attendance", __name__)


def _today_pending_counts(cursor):
    cursor.execute("SELECT COUNT(*) FROM leave_requests WHERE status='Pending'")
    pl = cursor.fetchone()[0]
    cursor.execute("SELECT COUNT(*) FROM resignation_requests WHERE status='Pending'")
    pr = cursor.fetchone()[0]
    cursor.execute("SELECT COUNT(*) FROM tickets WHERE status IN ('Open','In Progress')")
    pt = cursor.fetchone()[0]
    return pl, pr, pt


@attendance_bp.route("/today_present")
@admin_required
def today_present():
    db = get_db_connection()
    cursor = db.cursor(buffered=True)
    today = datetime.date.today()
    active_cid = session.get("active_company_id")
    _co, _co_args = co_scope_column(active_cid, alias="e")
    _args = (today,) + _co_args
    cursor.execute(f"""
        SELECT e.employee_id, e.name, e.role, a.login_time, a.logout_time,
               a.status, a.logout_status, a.attendance_type
        FROM employees e
        JOIN attendance a ON e.employee_id = a.employee_id AND a.date = %s
        WHERE a.login_time IS NOT NULL {_co}
        ORDER BY a.login_time
    """, _args)  # nosec B608
    rows = cursor.fetchall()
    pl, pr, pt = _today_pending_counts(cursor)
    cursor.close()
    db.close()
    return render_template("today_attendance.html",
                           filter_type="present", title="Present Today",
                           rows=rows, today=today.strftime("%d %b %Y"),
                           pending_leaves=pl, pending_resignations=pr, pending_tickets=pt)


@attendance_bp.route("/today_absent")
@admin_required
def today_absent():
    db = get_db_connection()
    cursor = db.cursor(buffered=True)
    today = datetime.date.today()
    active_cid = session.get("active_company_id")
    _co, _co_args = co_scope_column(active_cid, alias="e")
    _args = (today,) + _co_args
    cursor.execute(f"""
        SELECT e.employee_id, e.name, e.role
        FROM employees e
        LEFT JOIN attendance a ON e.employee_id = a.employee_id AND a.date = %s
        WHERE a.employee_id IS NULL {_co}
        ORDER BY e.name
    """, _args)  # nosec B608
    rows = cursor.fetchall()
    pl, pr, pt = _today_pending_counts(cursor)
    cursor.close()
    db.close()
    return render_template("today_attendance.html",
                           filter_type="absent", title="Absent Today",
                           rows=rows, today=today.strftime("%d %b %Y"),
                           pending_leaves=pl, pending_resignations=pr, pending_tickets=pt)


@attendance_bp.route("/today_late")
@admin_required
def today_late():
    db = get_db_connection()
    cursor = db.cursor(buffered=True)
    today = datetime.date.today()
    active_cid = session.get("active_company_id")
    _co, _co_args = co_scope_column(active_cid, alias="e")
    _args = (today,) + _co_args
    cursor.execute(f"""
        SELECT e.employee_id, e.name, e.role, a.login_time, a.status
        FROM employees e
        JOIN attendance a ON e.employee_id = a.employee_id AND a.date = %s
        WHERE a.status IN ('Late Login', 'Half Day Login') {_co}
        ORDER BY a.login_time
    """, _args)  # nosec B608
    rows = cursor.fetchall()
    pl, pr, pt = _today_pending_counts(cursor)
    cursor.close()
    db.close()
    return render_template("today_attendance.html",
                           filter_type="late", title="Late Logins Today",
                           rows=rows, today=today.strftime("%d %b %Y"),
                           pending_leaves=pl, pending_resignations=pr, pending_tickets=pt)


@attendance_bp.route("/shifts", methods=["GET"])
@admin_required
def shifts():
    return redirect("/settings?tab=shifts")


@attendance_bp.route("/add_shift", methods=["POST"])
@admin_required
def add_shift():
    name = (request.form.get("shift_name") or request.form.get("name", "")).strip()
    start = request.form.get("start_time", "").strip()
    half = request.form.get("half_time", "").strip()
    end = request.form.get("end_time", "").strip()
    dest = request.form.get("redirect") or (
        "/settings?tab=shifts" if request.form.get("redirect_to") == "settings" else "/settings?tab=shifts")
    cid_raw = request.form.get("company_id", "").strip()
    company_id = int(cid_raw) if cid_raw.isdigit() else None
    if not all([name, start, half, end]):
        return redirect(dest)
    db = get_db_connection()
    cursor = db.cursor(buffered=True)
    try:
        if company_id:
            cursor.execute(
                "INSERT INTO shifts (name, start_time, half_time, end_time, company_id) VALUES (%s,%s,%s,%s,%s)",
                (name, start, half, end, company_id)
            )
        else:
            cursor.execute(
                "INSERT INTO shifts (name, start_time, half_time, end_time) VALUES (%s,%s,%s,%s)",
                (name, start, half, end)
            )
        db.commit()
    except Exception:
        pass
    cursor.close()
    db.close()
    return redirect(dest)


@attendance_bp.route("/delete_shift", methods=["POST"])
@admin_required
def delete_shift_form():
    sid = request.form.get("shift_id", "").strip()
    dest = request.form.get("redirect") or "/settings?tab=shifts"
    if not sid:
        return redirect(dest)
    db = get_db_connection()
    cursor = db.cursor(buffered=True)
    cursor.execute("UPDATE employees SET shift_id=NULL WHERE shift_id=%s", (sid,))
    cursor.execute("UPDATE break_config SET shift_id=NULL WHERE shift_id=%s", (sid,))
    cursor.execute("DELETE FROM shifts WHERE id=%s", (sid,))
    db.commit()
    cursor.close()
    db.close()
    return redirect(dest)


@attendance_bp.route("/delete_shift/<int:sid>", methods=["POST"])
@admin_required
def delete_shift(sid):
    db = get_db_connection()
    cursor = db.cursor(buffered=True)
    cursor.execute("UPDATE employees SET shift_id=NULL WHERE shift_id=%s", (sid,))
    cursor.execute("UPDATE break_config SET shift_id=NULL WHERE shift_id=%s", (sid,))
    cursor.execute("DELETE FROM shifts WHERE id=%s", (sid,))
    db.commit()
    cursor.close()
    db.close()
    dest = request.form.get("redirect") or "/settings?tab=shifts"
    return redirect(dest)


@attendance_bp.route("/edit_shift", methods=["POST"])
@attendance_bp.route("/edit_shift/<int:sid>", methods=["POST"])
@admin_required
def edit_shift(sid=None):
    if sid is None:
        try:
            sid = int(request.form.get("shift_id", ""))
        except (ValueError, TypeError):
            return redirect("/employees?tab=schedule")
    name = (request.form.get("shift_name") or request.form.get("name", "")).strip()
    start = request.form.get("start_time", "").strip()
    half = request.form.get("half_time", "").strip()
    end = request.form.get("end_time", "").strip()
    dest = request.form.get("redirect") or "/settings?tab=shifts"
    if not all([name, start, half, end]):
        return redirect(dest)
    db = get_db_connection()
    cursor = db.cursor(buffered=True)
    cursor.execute(
        "UPDATE shifts SET name=%s, start_time=%s, half_time=%s, end_time=%s WHERE id=%s",
        (name, start, half, end, sid)
    )
    db.commit()
    cursor.close()
    db.close()
    return redirect(dest)


@attendance_bp.route("/bulk_assign_shift", methods=["POST"])
@admin_required
def bulk_assign_shift():
    shift_id = request.form.get("shift_id", "").strip()
    emp_ids = request.form.getlist("emp_ids")
    dept_filter = request.form.get("dept_filter", "").strip()
    dest = request.form.get("redirect") or "/employees?tab=schedule"
    db = get_db_connection()
    cursor = db.cursor(buffered=True)
    if emp_ids:
        cursor.execute(
            "UPDATE employees SET shift_id=%s WHERE employee_id = ANY(%s)",
            (shift_id if shift_id else None, emp_ids)
        )
    elif dept_filter:
        cursor.execute(
            "UPDATE employees SET shift_id=%s WHERE department=%s",
            (shift_id if shift_id else None, dept_filter)
        )
    else:
        cursor.execute("UPDATE employees SET shift_id=%s", (shift_id if shift_id else None,))
    db.commit()
    cursor.close()
    db.close()
    return redirect(dest)


@attendance_bp.route("/update_default_shift", methods=["POST"])
@admin_required
def update_default_shift():
    # No `global` needed — cfg.SHIFT_START/HALF/END live in utils.config now;
    # cfg.load_default_shift() below mutates that module's own globals via
    # its own correctly-scoped `global` statement, not this function's.
    start = request.form.get("shift_start", "").strip()
    half = request.form.get("shift_half", "").strip()
    end = request.form.get("shift_end", "").strip()
    if not all([start, half, end]):
        return redirect("/shifts?error=All+fields+required")
    db = get_db_connection()
    cursor = db.cursor(buffered=True)
    cursor.execute(
        "UPDATE company_settings SET shift_start=%s, shift_half=%s, shift_end=%s",
        (start, half, end)
    )
    db.commit()
    cursor.close()
    db.close()
    cfg.load_default_shift()
    return redirect("/shifts?default_saved=1")


@attendance_bp.route("/assign_shift", methods=["POST"])
@admin_required
def assign_shift():
    emp_id = request.form.get("emp_id", "").strip()
    shift_id = request.form.get("shift_id", "").strip()
    db = get_db_connection()
    cursor = db.cursor(buffered=True)
    cursor.execute(
        "UPDATE employees SET shift_id=%s WHERE employee_id=%s",
        (shift_id if shift_id else None, emp_id)
    )
    db.commit()
    cursor.close()
    db.close()
    return jsonify({"ok": True})


@attendance_bp.route("/submit_shift_swap", methods=["POST"])
@employee_required
def submit_shift_swap():
    requester_id = session["employee_id"]
    target_id = request.form.get("target_id", "").strip()
    reason = request.form.get("reason", "").strip()
    if not target_id or target_id == requester_id:
        return redirect("/employee_portal?swap_error=invalid_target#shift-swap")
    db = get_db_connection()
    cursor = db.cursor(buffered=True)
    # Fetch both employees' current shift_id (must both have shifts assigned)
    cursor.execute("SELECT shift_id FROM employees WHERE employee_id=%s", (requester_id,))
    row_r = cursor.fetchone()
    cursor.execute("SELECT shift_id FROM employees WHERE employee_id=%s", (target_id,))
    row_t = cursor.fetchone()
    if not row_r or not row_t or row_r[0] is None or row_t[0] is None:
        cursor.close()
        db.close()
        return redirect("/employee_portal?swap_error=no_shift#shift-swap")
    if row_r[0] == row_t[0]:
        cursor.close()
        db.close()
        return redirect("/employee_portal?swap_error=same_shift#shift-swap")
    # Check no open request already exists between them
    cursor.execute("""
        SELECT id FROM shift_swap_requests
        WHERE requester_id=%s AND target_id=%s
          AND status IN ('Pending_Target','Pending_Admin')
    """, (requester_id, target_id))
    if cursor.fetchone():
        cursor.close()
        db.close()
        return redirect("/employee_portal?swap_error=duplicate#shift-swap")
    cursor.execute("""
        INSERT INTO shift_swap_requests
            (requester_id, target_id, requester_shift_id, target_shift_id, reason)
        VALUES (%s, %s, %s, %s, %s)
    """, (requester_id, target_id, row_r[0], row_t[0], reason))
    db.commit()
    cursor.close()
    db.close()
    return redirect("/employee_portal?swap_sent=1#shift-swap")


@attendance_bp.route("/respond_shift_swap/<int:req_id>", methods=["POST"])
@employee_required
def respond_shift_swap(req_id):
    emp_id = session["employee_id"]
    action = request.form.get("action", "")
    response = request.form.get("response", "").strip()
    db = get_db_connection()
    cursor = db.cursor(buffered=True)
    cursor.execute("""
        SELECT id, requester_id, target_id, status
        FROM shift_swap_requests WHERE id=%s AND target_id=%s AND status='Pending_Target'
    """, (req_id, emp_id))
    row = cursor.fetchone()
    if not row:
        cursor.close()
        db.close()
        return redirect("/employee_portal?swap_error=not_found#shift-swap")
    if action == "accept":
        cursor.execute("""
            UPDATE shift_swap_requests SET status='Pending_Admin', target_response=%s WHERE id=%s
        """, (response or "Accepted", req_id))
    else:
        cursor.execute("""
            UPDATE shift_swap_requests SET status='Rejected', target_response=%s WHERE id=%s
        """, (response or "Rejected by employee", req_id))
    db.commit()
    cursor.close()
    db.close()
    return redirect("/employee_portal?swap_responded=1#shift-swap")


@attendance_bp.route("/admin_shift_swap/<int:req_id>", methods=["POST"])
@admin_required
def admin_shift_swap(req_id):
    action = request.form.get("action", "")
    response = request.form.get("admin_response", "").strip()
    db = get_db_connection()
    cursor = db.cursor(buffered=True)
    cursor.execute("""
        SELECT requester_id, target_id, requester_shift_id, target_shift_id
        FROM shift_swap_requests WHERE id=%s AND status='Pending_Admin'
    """, (req_id,))
    row = cursor.fetchone()
    if not row:
        cursor.close()
        db.close()
        return redirect("/admin_shift_swaps?error=not_found")
    requester_id, target_id, req_shift, tgt_shift = row
    if action == "approve":
        # Swap actual shift assignments
        cursor.execute("UPDATE employees SET shift_id=%s WHERE employee_id=%s", (tgt_shift, requester_id))
        cursor.execute("UPDATE employees SET shift_id=%s WHERE employee_id=%s", (req_shift, target_id))
        cursor.execute("""
            UPDATE shift_swap_requests SET status='Approved', admin_response=%s WHERE id=%s
        """, (response or "Approved by admin", req_id))
    else:
        cursor.execute("""
            UPDATE shift_swap_requests SET status='Rejected_Admin', admin_response=%s WHERE id=%s
        """, (response or "Rejected by admin", req_id))
    db.commit()
    cursor.close()
    db.close()
    return redirect("/admin_shift_swaps?ok=1")


@attendance_bp.route("/admin_shift_swaps")
@admin_required
def admin_shift_swaps():
    db = get_db_connection()
    cursor = db.cursor(buffered=True)
    cursor.execute("""
        SELECT ssr.id, ssr.requester_id, er.name, ssr.target_id, et.name,
               sr.name AS req_shift, st.name AS tgt_shift,
               ssr.reason, ssr.status, ssr.target_response, ssr.admin_response, ssr.created_at
        FROM shift_swap_requests ssr
        JOIN employees er ON er.employee_id = ssr.requester_id
        JOIN employees et ON et.employee_id = ssr.target_id
        JOIN shifts sr ON sr.id = ssr.requester_shift_id
        JOIN shifts st ON st.id = ssr.target_shift_id
        ORDER BY ssr.created_at DESC LIMIT 100
    """)
    swap_rows = cursor.fetchall()
    cursor.close()
    db.close()
    return render_template("admin_shift_swaps.html", swap_rows=swap_rows,
                           ok=request.args.get("ok"), error=request.args.get("error"))


@attendance_bp.route("/api/breaks")
@limiter.limit("30 per minute")
def api_breaks():
    if not (session.get("admin_logged_in") or session.get("employee_id")):
        auth = request.headers.get("Authorization", "")
        if not auth.startswith("Bearer "):
            return jsonify({"ok": False, "msg": "Unauthorized"}), 401
    db = get_db_connection()
    cursor = db.cursor(buffered=True)
    cursor.execute("SELECT id, break_name, break_time, duration_minutes FROM break_config WHERE is_active=1 ORDER BY break_time")
    rows = cursor.fetchall()
    cursor.close()
    db.close()
    result = []
    for row in rows:
        bt = row[2]
        if hasattr(bt, 'seconds'):
            total = bt.seconds
            h, m = divmod(total // 60, 60)
        else:
            h, m = bt.hour, bt.minute
        result.append({"id": row[0], "name": row[1],
                       "hour": h, "minute": m,
                       "duration": row[3]})
    return jsonify(result)


@attendance_bp.route("/break_config")
@admin_required
def view_break_config():
    return redirect("/settings?tab=shifts")


@attendance_bp.route("/add_break", methods=["POST"])
@admin_required
def add_break():
    name = request.form.get("break_name", "").strip()
    btime = request.form.get("break_time", "").strip()
    duration = int(request.form.get("duration_minutes", 10) or 10)
    dest = _safe_redirect(request.form.get("redirect", ""), _safe_referrer_redirect(
        request.referrer or "", "/employees?tab=schedule"))
    cid_raw = request.form.get("company_id", "").strip()
    company_id = int(cid_raw) if cid_raw.isdigit() else None
    sid_raw = request.form.get("shift_id", "").strip()
    shift_id = int(sid_raw) if sid_raw.isdigit() else None
    if not name or not btime:
        flash("Break name and time are required.", "error")
        return redirect(dest)
    db = get_db_connection()
    cursor = db.cursor(buffered=True)
    if company_id:
        cursor.execute(
            "INSERT INTO break_config (break_name, break_time, duration_minutes, company_id, shift_id) VALUES (%s,%s,%s,%s,%s)",
            (name, btime, duration, company_id, shift_id)
        )
    else:
        cursor.execute("INSERT INTO break_config (break_name, break_time, duration_minutes, shift_id) VALUES (%s,%s,%s,%s)",
                       (name, btime, duration, shift_id))
    db.commit()
    cursor.close()
    db.close()
    flash("Break added successfully.", "success")
    return redirect(dest)


@attendance_bp.route("/update_break", methods=["POST"])
@attendance_bp.route("/update_break/<int:bid>", methods=["POST"])
@admin_required
def update_break(bid=None):
    if bid is None:
        try:
            bid = int(request.form.get("break_id", ""))
        except (ValueError, TypeError):
            return redirect("/employees?tab=schedule")
    name = request.form.get("break_name", "").strip()
    btime = request.form.get("break_time", "").strip()
    duration = int(request.form.get("duration_minutes", 10) or 10)
    active = 1 if request.form.get("is_active") else 0
    dest = _safe_redirect(request.form.get("redirect", ""), _safe_referrer_redirect(
        request.referrer or "", "/employees?tab=schedule"))
    sid_raw = request.form.get("shift_id", "").strip()
    if not name or not btime:
        flash("Break name and time are required.", "error")
        return redirect(dest)
    db = get_db_connection()
    cursor = db.cursor(buffered=True)
    if sid_raw.isdigit():
        cursor.execute(
            "UPDATE break_config SET break_name=%s, break_time=%s, duration_minutes=%s, is_active=%s, shift_id=%s WHERE id=%s",
            (name, btime, duration, active, int(sid_raw), bid)
        )
    else:
        cursor.execute(
            "UPDATE break_config SET break_name=%s, break_time=%s, duration_minutes=%s, is_active=%s WHERE id=%s",
            (name, btime, duration, active, bid)
        )
    db.commit()
    cursor.close()
    db.close()
    flash("Break updated.", "success")
    return redirect(dest)


@attendance_bp.route("/delete_break", methods=["POST"])
@attendance_bp.route("/delete_break/<int:bid>", methods=["POST"])
@admin_required
def delete_break(bid=None):
    if bid is None:
        try:
            bid = int(request.form.get("break_id", ""))
        except (ValueError, TypeError):
            return redirect("/employees?tab=schedule")
    dest = request.form.get("redirect") or "/employees?tab=schedule"
    db = get_db_connection()
    cursor = db.cursor(buffered=True)
    cursor.execute("DELETE FROM break_config WHERE id=%s", (bid,))
    db.commit()
    cursor.close()
    db.close()
    flash("Break deleted.", "success")
    return redirect(dest)


@attendance_bp.route("/monthly_report")
@admin_required
def monthly_report():
    year = int(request.args.get("year", datetime.date.today().year))
    month = int(request.args.get("month", datetime.date.today().month))

    db = get_db_connection()
    cursor = db.cursor(buffered=True)

    active_cid = session.get("active_company_id")
    if active_cid:
        cursor.execute(
            "SELECT employee_id, name, COALESCE(role,''), COALESCE(phone,''), COALESCE(email,'') FROM employees WHERE company_id=%s ORDER BY name", (active_cid,))
    else:
        cursor.execute(
            "SELECT employee_id, name, COALESCE(role,''), COALESCE(phone,''), COALESCE(email,'') FROM employees ORDER BY name")
    employees = cursor.fetchall()

    _, last_day = calendar.monthrange(year, month)
    cursor.execute("""
        SELECT employee_id, date, login_time, logout_time, status, logout_status, attendance_type
        FROM attendance
        WHERE date BETWEEN %s AND %s
    """, (datetime.date(year, month, 1), datetime.date(year, month, last_day)))

    att_map = {}
    for row in cursor.fetchall():
        att_map.setdefault(row[0], {})[row[1]] = row

    holidays = fetch_holidays_set(year, month)
    working_days = get_working_days(year, month)
    today = datetime.date.today()

    report = []
    for emp_id, name, role, phone, email in employees:
        emp_att = att_map.get(emp_id, {})
        full_days = half_days = late_days = absent = 0

        for d in working_days:
            if d > today or d in holidays:
                continue
            row = emp_att.get(d)
            if row:
                _, _, login_t, logout_t, status, _logout_status, att_type = row
                final = att_type if att_type else infer_type_legacy(status, login_t, logout_t)
                if final == "Full Day":
                    full_days += 1
                elif final == "Late - Full Day":
                    late_days += 1
                elif final in ("Half Day", "Present"):
                    half_days += 1
                else:
                    absent += 1
            else:
                absent += 1

        billable = len([d for d in working_days if d <= today and d not in holidays])
        present_equiv = full_days + late_days + half_days * 0.5
        pct = round(present_equiv / billable * 100, 1) if billable > 0 else 0

        report.append({
            "emp_id": emp_id,
            "name": name,
            "role": role,
            "phone": phone,
            "email": email,
            "full_days": full_days,
            "half_days": half_days,
            "late_days": late_days,
            "absent": absent,
            "billable": billable,
            "pct": pct,
        })

    cursor.close()
    db.close()

    months = [(i, datetime.date(year, i, 1).strftime("%B")) for i in range(1, 13)]
    years = list(range(datetime.date.today().year - 2, datetime.date.today().year + 1))

    return render_template("monthly_report.html",
                           report=report,
                           month_name=datetime.date(year, month, 1).strftime("%B %Y"),
                           year=year, month=month,
                           months=months, years=years,
                           holiday_count=len(holidays),
                           total_working=len([d for d in working_days if d <= today and d not in holidays]),
                           )


@attendance_bp.route("/employee_attendance_detail/<emp_id>/<int:year>/<int:month>")
@admin_required
def employee_attendance_detail(emp_id, year, month):
    db = get_db_connection()
    cursor = db.cursor(buffered=True)

    cursor.execute(
        "SELECT employee_id, name, COALESCE(role,''), COALESCE(phone,''), COALESCE(email,'') "
        "FROM employees WHERE employee_id = %s", (emp_id,)
    )
    emp = cursor.fetchone()
    if not emp:
        cursor.close()
        db.close()
        return "Employee not found", 404

    _, last_day = calendar.monthrange(year, month)
    cursor.execute("""
        SELECT date, login_time, logout_time, status, logout_status, attendance_type
        FROM attendance
        WHERE employee_id = %s AND date BETWEEN %s AND %s
        ORDER BY date
    """, (emp_id, datetime.date(year, month, 1), datetime.date(year, month, last_day)))
    att_map = {row[0]: row for row in cursor.fetchall()}

    holidays_set = fetch_holidays_set(year, month)
    today = datetime.date.today()

    days = []
    full_days = half_days = late_days = absent = 0
    for d in range(1, last_day + 1):
        date = datetime.date(year, month, d)
        is_sunday = date.weekday() == 6
        is_holiday = date in holidays_set
        is_future = date > today
        row = att_map.get(date)

        if row:
            _, login_t, logout_t, status, logout_status, att_type = row
            final = att_type if att_type else infer_type_legacy(status, login_t, logout_t)
            login_str = _td_to_time(login_t).strftime("%I:%M %p") if login_t else "—"
            logout_str = _td_to_time(logout_t).strftime("%I:%M %p") if logout_t else "—"
            if not is_future:
                if final == "Full Day":
                    full_days += 1
                elif final == "Late - Full Day":
                    late_days += 1
                elif final in ("Half Day", "Present"):
                    half_days += 1
                else:
                    absent += 1
        else:
            final = "—"
            login_str = "—"
            logout_str = "—"
            if not is_sunday and not is_holiday and not is_future:
                absent += 1

        days.append({
            "date": date,
            "day_name": date.strftime("%a"),
            "login": login_str,
            "logout": logout_str,
            "status": final,
            "is_sunday": is_sunday,
            "is_holiday": is_holiday,
            "is_future": is_future,
        })

    cursor.close()
    db.close()

    months = [(i, datetime.date(year, i, 1).strftime("%B")) for i in range(1, 13)]
    years = list(range(datetime.date.today().year - 2, datetime.date.today().year + 1))

    return render_template("employee_attendance_detail.html",
                           emp=emp,
                           days=days,
                           month_name=datetime.date(year, month, 1).strftime("%B %Y"),
                           year=year, month=month,
                           months=months, years=years,
                           full_days=full_days,
                           late_days=late_days,
                           half_days=half_days,
                           absent=absent,
                           )


@attendance_bp.route("/correct_attendance", methods=["POST"])
@admin_required
def correct_attendance():
    emp_id = request.form.get("emp_id", "").strip()
    date_str = request.form.get("date", "").strip()
    login_str = request.form.get("login_time", "").strip()
    logout_str = request.form.get("logout_time", "").strip()
    att_type = request.form.get("attendance_type", "").strip()
    year = request.form.get("year", "")
    month = request.form.get("month", "")

    if not emp_id or not date_str or not att_type:
        flash("Missing required fields.", "error")
        return redirect(_safe_referrer_redirect(request.referrer or "", "/monthly_report"))

    try:
        date_obj = datetime.date.fromisoformat(date_str)
    except ValueError:
        flash("Invalid date.", "error")
        return redirect(_safe_referrer_redirect(request.referrer or "", "/monthly_report"))

    login_time = login_str if login_str else None
    logout_time = logout_str if logout_str else None

    db = get_db_connection()
    cursor = db.cursor(buffered=True)
    cursor.execute(
        "SELECT id FROM attendance WHERE employee_id=%s AND date=%s",
        (emp_id, date_obj)
    )
    existing = cursor.fetchone()

    if existing:
        cursor.execute(
            "UPDATE attendance SET login_time=%s, logout_time=%s, attendance_type=%s, "
            "status='Manual', logout_status='Manual' WHERE employee_id=%s AND date=%s",
            (login_time, logout_time, att_type, emp_id, date_obj)
        )
    else:
        cursor.execute(
            "INSERT INTO attendance (employee_id, date, login_time, logout_time, "
            "attendance_type, status, logout_status) VALUES (%s,%s,%s,%s,%s,'Manual','Manual')",
            (emp_id, date_obj, login_time, logout_time, att_type)
        )
    db.commit()
    cursor.close()
    db.close()

    flash(f"Attendance updated for {date_obj.strftime('%d %b %Y')}.", "success")
    return redirect(f"/employee_attendance_detail/{emp_id}/{year}/{month}")


@attendance_bp.route("/bulk_mark_attendance", methods=["GET", "POST"])
@admin_required
def bulk_mark_attendance():
    today = datetime.date.today()

    if request.method == "POST":
        date_str = request.form.get("date", "").strip()
        try:
            date_obj = datetime.date.fromisoformat(date_str)
        except ValueError:
            flash("Invalid date.", "error")
            return redirect("/bulk_mark_attendance")

        db = get_db_connection()
        cursor = db.cursor(buffered=True)
        cursor.execute("SELECT employee_id FROM employees WHERE is_active=1")
        emp_ids = [r[0] for r in cursor.fetchall()]

        rows = []
        for eid in emp_ids:
            att_type = request.form.get(f"att_{eid}", "").strip()
            if not att_type:
                continue
            login_t = request.form.get(f"login_{eid}", "").strip() or None
            logout_t = request.form.get(f"logout_{eid}", "").strip() or None
            rows.append((eid, date_obj, login_t, logout_t, att_type))

        # Single multi-row upsert instead of one INSERT round trip per
        # employee — placeholders are a fixed repeated pattern sized off
        # len(rows), never user input, so this stays parameterized (%s).
        saved = len(rows)
        if rows:
            placeholders = ",".join(["(%s,%s,%s,%s,%s,'Manual','Manual')"] * len(rows))
            params = [v for row in rows for v in row]
            cursor.execute(
                "INSERT INTO attendance (employee_id, date, login_time, logout_time, "  # nosec B608
                f"attendance_type, status, logout_status) VALUES {placeholders} "
                "ON CONFLICT (employee_id, date) DO UPDATE SET "
                "login_time=EXCLUDED.login_time, logout_time=EXCLUDED.logout_time, "
                "attendance_type=EXCLUDED.attendance_type, status='Manual', logout_status='Manual'",
                params
            )
        db.commit()
        cursor.close()
        db.close()
        flash(f"Attendance saved for {saved} employee(s) on {date_obj.strftime('%d %b %Y')}.", "success")
        return redirect(f"/bulk_mark_attendance?date={date_str}")

    date_str = request.args.get("date", today.isoformat())
    try:
        date_obj = datetime.date.fromisoformat(date_str)
    except ValueError:
        date_obj = today
        date_str = today.isoformat()

    db = get_db_connection()
    cursor = db.cursor(buffered=True)

    base_select = (
        "SELECT e.employee_id, e.name, COALESCE(e.department,''), COALESCE(e.designation,''), "
        "COALESCE(s.name,''), COALESCE(TO_CHAR(s.start_time,'HH24:MI'),''), "
        "COALESCE(TO_CHAR(s.end_time,'HH24:MI'),''), "
        "COALESCE(e.phone,''), COALESCE(e.email,''), "
        "COALESCE(e.work_mode,'office'), COALESCE(TO_CHAR(e.date_of_joining,'YYYY-MM-DD'),''), "
        "COALESCE(e.gender,''), COALESCE(e.role,'') "
        "FROM employees e LEFT JOIN shifts s ON s.id=e.shift_id "
    )
    active_cid = session.get("active_company_id")
    if active_cid:
        cursor.execute(base_select + "WHERE e.is_active=1 AND e.company_id=%s ORDER BY e.name", (active_cid,))
    else:
        cursor.execute(base_select + "WHERE e.is_active=1 ORDER BY e.name")
    # [11]=gender is Fernet-encrypted at rest — decrypt before display.
    employees = [row[:11] + (decrypt_pii(row[11]),) + row[12:] for row in cursor.fetchall()]

    cursor.execute(
        "SELECT employee_id, login_time, logout_time, attendance_type "
        "FROM attendance WHERE date=%s", (date_obj,)
    )
    att_map = {r[0]: r for r in cursor.fetchall()}

    # Monthly summary for the selected date's month
    cursor.execute(
        """SELECT employee_id,
             SUM(CASE WHEN attendance_type IN ('Full Day','Late - Full Day') THEN 1 ELSE 0 END) AS present_days,
             SUM(CASE WHEN attendance_type = 'Half Day' THEN 1 ELSE 0 END) AS half_days,
             SUM(CASE WHEN attendance_type = 'Absent' THEN 1 ELSE 0 END) AS absent_days,
             SUM(CASE WHEN attendance_type = 'Leave' THEN 1 ELSE 0 END) AS leave_days,
             COUNT(*) AS total_marked
           FROM attendance
           WHERE EXTRACT(YEAR FROM date)=%s AND EXTRACT(MONTH FROM date)=%s
           GROUP BY employee_id""",
        (date_obj.year, date_obj.month)
    )
    month_summary = {r[0]: r for r in cursor.fetchall()}

    co = get_company_settings()
    pending_leaves = 0
    pending_resignations = 0
    pending_tickets = 0
    try:
        cursor.execute("SELECT COUNT(*) FROM leave_requests WHERE status='Pending'")
        pending_leaves = cursor.fetchone()[0]
        cursor.execute("SELECT COUNT(*) FROM resignation_requests WHERE status='Pending'")
        pending_resignations = cursor.fetchone()[0]
        cursor.execute("SELECT COUNT(*) FROM tickets WHERE status='Open'")
        pending_tickets = cursor.fetchone()[0]
    except Exception:
        pass
    cursor.close()
    db.close()

    return render_template("bulk_attendance.html",
                           co=co, employees=employees, att_map=att_map,
                           month_summary=month_summary,
                           date_str=date_str, date_obj=date_obj,
                           today=today, pending_leaves=pending_leaves,
                           pending_resignations=pending_resignations,
                           pending_tickets=pending_tickets,
                           )


@attendance_bp.route("/monthly_report_export")
@admin_required
@limiter.limit("10 per minute")
def monthly_report_export():
    from flask import send_file
    year = int(request.args.get("year", datetime.date.today().year))
    month = int(request.args.get("month", datetime.date.today().month))

    db = get_db_connection()
    cursor = db.cursor(buffered=True)
    cursor.execute("SELECT employee_id, name FROM employees ORDER BY name")
    employees = cursor.fetchall()

    _, last_day = calendar.monthrange(year, month)
    cursor.execute("""
        SELECT employee_id, date, login_time, logout_time, status, logout_status, attendance_type
        FROM attendance WHERE date BETWEEN %s AND %s
    """, (datetime.date(year, month, 1), datetime.date(year, month, last_day)))

    att_map = {}
    for row in cursor.fetchall():
        att_map.setdefault(row[0], {})[row[1]] = row

    holidays = fetch_holidays_set(year, month)
    working_days = get_working_days(year, month)
    today = datetime.date.today()
    cursor.close()
    db.close()

    report = []
    for emp_id, name in employees:
        emp_att = att_map.get(emp_id, {})
        full_days = half_days = late_days = absent = 0
        for d in working_days:
            if d > today or d in holidays:
                continue
            row = emp_att.get(d)
            if row:
                _, _, login_t, logout_t, status, _ls, att_type = row
                final = att_type if att_type else infer_type_legacy(status, login_t, logout_t)
                if final == "Full Day":
                    full_days += 1
                elif final == "Late - Full Day":
                    late_days += 1
                elif final in ("Half Day", "Present"):
                    half_days += 1
                else:
                    absent += 1
            else:
                absent += 1
        billable = len([d for d in working_days if d <= today and d not in holidays])
        present_equiv = full_days + late_days + half_days * 0.5
        pct = round(present_equiv / billable * 100, 1) if billable > 0 else 0
        report.append({"emp_id": emp_id, "name": name, "full_days": full_days,
                       "late_days": late_days, "half_days": half_days,
                       "absent": absent, "billable": billable, "pct": pct})

    month_name = datetime.date(year, month, 1).strftime("%B %Y")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance Report"

    # ── styles ──
    hdr_fill = PatternFill("solid", fgColor="1E3A8A")
    hdr_font = Font(color="FFFFFF", bold=True, size=11)
    title_font = Font(bold=True, size=13, color="1E3A8A")
    center = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="DBEAFE")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    alt_fill = PatternFill("solid", fgColor="EFF6FF")

    # ── title row ──
    ws.merge_cells("A1:H1")
    ws["A1"] = f"Monthly Attendance Report — {month_name}"
    ws["A1"].font = title_font
    ws["A1"].alignment = center
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:H2")
    ws["A2"] = f"Working Days: {len([d for d in working_days if d <= today and d not in holidays])}   |   Holidays: {len(holidays)}   |   Employees: {len(report)}"
    ws["A2"].alignment = center
    ws["A2"].font = Font(size=10, color="64748B")
    ws.row_dimensions[2].height = 18

    # ── header row ──
    headers = ["Emp ID", "Name", "Full Days", "Late Days", "Half Days", "Absent", "Working Days", "Attendance %"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = center
        cell.border = border
    ws.row_dimensions[3].height = 22

    # ── data rows ──
    for i, r in enumerate(report, 4):
        row_fill = alt_fill if i % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        values = [r["emp_id"], r["name"], r["full_days"], r["late_days"],
                  r["half_days"], r["absent"], r["billable"], r["pct"]]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.fill = row_fill
            cell.alignment = center if col != 2 else Alignment(horizontal="left", vertical="center")
            cell.border = border
            if col == 8:  # Attendance %
                pct_val = val
                if pct_val >= 90:
                    cell.font = Font(color="15803D", bold=True)
                elif pct_val >= 70:
                    cell.font = Font(color="D97706", bold=True)
                else:
                    cell.font = Font(color="DC2626", bold=True)

    # ── column widths ──
    col_widths = [12, 24, 12, 12, 12, 10, 14, 14]
    for col, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"attendance_{year}_{month:02d}.xlsx"
    return send_file(buf, as_attachment=True, download_name=filename,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@attendance_bp.route("/send_absentee_report", methods=["POST"])
@admin_required
def send_absentee_report():
    cfg = get_email_config()
    if not cfg:
        return jsonify({"ok": False, "msg": "Email not configured. Go to Email Settings first."})

    today = datetime.date.today()
    db = get_db_connection()
    cursor = db.cursor(buffered=True)

    cursor.execute("SELECT employee_id, name FROM employees ORDER BY name")
    all_emp = cursor.fetchall()

    cursor.execute("SELECT DISTINCT employee_id FROM attendance WHERE date=%s", (today,))
    present_ids = {r[0] for r in cursor.fetchall()}
    cursor.close()
    db.close()

    absentees = [(eid, nm) for eid, nm in all_emp if eid not in present_ids]
    total = len(all_emp)
    absent = len(absentees)
    present = total - absent

    rows_html = "".join(
        f"<tr><td style='padding:8px 14px;border-bottom:1px solid #e2e8f0;'>{eid}</td>"
        f"<td style='padding:8px 14px;border-bottom:1px solid #e2e8f0;'>{nm}</td></tr>"
        for eid, nm in absentees
    ) or "<tr><td colspan='2' style='padding:14px;text-align:center;color:#16a34a;'>All employees present!</td></tr>"

    html = f"""
<div style="font-family:Segoe UI,sans-serif;max-width:600px;margin:auto;background:#f8fafc;border-radius:16px;overflow:hidden;border:1px solid #dbeafe;">
  <div style="background:#1e3a8a;padding:24px 28px;color:white;">
    <div style="font-size:20px;font-weight:700;">&#127970; Daily Absentee Report</div>
    <div style="font-size:13px;opacity:0.75;margin-top:4px;">{today.strftime('%A, %d %B %Y')}</div>
  </div>
  <div style="padding:24px;">
    <div style="display:flex;gap:16px;margin-bottom:24px;">
      <div style="flex:1;background:#dcfce7;border-radius:10px;padding:16px;text-align:center;">
        <div style="font-size:28px;font-weight:700;color:#15803d;">{present}</div>
        <div style="font-size:12px;color:#166534;">Present</div>
      </div>
      <div style="flex:1;background:#fee2e2;border-radius:10px;padding:16px;text-align:center;">
        <div style="font-size:28px;font-weight:700;color:#dc2626;">{absent}</div>
        <div style="font-size:12px;color:#991b1b;">Absent</div>
      </div>
      <div style="flex:1;background:#dbeafe;border-radius:10px;padding:16px;text-align:center;">
        <div style="font-size:28px;font-weight:700;color:#1d4ed8;">{total}</div>
        <div style="font-size:12px;color:#1e40af;">Total</div>
      </div>
    </div>
    <table style="width:100%;border-collapse:collapse;background:white;border-radius:10px;overflow:hidden;border:1px solid #dbeafe;">
      <thead>
        <tr style="background:#dbeafe;">
          <th style="padding:10px 14px;text-align:left;color:#1e3a8a;font-size:13px;">Employee ID</th>
          <th style="padding:10px 14px;text-align:left;color:#1e3a8a;font-size:13px;">Name</th>
        </tr>
      </thead>
      <tbody>{rows_html}</tbody>
    </table>
  </div>
</div>"""

    try:
        send_email_smtp(cfg.get("from_email", cfg["user"]),
                        f"Daily Absentee Report — {today.strftime('%d %b %Y')}", html, cfg)
        return jsonify({"ok": True, "msg": f"Report sent! {absent} absent out of {total} employees."})
    except Exception:
        app_log.error("Failed to send absentee report email", exc_info=True)
        return jsonify({"ok": False, "msg": "Failed to send email. Check email settings."})


@attendance_bp.route("/location", methods=["POST"])
def location():
    data = request.get_json(silent=True) or {}
    lat, lon = data.get("lat"), data.get("lon")
    if lat is None or lon is None:
        return jsonify({"status": "error", "msg": "lat and lon required"}), 400
    session["lat"] = lat
    session["lon"] = lon
    return jsonify({"status": "ok"})


@attendance_bp.route("/attendance", methods=["POST"])
def attendance():
    import base64
    import io
    import numpy as np
    from PIL import Image

    data = request.get_json() or {}
    emp_id = data.get("employee_id", "").strip()
    face_b64 = data.get("face_image", "")
    user_lat = data.get("lat")
    user_lon = data.get("lon")
    auth_combo = data.get("auth_combo", "qr_face")

    if auth_combo not in ("qr_face", "qr_only", "qr_fingerprint", "fingerprint_only"):
        return jsonify({"ok": False, "msg": "Invalid auth combination."})

    if not emp_id:
        err_msg = "Employee ID is required." if auth_combo == "fingerprint_only" else "No QR code data received."
        return jsonify({"ok": False, "msg": err_msg})

    auth_cfg = get_auth_config()

    if auth_combo in ("qr_fingerprint", "fingerprint_only"):
        if not auth_cfg["fingerprint_enabled"]:
            return jsonify({"ok": False, "msg": "Fingerprint not enabled. Ask your admin to enable it in Settings."}), 403
        # Real, server-verified, one-time, employee-bound proof from
        # /api/employee/webauthn-verify-challenge — not a client-supplied flag.
        if not _wa_fingerprint_recently_verified(emp_id):
            return jsonify({"ok": False, "msg": "Fingerprint verification failed. Please try again."}), 401

    needs_face = (auth_combo == "qr_face")
    if needs_face and not face_b64:
        return jsonify({"ok": False, "msg": "Face photo not captured."})

    frame = None
    if needs_face:
        try:
            img_bytes = base64.b64decode(face_b64)
            pil_img = Image.open(io.BytesIO(img_bytes)).convert("RGB")
            frame = np.array(pil_img)
        except Exception:
            return jsonify({"ok": False, "msg": "Invalid face image data."})

    db = get_db_connection()
    cursor = db.cursor(buffered=True)
    cursor.execute(
        "SELECT face_image, name, email, work_mode, work_lat, work_lon "
        "FROM employees WHERE employee_id=%s", (emp_id,))
    result = cursor.fetchone()

    if not result:
        cursor.close()
        db.close()
        not_found_msg = ("Employee ID not found. Please check your ID and try again."
                         if auth_combo == "fingerprint_only"
                         else "Employee not found. Please check your QR code.")
        return jsonify({"ok": False, "msg": not_found_msg})

    face_path, employee_name, employee_email, emp_work_mode, emp_work_lat, emp_work_lon = result

    # Location check
    if auth_cfg["location_enabled"] and (not user_lat or not user_lon):
        cursor.close()
        db.close()
        return jsonify({"ok": False, "msg": "Location not captured. Please allow location access."})
    if auth_cfg["location_enabled"] and user_lat and user_lon:
        if emp_work_mode == 'wfh':
            if emp_work_lat and emp_work_lon:
                if not is_within_range(float(user_lat), float(user_lon), float(emp_work_lat), float(emp_work_lon)):
                    cursor.close()
                    db.close()
                    return jsonify({"ok": False, "msg": "You are outside your registered home location."})
        else:
            if not is_within_range(float(user_lat), float(user_lon), cfg.OFFICE_LAT, cfg.OFFICE_LON):
                cursor.close()
                db.close()
                return jsonify({"ok": False, "msg": "You are outside the office premises."})

    # Face recognition (only for qr_face combo)
    known_encoding = None
    if needs_face:
        if not _face_recognition_available:
            cursor.close()
            db.close()
            return jsonify({"ok": False, "msg": "Face recognition is currently unavailable on this server. Contact your admin."})
        if not os.path.exists(face_path):
            cursor.close()
            db.close()
            return jsonify({"ok": False, "msg": "Face image missing. Please re-register."})
        known_encoding = _get_known_face_encoding(emp_id, face_path)
        if known_encoding is None:
            cursor.close()
            db.close()
            return jsonify({"ok": False, "msg": "Stored face image is invalid. Please re-register."})

    if needs_face:
        locs = face_recognition.face_locations(frame)
        encs = face_recognition.face_encodings(frame, locs)
        if not encs:
            cursor.close()
            db.close()
            return jsonify({"ok": False, "msg": "No face detected in photo. Look directly at the camera."})
        matched = any(
            True in face_recognition.compare_faces([known_encoding], enc)
            for enc in encs
        )
        if not matched:
            cursor.close()
            db.close()
            return jsonify({"ok": False, "msg": "Face does not match. Please try again."})

    now = datetime.datetime.now()
    today = now.date()
    current_time = now.time()

    cursor.execute(
        "SELECT login_time, logout_time, status, worked_minutes, last_relogin "
        "FROM attendance WHERE employee_id=%s AND date=%s",
        (emp_id, today)
    )
    record = cursor.fetchone()
    login_time = record[0] if record else None
    logout_time = record[1] if record else None
    login_status_stored = record[2] if record else None
    worked_mins_stored = (record[3] or 0) if record else 0
    last_relogin_stored = record[4] if record else None

    # Use employee's assigned shift, or global defaults
    s_start, s_half, s_end, shift_name = get_employee_shift(emp_id, cursor)

    if not login_time:
        grace_time = (datetime.datetime.combine(today, s_start) + datetime.timedelta(minutes=cfg.GRACE_MINUTES)).time()
        if current_time <= grace_time:
            login_status = "Full Day Login"
        elif current_time <= s_half:
            login_status = "Late Login"
        else:
            login_status = "Half Day Login"
        cursor.execute(
            "INSERT INTO attendance (employee_id, date, login_time, status) VALUES (%s,%s,%s,%s)",
            (emp_id, today, current_time, login_status)
        )
        db.commit()
        cursor.close()
        db.close()
        time_str = current_time.strftime("%H:%M:%S")
        return jsonify({"ok": True, "type": "login", "name": employee_name,
                        "status": login_status, "time": time_str, "shift": shift_name,
                        "work_mode": emp_work_mode})

    elif not logout_time:
        # Determine session start (re-login time if present, else first login)
        session_start = last_relogin_stored if last_relogin_stored else login_time
        if not isinstance(session_start, datetime.time):
            session_start = _td_to_time(session_start)
        cur_dt = datetime.datetime.combine(today, current_time)
        start_dt = datetime.datetime.combine(today, session_start)
        session_m = max(0, int((cur_dt - start_dt).total_seconds() / 60))
        total_m = worked_mins_stored + session_m

        if current_time < s_half:
            logout_status = "Half Day Logout"
        elif current_time < s_end:
            logout_status = "Early Logout"
        else:
            logout_status = "Completed"
        # Overtime: minutes beyond shift end
        now_mins = current_time.hour * 60 + current_time.minute
        end_mins = s_end.hour * 60 + s_end.minute
        overtime_m = max(0, now_mins - end_mins)
        att_type = classify_by_worked_minutes(login_status_stored, total_m, s_start, s_end)
        cursor.execute(
            "UPDATE attendance SET logout_time=%s, logout_status=%s, attendance_type=%s, worked_minutes=%s "
            "WHERE employee_id=%s AND date=%s",
            (current_time, logout_status, att_type, total_m, emp_id, today)
        )
        db.commit()
        cursor.close()
        db.close()
        detect_overtime(emp_id, today, current_time)
        time_str = current_time.strftime("%H:%M:%S")
        resp = {"ok": True, "type": "logout", "name": employee_name,
                "status": logout_status, "att_type": att_type,
                "time": time_str, "shift": shift_name, "work_mode": emp_work_mode}
        if overtime_m > 0:
            resp["overtime"] = f"{overtime_m // 60}h {overtime_m % 60}m" if overtime_m >= 60 else f"{overtime_m}m"
        return jsonify(resp)

    else:
        # Re-login after a break — re-open the session
        # worked_minutes was already saved on the previous logout, so just set last_relogin
        cursor.execute(
            "UPDATE attendance SET logout_time=NULL, last_relogin=%s "
            "WHERE employee_id=%s AND date=%s",
            (current_time, emp_id, today)
        )
        db.commit()
        cursor.close()
        db.close()
        time_str = current_time.strftime("%H:%M:%S")
        return jsonify({"ok": True, "type": "relogin", "name": employee_name,
                        "status": "Re-Login", "time": time_str, "shift": shift_name,
                        "work_mode": emp_work_mode})


@attendance_bp.route("/api/attendance/checkin", methods=["POST"])
@api_required
def api_checkin():
    data = request.get_json() or {}
    emp_id = data.get("employee_id")
    lat = data.get("lat")
    lon = data.get("lon")
    if not emp_id:
        return jsonify({"ok": False, "msg": "employee_id required"}), 400
    db = get_db_connection()
    cursor = db.cursor(buffered=True)
    cursor.execute(
        "SELECT name, work_mode, work_lat, work_lon FROM employees WHERE employee_id=%s", (emp_id,))
    result = cursor.fetchone()
    if not result:
        cursor.close()
        db.close()
        return jsonify({"ok": False, "msg": "Employee not found."})
    employee_name, emp_work_mode, emp_work_lat, emp_work_lon = result
    if lat and lon:
        if emp_work_mode == 'wfh':
            if emp_work_lat and emp_work_lon:
                if not is_within_range(float(lat), float(lon), float(emp_work_lat), float(emp_work_lon)):
                    cursor.close()
                    db.close()
                    return jsonify({"ok": False, "msg": "You are outside your registered home location."})
        else:
            if not is_within_range(float(lat), float(lon), cfg.OFFICE_LAT, cfg.OFFICE_LON):
                cursor.close()
                db.close()
                return jsonify({"ok": False, "msg": "You are outside the office premises."})
    now = datetime.datetime.now()
    today = now.date()
    current_time = now.time()
    cursor.execute(
        "SELECT login_time, logout_time, status, worked_minutes, last_relogin "
        "FROM attendance WHERE employee_id=%s AND date=%s",
        (emp_id, today)
    )
    record = cursor.fetchone()
    login_time = record[0] if record else None
    logout_time = record[1] if record else None
    login_status_stored = record[2] if record else None
    worked_mins_stored = (record[3] or 0) if record else 0
    last_relogin_stored = record[4] if record else None
    if not login_time:
        grace_time = (datetime.datetime.combine(today, cfg.SHIFT_START) +
                      datetime.timedelta(minutes=cfg.GRACE_MINUTES)).time()
        if current_time <= grace_time:
            login_status = "Full Day Login"
        elif current_time <= cfg.SHIFT_HALF:
            login_status = "Late Login"
        else:
            login_status = "Half Day Login"
        cursor.execute(
            "INSERT INTO attendance (employee_id, date, login_time, status) VALUES (%s,%s,%s,%s)",
            (emp_id, today, current_time, login_status)
        )
        db.commit()
        cursor.close()
        db.close()
        return jsonify({"ok": True, "type": "login", "name": employee_name,
                        "status": login_status, "time": current_time.strftime("%H:%M:%S")})
    elif not logout_time:
        session_start = last_relogin_stored if last_relogin_stored else login_time
        if not isinstance(session_start, datetime.time):
            session_start = _td_to_time(session_start)
        cur_dt = datetime.datetime.combine(today, current_time)
        start_dt = datetime.datetime.combine(today, session_start)
        session_m = max(0, int((cur_dt - start_dt).total_seconds() / 60))
        total_m = worked_mins_stored + session_m
        if current_time < cfg.SHIFT_HALF:
            logout_status = "Half Day Logout"
        elif current_time < cfg.SHIFT_END:
            logout_status = "Early Logout"
        else:
            logout_status = "Completed"
        att_type = classify_by_worked_minutes(login_status_stored, total_m, cfg.SHIFT_START, cfg.SHIFT_END)
        cursor.execute(
            "UPDATE attendance SET logout_time=%s, logout_status=%s, attendance_type=%s, worked_minutes=%s "
            "WHERE employee_id=%s AND date=%s",
            (current_time, logout_status, att_type, total_m, emp_id, today)
        )
        db.commit()
        cursor.close()
        db.close()
        detect_overtime(emp_id, today, current_time)
        return jsonify({"ok": True, "type": "logout", "name": employee_name,
                        "status": logout_status, "att_type": att_type,
                        "time": current_time.strftime("%H:%M:%S")})
    else:
        cursor.execute(
            "UPDATE attendance SET logout_time=NULL, last_relogin=%s "
            "WHERE employee_id=%s AND date=%s",
            (current_time, emp_id, today)
        )
        db.commit()
        cursor.close()
        db.close()
        return jsonify({"ok": True, "type": "relogin", "name": employee_name,
                        "status": "Re-Login", "time": current_time.strftime("%H:%M:%S")})


@attendance_bp.route("/api/shifts", methods=["GET"])
@api_required
def api_shifts_get():
    db = get_db_connection()
    cursor = db.cursor(buffered=True)
    cursor.execute("SELECT id, name, start_time, half_time, end_time FROM shifts ORDER BY start_time")
    shifts = [
        {"id": r[0], "name": r[1],
         "start": _td_to_time(r[2]).strftime("%H:%M") if r[2] else "--",
         "half": _td_to_time(r[3]).strftime("%H:%M") if r[3] else "--",
         "end": _td_to_time(r[4]).strftime("%H:%M") if r[4] else "--"}
        for r in cursor.fetchall()
    ]
    cursor.execute(
        "SELECT e.employee_id, e.name, e.role, s.id, s.name "
        "FROM employees e LEFT JOIN shifts s ON e.shift_id = s.id ORDER BY e.name"
    )
    employees = [
        {"emp_id": r[0], "name": r[1], "role": r[2] or "",
         "shift_id": r[3], "shift_name": r[4] or "Default"}
        for r in cursor.fetchall()
    ]
    cursor.close()
    db.close()
    return jsonify({"ok": True, "shifts": shifts, "employees": employees})


@attendance_bp.route("/api/shifts", methods=["POST"])
@api_required
def api_shifts_create():
    data = request.get_json(silent=True) or {}
    name = data.get("name", "").strip()
    start = data.get("start_time", "").strip()
    half = data.get("half_time", "").strip()
    end = data.get("end_time", "").strip()
    if not all([name, start, half, end]):
        return jsonify({"ok": False, "msg": "All fields required"}), 400
    db = get_db_connection()
    cursor = db.cursor(buffered=True)
    try:
        cursor.execute(
            "INSERT INTO shifts (name, start_time, half_time, end_time) VALUES (%s,%s,%s,%s) RETURNING id",
            (name, start, half, end)
        )
        sid = cursor.fetchone()[0]
        db.commit()
    except Exception:
        app_log.error("Failed to create shift", exc_info=True)
        cursor.close()
        db.close()
        return jsonify({"ok": False, "msg": "Failed to create shift. Check for duplicate names."}), 400
    cursor.close()
    db.close()
    return jsonify({"ok": True, "id": sid})


@attendance_bp.route("/api/shifts/<int:sid>", methods=["DELETE"])
@api_required
def api_shifts_delete(sid):
    db = get_db_connection()
    cursor = db.cursor(buffered=True)
    cursor.execute("UPDATE employees SET shift_id=NULL WHERE shift_id=%s", (sid,))
    cursor.execute("DELETE FROM shifts WHERE id=%s", (sid,))
    db.commit()
    cursor.close()
    db.close()
    return jsonify({"ok": True})


@attendance_bp.route("/api/shifts/assign", methods=["POST"])
@api_required
def api_shifts_assign():
    data = request.get_json(silent=True) or {}
    emp_id = data.get("emp_id", "").strip()
    shift_id = data.get("shift_id")
    if not emp_id:
        return jsonify({"ok": False, "msg": "emp_id required"}), 400
    db = get_db_connection()
    cursor = db.cursor(buffered=True)
    cursor.execute(
        "UPDATE employees SET shift_id=%s WHERE employee_id=%s",
        (shift_id if shift_id else None, emp_id)
    )
    db.commit()
    cursor.close()
    db.close()
    return jsonify({"ok": True})
